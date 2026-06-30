# -*- coding: utf-8 -*-
"""
Genera un workbook de ANALISIS no destructivo en DATA SENSIBLE\\ANALISIS_BP_<sello>.xlsx con:

  - BASE_GENERAL_NUEVA : la BASE GENERAL del Q2 actual + columnas recalculadas:
      * BP_NUEVO        (desde maestro, area->BP; sino dep->BP; sino REVISAR)
      * BP_REGLA        (regla aplicada: area/area~subcadena/dep/sin_regla)
      * BP_CAMBIA       ('SI'/'NO'/'-' si REVISAR)
      * META_DIAS       (= Objetivo cargado por RRHH; lo que se quiere lograr)
      * TIENE_META      ('Si'/'No') = Objetivo (RRHH) > 0   <-- ojo: NO es lo mismo que obligatorio
      * DIAS_LEGALES    (= Vencidas + Pendiente; lo que por LEY debe gozar)
      * ES_OBLIGATORIO  ('Si'/'No') = DIAS_LEGALES > 0      <-- Cesar SI puede mandarlo
      * SOLO_TRUNCOS    ('Si'/'No') = truncos>0 y no hay meta ni dias legales
      * CATEGORIA       (OBLIGATORIA / NO_OBLIGATORIA / OBLIG_SIN_META / SOLO_TRUNCOS / SIN_NADA)
      * CON_PROBLEMAS   ('Si'/'No') = tiene META y Registradas < META
      * DIAS_FALTANTES  = max(META - Registradas, 0)
      * COM_FLAG        (NO_OBLIGADO_LEY / PRIORIZAR / OTRO / -) leido del comentario
      * FUENTE_AREA     (MAESTRO / OBJETIVO si no estaba en maestro)
  - CAMBIOS_BP         : solo filas con BP_CAMBIA=SI (HRBP actual vs BP_NUEVO + dep/area)
  - REVISAR            : filas con BP_NUEVO=REVISAR (CSIR y casos sueltos)
  - RESUMEN_BP         : pivote BP_NUEVO x {N, Obligatoria, No Obligatoria, Solo Truncos,
                         Con Problemas, Meta(dias), Registradas, Avance%}
  - PARTICION          : matriz Meta x Obligatorio para entender el "target" visualmente
  - NO_EN_MAESTRO      : matriculas del Q2 que no estan en el maestro
  - MAPA_AREA_BP       : tabla del mapeo aplicado (auditable)

NO toca ningun archivo de produccion. Es un .xlsx aparte para revision.
"""
import sys, datetime as dt, json
from pathlib import Path
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, r"C:\Users\jlopezp\OneDrive - Universidad San Ignacio de Loyola\ACTIVIDADES\CESAR\SISTEMA DE VACACIONES")
from _bp_map import bp_de, norm, AREA_BP, AREA_CONTAINS, DEP_BP

DATA = Path(r"C:\Users\jlopezp\OneDrive - Universidad San Ignacio de Loyola\ACTIVIDADES\CESAR\SISTEMA DE VACACIONES\DATA SENSIBLE")
MAESTRO = DATA / "PersonalMaestroReporte_05_04_2026 16_37_07.xlsx"
OBJ2    = DATA / "Reporte Vacaciones Objetivo_Segundo Trimestre 2026.xlsx"

SELLO = dt.datetime.now().strftime("%Y%m%d_%H%M")
SALIDA = DATA / f"ANALISIS_BP_{SELLO}.xlsx"

# ---------- 1) Cargar maestro: matricula -> dep/area/sec/situacion ----------
print("Cargando maestro...")
wb = openpyxl.load_workbook(MAESTRO, read_only=True, data_only=True)
ws = wb["Hoja1"]
MAESTRO_MAP = {}
for row in ws.iter_rows(min_row=12, values_only=True):
    mat = row[0]
    if mat is None or not str(mat).strip().isdigit(): continue
    mat = str(mat).strip()
    MAESTRO_MAP[mat] = {
        "dep": str(row[198] or "").strip(),
        "area": str(row[199] or "").strip(),
        "sec": str(row[200] or "").strip(),
        "sit": str(row[26] or "").strip(),
    }
wb.close()
print(f"Maestro: {len(MAESTRO_MAP)} matriculas activas")

# ---------- 2) Leer BASE GENERAL del Q2 ----------
print("Leyendo BASE GENERAL del Q2...")
wb = openpyxl.load_workbook(OBJ2, read_only=True, data_only=True)
g = wb["BASE GENERAL"]
HDR = next(g.iter_rows(min_row=2, max_row=2, values_only=True))
COLS = [str(h).strip() if h is not None else "" for h in HDR]
filas = []
for row in g.iter_rows(min_row=3, values_only=True):
    mat = row[0]
    if mat is None or not str(mat).strip().isdigit(): continue
    filas.append(list(row))
wb.close()
print(f"BASE GENERAL: {len(filas)} filas, {len(COLS)} columnas")
# Indices clave segun nuestro inspect anterior
I_MAT, I_HRBP, I_DEP_O, I_AREA_O = 0, 6, 7, 8
I_VENC, I_PEND, I_TRUNC, I_SUMA, I_OBJ = 12, 13, 14, 15, 16
I_META, I_OBL, I_REG, I_PCT, I_DIASR, I_COM = 17, 18, 19, 20, 21, 22

# ---------- 3) Recalcular y armar filas enriquecidas ----------
nuevas = []
no_en_maestro = []
bp_count = Counter()
src_count = Counter()
cat_count = Counter()
problemas_count = 0
cambios = []
revisar = []
# acumuladores resumen
res = defaultdict(lambda: {"n":0,"meta":0,"reg":0,
                            "meta_forz":0,"reg_forz":0,  # solo OBLIGATORIA
                            "obligatoria":0,"no_obligatoria":0,
                            "solo_trunc":0,"sin_nada":0,"oblig_sin_meta":0,
                            "problemas":0,"no_obligado_ley":0})
# matriz partition: (TIENE_META, ES_OBLIGATORIO) -> count
particion = Counter()
# para REVISAR: agrupar por dep||area para sugerencias
revisar_grupos = Counter()

def f2num(v):
    return float(v) if isinstance(v,(int,float)) else 0.0

def clasif_comentario(com):
    """Lee el comentario tal cual viene del archivo y devuelve una etiqueta corta.
    Ojo: norm() devuelve MAYUSCULAS sin acentos, comparamos en MAYUSCULAS."""
    if not com or not str(com).strip(): return "-"
    n = norm(com)  # MAYUSCULAS sin acentos
    if "NO ESTA OBLIGADO" in n or "NO OBLIGADO" in n or "NO ESTA OBLIGADA" in n:
        return "NO_OBLIGADO_LEY"
    if "PRIORIZAR" in n or "VENCID" in n or "VENCIENDO" in n or "VENCER" in n:
        return "PRIORIZAR"
    if "PROGRAMAD" in n or "PROGRAMA" in n:
        return "PROGRAMADO"
    return "OTRO"

for row in filas:
    mat = str(row[I_MAT]).strip()
    cur_hrbp = str(row[I_HRBP] or "").strip()
    obj_dep, obj_area = str(row[I_DEP_O] or "").strip(), str(row[I_AREA_O] or "").strip()
    if mat in MAESTRO_MAP:
        m = MAESTRO_MAP[mat]
        dep_use, area_use, fuente_area = m["dep"], m["area"], "MAESTRO"
    else:
        dep_use, area_use, fuente_area = obj_dep, obj_area, "OBJETIVO"
        no_en_maestro.append([mat, str(row[1] or ""), obj_dep, obj_area, cur_hrbp])
    bp_new, regla = bp_de(dep_use, area_use)
    bp_count[bp_new] += 1
    src_count[fuente_area] += 1

    venc = f2num(row[I_VENC]); pend = f2num(row[I_PEND]); trunc = f2num(row[I_TRUNC])
    reg  = f2num(row[I_REG]);  obj  = f2num(row[I_OBJ])

    # --- META vs OBLIGATORIO LEGAL: dos cosas distintas ---
    meta_d = int(obj)              # META = Objetivo cargado por RRHH (target oficial)
    dias_leg = int(venc + pend)    # DIAS LEGALES = vencidas + pendiente
    tiene_meta = "Si" if meta_d > 0 else "No"
    es_oblig   = "Si" if dias_leg > 0 else "No"

    if   tiene_meta=="Si" and es_oblig=="Si": categoria = "OBLIGATORIA"        # se puede enviar
    elif tiene_meta=="Si" and es_oblig=="No": categoria = "NO_OBLIGATORIA"     # no se puede forzar
    elif tiene_meta=="No" and es_oblig=="Si": categoria = "OBLIG_SIN_META"     # raro: no hay target pero ley aplica
    elif trunc > 0:                            categoria = "SOLO_TRUNCOS"      # solo se paga, no se goza
    else:                                      categoria = "SIN_NADA"
    cat_count[categoria] += 1
    particion[(tiene_meta, es_oblig)] += 1

    dias_falt = max(meta_d - reg, 0)
    # "Con problemas" = tiene meta (target RRHH) y le faltan dias
    con_prob = "Si" if (meta_d > 0 and reg < meta_d) else "No"
    if con_prob == "Si": problemas_count += 1

    com_text = row[I_COM] if len(row)>I_COM else None
    com_flag = clasif_comentario(com_text)

    cambia = "-"
    if bp_new != "REVISAR":
        cambia = "NO" if norm(bp_new) == norm(cur_hrbp) else "SI"
    if cambia == "SI":
        cambios.append([mat, str(row[1] or ""), cur_hrbp, bp_new, dep_use, area_use, fuente_area, regla])
    if bp_new == "REVISAR":
        revisar.append([mat, str(row[1] or ""), cur_hrbp, dep_use, area_use, fuente_area, regla])
        revisar_grupos[(dep_use.strip().upper(), area_use.strip().upper())] += 1

    # acumulado por BP
    r = res[bp_new]
    r["n"] += 1
    r["meta"] += meta_d
    r["reg"]  += reg
    if categoria == "OBLIGATORIA":     # solo cuentan para "forzable"
        r["meta_forz"] += meta_d
        r["reg_forz"]  += reg
    r["obligatoria"]    += (1 if categoria=="OBLIGATORIA" else 0)
    r["no_obligatoria"] += (1 if categoria=="NO_OBLIGATORIA" else 0)
    r["solo_trunc"]     += (1 if categoria=="SOLO_TRUNCOS" else 0)
    r["sin_nada"]       += (1 if categoria=="SIN_NADA" else 0)
    r["oblig_sin_meta"] += (1 if categoria=="OBLIG_SIN_META" else 0)
    r["problemas"]      += (1 if con_prob=="Si" else 0)
    r["no_obligado_ley"]+= (1 if com_flag=="NO_OBLIGADO_LEY" else 0)

    nuevas.append(row + [bp_new, regla, cambia, meta_d, tiene_meta, dias_leg, es_oblig,
                          categoria, con_prob, dias_falt, com_flag, fuente_area])

print(f"\nBP_NUEVO: {dict(bp_count)}")
print(f"Fuente area: {dict(src_count)}")
print(f"Categoria: {dict(cat_count)}")
print(f"Con problemas: {problemas_count}")
print(f"Cambios de BP: {len(cambios)}  |  REVISAR: {len(revisar)}  |  No en maestro: {len(no_en_maestro)}")

# ---------- 4) Escribir workbook ----------
print(f"\nEscribiendo {SALIDA.name}...")
wbo = openpyxl.Workbook()
wbo.remove(wbo.active)

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def cab(ws, headers, row=1, freeze=True):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    if freeze: ws.freeze_panes = ws.cell(row=row+1, column=2)

def autosize(ws, max_w=42):
    for col in ws.columns:
        letter = col[0].column_letter
        m = 8
        for c in col:
            v = c.value
            if v is None: continue
            l = len(str(v))
            if l > m: m = l
        ws.column_dimensions[letter].width = min(m + 2, max_w)

# --- BASE_GENERAL_NUEVA ---
hojas_orden = ["BASE_GENERAL_NUEVA","RESUMEN_BP","PARTICION","REVISAR_SUGERENCIAS","CAMBIOS_BP","REVISAR","NO_EN_MAESTRO","MAPA_AREA_BP","NOTAS"]
# Sugerencias de BP para grupos no mapeados (editable, las pongo en su propia hoja).
# Estas son propuestas para que confirmes con Cesar.
SUGERENCIAS_REVISAR = {
    # (DEPARTAMENTO_NORM, AREA_NORM) -> (BP sugerido, justificacion)
    ("DIRECCION GENERAL DEL CSIR","DIRECCION GENERAL ACADEMICA"): ("Fatima Salazar","perfil academico (mas cercano a su cartera)"),
    ("DIRECCION GENERAL DEL CSIR","ADMINISTRATIVO"):              ("Lesley Reyes","admin/operaciones (cartera Lesley)"),
    ("DIRECCION GENERAL DEL CSIR",""):                            ("Fatima Salazar","CSIR sin area definida"),
    ("VICERRECTORADO ACADEMICO",""):                              ("Fatima Salazar","VRA sin area -> default academico"),
    ("PRESIDENCIA EJECUTIVA",""):                                 ("Melissa Higa","alineado a Asuntos Corporativos / Gerencia General"),
    ("RECTORADO",""):                                             ("Carlos Jara","Rectorado sin area -> Secretaria General default"),
    ("GERENCIA DE PROYECTOS - SEDE PERU",""):                     ("Lesley Reyes","operacion -> Lesley"),
}
sh = wbo.create_sheet("BASE_GENERAL_NUEVA")
extra = ["BP_NUEVO","BP_REGLA","BP_CAMBIA",
          "META_DIAS","TIENE_META",
          "DIAS_LEGALES","ES_OBLIGATORIO",
          "CATEGORIA","CON_PROBLEMAS","DIAS_FALTANTES",
          "COM_FLAG","FUENTE_AREA"]
cab(sh, COLS + extra)
# pintar la categoria
FILLS = {
    "OBLIGATORIA":      PatternFill("solid", fgColor="C6EFCE"),  # verde
    "NO_OBLIGATORIA":   PatternFill("solid", fgColor="FFEB9C"),  # amarillo
    "OBLIG_SIN_META":   PatternFill("solid", fgColor="FFC7CE"),  # rojo
    "SOLO_TRUNCOS":     PatternFill("solid", fgColor="DDEBF7"),  # azul claro
    "SIN_NADA":         PatternFill("solid", fgColor="F2F2F2"),  # gris
}
col_categoria = len(COLS) + extra.index("CATEGORIA") + 1
for r, f in enumerate(nuevas, start=2):
    for j, v in enumerate(f, start=1):
        sh.cell(row=r, column=j, value=v)
    cat_val = f[-5]   # CATEGORIA
    if cat_val in FILLS:
        sh.cell(row=r, column=col_categoria).fill = FILLS[cat_val]
autosize(sh, 36)

# --- CAMBIOS_BP ---
sh = wbo.create_sheet("CAMBIOS_BP")
cab(sh, ["Matricula","Nombre","HRBP_ACTUAL","BP_NUEVO","Departamento_Maestro","Area_Maestro","Fuente_Area","Regla"])
for r,f in enumerate(cambios, start=2):
    for j,v in enumerate(f, start=1):
        sh.cell(row=r, column=j, value=v)
autosize(sh, 40)

# --- REVISAR ---
sh = wbo.create_sheet("REVISAR")
cab(sh, ["Matricula","Nombre","HRBP_ACTUAL","Departamento","Area","Fuente_Area","Regla"])
for r,f in enumerate(revisar, start=2):
    for j,v in enumerate(f, start=1):
        sh.cell(row=r, column=j, value=v)
autosize(sh, 40)

# --- REVISAR_SUGERENCIAS ---
sh = wbo.create_sheet("REVISAR_SUGERENCIAS")
cab(sh, ["Departamento","Area","# Personas","BP_SUGERIDO","Justificacion","¿Aceptar? (Si/No/otro BP)"])
sh.cell(2,6,"<- editar esta columna").font = Font(italic=True, color="888888")
r = 2
for (dep, area), c in sorted(revisar_grupos.items(), key=lambda x:-x[1]):
    bp_sug, just = SUGERENCIAS_REVISAR.get((dep, area), ("REVISAR","sin sugerencia (decide Cesar)"))
    sh.cell(r,1,dep); sh.cell(r,2,area); sh.cell(r,3,c)
    sh.cell(r,4,bp_sug); sh.cell(r,5,just)
    if bp_sug == "REVISAR":
        sh.cell(r,4).fill = PatternFill("solid", fgColor="FFC7CE")
    else:
        sh.cell(r,4).fill = PatternFill("solid", fgColor="FFEB9C")
    r += 1
sh.cell(r+1,1,"Instrucciones:").font = Font(bold=True)
sh.cell(r+2,1,"  En col F escribe 'Si' para aceptar la sugerencia,")
sh.cell(r+3,1,"  o el nombre del BP que prefieres (Carlos Jara / Fatima Salazar / Lesley Reyes / Melissa Higa).")
sh.cell(r+4,1,"  Yo aplico tu decision en el pipeline y servidor.")
autosize(sh, 45)

# --- NO_EN_MAESTRO ---
sh = wbo.create_sheet("NO_EN_MAESTRO")
cab(sh, ["Matricula","Nombre","Departamento_Q2","Area_Q2","HRBP_actual"])
for r,f in enumerate(no_en_maestro, start=2):
    for j,v in enumerate(f, start=1):
        sh.cell(row=r, column=j, value=v)
autosize(sh, 40)

# --- RESUMEN_BP ---
sh = wbo.create_sheet("RESUMEN_BP")
cab(sh, ["Business Partner","N",
          "OBLIGATORIA","NO_OBLIGATORIA","OBLIG_SIN_META","SOLO_TRUNCOS","SIN_NADA",
          "Con Problemas","No Obligado (ley)",
          "Meta TOTAL","Reg TOTAL","Avance TOTAL %",
          "Meta FORZABLE","Reg FORZABLE","Avance FORZABLE %"])
orden_bp = ["Carlos Jara","Fatima Salazar","Lesley Reyes","Melissa Higa","REVISAR"]
KEYS = ["n","obligatoria","no_obligatoria","oblig_sin_meta","solo_trunc","sin_nada",
        "problemas","no_obligado_ley","meta","reg","meta_forz","reg_forz"]
total = {k:0 for k in KEYS}
r = 2
for bp in orden_bp:
    d = res.get(bp, {k:0 for k in KEYS})
    av_t = (d["reg"]/d["meta"]) if d["meta"] else 0
    av_f = (d["reg_forz"]/d["meta_forz"]) if d["meta_forz"] else 0
    sh.cell(r,1,bp); sh.cell(r,2,d["n"])
    sh.cell(r,3,d["obligatoria"]); sh.cell(r,4,d["no_obligatoria"])
    sh.cell(r,5,d["oblig_sin_meta"]); sh.cell(r,6,d["solo_trunc"]); sh.cell(r,7,d["sin_nada"])
    sh.cell(r,8,d["problemas"]); sh.cell(r,9,d["no_obligado_ley"])
    sh.cell(r,10,d["meta"]); sh.cell(r,11,d["reg"])
    c=sh.cell(r,12,av_t); c.number_format="0.0%"
    sh.cell(r,13,d["meta_forz"]); sh.cell(r,14,d["reg_forz"])
    c=sh.cell(r,15,av_f); c.number_format="0.0%"
    for k in total: total[k]+=d[k]
    r+=1
av_T = (total["reg"]/total["meta"]) if total["meta"] else 0
av_F = (total["reg_forz"]/total["meta_forz"]) if total["meta_forz"] else 0
sh.cell(r,1,"TOTAL"); sh.cell(r,2,total["n"])
sh.cell(r,3,total["obligatoria"]); sh.cell(r,4,total["no_obligatoria"])
sh.cell(r,5,total["oblig_sin_meta"]); sh.cell(r,6,total["solo_trunc"]); sh.cell(r,7,total["sin_nada"])
sh.cell(r,8,total["problemas"]); sh.cell(r,9,total["no_obligado_ley"])
sh.cell(r,10,total["meta"]); sh.cell(r,11,total["reg"])
c=sh.cell(r,12,av_T); c.number_format="0.0%"
sh.cell(r,13,total["meta_forz"]); sh.cell(r,14,total["reg_forz"])
c=sh.cell(r,15,av_F); c.number_format="0.0%"
for j in range(1,16):
    sh.cell(r,j).font = Font(bold=True)
# colorear las columnas de avance
for rr in range(2, r+1):
    sh.cell(rr,12).fill = PatternFill("solid", fgColor="E2EFDA")  # avance total - verde claro
    sh.cell(rr,15).fill = PatternFill("solid", fgColor="C6EFCE")  # avance forzable - verde mas fuerte
autosize(sh, 20)

# --- PARTICION (matriz Meta x Obligatorio) ---
sh = wbo.create_sheet("PARTICION")
cab(sh, ["", "ES_OBLIGATORIO = Si\n(vencidas/pend > 0)", "ES_OBLIGATORIO = No\n(sin vencidas/pend)", "Total"])
sh.cell(2,1,"TIENE_META = Si\n(Objetivo > 0)").alignment = Alignment(wrap_text=True, vertical="center")
sh.cell(3,1,"TIENE_META = No\n(Objetivo = 0)").alignment = Alignment(wrap_text=True, vertical="center")
sh.cell(4,1,"Total").font = Font(bold=True)
SS = particion.get(("Si","Si"),0); SN = particion.get(("Si","No"),0)
NS = particion.get(("No","Si"),0); NN = particion.get(("No","No"),0)
sh.cell(2,2,SS).fill = FILLS["OBLIGATORIA"]; sh.cell(2,3,SN).fill = FILLS["NO_OBLIGATORIA"]
sh.cell(2,4,SS+SN)
sh.cell(3,2,NS).fill = FILLS["OBLIG_SIN_META"]; sh.cell(3,3,NN).fill = FILLS["SIN_NADA"]
sh.cell(3,4,NS+NN)
sh.cell(4,2,SS+NS); sh.cell(4,3,SN+NN); sh.cell(4,4,SS+SN+NS+NN)
sh.cell(6,1,"Leyenda:").font = Font(bold=True)
sh.cell(7,1,"OBLIGATORIA (verde): tiene meta y ley obliga -> Cesar SI puede enviarlo")
sh.cell(8,1,"NO_OBLIGATORIA (amarillo): tiene meta pero ley NO obliga -> NO se puede forzar")
sh.cell(9,1,"OBLIG_SIN_META (rojo): caso raro, ley obliga pero no hay target -> revisar")
sh.cell(10,1,"SIN_NADA (gris): sin meta, sin obligacion legal")
sh.row_dimensions[2].height = 32
sh.row_dimensions[3].height = 32
sh.column_dimensions["A"].width = 30
for c in ("B","C","D"): sh.column_dimensions[c].width = 26

# --- MAPA_AREA_BP (auditoria) ---
sh = wbo.create_sheet("MAPA_AREA_BP")
cab(sh, ["NIVEL","CLAVE_NORMALIZADA","BP"])
r=2
for k,v in AREA_BP.items():
    sh.cell(r,1,"AREA exacta"); sh.cell(r,2,k); sh.cell(r,3,v); r+=1
for sub,v in AREA_CONTAINS:
    sh.cell(r,1,"AREA contiene"); sh.cell(r,2,sub); sh.cell(r,3,v); r+=1
for k,v in DEP_BP.items():
    sh.cell(r,1,"DEP exacto"); sh.cell(r,2,k); sh.cell(r,3,v); r+=1
autosize(sh, 60)

# --- NOTAS ---
sh = wbo.create_sheet("NOTAS")
notas = [
    "ANALISIS BP + PARTICION (no destructivo)",
    "",
    f"Generado: {dt.datetime.now():%Y-%m-%d %H:%M}",
    f"Maestro: {MAESTRO.name}",
    f"Objetivo Q2: {OBJ2.name}",
    "",
    "Reglas de asignacion de BP:",
    "  1. BP se busca por AREA exacta (maestro)",
    "  2. Si no, por subcadena de AREA (nombres truncados)",
    "  3. Si no, por DEPARTAMENTO",
    "  4. Si no, BP = REVISAR",
    "",
    "META vs OBLIGATORIO LEGAL (son DOS cosas distintas):",
    "  META_DIAS      = Objetivo (col Q) cargado por RRHH. Es el TARGET.",
    "  TIENE_META     = META_DIAS > 0  (RRHH le asigno objetivo)",
    "  DIAS_LEGALES   = Vencidas + Pendiente. Lo que la LEY obliga gozar.",
    "  ES_OBLIGATORIO = DIAS_LEGALES > 0  (Cesar SI puede enviarlo de vacaciones)",
    "",
    "Particion (CATEGORIA, con color en BASE_GENERAL_NUEVA):",
    "  OBLIGATORIA      VERDE   tiene meta y la ley obliga    -> se puede enviar",
    "  NO_OBLIGATORIA   AMARI.  tiene meta pero ley NO obliga -> NO se puede forzar",
    "  OBLIG_SIN_META   ROJO    raro: sin target pero ley aplica",
    "  SOLO_TRUNCOS     AZUL    solo truncos (se pagan, no se gozan)",
    "  SIN_NADA         GRIS    nada que cumplir",
    "",
    "  CON_PROBLEMAS = TIENE_META y Registradas < META_DIAS (le faltan dias).",
    "",
    "COM_FLAG (etiqueta del comentario en col W):",
    "  NO_OBLIGADO_LEY  - dice 'No esta obligado a salir por ley' (305 personas)",
    "  PRIORIZAR        - dice 'priorizar' / 'venciendo' / similar",
    "  OTRO             - hay comentario pero no entra en los anteriores",
    "  -                - no hay comentario",
    "",
    "Pendiente de tu decision:",
    "  - DIRECCION GENERAL DEL CSIR (~299 personas): definir BP.",
    "  - Casos sueltos en REVISAR (16 personas).",
    "  - Si quieres etiquetar 'no quiere salir' u otra senal,",
    "    Cesar debe registrarla en col W (Comentarios) y yo la leo aqui.",
]
for i, t in enumerate(notas, start=1):
    sh.cell(i, 1, t)
sh.column_dimensions["A"].width = 110

# reordenar
for i, name in enumerate(hojas_orden):
    if name in wbo.sheetnames:
        wbo.move_sheet(wbo[name], offset=i - wbo.sheetnames.index(name))

wbo.save(SALIDA)
print(f"\nOK: {SALIDA}")
