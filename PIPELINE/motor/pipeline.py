# -*- coding: utf-8 -*-
"""
pipeline.py  -  Motor del pipeline de Vacaciones USIL.

Hace, "como un humano pero automatico", todo el proceso que describio Cesar:
  1. Toma el VACRptMotivo_*.xlsx mas reciente de Adryan (carpeta de descargas).
  2. Lo limpia (quita metadatos y filas 'SIN GRUPO') y convierte tipos
     (Dias/Mes Pago -> numero; Fecha Ingreso/Inicio/Termino -> fecha; Matricula texto).
  3. Crea una COPIA FECHADA del objetivo (nunca toca el original) y la abre en Excel.
  4. Vuelca los datos limpios a la hoja 'base' (preservando OBSERVACION).
  5. Reajusta el rango del pivote y RefreshAll (DB_Vac).
  6. Regenera la tabla espejo L:S a partir del pivote refrescado.
  7. Rellena BASE GENERAL: Meta?, Obligatorio (regla verificada), Registradas, %, etc.
  8. RefreshAll de nuevo (R_Cumplimiento), recalcula, guarda y reporta reconciliacion.

Uso:
    python pipeline.py                 # corrida completa (Excel segun config)
    python pipeline.py --dry-run       # solo prepara datos y valida (NO abre Excel)
    python pipeline.py --oculto        # corre con Excel invisible (pruebas)
    python pipeline.py --no-cerrar     # deja Excel abierto al terminar (inspeccion)
    python pipeline.py --crudo RUTA    # forzar un crudo especifico
"""
import argparse
import datetime as dt
import json
import re
import shutil
import sys
import traceback
from collections import defaultdict
from pathlib import Path

import vac_lib as V

RX_MATRICULA = re.compile(r"^0*\d+$")


# ----------------------------------------------------------------- utilidades Excel
def col_letra(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def col_num(letra):
    n = 0
    for c in str(letra).upper():
        n = n * 26 + (ord(c) - 64)
    return n


# ----------------------------------------------------------------- objetivo / copia
def elegir_objetivo(cfg, log):
    sal = cfg["salida"]
    carpeta_salida = V.ruta_abs(sal["carpeta_salida"])
    fuente = V.ruta_abs(sal["objetivo_plantilla"])
    if sal.get("usar_salida_mas_reciente_como_fuente", True) and carpeta_salida.exists():
        previas = sorted(carpeta_salida.glob(sal["prefijo_salida"] + "*.xlsx"),
                         key=lambda p: p.stat().st_mtime, reverse=True)
        if previas:
            fuente = previas[0]
    if not fuente.exists():
        raise FileNotFoundError(f"No existe el objetivo fuente: {fuente}")
    log.info("Objetivo fuente : %s", fuente.name)
    return fuente


def crear_copia_fechada(cfg, fuente, log):
    sal = cfg["salida"]
    carpeta_salida = V.ruta_abs(sal["carpeta_salida"])
    carpeta_salida.mkdir(parents=True, exist_ok=True)
    sello = dt.datetime.now().strftime("%Y%m%d_%H%M")
    destino = carpeta_salida / f"{sal['prefijo_salida']}_{sello}.xlsx"
    shutil.copy2(fuente, destino)
    log.info("Copia fechada   : %s", destino.name)
    return destino


def publicar_salida(cfg, destino, log):
    """Copia la salida al/los archivo(s) que lee el dashboard (front), respaldando
    primero el archivo vivo. Asi el servidor refresca los KPIs (cache por mtime)."""
    intg = cfg.get("integracion_front", {})
    objetivos = intg.get("publicar_en", []) or []
    carpeta_resp = V.ruta_abs(cfg["salida"]["carpeta_respaldos"])
    carpeta_resp.mkdir(parents=True, exist_ok=True)
    publicados = []
    for tgt in objetivos:
        tgt = Path(tgt)
        try:
            tgt.parent.mkdir(parents=True, exist_ok=True)
            if tgt.exists():
                resp = carpeta_resp / f"{tgt.stem}__previo_{dt.datetime.now():%Y%m%d_%H%M%S}{tgt.suffix}"
                shutil.copy2(tgt, resp)
            shutil.copy2(destino, tgt)
            publicados.append(str(tgt))
            log.info("Publicado en   : %s", tgt)
        except PermissionError:
            log.error("No se pudo publicar (¿esta abierto en Excel?): %s", tgt)
        except Exception as e:
            log.error("Error publicando en %s: %s", tgt, e)
    return publicados


def escribir_estado(destino, resumen, log):
    """Escribe un JSON local con los KPIs (rapido, sin tocar OneDrive) que el dashboard lee."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(destino, data_only=True)
        kp = {}
        if "BASE GENERAL" in wb.sheetnames:
            g = wb["BASE GENERAL"]
            kp["meta_total"] = g["Q1"].value
            kp["registrado_total"] = g["T1"].value
            kp["avance"] = g["W1"].value
        if "R_Cumplimiento" in wb.sheetnames:
            kp["avance_cumplimiento"] = wb["R_Cumplimiento"]["E9"].value
        wb.close()
        kp["timestamp"] = dt.datetime.now().isoformat(timespec="seconds")
        kp["archivo"] = Path(destino).name
        kp["registros"] = resumen.get("registros")
        kp["faltan_en_bg"] = [list(x) for x in resumen.get("faltan_en_bg", [])]
        ruta = V.RAIZ / "estado_pipeline.json"
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(kp, f, ensure_ascii=False, indent=2, default=str)
        log.info("Estado KPIs    : %s", ruta.name)
        return kp
    except Exception as e:
        log.warning("No se pudo escribir estado_pipeline.json: %s", e)
        return None


# ----------------------------------------------------------------- datos limpios
def preparar_datos(cfg, ruta_crudo, log):
    esquema = cfg["esquema_base"]
    res = V.leer_crudo(ruta_crudo, esquema, cfg["limpieza_crudo"]["regex_matricula_valida"])
    log.info("Crudo: cabecera fila %d | %d filas validas | %d descartadas (%s)",
             res["fila_cabecera"], len(res["filas"]), res["descartadas"],
             ", ".join(f"f{r}:{b}" for r, _, b in res["muestras_descarte"]) or "ninguna")
    if res["faltantes"]:
        log.warning("Columnas del esquema NO halladas en el crudo: %s", res["faltantes"])
    inc = V.convertir_tipos(res["filas"], esquema)
    if inc:
        log.warning("Incidencias de conversion (%d). Ej: %s", len(inc), inc[:5])
    else:
        log.info("Conversion de tipos: 0 incidencias.")
    return res["filas"]


def construir_matriz(filas, esquema):
    nombres = [c["nombre"] for c in esquema]
    matriz = []
    for reg in filas:
        matriz.append([("" if reg.get(n) is None else reg.get(n)) for n in nombres])
    return matriz, nombres


# ----------------------------------------------------------------- motor COM
def ejecutar_com(cfg, destino, filas, oculto, no_cerrar, log):
    import xlwings as xw

    esquema = cfg["esquema_base"]
    H = cfg["hojas"]
    piv_cfg = cfg["pivote"]
    bg = cfg["base_general"]
    n_filas = len(filas)
    matriz, nombres = construir_matriz(filas, esquema)

    ult_base = 1 + n_filas
    tope = max(piv_cfg["tope_filas_minimo"], ult_base + piv_cfg["buffer_filas"])
    log.info("Registros=%d | ultima fila base=%d | tope rangos=%d", n_filas, ult_base, tope)

    visible = (not oculto) and cfg["excel"]["visible"]
    cerrar = (not no_cerrar) and cfg["excel"].get("cerrar_al_terminar", True)
    app = xw.App(visible=visible, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = None
    try:
        wb = app.books.open(str(destino), update_links=False)
        try:
            app.api.AskToUpdateLinks = False
        except Exception:
            pass
        # Excel puede devolver OLE 0x800ac472 ("busy") los primeros segundos tras abrir
        # un libro con macros/links. Reintentamos con backoff antes de fallar.
        for _intento in range(6):
            try:
                app.calculation = "manual"
                break
            except Exception as _e:
                codigo = getattr(_e, 'args', [None])[0]
                if codigo == -2146777998 and _intento < 5:
                    espera = 2 ** _intento          # 1,2,4,8,16,32 seg
                    log.warning("Excel ocupado (OLE 0x800ac472), reintentando en %ds... (%d/6)",
                                espera, _intento + 1)
                    import time as _t; _t.sleep(espera)
                else:
                    raise

        sh_base = wb.sheets[H["base"]]
        sh_piv = wb.sheets[H["pivote"]]
        sh_g = wb.sheets[H["general"]]
        sh_cump = wb.sheets[H["cumplimiento"]]

        gen_last = ultima_fila_empleado(sh_g, bg)
        log.info("BASE GENERAL: empleados filas %d-%d (%d).",
                 bg["fila_primer_dato"], gen_last, gen_last - bg["fila_primer_dato"] + 1)

        # (A) preservar OBSERVACION previa y fusionar (crudo gana; si vacio, nota previa)
        prev_obs = leer_observacion_previa(sh_base, esquema, cfg["clave_observacion"], log)
        i_obs = nombres.index("OBSERVACIÓN")
        reusadas = 0
        for fila, reg in zip(matriz, filas):
            val = fila[i_obs]
            if not (isinstance(val, str) and val.strip()):
                k = V.clave_registro(reg, cfg["clave_observacion"])
                if prev_obs.get(k):
                    fila[i_obs] = prev_obs[k]
                    reusadas += 1
        log.info("OBSERVACION: %d notas previas reutilizadas.", reusadas)

        # (B) volcar la hoja base
        volcar_base(sh_base, matriz, esquema, n_filas, log)

        # (C) reajustar fuente del pivote + RefreshAll #1 (DB_Vac visual toma la nueva base)
        ajustar_fuente_pivote(sh_piv, H["base"], tope, log)
        log.info("RefreshAll #1 (DB_Vac visual)...")
        wb.api.RefreshAll()
        app.api.CalculateUntilAsyncQueriesDone()

        # (D) calcular y escribir la tabla espejo L:S EN PYTHON (matricula como texto;
        #     fuente de verdad de los numeros, independiente del pivote visual)
        meses = sorted({f["Mes Pago"] for f in filas if isinstance(f.get("Mes Pago"), int)})
        bloque = calcular_espejo(filas, meses)
        ncols, ult_espejo = escribir_espejo(sh_piv, piv_cfg, bloque, tope, gen_last, log)
        log.info("Meses detectados=%s | empleados con vacaciones=%d", meses, len(bloque) - 1)

        # (E) actualizar rangos SUM del pivote
        actualizar_rangos_pivote(sh_piv, piv_cfg, ncols, tope, log)

        # (F) BASE GENERAL: formulas y totales
        rellenar_base_general(sh_g, bg, H["pivote"], tope, ncols, gen_last, log)

        # (G) recalcular, luego RefreshAll #2 para que R_Cumplimiento lea la nueva BG
        app.api.Calculate()
        app.api.CalculateUntilAsyncQueriesDone()
        log.info("RefreshAll #2 (R_Cumplimiento)...")
        wb.api.RefreshAll()
        app.api.CalculateUntilAsyncQueriesDone()
        app.api.Calculate()

        # (H) guardar
        wb.save()
        log.info("Guardado OK: %s", Path(destino).name)

        # (I) reporte de reconciliacion
        resumen = reporte_reconciliacion(sh_piv, sh_cump, sh_g, piv_cfg, ncols, ult_espejo, log)
        resumen.update({"registros": n_filas,
                        "empleados_bg": gen_last - bg["fila_primer_dato"] + 1,
                        "archivo": str(destino)})
        return resumen
    finally:
        try:
            if wb is not None and cerrar:
                wb.close()
        except Exception:
            pass
        try:
            if cerrar:
                app.quit()
        except Exception:
            pass


def ultima_fila_empleado(sh_g, bg):
    f0 = bg["fila_primer_dato"]
    cA = col_num(bg["columnas"]["matricula"])
    col = sh_g.range((f0, cA), (f0 + 6000, cA)).value
    last = f0 - 1
    for i, v in enumerate(col or []):
        if v is not None and RX_MATRICULA.fullmatch(str(v).strip()):
            last = f0 + i
    return last


def leer_observacion_previa(sh_base, esquema, campos_clave, log):
    nombres = [c["nombre"] for c in esquema]
    col_obs = nombres.index("OBSERVACIÓN")
    nrows = sh_base.used_range.rows.count
    if nrows < 2:
        return {}
    datos = sh_base.range((2, 1), (nrows, len(nombres))).value
    if datos and not isinstance(datos[0], list):
        datos = [datos]
    idx = {n: nombres.index(n) for n in campos_clave}
    mapa = {}
    for fila in datos or []:
        obs = fila[col_obs]
        if obs in (None, ""):
            continue
        reg = {n: fila[idx[n]] for n in campos_clave}
        mapa[V.clave_registro(reg, campos_clave)] = obs
    log.info("OBSERVACION previa: %d notas en la base anterior.", len(mapa))
    return mapa


def volcar_base(sh_base, matriz, esquema, n_filas, log):
    nombres = [c["nombre"] for c in esquema]
    ncol = len(nombres)
    viejo_last = max(sh_base.used_range.rows.count, 2)
    sh_base.range((2, 1), (max(viejo_last, n_filas + 1) + 10, ncol)).clear_contents()
    if n_filas == 0:
        log.warning("No hay filas para volcar.")
        return
    last = 1 + n_filas
    for j, campo in enumerate(esquema, start=1):
        rng = sh_base.range((2, j), (last, j))
        if campo["tipo"] == "texto":
            rng.number_format = "@"
        elif campo["tipo"] == "fecha":
            rng.number_format = "DD/MM/YYYY"
        elif campo["tipo"] == "entero":
            rng.number_format = "0"
    sh_base.range((2, 1)).value = matriz
    log.info("Base volcada: %d filas x %d columnas (hasta %s%d).", n_filas, ncol, col_letra(ncol), last)


def ajustar_fuente_pivote(sh_piv, nombre_hoja_base, tope, log):
    """Expande la fuente del pivote VISUAL solo si la base supera las ~2000 filas que ya
    cubre el cache de la plantilla. Es best-effort: los numeros de cumplimiento NO dependen
    de esto (vienen del espejo calculado en Python)."""
    if tope <= 2000:
        return
    try:
        wbapi = sh_piv.api.Parent
        pt = sh_piv.api.PivotTables(1)
        pc = wbapi.PivotCaches().Create(SourceType=1,  # xlDatabase
                                        SourceData=f"{nombre_hoja_base}!R1C1:R{tope}C17")
        pt.ChangePivotCache(pc)
        log.info("Fuente del pivote visual expandida a %d filas.", tope)
    except Exception as e:
        log.warning("No se pudo expandir la fuente del pivote visual a %d filas. "
                    "El pivote DB_Vac puede quedar corto visualmente, pero el cumplimiento "
                    "es correcto (espejo calculado en Python). Extiende el pivote a mano si lo "
                    "necesitas. Detalle: %s", tope, e)


def calcular_espejo(filas, meses):
    """Agrupa Suma de Dias por (Matricula x Mes Pago). Devuelve bloque 2D:
    fila 0 = cabecera [BP, Nombre, *meses, Total general]; resto = datos por matricula."""
    agg = {}
    for f in filas:
        m = f.get("Matrícula")
        if not m:
            continue
        d = agg.setdefault(m, {"nombre": f.get("Nombre") or "", "m": defaultdict(float)})
        d["m"][f.get("Mes Pago")] += (f.get("Días") or 0)
    bloque = [["BP", "Nombre"] + list(meses) + ["Total general"]]
    for m in sorted(agg):
        d = agg[m]
        fila = [m, d["nombre"]]
        for mes in meses:
            v = d["m"].get(mes, 0)
            fila.append(int(v) if float(v).is_integer() else v)
        tot = sum(d["m"].values())
        fila.append(int(tot) if float(tot).is_integer() else tot)
        bloque.append(fila)
    return bloque


def escribir_espejo(sh_piv, piv_cfg, bloque, tope, gen_last, log):
    """Escribe el bloque (cabecera + datos) en la tabla espejo L:S; matricula como TEXTO.
    Anade la columna de reconciliacion (VLOOKUP a BASE GENERAL) tras el Total general."""
    esp = piv_cfg["tabla_espejo"]
    c_ini = col_num(esp["col_inicio"])        # L = 12
    fila_cab = esp["fila_cabecera"]           # 4
    ncols = len(bloque[0])                     # 2 + nmeses + 1
    n_datos = len(bloque) - 1
    fila_ini = fila_cab + 1                    # 5
    ult_espejo = fila_cab + n_datos
    # limpiar zona vieja (generosa) y fijar formatos antes de escribir
    sh_piv.range((fila_cab, c_ini), (tope + 10, c_ini + ncols + 1)).clear_contents()
    sh_piv.range((fila_cab, c_ini), (tope + 10, c_ini + 1)).number_format = "@"          # BP, Nombre
    sh_piv.range((fila_cab, c_ini + 2), (tope + 10, c_ini + ncols - 1)).number_format = "General"
    sh_piv.range((fila_cab, c_ini)).value = bloque
    # columna de reconciliacion (justo despues del Total general)
    c_rec = c_ini + ncols
    L = esp["col_inicio"]
    formulas = [[f'=IFERROR(VLOOKUP({L}{r},\'BASE GENERAL\'!$A$2:$B${gen_last},1,0),"FALTA EN BG")']
                for r in range(fila_ini, ult_espejo + 1)]
    if formulas:
        sh_piv.range((fila_ini, c_rec), (ult_espejo, c_rec)).formula = formulas
    log.info("Espejo L:S escrito: %d filas datos (%d-%d), ncols=%d, recon col %s.",
             n_datos, fila_ini, ult_espejo, ncols, col_letra(c_rec))
    return ncols, ult_espejo


def actualizar_rangos_pivote(sh_piv, piv_cfg, ncols, tope, log):
    esp = piv_cfg["tabla_espejo"]
    c_ini = col_num(esp["col_inicio"])
    c_tot = c_ini + ncols - 1
    letra_tot = col_letra(c_tot)
    for c in range(c_ini + 2, c_tot + 1):       # N..S
        L = col_letra(c)
        sh_piv.range((3, c)).formula = f"=SUM({L}5:{L}{tope})"
        sh_piv.range((2, c)).formula = f"={L}3/${letra_tot}$3"
    log.info("Rangos SUM del pivote -> fila %d (cols %s..%s).",
             tope, col_letra(c_ini + 2), letra_tot)


def rellenar_base_general(sh_g, bg, hoja_pivote, tope, ncols, gen_last, log):
    cols = bg["columnas"]
    f0 = bg["fila_primer_dato"]
    Q, M, N = cols["objetivo"], cols["vac_vencidas"], cols["vac_pendiente"]
    R, S, T = cols["meta"], cols["obligatorio"], cols["registradas"]
    U, P, Vd = cols["meta_pct"], cols["suma_total"], cols["dias_restantes"]
    mat = cols["matricula"]
    idx_total = ncols  # columna 'Total general' dentro del espejo L:S (8)
    fR, fS, fT, fU, fV = [], [], [], [], []
    for r in range(f0, gen_last + 1):
        fR.append([f'=IF({Q}{r}>0,"Sí","No")'])
        fS.append([f'=IF(OR(N({M}{r})>0,N({N}{r})>0),"Sí","No")'])
        fT.append([f'=IFERROR(VLOOKUP({mat}{r},{hoja_pivote}!$L$4:$S${tope},{idx_total},0),0)'])
        fU.append([f'=IFERROR({T}{r}/{Q}{r},0)'])
        fV.append([f'={P}{r}-{T}{r}'])
    if gen_last >= f0:
        sh_g.range((f0, col_num(R)), (gen_last, col_num(R))).formula = fR
        sh_g.range((f0, col_num(S)), (gen_last, col_num(S))).formula = fS
        sh_g.range((f0, col_num(T)), (gen_last, col_num(T))).formula = fT
        sh_g.range((f0, col_num(U)), (gen_last, col_num(U))).formula = fU
        sh_g.range((f0, col_num(U)), (gen_last, col_num(U))).number_format = "0%"
        sh_g.range((f0, col_num(Vd)), (gen_last, col_num(Vd))).formula = fV
    sh_g.range(f"{Q}1").formula = f"=SUM({Q}{f0}:{Q}{gen_last})"
    sh_g.range(f"{T}1").formula = f"=SUM({T}{f0}:{T}{gen_last})"
    sh_g.range("W1").formula = f"={T}1/{Q}1"
    log.info("BASE GENERAL: R,S,T,U,V rellenadas %d-%d; totales fila1 SUM(...:%d).",
             f0, gen_last, gen_last)


def reporte_reconciliacion(sh_piv, sh_cump, sh_g, piv_cfg, ncols, ult_espejo, log):
    esp = piv_cfg["tabla_espejo"]
    c_ini = col_num(esp["col_inicio"])
    c_rec = c_ini + ncols
    faltan = []
    if ult_espejo >= 5:
        vals = sh_piv.range((5, c_ini), (ult_espejo, c_rec)).value
        for fila in (vals or []):
            if not isinstance(fila, list):
                fila = [fila]
            flag = fila[-1]
            if isinstance(flag, str) and "FALTA" in flag.upper():
                faltan.append((fila[0], fila[1] if len(fila) > 1 else ""))
    avance = None
    for sh, celda in ((sh_cump, "E9"), (sh_g, "W1")):
        try:
            v = sh.range(celda).value
            if isinstance(v, (int, float)):
                avance = v
                break
        except Exception:
            pass
    if faltan:
        log.warning("RECONCILIACION: %d con vacaciones registradas NO estan en BASE GENERAL (posibles nuevos ingresos):", len(faltan))
        for m, n in faltan[:30]:
            log.warning("   - %s  %s", m, n)
    else:
        log.info("RECONCILIACION: todos los que registraron estan en BASE GENERAL.")
    return {"faltan_en_bg": faltan, "avance_global": avance}


# ----------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Pipeline Vacaciones USIL")
    ap.add_argument("--dry-run", action="store_true", help="Solo prepara datos (no abre Excel)")
    ap.add_argument("--oculto", action="store_true", help="Excel invisible")
    ap.add_argument("--no-cerrar", action="store_true", help="Deja Excel abierto al terminar")
    ap.add_argument("--no-publicar", action="store_true", help="No publica al archivo del front")
    ap.add_argument("--crudo", help="Ruta a un crudo especifico")
    ap.add_argument("--config", help="config.json alternativo")
    args = ap.parse_args()

    cfg = V.cargar_config(args.config)
    log = V.configurar_log(cfg["salida"]["carpeta_logs"])
    log.info("=" * 70)
    log.info("PIPELINE VACACIONES USIL  -  %s", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    try:
        if args.crudo:
            ruta_crudo = args.crudo
            log.info("Crudo (forzado): %s", ruta_crudo)
        else:
            ent = cfg["entrada"]
            ruta_crudo, cands = V.encontrar_crudo(ent["carpetas_descarga"], ent["patron_crudo"],
                                                  ent.get("minutos_antiguedad_maxima"))
            log.info("Crudos encontrados: %d. Elegido: %s",
                     len(cands), Path(ruta_crudo).name if ruta_crudo else "NINGUNO")
        if not ruta_crudo:
            log.error("No se encontro crudo VACRptMotivo_*.xlsx. Abortando.")
            return 2

        filas = preparar_datos(cfg, ruta_crudo, log)
        if args.dry_run:
            log.info("DRY-RUN: %d registros listos. No se abre Excel.", len(filas))
            return 0

        fuente = elegir_objetivo(cfg, log)
        destino = crear_copia_fechada(cfg, fuente, log)
        resumen = ejecutar_com(cfg, destino, filas, args.oculto, args.no_cerrar, log)

        # publicar al archivo que lee el dashboard (front), para refrescar KPIs
        intg = cfg.get("integracion_front", {})
        if not args.no_publicar and intg.get("publicar_por_defecto", False):
            resumen["publicados"] = publicar_salida(cfg, destino, log)
        escribir_estado(destino, resumen, log)

        av = resumen["avance_global"]
        av_txt = f"{av:.1%}" if isinstance(av, (int, float)) else str(av)
        log.info("-" * 70)
        log.info("RESUMEN: %d registros | %d empleados BG | avance global=%s | faltan en BG=%d",
                 resumen["registros"], resumen["empleados_bg"], av_txt, len(resumen["faltan_en_bg"]))
        log.info("Archivo final: %s", resumen["archivo"])
        log.info("PIPELINE OK.")
        return 0
    except Exception as e:
        log.error("ERROR: %s", e)
        log.error(traceback.format_exc())
        return 1


if __name__ == "__main__":
    sys.exit(main())
