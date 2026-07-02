# -*- coding: utf-8 -*-
"""
vac_lib.py  -  Utilidades del pipeline de Vacaciones USIL.
Lectura/limpieza del crudo de Adryan, normalizacion de cabeceras y conversion de tipos.
NO toca Excel por COM: solo lee el crudo (xlsx) con openpyxl. Es seguro y testeable.
"""
import os, re, glob, json, time, unicodedata, datetime, logging
from pathlib import Path
import openpyxl

# ---------------------------------------------------------------- rutas / config
RAIZ = Path(__file__).resolve().parent.parent      # carpeta PIPELINE
MOTOR = Path(__file__).resolve().parent             # carpeta PIPELINE/motor

def cargar_config(ruta=None):
    ruta = Path(ruta) if ruta else (MOTOR / "config.json")
    with open(ruta, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    cfg["_raiz"] = str(RAIZ)
    return cfg

def ruta_abs(rel):
    """Resuelve una ruta relativa respecto a la carpeta PIPELINE.
    Si no existe y estamos dentro de dist_electron, intenta buscar en el area de desarrollo (carpeta principal)."""
    p = Path(rel)
    if p.is_absolute():
        return p
    res = RAIZ / p
    if res.exists():
        return res
    if "dist_electron" in str(RAIZ):
        curr = RAIZ
        for _ in range(4):
            parent_pipeline = curr / "PIPELINE"
            if parent_pipeline.is_dir() and (parent_pipeline / p).exists():
                return parent_pipeline / p
            curr = curr.parent
    return res

# ---------------------------------------------------------------- normalizacion
def normalizar(s):
    """minusculas, sin acentos, espacios colapsados. 'Fecha Início' -> 'fecha inicio'."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s

# ---------------------------------------------------------------- parsers de tipo
_FMT_FECHA = ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d")

def parse_fecha(v):
    """Texto dd/mm/aaaa (formato peruano, dia primero) -> datetime. None si no aplica."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, datetime.date):
        return datetime.datetime(v.year, v.month, v.day)
    s = str(v).strip()
    if not s:
        return None
    for fmt in _FMT_FECHA:
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def parse_numero(v):
    """Texto/num -> int si es entero, float si tiene decimales. None si no aplica."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else float(v)
    s = str(v).strip().replace(" ", "")
    if not s:
        return None
    # admite coma decimal o de miles segun forma
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None

# ---------------------------------------------------------------- localizar crudo
def encontrar_crudo(carpetas, patron, minutos_max=None):
    """Devuelve (ruta_mas_reciente, lista_candidatos_ordenada). None si no hay."""
    cands = []
    for carp in carpetas:
        cands.extend(glob.glob(os.path.join(carp, patron)))
    cands = sorted(set(cands), key=lambda p: os.path.getmtime(p), reverse=True)
    if minutos_max:
        corte = time.time() - minutos_max * 60
        cands = [c for c in cands if os.path.getmtime(c) >= corte]
    return (cands[0] if cands else None), cands

# ---------------------------------------------------------------- leer y limpiar
def detectar_cabecera(ws, ancla="matricula", max_filas=40):
    """Busca la fila cuyo contenido normalizado contiene el ancla. Devuelve (fila, {norm:col})."""
    for r in range(1, min(max_filas, ws.max_row) + 1):
        norms = {}
        for c in range(1, ws.max_column + 1):
            nn = normalizar(ws.cell(r, c).value)
            if nn:
                norms.setdefault(nn, c)   # primera aparicion gana
        if ancla in norms:
            return r, norms
    return None, None

def leer_crudo(path, esquema, regex_matricula="^0*\\d+$"):
    """
    Lee el crudo de Adryan. Detecta cabecera, mapea columnas del esquema por NOMBRE
    normalizado y conserva solo filas con matricula valida (descarta metadatos y
    filas de grupo tipo 'SIN GRUPO').
    Devuelve dict con: filas (list[dict nombre->valor]), faltantes, descartadas, etc.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    fila_cab, header_map = detectar_cabecera(ws)
    if fila_cab is None:
        wb.close()
        raise ValueError(f"No se encontro la fila de cabecera (ancla 'Matrícula') en {path}")

    # mapear cada campo del esquema a una columna de origen, por nombre normalizado
    col_de = {}
    faltantes = []
    for campo in esquema:
        nn = normalizar(campo["nombre"])
        if nn in header_map:
            col_de[campo["nombre"]] = header_map[nn]
        else:
            faltantes.append(campo["nombre"])

    rx = re.compile(regex_matricula)
    col_matricula = col_de.get("Matrícula")
    if not col_matricula:
        wb.close()
        raise ValueError("No se ubico la columna 'Matrícula' en el crudo.")

    filas, descartadas, muestras_descarte = [], 0, []
    for r in range(fila_cab + 1, ws.max_row + 1):
        matr = ws.cell(r, col_matricula).value
        matr_s = "" if matr is None else str(matr).strip()
        if not rx.fullmatch(matr_s):
            descartadas += 1
            if len(muestras_descarte) < 5 and (matr_s or ws.cell(r, 2).value):
                muestras_descarte.append((r, matr_s, str(ws.cell(r, 2).value)[:25]))
            continue
        reg = {}
        for campo in esquema:
            ci = col_de.get(campo["nombre"])
            reg[campo["nombre"]] = ws.cell(r, ci).value if ci else None
        filas.append(reg)
    wb.close()
    return {
        "filas": filas,
        "fila_cabecera": fila_cab,
        "col_de": col_de,
        "faltantes": faltantes,
        "descartadas": descartadas,
        "muestras_descarte": muestras_descarte,
        "cabeceras_detectadas": sorted(header_map.keys()),
    }

def convertir_tipos(filas, esquema):
    """Convierte in-place segun el esquema. Devuelve lista de incidencias (no convertibles)."""
    incidencias = []
    for i, reg in enumerate(filas):
        for campo in esquema:
            nombre, tipo = campo["nombre"], campo["tipo"]
            # OBSERVACIÓN (preservar) tambien viene del crudo de Adryan; se limpia como
            # texto aqui y luego en el motor se fusiona con notas manuales previas (el
            # valor del crudo gana si no esta vacio).
            v = reg.get(nombre)
            if tipo == "fecha":
                nv = parse_fecha(v)
                if v not in (None, "") and nv is None:
                    incidencias.append((i + 1, nombre, repr(v)))
                reg[nombre] = nv
            elif tipo == "entero":
                nv = parse_numero(v)
                if v not in (None, "") and nv is None:
                    incidencias.append((i + 1, nombre, repr(v)))
                reg[nombre] = nv
            else:  # texto
                reg[nombre] = "" if v is None else (v.strip() if isinstance(v, str) else str(v))
    return incidencias

def clave_registro(reg, campos_clave):
    """Clave estable para re-adjuntar OBSERVACIÓN tras refrescar la base."""
    partes = []
    for k in campos_clave:
        v = reg.get(k)
        if isinstance(v, datetime.datetime):
            v = v.strftime("%Y-%m-%d")
        partes.append("" if v is None else str(v).strip())
    return "|".join(partes)

# ---------------------------------------------------------------- logging
def configurar_log(carpeta_logs):
    carpeta_logs = ruta_abs(carpeta_logs)
    carpeta_logs.mkdir(parents=True, exist_ok=True)
    archivo = carpeta_logs / f"pipeline_{datetime.datetime.now():%Y%m%d_%H%M%S}.log"
    logger = logging.getLogger("vac")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s  %(levelname)-7s  %(message)s", "%H:%M:%S")
    fh = logging.FileHandler(archivo, encoding="utf-8")
    fh.setFormatter(fmt)
    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info("Log: %s", archivo)
    return logger


# ---------------------------------------------------------------- autoprueba
if __name__ == "__main__":
    cfg = cargar_config()
    esquema = cfg["esquema_base"]
    ent = cfg["entrada"]
    ruta, cands = encontrar_crudo(ent["carpetas_descarga"], ent["patron_crudo"],
                                  ent.get("minutos_antiguedad_maxima"))
    print("Candidatos crudo:", len(cands))
    print("Crudo elegido   :", ruta)
    if not ruta:
        raise SystemExit("No hay crudo.")
    res = leer_crudo(ruta, esquema, cfg["limpieza_crudo"]["regex_matricula_valida"])
    print(f"Fila cabecera   : {res['fila_cabecera']}")
    print(f"Filas validas   : {len(res['filas'])}")
    print(f"Filas descartadas: {res['descartadas']}  muestras={res['muestras_descarte']}")
    print(f"Columnas faltantes en crudo: {res['faltantes']}")
    inc = convertir_tipos(res["filas"], esquema)
    print(f"Incidencias de conversion: {len(inc)}  (primeras 5: {inc[:5]})")
    print("--- 2 registros de muestra (ya tipados) ---")
    for reg in res["filas"][:2]:
        for campo in esquema:
            n = campo["nombre"]
            print(f"   {n:<20} = {reg[n]!r}   [{type(reg[n]).__name__}]")
        print("   " + "-" * 40)
