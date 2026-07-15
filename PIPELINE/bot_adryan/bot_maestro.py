# -*- coding: utf-8 -*-
"""
Bot que descarga el Maestro de Personal desde Adryan y lo sanitiza.

Pasos:
  1. Login en Adryan (login fresco de un solo intento, igual que diagnostico_login.py)
  2. Personal > Maestro del Personal sin filtro > descargar
  3. Sanitizar el Excel descargado:
     - Eliminar los 3 primeros registros (dueños de la corporación)
     - Sustituir TODOS los emails (col DIRECCIÓN ELECTRÓNICA) por jlopezp@usil.edu.pe

Uso:
    python bot_maestro.py            (headless según config)
    python bot_maestro.py --visible  (fuerza ver el navegador)
"""
import os
import sys
import json
import time
import shutil
import datetime
import traceback

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    print("ERROR: playwright no esta instalado en este entorno de Python.")
    print("Esta funcion solo esta disponible desde la maquina administrativa.")
    print("Para instalar: pip install playwright && playwright install chromium")
    sys.exit(1)

AQUI = os.path.dirname(os.path.abspath(__file__))
CONFIG = os.path.join(AQUI, "config_bot.json")
LOG_DIR = os.path.join(AQUI, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

import guardar_password
import bot_adryan  # reusa el login robusto (campo oculto duplicado + verificacion + reintentos)

# Defaults de respaldo si config_bot.json no trae estas claves (compatibilidad
# hacia atras); el valor real y editable vive en config_bot.json.
FILAS_DUENOS_DEFAULT = 3
EMAIL_SEGURO_DEFAULT = "jlopezp@usil.edu.pe"
FILA_HEADER_DEFAULT = 11
FILA_DATA_INICIO_DEFAULT = 12


def log(msg: str):
    linea = f"[{datetime.datetime.now():%Y-%m-%d %H:%M:%S}] {msg}"
    print(linea, flush=True)
    with open(os.path.join(LOG_DIR, f"maestro_{datetime.date.today():%Y%m%d}.log"),
              "a", encoding="utf-8") as f:
        f.write(linea + "\n")


def cargar_config() -> dict:
    with open(CONFIG, encoding="utf-8") as f:
        return json.load(f)


def _navegar_menu_maestro(page):
    """Navegar por el menu: Personal > Maestro del Personal sin filtro."""
    page.goto("https://adryancloudusil.sapia.com.pe/Home/IndexAdminDashboard",
              wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(2000)

    # Clic en el item de menu "Personal" (igual que el script grabado)
    try:
        page.get_by_role("link", name="Personal").first.click()
        page.wait_for_timeout(1500)
    except Exception:
        # Fallback JS
        page.evaluate("""() => {
            const all = [...document.querySelectorAll('a')];
            const a = all.find(el => (el.textContent || '').trim() === 'Personal');
            if (a) a.click();
        }""")
        page.wait_for_timeout(1500)

    # Clic en "Maestro del Personal sin [filtro]"
    try:
        page.get_by_role("link", name="Maestro del Personal sin").click()
    except Exception:
        # Fallback: buscar por texto parcial
        page.get_by_role("link", name="Maestro del Personal").first.click()

    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(3000)
    log("Maestro abierto.")


def descargar_maestro(page, cfg) -> str:
    log("Navegando a Maestro del Personal...")

    # Intentar URL directa primero (más robusto que navegar por menú)
    url_maestro = cfg.get("url_maestro", "").strip()
    if url_maestro:
        page.goto(url_maestro, wait_until="domcontentloaded")
        page.wait_for_load_state("networkidle")
        time.sleep(2)
        if page.locator(".ctn-iconos-edit-tabla > a").first.is_visible(timeout=5000):
            log("Maestro abierto por URL directa.")
        else:
            log("URL directa no mostro el maestro; intento por menu lateral.")
            _navegar_menu_maestro(page)
    else:
        _navegar_menu_maestro(page)

    # Esperar a que la grilla cargue datos (filas en tbody)
    log("Esperando a que la grilla cargue datos (filas en tbody)...")
    try:
        page.locator("tbody tr").first.wait_for(state="visible", timeout=45000)
        log("Grilla con datos detectada. Esperando 3 segundos para estabilizacion...")
        page.wait_for_timeout(3000)
    except Exception as e:
        log(f"Advertencia: tiempo de espera de datos agotado o error: {e}. Se intentara descargar igual.")

    log("Descargando maestro...")
    timeout_dl = max(cfg.get("timeout_ms", 60000), 120000)  # minimo 2 minutos

    # La descarga se dispara via JS (reportpersonallistjs.export(0)),
    # no por navegacion. Usamos expect_download + evaluate para capturarla.
    try:
        with page.expect_download(timeout=timeout_dl) as dl_info:
            page.evaluate("reportpersonallistjs.export(0)")
        download = dl_info.value
    except PWTimeout:
        # Fallback: intentar con click en el boton de descarga
        log("evaluate no disparo descarga, reintentando con click...")
        with page.expect_download(timeout=timeout_dl) as dl_info:
            page.locator(".ctn-iconos-edit-tabla > a").first.click(no_wait_after=True)
        download = dl_info.value

    nombre = download.suggested_filename or f"PersonalMaestro_{datetime.datetime.now():%m_%d_%Y %H_%M_%S}.xlsx"
    carpeta = cfg.get("carpeta_descarga", "").strip()
    if not carpeta or not os.path.isdir(carpeta):
        carpeta = os.path.join(os.path.expanduser("~"), "Downloads")
        os.makedirs(carpeta, exist_ok=True)
        log(f"carpeta_descarga no encontrada; usando {carpeta}")
    destino = os.path.join(carpeta, nombre)
    download.save_as(destino)
    log(f"Descargado: {destino}")
    return destino


def sanitizar_maestro(ruta_xlsx: str, cfg: dict) -> str:
    """Elimina los primeros registros (dueños de la corporacion, cantidad
    configurable via filas_duenos en config_bot.json) y reemplaza todos los emails."""
    import openpyxl

    filas_duenos = int(cfg.get("filas_duenos", FILAS_DUENOS_DEFAULT))
    email_seguro = cfg.get("email_seguro", EMAIL_SEGURO_DEFAULT)
    fila_header = int(cfg.get("fila_header_maestro", FILA_HEADER_DEFAULT))
    fila_data_inicio = int(cfg.get("fila_data_inicio_maestro", FILA_DATA_INICIO_DEFAULT))

    log(f"Sanitizando {os.path.basename(ruta_xlsx)}...")
    wb = openpyxl.load_workbook(ruta_xlsx)
    ws = wb[wb.sheetnames[0]]

    # Leer todas las filas en memoria para procesarlas rápidamente
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(row)

    # El encabezado está en la fila fila_header (índice fila_header-1)
    # Los datos comienzan en la fila siguiente
    metadata_and_header = all_rows[:fila_header]
    header_row = all_rows[fila_header - 1]
    data_rows = all_rows[fila_header:]

    # Columnas de email por NOMBRE de cabecera, no por posicion fija: la plantilla
    # de Adryan ha ganado/movido columnas entre corridas y una posicion fija
    # (ej. "columna 20") deja de apuntar al email real sin avisar.
    cols_email = [i for i, h in enumerate(header_row)
                  if h and 'email' in str(h).strip().lower()]
    log(f"  Columnas de email detectadas por nombre: {cols_email} "
        f"({[header_row[i] for i in cols_email]})")

    # Columnas de matricula/nombre por header, solo para el log auditable de abajo
    idx_mat = next((i for i, h in enumerate(header_row) if h and 'matr' in str(h).strip().lower()), None)
    idx_ap = next((i for i, h in enumerate(header_row) if h and 'apellido paterno' in str(h).strip().lower()), None)
    idx_am = next((i for i, h in enumerate(header_row) if h and 'apellido materno' in str(h).strip().lower()), None)
    idx_no = next((i for i, h in enumerate(header_row) if h and str(h).strip().lower() == 'nombre'), None)

    def _describir(row):
        mat = row[idx_mat] if idx_mat is not None and idx_mat < len(row) else '?'
        partes = [row[i] for i in (idx_ap, idx_am, idx_no) if i is not None and i < len(row) and row[i]]
        return f"{mat} - {' '.join(str(p) for p in partes) or '(sin nombre)'}"

    # 1) Eliminar los primeros N registros de data (asumidos "dueños de la corporacion")
    descartados = data_rows[:filas_duenos]
    filtered_data_rows = data_rows[filas_duenos:]
    log(f"  Descartando {filas_duenos} registro(s) como 'dueños de la corporación' "
        f"(configurable en config_bot.json -> filas_duenos):")
    for r in descartados:
        if r:
            log(f"    - {_describir(r)}")

    # 2) Sustituir todos los emails por el email seguro
    emails_reemplazados = 0
    sanitized_data = []
    for row in filtered_data_rows:
        if not row:
            continue
        row_list = list(row)
        for col in cols_email:
            if col < len(row_list):
                val = row_list[col]
                if val and "@" in str(val):
                    row_list[col] = email_seguro
                    emails_reemplazados += 1
        sanitized_data.append(row_list)
    log(f"  Emails reemplazados: {emails_reemplazados} -> {email_seguro}")

    # Recrear la hoja para evitar la lentitud extrema de delete_rows
    sheet_title = ws.title
    new_ws = wb.create_sheet(title="SanitizadoTemp")

    # Escribir cabecera y metadata
    for r_idx, row_val in enumerate(metadata_and_header, start=1):
        for c_idx, val in enumerate(row_val, start=1):
            new_ws.cell(row=r_idx, column=c_idx, value=val)

    # Escribir filas de datos sanitizadas
    for r_idx, row_val in enumerate(sanitized_data, start=fila_data_inicio):
        for c_idx, val in enumerate(row_val, start=1):
            new_ws.cell(row=r_idx, column=c_idx, value=val)

    # Eliminar hoja antigua y renombrar la nueva
    wb.remove(ws)
    new_ws.title = sheet_title

    # Guardar
    sanitizado = ruta_xlsx.replace(".xlsx", "_sanitizado.xlsx")
    wb.save(sanitizado)
    wb.close()
    log(f"  Archivo sanitizado: {sanitizado}")

    wb2 = openpyxl.load_workbook(sanitizado, read_only=True, data_only=True)
    ws2 = wb2[wb2.sheetnames[0]]
    total = sum(1 for row in ws2.iter_rows(min_row=fila_data_inicio, values_only=True) if row and row[0])
    wb2.close()
    log(f"  Registros finales: {total}")

    return sanitizado


def main():
    visible = "--visible" in sys.argv
    cfg = cargar_config()
    headless = cfg.get("headless", True) and not visible
    sesion = os.path.join(AQUI, cfg["archivo_sesion"])

    try:
        password = guardar_password.cargar()
    except Exception as e:
        log(f"ERROR: no hay contrasena guardada. Corre guardar_password.py primero. ({e})")
        return 2

    with sync_playwright() as pw:
        # Mismos args que bot_adryan.py: si headless=False (necesario para el login,
        # ver _comentario_headless en config_bot.json), la ventana se abre real pero
        # posicionada fuera de pantalla para que nadie la toque por accidente mientras
        # el bot descarga el maestro.
        browser = pw.chromium.launch(
            channel=cfg["canal_navegador"],
            headless=headless,
            args=["--disable-gpu", "--window-position=-32000,-32000", "--hide-scrollbars"]
        )
        # Sin storage_state: igual que diagnostico_login.py, siempre login fresco
        # directo a la pagina de login (reusar sesion guardada era lo sospechoso
        # de causar los fallos de login).
        context = browser.new_context(accept_downloads=True)
        context.set_default_timeout(cfg["timeout_ms"])
        page = context.new_page()

        try:
            bot_adryan.hacer_login(page, cfg, password)
            context.storage_state(path=sesion)
            log("Sesion guardada/actualizada.")

            destino = descargar_maestro(page, cfg)
            context.storage_state(path=sesion)

            sanitizado = sanitizar_maestro(destino, cfg)
            log("OK - maestro descargado y sanitizado.")
            print(f"ARCHIVO_MAESTRO={sanitizado}")
            return 0

        except Exception as e:
            log(f"ERROR: {e}")
            log(traceback.format_exc())
            try:
                shot = os.path.join(LOG_DIR, f"error_maestro_{datetime.datetime.now():%Y%m%d_%H%M%S}.png")
                page.screenshot(path=shot, full_page=True)
                log(f"Captura del error: {shot}")
            except Exception:
                pass
            return 1
        finally:
            context.close()
            browser.close()


if __name__ == "__main__":
    sys.exit(main())
