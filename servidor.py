# -*- coding: utf-8 -*-

"""

Sistema de Vacaciones USIL

People Analytics Â· USIL Â· Puerto 5002

"""

import sys, os



if sys.stdout and hasattr(sys.stdout, 'reconfigure'):

    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

if sys.stderr and hasattr(sys.stderr, 'reconfigure'):

    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

os.environ.setdefault('PYTHONIOENCODING', 'utf-8')



from flask import Flask, jsonify, request, send_from_directory, send_file, session

import pandas as pd

from datetime import datetime, date, timedelta

import html as html_lib

import glob, json, threading, unicodedata, uuid, shutil, tempfile, re, time, pickle
from concurrent.futures import ThreadPoolExecutor

def _safe_read_excel(ruta, **kwargs):
    """Lee un Excel copiando primero a temp para evitar PermissionError cuando está abierto en Excel."""
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        shutil.copy2(ruta, tmp_path)
        return pd.read_excel(tmp_path, **kwargs)
    finally:
        if tmp_path:
            try: os.unlink(tmp_path)
            except: pass

def medir_tiempo(func):

    """Decorador para medir el tiempo de ejecucion de funciones criticas."""

    import time, functools

    @functools.wraps(func)

    def wrapper(*args, **kwargs):

        t0 = time.time()

        res = func(*args, **kwargs)

        print(f"[PERF] {func.__name__} tomo {time.time()-t0:.3f}s")

        return res

    return wrapper




app = Flask(__name__, static_folder='.', static_url_path='')



SCRIPT_DIR      = os.path.dirname(os.path.abspath(__file__))

def _buscar_directorio(nombre_dir):
    d = SCRIPT_DIR
    for _ in range(4):
        p = os.path.join(d, nombre_dir)
        if os.path.isdir(p):
            return p
        d = os.path.dirname(d)
    return os.path.join(SCRIPT_DIR, nombre_dir)

DATAS_DIR       = _buscar_directorio('DATAS')

DATA_SENSIBLE_DIR = _buscar_directorio('DATA SENSIBLE')

COLAB_EDIT_FILE = os.path.join(DATAS_DIR, 'colaboradores_overrides.json')



# Nombre del archivo de vacaciones – busca el más reciente si el nombre exacto no existe

_VAC_NOMBRE_EXACTO = 'Reporte Vacaciones Objetivo_Segundo Trimestre 2026.xlsx'

VACACIONES_DATA_FILE = os.path.join(DATA_SENSIBLE_DIR, _VAC_NOMBRE_EXACTO)

if not os.path.isfile(VACACIONES_DATA_FILE):

    # Fallback: cualquier Reporte Vacaciones Objetivo en DATA SENSIBLE (orden alfa descendente)

    _candidatos_vac = sorted(

        glob.glob(os.path.join(DATA_SENSIBLE_DIR, 'Reporte Vacaciones Objetivo*.xlsx')),

        key=os.path.basename,

        reverse=True

    )

    VACACIONES_DATA_FILE = _candidatos_vac[0] if _candidatos_vac else VACACIONES_DATA_FILE

_PATRONES_VACACIONES_DATAS = [

    'Reporte Vacaciones Objetivo*.xlsx',

    'Vacaciones_Talento_Cultura*.xlsx',

]



def _cargar_con_cache(ruta, cache_name, loader_func, **kwargs):

    """Carga datos usando un cache en disco (pickle) si el archivo no ha cambiado."""

    cache_dir = os.path.join(DATAS_DIR, '__cache__')

    os.makedirs(cache_dir, exist_ok=True)

    cache_path = os.path.join(cache_dir, f'{cache_name}.pkl')

    

    if not os.path.exists(ruta):

        return None, f"Archivo no encontrado: {ruta}"

    

    mtime = os.path.getmtime(ruta)

    

    # Intentar cargar desde cache

    if os.path.exists(cache_path):

        try:

            with open(cache_path, 'rb') as f:

                c_data = pickle.load(f)

            if c_data.get('mtime') == mtime:

                print(f"[CACHE] Cargado {cache_name} desde disco (mtime coincide)")

                return c_data['data'], None

        except Exception as e:

            print(f"[CACHE] Error leyendo cache {cache_name}: {e}")

            

    # Si no hay cache o mtime cambio, cargar de verdad

    print(f"[CACHE] Leyendo {cache_name} desde Excel (fuerza bruta)...")

    data, err = loader_func(ruta, **kwargs)

    if not err:

        try:

            with open(cache_path, 'wb') as f:

                pickle.dump({'mtime': mtime, 'data': data}, f)

        except Exception as e:

            print(f"[CACHE] Error guardando cache {cache_name}: {e}")

            

    return data, err




# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# CONFIGURACIÃ“N (pa_config.json)

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _buscar_pa_config():

    d = SCRIPT_DIR

    for _ in range(5):

        p = os.path.join(d, 'pa_config.json')

        if os.path.isfile(p):

            return p

        d = os.path.dirname(d)

    return None



_PA_CONFIG_PATH = _buscar_pa_config()

_PA_CONFIG = {}

if _PA_CONFIG_PATH:

    try:

        with open(_PA_CONFIG_PATH, 'r', encoding='utf-8') as f:

            _PA_CONFIG = json.load(f)

        print(f'[OK] pa_config.json: {_PA_CONFIG_PATH}')

    except Exception as e:

        print(f'[WARN] pa_config.json: {e}')

app.secret_key = os.environ.get('PA_SESSION_SECRET', _PA_CONFIG.get('session_secret', 'dev-secret-123'))

@app.before_request
def log_request_info():
    # Silenciar endpoints de polling frecuente que generan spam en la consola
    _SILENT_PATHS = {
        '/api/cola-pa/ultima-pendiente',
        '/api/status/ready',
        '/api/scheduler/estado',
        '/api/confirmaciones/resumen',
    }
    if request.path not in _SILENT_PATHS:
        print(f'[REQ] {request.method} {request.path} {request.args.to_dict()}')





def _guardar_pa_config(patch_dict):

    """Actualiza pa_config en memoria y en disco para reflejar cambios sin reiniciar."""

    global _PA_CONFIG

    if not isinstance(patch_dict, dict):

        return False, 'Patch invalido'



    try:

        _PA_CONFIG.update(patch_dict)

        if _PA_CONFIG_PATH and os.path.isfile(_PA_CONFIG_PATH):

            with open(_PA_CONFIG_PATH, 'r', encoding='utf-8') as f:

                cfg = json.load(f)

        else:

            cfg = {}

        cfg.update(patch_dict)

        if _PA_CONFIG_PATH:

            with open(_PA_CONFIG_PATH, 'w', encoding='utf-8') as f:

                json.dump(cfg, f, indent=2, ensure_ascii=False)

        return True, None

    except Exception as e:

        return False, str(e)



TEAMS_WEBHOOK_URL          = os.environ.get('TEAMS_WEBHOOK_URL', '') or _PA_CONFIG.get('teams_webhook_url', '')

TEAMS_WEBHOOK_PERSONAL_URL = os.environ.get('TEAMS_WEBHOOK_PERSONAL_URL', '') or _PA_CONFIG.get('teams_webhook_personal_url', '') or TEAMS_WEBHOOK_URL

POWER_AUTOMATE_URL         = os.environ.get('POWER_AUTOMATE_URL', '') or _PA_CONFIG.get('power_automate_url', '')

SMTP_EMAIL         = _PA_CONFIG.get('smtp_email', '')

SMTP_PASSWORD      = _PA_CONFIG.get('smtp_password', '')

_scheduler_activo  = bool(_PA_CONFIG.get('scheduler_activo', True))





def _cfg_bool(value, default=False):

    if value is None:

        return default

    if isinstance(value, bool):

        return value

    if isinstance(value, (int, float)):

        return bool(value)

    txt = str(value).strip().lower()

    if txt in {'1', 'true', 't', 'yes', 'y', 'si', 's', 'on'}:

        return True

    if txt in {'0', 'false', 'f', 'no', 'n', 'off'}:

        return False

    return default





_ENCOLAR_ALERTAS_PA = _cfg_bool(

    os.environ.get('PA_ENCOLAR_ALERTAS', None),

    _cfg_bool(_PA_CONFIG.get('encolar_alertas_pa', True), True)

)



# PersonalMaestroReporte como universo completo TyC (fuente de verdad de personas)

_UNIVERSO_COMPLETO_ENABLED = True

_MAESTRO_FULL_PATH = next(iter(sorted(

    glob.glob(os.path.join(DATAS_DIR, 'PersonalMaestroReporte*.xlsx')) +

    glob.glob(os.path.join(DATA_SENSIBLE_DIR, 'PersonalMaestroReporte*.xlsx')),

    reverse=True

)), '')





_TEST_EMAIL_DESTINO = _PA_CONFIG.get('vacaciones_test_email', '')



_DEMO_JEFE_ENABLED = _cfg_bool(_PA_CONFIG.get('vacaciones_demo_jefe_enabled', False), False)

_DEMO_JEFE_NOMBRE = str(_PA_CONFIG.get('vacaciones_demo_jefe_nombre', 'Joshua Lopez') or 'Joshua Lopez').strip()

_DEMO_JEFE_EMAIL = str(_PA_CONFIG.get('vacaciones_demo_jefe_email', 'jlopezp@usil.edu.pe') or 'jlopezp@usil.edu.pe').strip().lower()



_PA_DELAY_SECONDS = 60



app.config['SECRET_KEY'] = (

    os.environ.get('PA_SESSION_SECRET')

    or _PA_CONFIG.get('session_secret')

    or 'pa_vacaciones_session_secret_change_me'

)

app.config['SESSION_COOKIE_HTTPONLY'] = True

app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

app.config['SESSION_COOKIE_SECURE'] = False



_INTERNAL_API_TOKEN = (

    os.environ.get('PA_INTERNAL_API_TOKEN')

    or _PA_CONFIG.get('internal_api_token')

    or str(uuid.uuid4())

)



# MatrÃ­culas NUNCA notificadas por Teams (directivos / esposa)

_TEAMS_EXCLUIR_MAT = {'0000000013', '0000000486'}   # DIEZ CANSECO TERRY RAUL + DANIEL HUGH



def _resolver_cola_dir():

    cfg_dir = str(_PA_CONFIG.get('alertas_cola_dir', '') or '').strip()

    if cfg_dir:

        return cfg_dir



    base_od = os.path.join(

        os.path.expanduser('~'),

        'OneDrive - Universidad San Ignacio de Loyola'

    )

    candidatos = [

        os.path.join(base_od, 'alertas_cola'),

        os.path.join(base_od, 'Documents', 'alertas_cola'),

    ]

    for cand in candidatos:

        if os.path.isdir(cand):

            return cand

    return candidatos[0]





def _aplicar_rutas_cola(base_dir):

    global COLA_DIR, COLA_IN_DIR, COLA_PROCESANDO_DIR, COLA_ARCHIVADOS_DIR

    global COLA_PENDIENTES_DIR, COLA_CANCELADOS_DIR

    global COLA_FUENTES_DIR, COLA_FUENTES_MAESTRA_DIR, COLA_FUENTES_OBJETIVOS_DIR



    base = os.path.abspath(os.path.expanduser(str(base_dir or '').strip() or _resolver_cola_dir()))

    COLA_DIR = base

    # PA trigger: /alertas_cola/in â€” Python deposita en esa subcarpeta.
    # Copia de auditorÃ­a se guarda en archivados\ inmediatamente.
    COLA_IN_DIR = os.path.join(COLA_DIR, 'in')
    COLA_PROCESANDO_DIR = os.path.join(COLA_DIR, 'procesando')
    COLA_ARCHIVADOS_DIR = os.path.join(COLA_DIR, 'archivados')
    COLA_PENDIENTES_DIR = os.path.join(COLA_DIR, 'pendientes')
    COLA_CANCELADOS_DIR = os.path.join(COLA_DIR, 'cancelados')
    COLA_FUENTES_DIR = os.path.join(COLA_DIR, 'fuentes_vacaciones')
    COLA_FUENTES_MAESTRA_DIR = os.path.join(COLA_FUENTES_DIR, 'maestra')
    COLA_FUENTES_OBJETIVOS_DIR = os.path.join(COLA_FUENTES_DIR, 'objetivos')

    try:
        for path in (
            COLA_DIR,
            COLA_IN_DIR,
            COLA_PROCESANDO_DIR,
            COLA_ARCHIVADOS_DIR,
            COLA_PENDIENTES_DIR,
            COLA_CANCELADOS_DIR,
            COLA_FUENTES_MAESTRA_DIR,
            COLA_FUENTES_OBJETIVOS_DIR,
        ):
            os.makedirs(path, exist_ok=True)
    except Exception as e:
        print(f'[WARN] Error creando directorios de cola en {COLA_DIR}: {e}. Usando fallback local...')
        COLA_DIR = os.path.join(SCRIPT_DIR, 'alertas_cola')
        COLA_IN_DIR = os.path.join(COLA_DIR, 'in')
        COLA_PROCESANDO_DIR = os.path.join(COLA_DIR, 'procesando')
        COLA_ARCHIVADOS_DIR = os.path.join(COLA_DIR, 'archivados')
        COLA_PENDIENTES_DIR = os.path.join(COLA_DIR, 'pendientes')
        COLA_CANCELADOS_DIR = os.path.join(COLA_DIR, 'cancelados')
        COLA_FUENTES_DIR = os.path.join(COLA_DIR, 'fuentes_vacaciones')
        COLA_FUENTES_MAESTRA_DIR = os.path.join(COLA_FUENTES_DIR, 'maestra')
        COLA_FUENTES_OBJETIVOS_DIR = os.path.join(COLA_FUENTES_DIR, 'objetivos')
        for path in (
            COLA_DIR,
            COLA_IN_DIR,
            COLA_PROCESANDO_DIR,
            COLA_ARCHIVADOS_DIR,
            COLA_PENDIENTES_DIR,
            COLA_CANCELADOS_DIR,
            COLA_FUENTES_MAESTRA_DIR,
            COLA_FUENTES_OBJETIVOS_DIR,
        ):
            os.makedirs(path, exist_ok=True)

    return COLA_DIR



_aplicar_rutas_cola(_resolver_cola_dir())

_ONEDRIVE_JSON    = os.path.join(SCRIPT_DIR, 'alertas_pa_onedrive.json')

_LOG_ENVIOS_PATH  = os.path.join(SCRIPT_DIR, 'log_envios.json')

_CONFIRMACIONES_PATH = os.path.join(SCRIPT_DIR, 'confirmaciones_vacaciones.json')

_NOTIF_STATE_FILE = os.path.join(SCRIPT_DIR, 'notif_estado.json')

_log_lock = threading.Lock()

_confirm_lock = threading.Lock()





def _meta_envio_pa(archivo, estado='pendiente'):

    creado = datetime.now()

    liberar = creado + timedelta(seconds=_PA_DELAY_SECONDS)

    return {

        'archivo': archivo,

        'estado': estado,

        'retraso_segundos': int(_PA_DELAY_SECONDS),

        'creado_en': creado.isoformat(),

        'liberar_en': liberar.isoformat(),

    }


def _leer_confirmaciones_vacaciones():

    with _confirm_lock:

        try:

            if not os.path.exists(_CONFIRMACIONES_PATH):

                return []

            with open(_CONFIRMACIONES_PATH, 'r', encoding='utf-8') as f:

                data = json.load(f)

            return data if isinstance(data, list) else []

        except Exception:

            return []


def _guardar_confirmaciones_vacaciones(registros):

    with _confirm_lock:

        try:

            with open(_CONFIRMACIONES_PATH, 'w', encoding='utf-8') as f:

                json.dump(registros, f, ensure_ascii=False, indent=2)

            return True

        except Exception as e:

            print(f'[CONFIRMACIONES] Error guardando: {e}')

            return False


def _clave_confirmacion_persona(persona):

    persona = persona or {}

    mat = _id_key(persona.get('matricula', ''))

    if mat:

        return mat

    return _norm(str(persona.get('nombre', '') or '')).upper()


def _resolver_public_base_url(default_url=''):

    base = str(_PA_CONFIG.get('public_base_url', '') or '').strip()

    if not base:

        base = str(default_url or '').strip()

    return base.rstrip('/')


def _registrar_confirmaciones_contenido(contenido, public_base_url=''):

    if not isinstance(contenido, dict):

        return contenido

    return contenido

    equipo = contenido.get('detalle_equipo_full') or []

    base = _resolver_public_base_url(public_base_url)

    if not equipo or not base:

        return contenido

    campaign_id = uuid.uuid4().hex
    created_at = datetime.now().isoformat()
    supervisor_nombre = contenido.get('nombre_jefe', '')
    supervisor_email = contenido.get('email_destino_real', '') or contenido.get('email_jefe', '')
    registros = _leer_confirmaciones_vacaciones()
    confirmaciones = {}

    for persona in equipo:

        nombre = str((persona or {}).get('nombre', '') or '').strip()

        if not nombre:

            continue

        token = uuid.uuid4().hex
        clave = _clave_confirmacion_persona(persona)
        confirmaciones[clave] = {
            'token': token,
            'url': f'{base}/confirmar-vacaciones?token={token}',
            'status': 'pendiente',
        }
        registros.append({
            'token': token,
            'campaign_id': campaign_id,
            'tipo': 'salida_vacaciones',
            'status': 'pendiente',
            'created_at': created_at,
            'confirmed_at': None,
            'supervisor_nombre': supervisor_nombre,
            'supervisor_email': supervisor_email,
            'email_destino': contenido.get('email_jefe', ''),
            'modo_prueba': bool(contenido.get('modo_prueba', False)),
            'asunto': contenido.get('asunto', ''),
            'colaborador_matricula': str((persona or {}).get('matricula', '') or '').strip(),
            'colaborador_nombre': nombre,
            'colaborador_area': str((persona or {}).get('area', '') or '').strip(),
            'colaborador_departamento': str((persona or {}).get('departamento', '') or '').strip(),
        })

    registros = registros[-10000:]
    if not _guardar_confirmaciones_vacaciones(registros):

        return contenido

    contenido['confirmacion_campania_id'] = campaign_id
    contenido['confirmacion_total'] = len(confirmaciones)
    contenido['mensaje_html'] = _build_html_jefe(
        contenido.get('nombre_jefe', ''),
        contenido.get('detalle_retraso', []),
        contenido.get('detalle_proximos', []),
        contenido.get('detalle_sin_cumplir', []),
        contenido.get('fecha_generacion', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        contenido.get('mensaje_cfg', ''),
        contenido.get('aviso_cfg', ''),
        contenido.get('recomendacion_cfg', ''),
        contenido.get('detalle_adicionales', []),
        contenido.get('campania', {}),
        contenido.get('meta_resumen', {}),
        hrbp_nombre=contenido.get('hrbp_nombre', ''),
        fecha_limite=contenido.get('fecha_limite_cfg', ''),
        equipo_full=contenido.get('detalle_equipo_full', []),
        confirmaciones=confirmaciones,
    )
    return contenido


def _render_confirmacion_html(title, message, accent='#0f766e'):

    safe_title = html_lib.escape(title or 'Confirmacion registrada')
    safe_message = html_lib.escape(message or '')
    return (
        '<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
        f'<title>{safe_title}</title></head>'
        '<body style="margin:0;padding:0;background:#eff6ff;font-family:Segoe UI,Arial,sans-serif">'
        '<div style="max-width:680px;margin:48px auto;padding:0 18px">'
        '<div style="background:#fff;border:1px solid #dbeafe;border-radius:18px;box-shadow:0 12px 40px rgba(15,23,42,.08);overflow:hidden">'
        f'<div style="padding:20px 24px;background:{accent};color:#fff">'
        f'<h1 style="margin:0;font-size:24px">{safe_title}</h1>'
        '</div>'
        '<div style="padding:24px">'
        f'<p style="margin:0;color:#1e293b;font-size:16px;line-height:1.6">{safe_message}</p>'
        '<p style="margin:18px 0 0;color:#64748b;font-size:13px">People Analytics USIL</p>'
        '</div></div></div></body></html>'
    )





def _resolver_liberar_por_archivo(archivo):

    p = os.path.join(COLA_PENDIENTES_DIR, archivo)

    if not os.path.isfile(p):

        return None

    mt = os.path.getmtime(p)

    return datetime.fromtimestamp(mt) + timedelta(seconds=_PA_DELAY_SECONDS)





def _mover_pendientes_listos():

    os.makedirs(COLA_PENDIENTES_DIR, exist_ok=True)

    os.makedirs(COLA_IN_DIR, exist_ok=True)

    os.makedirs(COLA_ARCHIVADOS_DIR, exist_ok=True)

    ahora = datetime.now()

    movidos = 0

    for nombre in os.listdir(COLA_PENDIENTES_DIR):

        if not nombre.lower().endswith('.json'):

            continue

        src = os.path.join(COLA_PENDIENTES_DIR, nombre)

        if not os.path.isfile(src):

            continue

        liberar_en = _resolver_liberar_por_archivo(nombre)

        if liberar_en is None or ahora < liberar_en:

            continue

        dst_in = os.path.join(COLA_IN_DIR, nombre)

        dst_arc = os.path.join(COLA_ARCHIVADOS_DIR, nombre)

        try:
            # Escribir contenido nuevo en in/ en vez de mover el archivo.
            # shutil.move() en OneDrive genera un evento "rename", no "file created",
            # y el trigger de Power Automate no se dispara. Escribiendo el archivo
            # desde cero OneDrive lo sincroniza como creacion nueva y PA sí reacciona.
            with open(src, 'rb') as f_src:
                contenido = f_src.read()
            with open(dst_in, 'wb') as f_dst:
                f_dst.write(contenido)
            os.remove(src)

            if not os.path.isfile(dst_arc):
                shutil.copy2(dst_in, dst_arc)

            movidos += 1
            print(f'[COLA-PA] Liberado a in/ (write-new): {nombre}')

        except Exception as e:

            print(f'[COLA-PA][WARN] No se pudo liberar {nombre}: {e}')

    return movidos



def _liberar_todo_pa_ahora():

    """Mueve todos los archivos de PENDIENTES a IN inmediatamente."""

    os.makedirs(COLA_PENDIENTES_DIR, exist_ok=True)

    os.makedirs(COLA_IN_DIR, exist_ok=True)

    movidos = 0

    for nombre in os.listdir(COLA_PENDIENTES_DIR):

        if not nombre.lower().endswith('.json'):

            continue

        src = os.path.join(COLA_PENDIENTES_DIR, nombre)

        dst_in = os.path.join(COLA_IN_DIR, nombre)

        try:

            shutil.move(src, dst_in)

            movidos += 1

        except:

            pass

    return movidos







def _smtp_procesar_in():
    """Fallback: si un archivo lleva más de 3 min en in/ sin que PA lo levante,
    lo enviamos por SMTP directamente y lo movemos a procesados/."""
    ESPERA = 3 * 60  # segundos antes de asumir que PA no va a procesar
    procesados_dir = os.path.join(COLA_DIR, 'procesados')
    errores_dir    = os.path.join(COLA_DIR, 'errores')
    os.makedirs(procesados_dir, exist_ok=True)
    os.makedirs(errores_dir, exist_ok=True)

    ahora = time.time()
    for nombre in list(os.listdir(COLA_IN_DIR)):
        if not nombre.lower().endswith('.json'):
            continue
        src = os.path.join(COLA_IN_DIR, nombre)
        if not os.path.isfile(src):
            continue
        if ahora - os.path.getmtime(src) < ESPERA:
            continue  # todavía joven, PA puede estar a punto de tomarlo

        try:
            with open(src, 'r', encoding='utf-8') as f:
                entradas = json.load(f)
            if not isinstance(entradas, list):
                entradas = [entradas]

            enviados, errores = 0, []
            for entrada in entradas:
                email  = (entrada.get('email_jefe') or entrada.get('email_destino_real') or '').strip()
                nombre_dest = entrada.get('nombre_jefe', '')
                asunto = entrada.get('asunto', 'Alertas Vacaciones USIL')
                html   = entrada.get('mensaje_html', '')
                if not email or '@' not in email or not html:
                    continue
                ok, err = _enviar_correo_smtp(email, nombre_dest, asunto, html)
                if ok:
                    enviados += 1
                else:
                    errores.append(f'{email}: {err}')
                    print(f'[SMTP-FB] ERROR enviando a {email}: {err}')

            destino = procesados_dir if not errores else errores_dir
            shutil.move(src, os.path.join(destino, nombre))
            print(f'[SMTP-FB] {nombre}: {enviados} enviados, {len(errores)} errores -> {os.path.basename(destino)}/')

        except Exception as e:
            print(f'[SMTP-FB] Error procesando {nombre}: {e}')


def _outlook_procesar_in():
    """Fallback automático: si un archivo lleva >15 min en in/ sin que PA lo procese,
    lo enviamos por Outlook COM y lo movemos a procesados/.
    No requiere SMTP AUTH — usa la sesión de Outlook ya abierta."""
    cfg = {}
    if _PA_CONFIG_PATH and os.path.isfile(_PA_CONFIG_PATH):
        try:
            with open(_PA_CONFIG_PATH, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
        except Exception:
            pass
    if not cfg:
        cfg = _PA_CONFIG

    fallback_enabled = cfg.get('outlook_fallback_enabled', True)
    if not fallback_enabled:
        return

    minutos_espera = cfg.get('outlook_fallback_wait_minutes', 15)
    if minutos_espera <= 0:
        return

    import subprocess as _sp
    ESPERA = minutos_espera * 60
    ahora = time.time()
    try:
        pendientes = [
            f for f in os.listdir(COLA_IN_DIR)
            if f.lower().endswith('.json')
            and os.path.isfile(os.path.join(COLA_IN_DIR, f))
            and ahora - os.path.getmtime(os.path.join(COLA_IN_DIR, f)) >= ESPERA
        ]
    except Exception:
        return
    if not pendientes:
        return
    print(f'[OUTLOOK-FB] {len(pendientes)} archivo(s) en in/ sin procesar (>3 min) — enviando via Outlook COM...')
    script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'enviar_cola_outlook.py')
    if not os.path.isfile(script):
        print('[OUTLOOK-FB] enviar_cola_outlook.py no encontrado, saltando fallback.')
        return
    try:
        result = _sp.run(
            [sys.executable, script, '--solo-in'],
            capture_output=True, text=True, timeout=120, encoding='utf-8', errors='replace'
        )
        for linea in (result.stdout or '').splitlines():
            print(f'[OUTLOOK-FB] {linea}')
        if result.stderr:
            print(f'[OUTLOOK-FB][STDERR] {result.stderr[:300]}')
    except _sp.TimeoutExpired:
        print('[OUTLOOK-FB] Timeout: Outlook tardó más de 2 min.')
    except Exception as e:
        print(f'[OUTLOOK-FB][ERR] {e}')


def _cola_pa_loop():

    import time

    print(f'[COLA-PA] Delay activo: {_PA_DELAY_SECONDS}s | pendientes: {COLA_PENDIENTES_DIR}')

    _ultimo_reintento_errores = 0  # timestamp de la última recuperación de errores/

    while True:

        try:
            _mover_pendientes_listos()
        except Exception as e:
            print(f'[COLA-PA][ERR] {e}')

        # Fallback Outlook COM: si PA no procesó el archivo en >3 min, lo enviamos nosotros.
        try:
            _outlook_procesar_in()
        except Exception as e:
            print(f'[COLA-PA][OUTLOOK-FB][ERR] {e}')

        # Recuperar errores/ cada 10 minutos (no solo al arrancar)
        ahora_loop = time.time()
        if ahora_loop - _ultimo_reintento_errores >= 600:
            try:
                errores_dir = os.path.join(COLA_DIR, 'errores')
                if os.path.isdir(errores_dir):
                    recuperados = 0
                    for f in os.listdir(errores_dir):
                        if f.lower().endswith('.json'):
                            src = os.path.join(errores_dir, f)
                            dst = os.path.join(COLA_IN_DIR, f)
                            if os.path.isfile(src) and not os.path.exists(dst):
                                shutil.move(src, dst)
                                recuperados += 1
                    if recuperados:
                        print(f'[COLA-PA] Recuperados {recuperados} archivo(s) de errores/ -> in/ para reintento')
            except Exception as e:
                print(f'[COLA-PA][RECOVER-ERR] {e}')
            _ultimo_reintento_errores = ahora_loop

        time.sleep(30)





def _cancelar_pendiente_pa(archivo, motivo='cancelado desde sistema'):

    nombre = os.path.basename(str(archivo or '').strip())

    if not nombre:

        return False, 'Archivo requerido', None

    src = os.path.join(COLA_PENDIENTES_DIR, nombre)

    if not os.path.isfile(src):

        return False, 'No se encontro en pendientes o ya fue liberado', None

    os.makedirs(COLA_CANCELADOS_DIR, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')

    dst = os.path.join(COLA_CANCELADOS_DIR, f'cancelado_{ts}_{nombre}')

    try:

        shutil.move(src, dst)

        info = {

            'archivo_origen': nombre,

            'archivo_cancelado': os.path.basename(dst),

            'cancelado_en': datetime.now().isoformat(),

            'motivo': motivo,

        }

        with open(dst + '.meta.json', 'w', encoding='utf-8') as f:

            json.dump(info, f, ensure_ascii=False, indent=2)

        return True, None, info

    except Exception as e:

        return False, str(e), None





def _firma_entradas_pa(entradas):

    firma = []

    if not isinstance(entradas, list):

        return tuple()

    for it in entradas:

        if not isinstance(it, dict):

            continue

        em = _norm_email(it.get('email_jefe', '') or it.get('email', ''))

        asu = str(it.get('asunto', '') or '').strip().upper()

        nom = str(it.get('nombre_jefe', '') or it.get('nombre', '') or '').strip().upper()

        firma.append(f'{em}|{nom}|{asu}')

    firma.sort()

    return tuple(firma)


def _snapshot_entradas_pa(entradas):

    if not isinstance(entradas, list):

        return ''

    try:

        return json.dumps(entradas, ensure_ascii=False, sort_keys=True, separators=(',', ':'))

    except Exception:

        return ''





def _buscar_duplicado_pendiente(entradas):

    os.makedirs(COLA_PENDIENTES_DIR, exist_ok=True)

    objetivo = _firma_entradas_pa(entradas)

    if not objetivo:

        return None

    for nombre in os.listdir(COLA_PENDIENTES_DIR):

        if not nombre.lower().endswith('.json'):

            continue

        p = os.path.join(COLA_PENDIENTES_DIR, nombre)

        if not os.path.isfile(p):

            continue

        try:

            with open(p, 'r', encoding='utf-8') as f:

                data = json.load(f)

            if _firma_entradas_pa(data) == objetivo:

                return nombre

        except Exception:

            continue

    return None





def _guardar_json_cola(entradas, prefijo='alerta_vacaciones'):

    os.makedirs(COLA_PENDIENTES_DIR, exist_ok=True)

    os.makedirs(COLA_PROCESANDO_DIR, exist_ok=True)

    os.makedirs(COLA_ARCHIVADOS_DIR, exist_ok=True)

    os.makedirs(COLA_CANCELADOS_DIR, exist_ok=True)



    dup = _buscar_duplicado_pendiente(entradas)

    if dup:

        p_dup = os.path.join(COLA_PENDIENTES_DIR, dup)

        previas = None

        try:

            with open(p_dup, 'r', encoding='utf-8') as f:

                previas = json.load(f)

        except Exception:

            previas = None

        if _snapshot_entradas_pa(previas) != _snapshot_entradas_pa(entradas):

            with open(p_dup, 'w', encoding='utf-8') as f:

                json.dump(entradas, f, ensure_ascii=False, indent=2)

            meta = _meta_envio_pa(dup, 'actualizado')

            meta['segundos_restantes'] = int(_PA_DELAY_SECONDS)

            return dup, os.path.join(COLA_ARCHIVADOS_DIR, dup), meta

        meta = _meta_envio_pa(dup, 'duplicado')

        liberar = _resolver_liberar_por_archivo(dup)

        if liberar:
            meta['segundos_restantes'] = max(0, int((liberar - datetime.now()).total_seconds()))

        return dup, os.path.join(COLA_ARCHIVADOS_DIR, dup), meta



    ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')

    fname = f'{prefijo}_{ts}_{uuid.uuid4().hex[:8]}.json'

    fpath = os.path.join(COLA_PENDIENTES_DIR, fname)

    with open(fpath, 'w', encoding='utf-8') as f:

        json.dump(entradas, f, ensure_ascii=False, indent=2)

    meta = _meta_envio_pa(fname, 'pendiente')

    meta['segundos_restantes'] = int(_PA_DELAY_SECONDS)

    return fname, os.path.join(COLA_ARCHIVADOS_DIR, fname), meta



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# CORS

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.after_request

def after_request(response):

    response.headers['Access-Control-Allow-Origin']  = '*'

    response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization'

    response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,DELETE,OPTIONS'

    return response



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# CACHÃ‰

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

_df_cache  = {'df': None, 'mtime': None, 'ruta': None, 'fecha': None}

_cache_lock = threading.Lock()

_obj_cache = {'df': None, 'mtime': None, 'ruta': None}

_OBJ_CACHE_VERSION = 'obj_hrbp_v3'

_filtros_cache = {'key': None, 'data': None}

_filtros_lock = threading.Lock()

_overrides_lock = threading.Lock()



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# FUNCIONES AUXILIARES

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _norm(t):

    if pd.isna(t) or t is None: return ''

    return unicodedata.normalize('NFKD', str(t)).encode('ascii', 'ignore').decode('ascii').strip()





def _nombre_cmp_key(nombre):

    """Normaliza nombre para comparaciones robustas (ignora comas, puntos y dobles espacios)."""

    base = _norm(nombre).upper()

    base = re.sub(r'[^A-Z0-9]+', ' ', base)

    return re.sub(r'\s+', ' ', base).strip()


def _primer_nombre_usil(nombre_completo):
    """'APELLIDOS, NOMBRES' → primer nombre title-cased. Ej: 'CHANG CHANG, GABRIEL ANDRES' → 'Gabriel'"""
    if not nombre_completo:
        return ''
    nombre = str(nombre_completo).strip()
    if ',' in nombre:
        nombres_part = nombre.split(',', 1)[1].strip()
        tokens = nombres_part.split()
    else:
        tokens = nombre.split()
    return tokens[0].capitalize() if tokens else nombre.title()


def _nombre_firma_usil(nombre_completo):
    """'APELLIDOS, NOMBRES' → 'Primer Nombre Primer Apellido'. Ej: 'JARA ORTIZ, CARLOS HUMBERTO' → 'Carlos Jara'"""
    if not nombre_completo:
        return ''
    nombre = str(nombre_completo).strip()
    if ',' in nombre:
        partes = nombre.split(',', 1)
        apellidos_tokens = partes[0].strip().split()
        nombres_tokens = partes[1].strip().split()
        primer_apellido = apellidos_tokens[0].capitalize() if apellidos_tokens else ''
        primer_nombre = nombres_tokens[0].capitalize() if nombres_tokens else ''
        if primer_nombre and primer_apellido:
            return f'{primer_nombre} {primer_apellido}'
        return (primer_nombre or primer_apellido or nombre).title()
    else:
        tokens = nombre.split()
        if len(tokens) >= 2:
            return f'{tokens[0].capitalize()} {tokens[1].capitalize()}'
        return nombre.title()





_HRBP_TYC_DEFAULT = str(

    _PA_CONFIG.get('vacaciones_hrbp_tyc_nombre', 'JARA ORTIZ, CARLOS HUMBERTO')

    or 'JARA ORTIZ, CARLOS HUMBERTO'

).strip()



_SUPERVISOR_GERENCIAS_EXCLUIDAS = {

    _nombre_cmp_key('SANCHEZ VASQUEZ MARIELA PATRICIA'): {'GERENCIA DE TALENTO Y CULTURA'},

}





def _es_talento_cultura(*vals):

    txt = _norm(' '.join(_safe(v) for v in vals)).upper()

    return any(k in txt for k in ('TALENTO', 'CULTURA', 'GESTION HUMANA', 'ATRACCION'))





def _resolver_hrbp_tyc(hrbp_actual='', area='', departamento='', division='', puesto=''):

    # Regla de negocio confirmada: en Talento y Cultura el BP es Carlos Jara.

    if _es_talento_cultura(area, departamento, division, puesto):

        return _HRBP_TYC_DEFAULT

    hrbp_txt = _safe(hrbp_actual)

    if 'GABRIEL CHANG' in _norm(hrbp_txt).upper():

        return ''

    return hrbp_txt



def _filtrar_gerencias_supervisor(nombre_supervisor, gerencias):

    excluidas = _SUPERVISOR_GERENCIAS_EXCLUIDAS.get(_nombre_cmp_key(nombre_supervisor), set())

    if not excluidas:

        return sorted(list(gerencias or []), key=lambda x: x.upper())

    return sorted([g for g in (gerencias or []) if g not in excluidas], key=lambda x: x.upper())



def _col(df, *keys):

    for k in keys:

        found = [c for c in df.columns if k.lower() in c.lower()]

        if found: return found[0]

    return None



def _safe(v):

    if pd.isna(v) or v is None: return ''

    return str(v).strip()



def _norm_email(v):

    if not v:

        return ''

    return str(v).strip().lower()



def _norm_id(v):

    if v is None or pd.isna(v):

        return ''

    txt = str(v).strip()

    if not txt or txt.lower() == 'nan':

        return ''

    if txt.endswith('.0') and txt[:-2].isdigit():

        txt = txt[:-2]

    return txt



def _id_key(v):

    txt = _norm_id(v)

    if not txt:

        return ''

    txt2 = txt.lstrip('0')

    return txt2 or '0'



def _to_int0(v):

    try:

        if v is None or pd.isna(v):

            return 0

        return int(float(v))

    except Exception:

        return 0



def _to_float0(v):

    try:

        if v is None or pd.isna(v):

            return 0.0

        return float(v)

    except Exception:

        return 0.0



def _col_correo(df):

    return _col(

        df,

        'Email Institucional',

        'Email Trabajo',

        'Correo Institucional',

        'Correo Corporativo',

        'Correo Electronico',

        'Correo',

        'Email',

        'E-mail',

        'Mail'

    )





def _norm_name_key(v):

    txt = _safe(v)

    if not txt:

        return ''

    return _norm(txt).upper().replace(',', ' ').replace('.', ' ').replace('  ', ' ').strip()



def _detectar_llave_union(df_obj, df_m):

    """Detecta la llave de unión más confiable entre objetivo y maestro."""

    if df_obj is None or df_m is None or len(df_obj) == 0 or len(df_m) == 0:

        return {'llave': None, 'col_obj': None, 'col_maestro': None, 'overlap': 0}

    candidatos = []

    c_obj_mat = _col(df_obj, 'Matricula')
    c_m_mat = _col(df_m, 'Matricula')
    if c_obj_mat and c_m_mat:
        candidatos.append(('matricula', c_obj_mat, c_m_mat, _id_key))

    c_obj_dni = _col(df_obj, 'DNI', 'Numero Documento', 'Documento')
    c_m_dni = _col(df_m, 'DNI', 'Numero Documento', 'Documento')
    if c_obj_dni and c_m_dni:
        candidatos.append(('dni', c_obj_dni, c_m_dni, _id_key))

    c_obj_mail = _col_correo(df_obj)
    c_m_mail = _col_correo(df_m)
    if c_obj_mail and c_m_mail:
        candidatos.append(('correo', c_obj_mail, c_m_mail, _norm_email))

    c_obj_nom = _col(df_obj, 'Apellidos y Nombres', 'Nombre Completo', 'Colaborador', 'Nombre')
    c_m_nom = _col(df_m, 'Apellidos y Nombres', 'Nombre Completo', 'Colaborador', 'Nombre')
    if c_obj_nom and c_m_nom:
        candidatos.append(('nombre', c_obj_nom, c_m_nom, _norm_name_key))

    best = {'llave': None, 'col_obj': None, 'col_maestro': None, 'overlap': 0}

    for llave, c_obj, c_m, norm_fn in candidatos:
        try:
            s_obj = {norm_fn(v) for v in df_obj[c_obj].dropna().tolist() if norm_fn(v)}
            s_m = {norm_fn(v) for v in df_m[c_m].dropna().tolist() if norm_fn(v)}
            overlap = len(s_obj & s_m)
            if overlap > best['overlap']:
                best = {'llave': llave, 'col_obj': c_obj, 'col_maestro': c_m, 'overlap': overlap}
        except Exception:
            continue

    return best



def _invalidate_data_caches():

    _df_cache.update({'df': None, 'mtime': None, 'ruta': None, 'fecha': None})

    _obj_cache.update({'df': None, 'mtime': None, 'ruta': None})

    _filtros_cache.update({'key': None, 'data': None})

    _MAESTRO_CONTACTOS_CACHE.update({'ruta': None, 'mtime': None, 'df': None})

    _MAESTRO_UNIV_CACHE.update({'ruta': None, 'mtime': None, 'df': None})





def _leer_colaboradores_overrides():

    if not os.path.isfile(COLAB_EDIT_FILE):

        return {}

    try:

        with open(COLAB_EDIT_FILE, 'r', encoding='utf-8') as f:

            data = json.load(f)

        if isinstance(data, dict):

            return data

    except Exception as e:

        print(f'[OVERRIDES] Error leyendo archivo: {e}')

    return {}





def _guardar_colaboradores_overrides(data):

    os.makedirs(os.path.dirname(COLAB_EDIT_FILE), exist_ok=True)

    tmp = COLAB_EDIT_FILE + '.tmp'

    with open(tmp, 'w', encoding='utf-8') as f:

        json.dump(data, f, indent=2, ensure_ascii=False)

    os.replace(tmp, COLAB_EDIT_FILE)





def _resolver_test_email_destino():

    return str(

        _PA_CONFIG.get('vacaciones_test_email', '')

        or _TEST_EMAIL_DESTINO

        or ''

    ).strip().lower()





def _aplicar_overrides_colaboradores(df, modo='datos'):

    if df is None or df.empty:

        return df

    overrides = _leer_colaboradores_overrides()

    if not overrides:

        return df



    c_mat = _col(df, 'Matricula')

    if not c_mat:

        return df



    maps = {

        'nombre': _col(df, 'Apellidos y Nombres', 'Nombre Colaborador', 'Nombre'),

        'email': _col(df, 'Correo', 'Correo Institucional', 'Email Trabajo', 'Email'),

        'departamento': _col(df, 'Nombre Departamento', 'Departamento', 'Division', 'Subgerencia'),

        'area': _col(df, 'Nombre Area', 'Area', 'Seccion'),

        'puesto': _col(df, 'Nombre Puesto', 'Puesto', 'Cargo'),

        'supervisor': _col(df, 'Jefe_Directo', 'Jefe Directo', 'Supervisor', 'Jefe', 'Business Partner', 'HRBP'),

    }



    if modo == 'objetivos' and not maps.get('supervisor'):

        df['Supervisor Override'] = ''

        maps['supervisor'] = 'Supervisor Override'



    df = df.copy()

    for idx, row in df.iterrows():

        key = _id_key(row.get(c_mat, ''))

        if not key:

            continue

        ov = overrides.get(key)

        if not isinstance(ov, dict):

            continue

        for field, col in maps.items():

            if not col:

                continue

            if field in ov and ov.get(field) not in (None, ''):

                df.at[idx, col] = str(ov.get(field, '')).strip()

    return df





def _overrides_a_items_busqueda():

    """Convierte overrides a lista de colaboradores buscables (incluye registros manuales no presentes en Excel)."""

    out = []

    raw = _leer_colaboradores_overrides()

    if not isinstance(raw, dict):

        return out



    for mk, ov in raw.items():

        if not isinstance(ov, dict):

            continue

        mat = _id_key(mk or ov.get('matricula', ''))

        nombre = str(ov.get('nombre', '') or '').strip()

        if not mat or not nombre:

            continue

        out.append({

            'matricula': mat,

            'nombre': nombre,

            'email': _norm_email(ov.get('email', '') or ''),

            'departamento': str(ov.get('departamento', '') or '').strip(),

            'area': str(ov.get('area', '') or '').strip(),

            'puesto': str(ov.get('puesto', '') or '').strip(),

            'supervisor': str(ov.get('supervisor', '') or '').strip(),

        })

    return out



def _emails_autorizados():

    df, err = cargar_datos()

    if err or df is None:

        return set(), err or 'No se pudo cargar la tabla maestra'



    cc = _col_correo(df)

    if not cc:

        return set(), 'No se encontro una columna de correo en la tabla maestra'



    emails = set()

    for v in df[cc].dropna().tolist():

        em = _norm_email(v)

        if em and '@' in em:

            emails.add(em)



    for em in _PA_CONFIG.get('login_correos_extra', []) or []:

        emn = _norm_email(em)

        if emn and '@' in emn:

            emails.add(emn)

    emails.add(_norm_email(_ADMIN_EMAIL))



    return emails, None



def _usuario_session():

    return {

        'email': session.get('usuario_email', ''),

        'nombre': session.get('usuario_nombre', ''),

    }



def _es_llamada_interna():

    token = request.headers.get('X-PA-Internal-Token', '')

    return bool(token) and token == _INTERNAL_API_TOKEN





def _ncols_fast(ruta):

    import zipfile, re as _r

    try:

        with zipfile.ZipFile(ruta, 'r') as z:

            sheet = next((n for n in z.namelist() if _r.match(r'xl/worksheets/sheet\d+\.xml', n)), None)

            if not sheet: return 0

            with z.open(sheet) as f:

                chunk = f.read(65536).decode('utf-8', errors='ignore')  # 64 KB â€” cubre headers profundos

            m = _r.search(r'dimension\s+ref="[A-Z]+\d+:([A-Z]+)\d+"', chunk)

            if m:

                s, n = m.group(1), 0

                for c in s: n = n*26 + (ord(c)-ord('A')+1)

                return n

            return len(_r.findall(r'<c\s+r="[A-Z]+1"', chunk))

    except: return 0





def _header_keywords(modo='datos'):

    if str(modo).lower() == 'objetivos':

        primary = ['matric', 'apellido', 'nombre', 'objetivo', 'meta', 'gozado',

                   'cumpl', 'pend', 'trunc', 'venc', 'restante', 'registro']

        secondary = ['hrbp', 'area', 'depart', 'activo', 'cesado', 'puesto', 'seccion']

    else:

        primary = ['matric', 'apellido', 'nombre', 'ingreso', 'area', 'cargo',

                   'situac', 'division', 'puesto', 'sexo', 'regimen']

        secondary = ['email', 'supervisor', 'jefe', 'depart', 'activo', 'seccion']

    return primary, secondary



def _score_header_row(vals, modo='datos'):

    primary, secondary = _header_keywords(modo)

    primary_hits = sum(1 for kw in primary if any(kw in v for v in vals))

    secondary_hits = sum(1 for kw in secondary if any(kw in v for v in vals))

    non_empty = sum(1 for v in vals if str(v).strip() not in {'', 'nan'})

    score = primary_hits * 12 + secondary_hits * 4 + min(non_empty, 20)

    if str(modo).lower() == 'objetivos':

        has_identity = any(any(tok in v for v in vals) for tok in ('matric', 'apellido', 'nombre'))

        has_vac = any(any(tok in v for v in vals) for tok in ('objetivo', 'meta', 'gozado', 'cumpl', 'pend', 'trunc', 'venc', 'restante', 'registro'))

        if has_identity:

            score += 25

        if has_vac:

            score += 20

        if not has_identity:

            score -= 20

        if not has_vac:

            score -= 10

    return score



def _resolver_hoja_excel(ruta, modo='datos'):

    """Selecciona la hoja mas probable para datos de vacaciones/objetivos."""

    try:

        xls = pd.ExcelFile(ruta)

    except Exception:

        return 0



    if not xls.sheet_names:

        return 0

    if len(xls.sheet_names) == 1:

        return xls.sheet_names[0]



    # Regla explicita para este proyecto: preferir "Base Completa" o "BASE GENERAL" si existe.

    for sn in xls.sheet_names:

        if str(sn).strip().lower() in ('base completa', 'base general'):

            return sn





    if str(modo).lower() == 'objetivos':

        preferred = ['base completa', 'base general', 'con vacaciones', 'objetivo', 'detalle', 'registro']

    else:

        preferred = ['base completa', 'base general', 'con vacaciones']



    best_sheet = xls.sheet_names[0]

    best_score = -1

    for sn in xls.sheet_names:

        score = 0

        sn_low = str(sn).strip().lower()

        if any(p in sn_low for p in preferred):

            score += 40

        try:

            df_raw = _safe_read_excel(ruta, sheet_name=sn, nrows=20, header=None, dtype=str)

            for i in range(len(df_raw)):

                vals = df_raw.iloc[i].fillna('').astype(str).str.lower().tolist()

                score = max(score, _score_header_row(vals, modo))

        except Exception:

            pass

        if score > best_score:

            best_score = score

            best_sheet = sn



    return best_sheet



def _resolver_hoja_excel_from_xls(xls, modo='datos'):

    """Optimizado: Usa un objeto pd.ExcelFile ya abierto para detectar la mejor hoja."""

    sheets = xls.sheet_names

    if not sheets: return 0

    if len(sheets) == 1: return sheets[0]

    # Regla explícita: preferir "Base Completa" o "BASE GENERAL" si existe (igual que _resolver_hoja_excel)
    for sn in sheets:
        if str(sn).strip().lower() in ('base completa', 'base general'):
            print(f'[HOJA] Detectada hoja prioritaria: {sn}')
            return sn

    preferred = ['DATA', 'REPORTE', 'MAESTRO', 'ACTIVOS'] if modo == 'datos' else ['OBJETIVO', 'META', 'VACACIONES', 'BASE', 'DETALLE']

    best_sheet = sheets[0]

    best_score = -1

    for sn in sheets:

        score = 0

        sn_low = str(sn).strip().lower()

        if any(p in sn_low for p in preferred): score += 40

        try:

            df_raw = pd.read_excel(xls, sheet_name=sn, nrows=10, header=None, dtype=str)

            for i in range(len(df_raw)):

                vals = df_raw.iloc[i].fillna('').astype(str).str.lower().tolist()

                score = max(score, _score_header_row(vals, modo))

        except Exception: pass

        if score > best_score:

            best_score = score

            best_sheet = sn

    return best_sheet





def _detectar_header_row_from_df(df_raw, modo='datos'):

    """Detecta en qué fila están los headers reales desde un dataframe parcial."""

    best_i = 0

    best_hits = -1

    for i in range(len(df_raw)):

        vals = df_raw.iloc[i].fillna('').astype(str).str.lower().values

        score = _score_header_row(vals, modo)

        if score > best_hits:

            best_hits = score

            best_i = i

    return best_i if best_hits > 0 else 0



def _detectar_header_row(ruta, sheet_name=0, modo='datos'):

    """Detecta en quÃ© fila estÃ¡n los headers reales.

    Requiere â‰¥3 keywords de RRHH para evitar confundir sub-encabezados con el header real.

    Soporta formato nuevo (header en fila 10) y viejo (header en fila 0).

    """

    try:

        df_raw = _safe_read_excel(ruta, sheet_name=sheet_name, nrows=25, header=None, dtype=str)

        best_i = 0

        best_hits = -1

        for i in range(len(df_raw)):

            vals = df_raw.iloc[i].fillna('').astype(str).str.lower().values

            score = _score_header_row(vals, modo)

            if score > best_hits:

                best_hits = score

                best_i = i

        if best_hits > 0:

            return best_i

    except:

        pass

    return 0  # fallback



def _fecha_key_archivo(path):

    base = os.path.basename(path)



    # PersonalMaestroReporte_MM_DD_YYYY[ HH_MM_SS].xlsx

    m = re.search(r'PersonalMaestroReporte_(\d{2})_(\d{2})_(\d{4})(?:\s+(\d{2})_(\d{2})_(\d{2}))?', base, re.IGNORECASE)

    if m:

        mm, dd, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))

        hh = int(m.group(4) or 0)

        mi = int(m.group(5) or 0)

        ss = int(m.group(6) or 0)

        return (yyyy, mm, dd, hh, mi, ss, base.upper())



    # YYYY-MM-DD o YYYY_MM_DD

    m = re.search(r'(\d{4})[-_](\d{2})[-_](\d{2})', base)

    if m:

        yyyy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))

        return (yyyy, mm, dd, 0, 0, 0, base.upper())



    # DD-MM-YYYY o DD_MM_YYYY

    m = re.search(r'(\d{2})[-_](\d{2})[-_](\d{4})', base)

    if m:

        dd, mm, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))

        return (yyyy, mm, dd, 0, 0, 0, base.upper())



    # "Primer/Segundo/Tercer/Cuarto Trimestre YYYY"

    q = re.search(r'(primer|segundo|tercer|cuarto)\s+trimestre\s*(\d{4})', base, re.IGNORECASE)

    if q:

        q_map = {'primer': 1, 'segundo': 2, 'tercer': 3, 'cuarto': 4}

        qq = q_map.get((q.group(1) or '').lower(), 1)

        yyyy = int(q.group(2))

        mm = qq * 3

        return (yyyy, mm, 31, 0, 0, 0, base.upper())



    return (0, 0, 0, 0, 0, 0, base.upper())





def _trimestre_desde_fecha(ref=None):

    d = ref or date.today()

    q = ((d.month - 1) // 3) + 1

    return {

        'q': q,

        'anio': d.year,

        'label': f'Q{q}-{d.year}',

        'descripcion': f'Trimestre {q} {d.year}'

    }





def _normalizar_trimestre_txt(txt):

    raw = _norm(txt).upper().replace(' ', '')

    if not raw:

        return None

    m = re.search(r'Q([1-4])[-_/]?(20\d{2})', raw)

    if m:

        q, y = int(m.group(1)), int(m.group(2))

        return {'q': q, 'anio': y, 'label': f'Q{q}-{y}', 'descripcion': f'Trimestre {q} {y}'}

    m = re.search(r'([1-4])TRIMESTRE[-_/]?(20\d{2})', raw)

    if m:

        q, y = int(m.group(1)), int(m.group(2))

        return {'q': q, 'anio': y, 'label': f'Q{q}-{y}', 'descripcion': f'Trimestre {q} {y}'}

    return None





def _extraer_trimestre_de_nombre(nombre_archivo):

    base = os.path.basename(str(nombre_archivo or ''))

    if not base:

        return None



    hit = _normalizar_trimestre_txt(base)

    if hit:

        return hit



    q_txt = re.search(r'(primer|segundo|tercer|cuarto)\s+trimestre\s*(20\d{2})', base, re.IGNORECASE)

    if q_txt:

        q_map = {'primer': 1, 'segundo': 2, 'tercer': 3, 'cuarto': 4}

        q = q_map.get((q_txt.group(1) or '').lower(), 1)

        y = int(q_txt.group(2))

        return {'q': q, 'anio': y, 'label': f'Q{q}-{y}', 'descripcion': f'Trimestre {q} {y}'}



    return None





def _trimestre_vigente_info(fuente_archivo=''):

    manual_txt = str(_PA_CONFIG.get('vacaciones_trimestre_actual', '') or '').strip()

    manual = _normalizar_trimestre_txt(manual_txt)

    if manual:

        manual['modo'] = 'manual'

        manual['origen'] = 'pa_config'

        return manual



    auto = _extraer_trimestre_de_nombre(fuente_archivo or '')

    if auto:

        auto['modo'] = 'auto'

        auto['origen'] = 'archivo'

        return auto



    d = _trimestre_desde_fecha()

    d['modo'] = 'auto'

    d['origen'] = 'fecha'

    return d





_CANDIDATOS_CACHE_DATA = None
_CANDIDATOS_CACHE_TIME = 0

def _candidatos_vacaciones():
    global _CANDIDATOS_CACHE_DATA, _CANDIDATOS_CACHE_TIME
    import time
    if _CANDIDATOS_CACHE_DATA is not None and (time.time() - _CANDIDATOS_CACHE_TIME < 60):
        return _CANDIDATOS_CACHE_DATA

    items = []

    vistos = set()

    datas_root = os.path.abspath(DATAS_DIR).lower()

    sensible_root = os.path.abspath(DATA_SENSIBLE_DIR).lower()



    def _valida_fuente_vacaciones(path):

        """Doble verificacion para usar un Excel de vacaciones."""

        ap = os.path.abspath(path)

        nm = _norm(os.path.basename(ap)).upper()



        in_datas = ap.lower().startswith(datas_root + os.sep.lower())

        in_sensible = ap.lower().startswith(sensible_root + os.sep.lower())

        

        if not (in_datas or in_sensible) or not ap.lower().endswith('.xlsx'):

            return False

        # Archivos con nombre exacto del patron del proyecto: confiar sin validacion de contenido
        import fnmatch as _fnmatch
        _nm_base = os.path.basename(ap)
        if _fnmatch.fnmatch(_nm_base, 'Reporte Vacaciones Objetivo*.xlsx') or \
           _fnmatch.fnmatch(_nm_base, 'Vacaciones_Talento_Cultura*.xlsx') or \
           _fnmatch.fnmatch(_nm_base, 'Copia de Reporte Vacaciones Objetivo*.xlsx'):
            return True

        # Intentar validar en TODAS las hojas si la primera falla

        tmp_path = None

        try:

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:

                tmp_path = tmp.name

            shutil.copy2(ap, tmp_path)

            xls = pd.ExcelFile(tmp_path)

            for sn in xls.sheet_names:

                try:

                    sample = pd.read_excel(xls, sheet_name=sn, header=None, dtype=str, nrows=50)

                    joined = ' | '.join(_norm(v).upper() for v in sample.fillna('').astype(str).values.ravel())

                    

                    check_1 = ('MATRICULA' in joined)

                    check_2 = (('OBJETIVO' in joined) or ('CUMPLIMIENTO' in joined) or ('APELLIDOS' in joined) or ('VENCIDA' in joined))

                    

                    if check_1 and check_2:

                        try: xls.close()

                        except: pass

                        return True

                except Exception:

                    continue

        except Exception:

            return False

        finally:

            if tmp_path and os.path.exists(tmp_path):

                try: os.unlink(tmp_path)

                except: pass





        # Evitar explícitamente artefactos de ejemplo/demo.

        if 'EJEMPLO' in nm or 'DEMO' in nm:

            return False

        return False





    def _add(path, origen):

        if not path:

            return

        ap = os.path.abspath(path)

        if ap in vistos or not os.path.isfile(ap):

            return

        if not _valida_fuente_vacaciones(ap):

            return

        vistos.add(ap)

        items.append({

            'path': ap,

            'archivo': os.path.basename(ap),

            'origen': origen,

            'fecha_key': _fecha_key_archivo(ap),

        })



    # Prioridad 0: copia operativa cargada manualmente por Talento y Cultura.

    for path in glob.glob(os.path.join(DATAS_DIR, 'Copia de Reporte Vacaciones Objetivo*.xlsx')):

        _add(path, 'prioridad_copia')



    # Prioridad 1: archivo principal de produccion (Talento y Cultura)

    _add(VACACIONES_DATA_FILE, 'principal')



    # Prioridad 2: candidatos dentro de DATAS y DATA SENSIBLE

    for pat in _PATRONES_VACACIONES_DATAS:

        for path in glob.glob(os.path.join(DATAS_DIR, pat)):

            _add(path, 'datas')

        for path in glob.glob(os.path.join(DATA_SENSIBLE_DIR, pat)):

            _add(path, 'sensible')





    # Orden: principal primero; resto por fecha en nombre

    def _sort_key(x):

        origen_rank = 3 if x['origen'] == 'prioridad_copia' else (2 if x['origen'] == 'principal' else (1 if x['origen'] == 'datas' else 0))

        return (origen_rank, *x['fecha_key'])



    items.sort(key=_sort_key, reverse=True)

    _CANDIDATOS_CACHE_DATA = items
    _CANDIDATOS_CACHE_TIME = time.time()
    return items





def _archivo_data_prueba_unico():

    candidatos = _candidatos_vacaciones()

    if not candidatos:

        return None

    tri_manual = _normalizar_trimestre_txt(str(_PA_CONFIG.get('vacaciones_trimestre_actual', '') or '').strip())
    if tri_manual:
        tri_label = tri_manual.get('label', '')
        for item in candidatos:
            tri_item = (_extraer_trimestre_de_nombre(item.get('archivo', '')) or {}).get('label', '')
            if tri_item == tri_label:
                return item['path']
    return candidatos[0]['path']





_MAESTRO_CONTACTOS_CACHE = {'ruta': None, 'mtime': None, 'df': None}





def _maestro_path_candidates():

    """Retorna candidatos del archivo PersonalMaestroReporte ordenados por fecha en nombre."""
    paths = []
    vistos = set()

    def _add(path):
        if not path:
            return
        ap = os.path.abspath(path)
        if not os.path.isfile(ap):
            return
        if ap in vistos:
            return
        vistos.add(ap)
        paths.append(ap)

    _add(_MAESTRO_FULL_PATH)
    for p in glob.glob(os.path.join(DATAS_DIR, 'PersonalMaestroReporte*.xlsx')):
        _add(p)
    for p in glob.glob(os.path.join(DATA_SENSIBLE_DIR, 'PersonalMaestroReporte*.xlsx')):
        _add(p)

    paths.sort(key=_maestro_fecha_key, reverse=True)
    return paths





def _maestro_fecha_key(path):

    return _fecha_key_archivo(path)





def _resumen_fuente_datos():

    candidatos = _candidatos_vacaciones()

    actual_path = _archivo_data_prueba_unico()
    actual = next((item for item in candidatos if item.get('path') == actual_path), None) if actual_path else None
    if actual is None:
        actual = candidatos[0] if candidatos else None

    trim_arch = _extraer_trimestre_de_nombre((actual or {}).get('archivo', '')) if actual else None

    trim_vig = _trimestre_vigente_info((actual or {}).get('archivo', ''))

    return {

        'archivo': (actual or {}).get('archivo', ''),

        'origen': (actual or {}).get('origen', ''),

        'ruta': (actual or {}).get('path', ''),

        'trimestre_archivo': (trim_arch or {}).get('label', ''),

        'trimestre_vigente': (trim_vig or {}).get('label', ''),

        'trimestre_modo': (trim_vig or {}).get('modo', 'auto'),

        'total_candidatos': len(candidatos),

        'candidatos': [

            {

                'archivo': x.get('archivo', ''),

                'origen': x.get('origen', ''),

                'ruta': x.get('path', ''),

                'trimestre': (_extraer_trimestre_de_nombre(x.get('archivo', '')) or {}).get('label', ''),

            }

            for x in candidatos[:10]

        ],

        'trimestre_control': {

            'trimestre': (trim_vig or {}).get('label', ''),

            'modo': (trim_vig or {}).get('modo', 'auto'),

            'origen': (trim_vig or {}).get('origen', 'fecha')

        }

    }





def _cargar_maestro_contactos():

    paths = [p for p in _maestro_path_candidates() if p and os.path.isfile(p)]

    if not paths:

        return None, 'No se encontro PersonalMaestroReporte en DATAS ni DATA SENSIBLE'



    ordered = sorted(paths, key=_maestro_fecha_key, reverse=True)



    ultimo_error = None

    for ruta in ordered:

        try:

            mt = os.path.getmtime(ruta)

        except Exception:

            mt = None



        if (

            _MAESTRO_CONTACTOS_CACHE.get('df') is not None

            and _MAESTRO_CONTACTOS_CACHE.get('ruta') == ruta

            and _MAESTRO_CONTACTOS_CACHE.get('mtime') == mt

        ):

            return _MAESTRO_CONTACTOS_CACHE.get('df'), None



        try:

            hoja = _resolver_hoja_excel(ruta, modo='datos')

            hdr = _detectar_header_row(ruta, sheet_name=hoja)

            df = _safe_read_excel(ruta, sheet_name=hoja, header=hdr, dtype=str)

            df.columns = [_norm(c) for c in df.columns]

            c_mail = _col_correo(df)

            c_nom = _col(df, 'Apellido Paterno', 'Apellidos y Nombres', 'Nombre Completo', 'Colaborador')

            if not c_mail or not c_nom:

                ultimo_error = f'Sin columnas requeridas en {os.path.basename(ruta)}'

                continue



            _MAESTRO_CONTACTOS_CACHE.update({'ruta': ruta, 'mtime': mt, 'df': df})

            print(f'[CONTACTOS] Maestro para autocompletar: {os.path.basename(ruta)}')

            return df, None

        except Exception as e:

            ultimo_error = str(e)

            continue



    return None, f'No se pudo cargar un maestro con nombre+correo ({ultimo_error or "sin detalle"})'





_MAESTRO_UNIV_CACHE = {'ruta': None, 'mtime': None, 'df': None}
_MAESTRO_UNIV_LOCK = threading.Lock()


def _cargar_maestro_universo():

    """Carga la tabla maestra completa (PersonalMaestroReporte) desde DATAS o DATA SENSIBLE.

    NO filtra por TyC para que sea el universo completo de personas y poder encontrar jefes.

    Cachea en memoria por (ruta, mtime). Devuelve SIEMPRE una copia para que los callers
    que mutan el DataFrame no corrompan el cache compartido."""

    candidates = _maestro_path_candidates()
    ruta = candidates[0] if candidates else ''

    if not ruta or not os.path.isfile(ruta):

        return None, 'No se encontro PersonalMaestroReporte en DATAS ni DATA SENSIBLE'

    _mt = os.path.getmtime(ruta)
    with _MAESTRO_UNIV_LOCK:
        if (_MAESTRO_UNIV_CACHE['df'] is not None and
                _MAESTRO_UNIV_CACHE['ruta'] == ruta and
                _MAESTRO_UNIV_CACHE['mtime'] == _mt):
            return _MAESTRO_UNIV_CACHE['df'].copy(), None

    try:

        hdr = _detectar_header_row(ruta)

        df = _safe_read_excel(ruta, header=hdr, dtype=str)

        df.columns = [_norm(c) for c in df.columns]

        # Eliminar los primeros 3 registros del reporte SAP (filas fantasma/de sistema)
        if len(df) > 3:
            df = df.iloc[3:].reset_index(drop=True)
            print(f'[MAESTRO-UNIVERSO] Primeros 3 registros eliminados (filas fantasma SAP)')

        # IMPORTANTE: NO filtrar por TyC — la tabla maestra es el universo completo
        # necesitamos poder encontrar jefes de cualquier departamento.

        # Normalizar matricula: convertir float a int string con zfill(10)
        c_mat = _col(df, 'Matricula')

        if c_mat:
            # Filtrar filas donde matricula sea vacía o nula antes de normalizar
            df = df[df[c_mat].notna() & (df[c_mat].str.strip() != '') & (df[c_mat].str.strip() != 'nan')].copy()
            df.reset_index(drop=True, inplace=True)

            df[c_mat] = df[c_mat].apply(lambda v: str(int(float(str(v).strip()))).zfill(10)

                                         if str(v).strip() not in ('', 'nan', 'None') and

                                         str(v).strip().replace('.','').replace('-','').isdigit()

                                         else _norm_id(str(v)))

        # Normalizar Matricula Supervisor
        c_sup_mat = _col(df, 'Matricula Supervisor')

        if c_sup_mat:

            df[c_sup_mat] = df[c_sup_mat].apply(lambda v: str(int(float(str(v).strip()))).zfill(10)

                                                  if str(v).strip() not in ('', 'nan', 'None') and

                                                  str(v).strip().replace('.','').replace('-','').isdigit()

                                                  else _norm_id(str(v)))

        # Columna AD = Email Trabajo — fuente principal del correo del colaborador
        # Si no existe 'Correo', lo creamos como alias de Email Trabajo
        c_mail_trabajo = _col(df, 'Email Trabajo')
        if c_mail_trabajo:
            df['email trabajo'] = df[c_mail_trabajo].apply(lambda v: _norm_email(str(v)) if v and str(v).strip() not in ('', 'nan', 'None') else '')
            if not _col(df, 'Correo'):
                df['Correo'] = df['email trabajo']

        c_mail = _col(df, 'Email Trabajo')

        if c_mail and not _col(df, 'Correo'):

            df['Correo'] = df[c_mail]

        # Construir columna combinada de nombre
        c_ap = _col(df, 'Apellido Paterno')

        c_am = _col(df, 'Apellido Materno')

        c_nom = _col(df, 'Nombre')

        if c_ap and c_am and c_nom and not _col(df, 'Apellidos y Nombres'):

            df['Apellidos y Nombres'] = (

                df[c_ap].fillna('') + ' ' + df[c_am].fillna('') + ' ' + df[c_nom].fillna('')

            ).str.strip()

        print(f'[MAESTRO-UNIVERSO] {os.path.basename(ruta)} | {len(df)} personas (universo completo)')

        with _MAESTRO_UNIV_LOCK:
            _MAESTRO_UNIV_CACHE.update({'ruta': ruta, 'mtime': _mt, 'df': df})

        return df.copy(), None

    except Exception as e:

        return None, str(e)





_MAESTRO_JEFES_CACHE = {'ruta': None, 'mtime': None, 'data': None}



@medir_tiempo

def _cargar_tabla_maestra_jefes():

    """Optimizado: Carga el mapa de subordinado -> jefe y matricula -> info desde la maestra."""

    candidates = _maestro_path_candidates()
    ruta = candidates[0] if candidates else ''

    if not ruta or not os.path.isfile(ruta):

        return {}, {}, 'No se encontro PersonalMaestroReporte'

    

    try:

        mt = os.path.getmtime(ruta)

        if (_MAESTRO_JEFES_CACHE['ruta'] == ruta and

                _MAESTRO_JEFES_CACHE['mtime'] == mt and

                _MAESTRO_JEFES_CACHE['data'] is not None):

            d = _MAESTRO_JEFES_CACHE['data']

            return d['colab_a_jefe'], d['mat_a_info'], None

    except Exception: mt = None



    df_m, err = _cargar_maestro_universo()

    if err or df_m is None:

        return {}, {}, err or 'Error al cargar maestro'

    

    c_mat = _col(df_m, 'Matricula')

    c_sup_nom = _col(df_m, 'Supervisor', 'Jefe Directo', 'Jefe_Directo')

    c_sup_mat = _col(df_m, 'Matricula Supervisor')

    c_mail = _col_correo(df_m)

    c_hrbp = _col(df_m, 'HRBP', 'Business Partner', 'BP')

    c_pue = _col(df_m, 'Puesto', 'Nombre Puesto')

    c_area = _col(df_m, 'Area', 'Nombre Area')

    c_dep = _col(df_m, 'Departamento', 'Nombre Departamento', 'Division')

    c_nom_ap = _col(df_m, 'Apellido Paterno')

    c_nom_am = _col(df_m, 'Apellido Materno')

    c_nom_n = _col(df_m, 'Nombre')

    c_comb = _col(df_m, 'Apellidos y Nombres', 'Apellidos_Nombres')



    # --- OPTIMIZACION: Usar diccionarios en lugar de iterrows() ---

    data_list = df_m.to_dict('records')

    subordinado_map = {}

    matricula_map = {}



    for row in data_list:

        mat = _norm_id(str(row.get(c_mat, '')))

        if not mat: continue

        

        nombre_full = ''

        if c_nom_ap and c_nom_n:

            ap = _safe(row.get(c_nom_ap, ''))

            am = _safe(row.get(c_nom_am, ''))

            no = _safe(row.get(c_nom_n, ''))

            nombre_full = f"{ap} {am} {no}".strip().replace('  ', ' ')

        elif c_comb:

            nombre_full = _safe(row.get(c_comb, ''))

        

        email = _norm_email(row.get(c_mail, ''))

        mat_sup = _norm_id(str(row.get(c_sup_mat, '')))

        nom_sup = _safe(row.get(c_sup_nom, ''))

        hrbp_v = _safe(row.get(c_hrbp, ''))

        

        info = {

            'matricula': mat,

            'nombre': nombre_full,

            'email': email,

            'puesto': _safe(row.get(c_pue, '')),

            'area': _safe(row.get(c_area, '')),

            'departamento': _safe(row.get(c_dep, '')),

            'mat_jefe': mat_sup,

            'nombre_jefe': nom_sup,

            'hrbp': hrbp_v

        }

        matricula_map[mat] = info

        if mat_sup:

            subordinado_map[mat] = {

                'mat_jefe': mat_sup,

                'nombre_jefe': nom_sup,

                'email_jefe': '',

                'hrbp': hrbp_v

            }

            

    for mat, jefe_info in subordinado_map.items():

        mj = jefe_info['mat_jefe']

        if mj in matricula_map:

            jefe_info['email_jefe'] = matricula_map[mj]['email']

            nombre_jefe_actual = _safe(jefe_info.get('nombre_jefe', ''))

            nombre_parece_matricula = _norm_id(nombre_jefe_actual) == _norm_id(mj)

            nombre_sin_letras = not re.search(r'[A-Za-zÁÉÍÓÚÑáéíóúñ]', nombre_jefe_actual)

            if (not nombre_jefe_actual) or nombre_parece_matricula or nombre_sin_letras:

                jefe_info['nombre_jefe'] = matricula_map[mj]['nombre']



    if mt:

        _MAESTRO_JEFES_CACHE.update({'ruta': ruta, 'mtime': mt,

                                      'data': {'colab_a_jefe': subordinado_map, 'mat_a_info': matricula_map}})

    return subordinado_map, matricula_map, None





def _obtener_cadena_mando_recursiva(identificador, max_niveles=15):

    """Descubre recursivamente hacia arriba quiénes son los jefes."""

    colab_a_jefe, mat_a_info, err = _cargar_tabla_maestra_jefes()

    if err: return []

    

    # Buscar punto de partida

    start_mat = _norm_id(identificador)

    if start_mat not in mat_a_info:

        # Intentar por nombre exacto o contenido si no es matricula

        for m, info in mat_a_info.items():

            if identificador.upper() in info['nombre'].upper():

                start_mat = m

                break

    

    cadena = []

    actual = start_mat

    visitados = set()

    

    for _ in range(max_niveles):

        if not actual or actual in visitados: break

        visitados.add(actual)

        

        jefe_data = colab_a_jefe.get(actual)

        if not jefe_data: break

        

        mat_jefe = jefe_data.get('mat_jefe')

        if not mat_jefe: break

        

        # Info detallada del jefe

        info_jefe = mat_a_info.get(mat_jefe, {

            'nombre': jefe_data.get('nombre_jefe', 'Desconocido'),

            'email': jefe_data.get('email_jefe', ''),

            'matricula': mat_jefe

        })

        

        cadena.append(info_jefe)

        actual = mat_jefe

        

    return cadena





def _obtener_equipo_recursivo(identificador_jefe, max_profundidad=10):

    """Descubre recursivamente hacia abajo todos los colaboradores que reportan a alguien."""

    colab_a_jefe, mat_a_info, err = _cargar_tabla_maestra_jefes()

    if err: return []

    

    # Resolver jefe inicial

    jefe_mat = _norm_id(identificador_jefe)

    if jefe_mat not in mat_a_info:

        for m, info in mat_a_info.items():

            if identificador_jefe.upper() in info['nombre'].upper():

                jefe_mat = m

                break

    

    if not jefe_mat: return []

    

    # Construir mapa inverso (jefe -> lista de subordinados)

    jefe_a_subs = {}

    for sub_mat, jefe_data in colab_a_jefe.items():

        mj = jefe_data.get('mat_jefe')

        if not mj: continue

        if mj not in jefe_a_subs: jefe_a_subs[mj] = []

        jefe_a_subs[mj].append(sub_mat)

        

    equipo_total = []

    cola = [(jefe_mat, 0)]

    visitados = {jefe_mat}

    

    while cola:

        actual_mat, nivel = cola.pop(0)

        if nivel >= max_profundidad: continue

        

        subs = jefe_a_subs.get(actual_mat, [])

        for s_mat in subs:

            if s_mat in visitados: continue

            visitados.add(s_mat)

            

            info = mat_a_info.get(s_mat, {'matricula': s_mat, 'nombre': 'Desconocido'})

            info['nivel_jerarquico'] = nivel + 1

            equipo_total.append(info)

            cola.append((s_mat, nivel + 1))

            

    return equipo_total





def _fusionar_universo_maestro(df):

    if df is None or not _UNIVERSO_COMPLETO_ENABLED:

        return df



    df_m, err_m = _cargar_maestro_universo()

    if err_m or df_m is None:

        print(f'[WARN] Cruce maestro full omitido: {err_m}')

        return df



    llave_union = _detectar_llave_union(df, df_m)
    if llave_union.get('llave'):
        print(
            f"[MERGE] Llave detectada: {llave_union.get('llave')} | "
            f"objetivo='{llave_union.get('col_obj')}' | maestro='{llave_union.get('col_maestro')}' | "
            f"coincidencias={llave_union.get('overlap')}"
        )
    else:
        print('[MERGE] No se detectó llave de unión confiable entre objetivo y maestro')

    c_mat_base = _col(df, 'Matricula')

    c_mat_m = _col(df_m, 'Matricula')

    if not c_mat_base or not c_mat_m:

        print('[WARN] Cruce maestro full omitido: sin columna Matricula')

        return df



    existentes = {_id_key(v) for v in df[c_mat_base].tolist() if _id_key(v)}

    c_act_m = _col(df_m, 'Situacion Trabajador', 'Activo/Cesado', 'Activo')



    # --- NUEVO: Mapa de HRBP por Division/Departamento/Area desde el df de objetivos ---

    hrbp_map = {}

    c_hrbp_base = _col(df, 'HRBP', 'BP')

    c_div_base = _col(df, 'Division', 'Departamento', 'Nombre Departamento')

    c_dep_base = _col(df, 'Departamento', 'Nombre Departamento')

    c_area_base = _col(df, 'Area', 'Nombre Area')



    if c_hrbp_base:

        if c_area_base:

            for k, v in df.dropna(subset=[c_area_base, c_hrbp_base]).drop_duplicates(c_area_base).set_index(c_area_base)[c_hrbp_base].to_dict().items():

                hrbp_map[f'AREA:{_norm(k).upper()}'] = v

        if c_dep_base:

            for k, v in df.dropna(subset=[c_dep_base, c_hrbp_base]).drop_duplicates(c_dep_base).set_index(c_dep_base)[c_hrbp_base].to_dict().items():

                hrbp_map[f'DEP:{_norm(k).upper()}'] = v

        if c_div_base:

            for k, v in df.dropna(subset=[c_div_base, c_hrbp_base]).drop_duplicates(c_div_base).set_index(c_div_base)[c_hrbp_base].to_dict().items():

                hrbp_map[f'DIV:{_norm(k).upper()}'] = v



    if c_hrbp_base not in df.columns:

        df['HRBP'] = ''

        c_hrbp_base = 'HRBP'





    maps = {

        'mat': c_mat_base,

        'ap': _col(df, 'Apellido Paterno'),

        'am': _col(df, 'Apellido Materno'),

        'nom': _col(df, 'Nombre', 'Nombre Colaborador'),

        'comb': _col(df, 'Apellidos y Nombres', 'Nombre Colaborador'),

        'mail': _col(df, 'Correo', 'Correo Institucional', 'Email'),

        'dep': _col(df, 'Nombre Departamento', 'Departamento', 'Division'),

        'area': _col(df, 'Nombre Area', 'Area'),

        'puesto': _col(df, 'Nombre Puesto', 'Puesto'),

        'sup': _col(df, 'Jefe_Directo', 'Jefe Directo', 'Supervisor', 'Jefe', 'Business Partner', 'HRBP'),

        'ing': _col(df, 'Ingreso Compan', 'Ingreso Compa', 'F. Ingreso', 'Fecha Ingreso Compania', 'Fecha Ingreso'),

        'act': _col(df, 'Activo/Cesado', 'Situacion Trabajador', 'Activo'),

    }



    mcols = {

        'mat': c_mat_m,

        'ap': _col(df_m, 'Apellido Paterno'),

        'am': _col(df_m, 'Apellido Materno'),

        'nom': _col(df_m, 'Nombre'),

        'comb': _col(df_m, 'Apellidos y Nombres', 'Apellidos_Nombres', 'Nombre Completo'),

        'mail': _col(df_m, 'Correo Institucional', 'Correo Corporativo', 'Email Trabajo', 'Correo', 'Email'),

        'dep': _col(df_m, 'Nombre Departamento', 'Departamento', 'Division', 'Subgerencia'),

        'area': _col(df_m, 'Nombre Area', 'Area', 'Seccion'),

        'puesto': _col(df_m, 'Nombre Puesto', 'Puesto', 'Cargo'),

        'sup': _col(df_m, 'Jefe_Directo', 'Jefe Directo', 'Supervisor', 'Jefe', 'Business Partner', 'HRBP'),

        'ing': _col(df_m, 'Ingreso Compan', 'Ingreso Compa', 'Fecha Ingreso', 'F. Ingreso'),

        'act': c_act_m,

    }



    # Asegurar columna Correo en df principal para poder rellenarla desde maestro

    c_correo_m = mcols.get('mail') or _col(df_m, 'Correo', 'Email Trabajo', 'Correo Institucional', 'Email')

    if 'Correo' not in df.columns:

        df['Correo'] = ''

        maps['mail'] = 'Correo'



    nuevas = []

    for _, row in df_m.iterrows():

        if mcols['act']:

            act_txt = _safe(row.get(mcols['act'], '')).upper()

            if act_txt and 'ACTIVO' not in act_txt and 'CESADO' in act_txt:

                continue

        mkey = _id_key(row.get(mcols['mat'], ''))

        if not mkey or mkey in existentes:

            continue



        item = {c: (None if c.startswith('_') else '') for c in df.columns}

        item[maps['mat']] = _norm_id(row.get(mcols['mat'], ''))

        if maps['ap'] and mcols['ap']:

            item[maps['am']] = _safe(row.get(mcols['am'], ''))

        if maps['nom'] and mcols['nom']:

            item[maps['nom']] = _safe(row.get(mcols['nom'], ''))

        if maps['comb'] and mcols['comb']:

            item[maps['comb']] = _safe(row.get(mcols['comb'], ''))

        # Email: intentar desde columna mail mapeada o desde Correo/Email Trabajo del maestro

        if maps['mail']:

            email_val = ''

            if mcols['mail']:

                email_val = _safe(row.get(mcols['mail'], ''))

            if not email_val and c_correo_m:

                email_val = _safe(row.get(c_correo_m, ''))

            item[maps['mail']] = email_val

        if maps['dep'] and mcols['dep']:

            item[maps['dep']] = _safe(row.get(mcols['dep'], ''))

        if maps['area'] and mcols['area']:

            item[maps['area']] = _safe(row.get(mcols['area'], ''))

        if maps['puesto'] and mcols['puesto']:

            item[maps['puesto']] = _safe(row.get(mcols['puesto'], ''))

        if maps['sup'] and mcols['sup']:

            item[maps['sup']] = _safe(row.get(mcols['sup'], ''))

        if maps['ing'] and mcols['ing']:

            item[maps['ing']] = _safe(row.get(mcols['ing'], ''))

        if maps['act'] and mcols['act']:

            item[maps['act']] = _safe(row.get(mcols['act'], ''))

        # Marcar origen para identificar registros sin datos de vacaciones

        item['_origen_maestro'] = 'maestro'



        # Asignar HRBP segun el mapa

        val_dep = _safe(row.get(mcols['dep'], ''))

        val_area = _safe(row.get(mcols['area'], ''))

        val_puesto = _safe(row.get(mcols['puesto'], ''))

        item_hrbp = ''

        if val_area and f'AREA:{_norm(val_area).upper()}' in hrbp_map:

            item_hrbp = hrbp_map[f'AREA:{_norm(val_area).upper()}']

        elif val_dep and f'DEP:{_norm(val_dep).upper()}' in hrbp_map:

            item_hrbp = hrbp_map[f'DEP:{_norm(val_dep).upper()}']

        

        # Regla especial TyC: si es de Talento y Cultura, el HRBP es Carlos Jara

        if not item_hrbp or _es_talento_cultura(val_dep, val_area, val_puesto):

             item_hrbp = _resolver_hrbp_tyc(item_hrbp, val_area, val_dep, '', val_puesto)

        

        item[c_hrbp_base] = item_hrbp





        nuevas.append(item)



    if not nuevas:

        print('[CRUCE-MAESTRA] Sin filas nuevas para agregar')

        return df



    df_out = pd.concat([df, pd.DataFrame(nuevas)], ignore_index=True)

    print(f'[CRUCE-MAESTRA] Filas agregadas desde maestro full: {len(nuevas)}')

    return df_out



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# CARGA DE DATOS

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _encontrar_excel():

    ruta = _archivo_data_prueba_unico()

    if ruta:

        info = _resumen_fuente_datos()

        origen = info.get('origen', 'auto')

        print(f'[FUENTE] {origen}: {os.path.basename(ruta)}')

    return ruta





def _registros_demo_jefe():

    if not _DEMO_JEFE_ENABLED or not _DEMO_JEFE_NOMBRE:

        return []

    nombres = [

        'PERSONA A', 'PERSONA B', 'PERSONA C', 'PERSONA D', 'PERSONA E',

        'PERSONA F', 'PERSONA G', 'PERSONA H', 'PERSONA I', 'PERSONA J'

    ]

    out = []

    for i, nombre in enumerate(nombres, start=1):

        pendientes = float(3 + (i % 5))

        truncos = float(i % 3)

        total = round(pendientes + truncos + 8 + (i % 4), 1)

        out.append({

            'matricula': f'9900{i:02d}',

            'nombre': nombre,

            'correo': f'persona_{chr(64+i).lower()}@usil.edu.pe',

            'meta': 10 + (i % 6),

            'fecha_ingreso': f'{(i % 28) + 1:02d}/0{((i - 1) % 9) + 1}/2021',

            'puesto': 'Colaborador Ficticio de Prueba',

            'pendientes': pendientes,

            'truncos': truncos,

            'total_dias': total,

            'hrbp': _DEMO_JEFE_NOMBRE,

            'cant_pendiente': round(pendientes + truncos, 1),

            'estado': 'En alerta' if i % 2 == 0 else 'Proximo',

            'aviso': 'Registro ficticio para pruebas controladas',

            'recomendacion': 'Validar envio y confirmacion antes de campana real',

            'departamento': 'TALENTO Y CULTURA',

            'area': 'BUSINESS PARTNER',

            'seccion': 'PEOPLE ANALYTICS',

            'activo': 'ACTIVO',

        })

    return out





def _inyectar_registros_demo(df):

    if df is None or not _DEMO_JEFE_ENABLED:

        return df

    demo_rows = _registros_demo_jefe()

    if not demo_rows:

        return df



    c_mat = _col(df, 'Matricula')

    if not c_mat:

        return df



    cols = {

        'matricula': c_mat,

        'nombre': _col(df, 'Nombre Colaborador', 'Apellidos y Nombres', 'Nombre'),

        'correo': _col(df, 'Correo'),

        'meta': _col(df, 'Meta', 'Objetivo'),

        'fecha_ingreso': _col(df, 'F. Ingreso', 'Fecha Ingreso', 'Ingreso'),

        'puesto': _col(df, 'Puesto'),

        'pendientes': _col(df, 'Pendientes'),

        'truncos': _col(df, 'Truncos', 'Truncas'),

        'total_dias': _col(df, 'Total Dias', 'Suma de Dias Total'),

        'hrbp': _col(df, 'Business Partner', 'HRBP', 'Supervisor', 'Jefe_Directo', 'Jefe'),

        'cant_pendiente': _col(df, 'Cant. Pendiente', 'Cantidad Pendiente'),

        'estado': _col(df, 'Estado', 'Elegibilidad'),

        'aviso': _col(df, 'Aviso'),

        'recomendacion': _col(df, 'Recomendacion'),

        'departamento': _col(df, 'Departamento', 'Nombre Departamento'),

        'area': _col(df, 'Area', 'Nombre Area'),

        'seccion': _col(df, 'Seccion'),

        'activo': _col(df, 'Activo/Cesado', 'Activo'),

    }



    existentes = {

        _id_key(v) for v in df[c_mat].tolist()

        if _id_key(v)

    }

    nuevas = []

    for item in demo_rows:

        if _id_key(item.get('matricula')) in existentes:

            continue

        row_data = {c: '' for c in df.columns}

        for key, col in cols.items():

            if col:

                row_data[col] = item.get(key, '')

        nuevas.append(row_data)



    if not nuevas:

        return df



    df_out = pd.concat([df, pd.DataFrame(nuevas)], ignore_index=True)

    print(f'[DEMO] Registros de prueba inyectados: {len(nuevas)} para jefe {_DEMO_JEFE_NOMBRE}')

    return df_out



def _meses(fi, hoy):

    if pd.isna(fi): return None

    d = fi.date() if hasattr(fi, 'date') else fi

    return (hoy.year - d.year)*12 + (hoy.month - d.month)



def _dias_aniv(fi, hoy):

    d = fi.date() if hasattr(fi, 'date') else fi

    try:

        p = d.replace(year=hoy.year)

        if p < hoy:

            r = (hoy - p).days

            if r <= 30: return -r

            p = d.replace(year=hoy.year+1)

        return (p - hoy).days

    except ValueError: return None



def _prox_periodo(fi, hoy):

    d = fi.date() if hasattr(fi, 'date') else fi

    try:

        p = d.replace(year=hoy.year)

        if p < hoy:

            if (hoy - p).days <= 30: return p.strftime('%d/%m/%Y')

            p = d.replace(year=hoy.year+1)

        return p.strftime('%d/%m/%Y')

    except ValueError: return None



def _elegib(m):

    if m is None:  return 'Sin dato'

    if m >= 12:    return 'Elegible'

    if m >= 11:    return 'Proximo (< 1 mes)'

    if m >= 10:    return 'En alerta (1-2 meses)'

    return 'No aplica'



def _calc_vac(fi, hoy, esq_str=''):

    import re

    m = re.search(r'(\d+)', str(esq_str) if esq_str else '')

    dias_esq = int(m.group(1)) if m else 30

    if pd.isna(fi): return None, None, None, None

    d = fi.date() if hasattr(fi, 'date') else fi

    tot_m = (hoy.year-d.year)*12 + (hoy.month-d.month)

    anios = tot_m // 12

    parcial = tot_m % 12

    truncas = round((parcial/12)*dias_esq, 2)

    pend, venc = 0.0, 0.0

    for n in range(1, anios+1):

        try: v = d.replace(year=d.year+n+1)

        except ValueError: v = d.replace(year=d.year+n+1, day=28)

        if hoy <= v: pend += dias_esq

        else: venc += dias_esq

    return truncas, float(pend), float(venc), round(truncas+pend+venc, 2)



@medir_tiempo

def cargar_datos():

    hoy = date.today()

    ruta = _encontrar_excel()

    if not ruta: return None, "No se encontro archivo Excel"



    _mt = None

    try:

        _mt = os.path.getmtime(ruta)

        _pkl = os.path.join(DATAS_DIR, '__cache__', 'df_final_v2.pkl')
        try:
            os.makedirs(os.path.dirname(_pkl), exist_ok=True)
            if os.path.exists(_pkl) and os.path.getmtime(_pkl) >= _mt:
                with open(_pkl, 'rb') as fp:
                    c_data = pickle.load(fp)
                    if c_data.get('mtime') == _mt and c_data.get('ruta') == ruta and c_data.get('fecha') == hoy:
                        _df_cache.update(c_data)
                        print('[CACHE] df_final cargado de pkl')
                        return _df_cache['df'], None
        except Exception as e: print("Cache Pkl err:", e)
        if _df_cache['df'] is not None and _df_cache['fecha'] == hoy and _df_cache['ruta'] == ruta and _df_cache['mtime'] == _mt:

            return _df_cache['df'], None

    except OSError: pass



    with _cache_lock:

        if _df_cache['df'] is not None and _df_cache['fecha'] == hoy and _df_cache['ruta'] == ruta and _df_cache['mtime'] == _mt:

            return _df_cache['df'], None



        try:

            # --- OPTIMIZACION: Abrir ExcelFile una sola vez ---

            tmp_path = None

            try:

                import tempfile, shutil

                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp: tmp_path = tmp.name

                shutil.copy2(ruta, tmp_path)

                with pd.ExcelFile(tmp_path) as xls:

                    _hoja = _resolver_hoja_excel_from_xls(xls, modo='datos')

                    # Leer primeras 25 filas para detectar header sin leer todo el archivo

                    df_preview = pd.read_excel(xls, sheet_name=_hoja, nrows=25, header=None, dtype=str)

                    _hrow = _detectar_header_row_from_df(df_preview)

                    

                    print(f'[LOAD] {os.path.basename(ruta)} (fila {_hrow})')

                    # Ahora si, leer todo con el header detectado

                    df = pd.read_excel(xls, sheet_name=_hoja, header=_hrow, dtype=str)

            finally:

                if tmp_path and os.path.exists(tmp_path): os.unlink(tmp_path)

        except Exception as e:

            return None, f"Error al leer Excel: {e}"



    df.columns = [_norm(c) for c in df.columns]

    # Formato viejo (header en fila 0): las 2 primeras filas de datos son basura

    # Formato nuevo (header detectado en fila >0): datos empiezan directamente

    if _hrow == 0:

        df = df.iloc[2:].reset_index(drop=True)

    else:

        df = df.reset_index(drop=True)



    tc = _col(df, 'Tipo Trabajador')

    mc = _col(df, 'Modalidad Formativa', 'Modalidad Forma')

    if tc: df = df[~df[tc].str.upper().str.contains('PRACTICANTE|TRAINEE', na=False)]

    if mc: df = df[~df[mc].str.upper().str.contains('PRACTIC|FORMATIVA|TRAINEE', na=False)]



    ec = _col(df, 'Esquema Vacac')

    if ec: df = df[~df[ec].str.upper().str.contains(r'ESQUEMA 0 ', na=False)]



    sc = _col(df, 'Situacion Trabajador', 'ituaci')

    if sc: df = df[df[sc].str.upper().str.contains('ACTIVO', na=False)]



    fc = _col(df, 'Ingreso Compan', 'Ingreso Compa', 'ngreso Compa')

    if not fc:

        cands = [c for c in df.columns if 'Ingreso' in c and 'Afp' not in c and 'AFP' not in c]

        fc = cands[0] if cands else None

    if fc:

        df['_fecha_ingreso'] = pd.to_datetime(df[fc], dayfirst=True, errors='coerce')

        df['_meses'] = df['_fecha_ingreso'].apply(lambda f: _meses(f, hoy) if pd.notna(f) else None)

        df['_dias_aniv'] = df['_fecha_ingreso'].apply(lambda f: _dias_aniv(f, hoy) if pd.notna(f) else None)

        df['_elegibilidad'] = df['_meses'].apply(_elegib)

        mask = df['_dias_aniv'].notna() & (df['_dias_aniv'] < 0)

        df.loc[mask, '_elegibilidad'] = 'En retraso'

        df['_prox_periodo'] = df['_fecha_ingreso'].apply(lambda f: _prox_periodo(f, hoy) if pd.notna(f) else None)



    # â”€â”€ FILTRO TALENTO Y CULTURA (opcional por configuracion) â”€â”€

    limitar_tyc = _cfg_bool(_PA_CONFIG.get('vacaciones_limitar_tyc', True), True)

    prueba_file = _archivo_data_prueba_unico()

    if prueba_file and limitar_tyc:

            cdept = _col(df, 'Nombre Departamento', 'Division', 'Departamento')

            if cdept:

                df = df[df[cdept].str.upper().str.contains('TALENTO|CULTURA|GESTION HUMANA|ATRACCION', na=False)].reset_index(drop=True)

            print(f'[FILTRO-TYC] {len(df)} personas de Talento y Cultura')

            # Agregar Jefe_Directo desde DATA DE PRUEBA (columna Jefe Directo si existe)

            try:

                _hoja_pd = _resolver_hoja_excel(prueba_file, modo='datos')
                _hrow_pd = _detectar_header_row(prueba_file, sheet_name=_hoja_pd)
                df_pd = _safe_read_excel(prueba_file, sheet_name=_hoja_pd, header=_hrow_pd, dtype=str)

                df_pd.columns = [_norm(c) for c in df_pd.columns]

                c_mat_pd  = _col(df_pd, 'Matricula')

                c_jefe_pd = _col(df_pd, 'Jefe_Directo', 'Jefe Directo')

                if c_mat_pd and c_jefe_pd:

                    def _mat_key(x):

                        return _norm_id(str(x)).lstrip('0') or '0'

                    jefe_map = dict(zip(

                        df_pd[c_mat_pd].apply(_mat_key),

                        df_pd[c_jefe_pd].apply(lambda x: _safe(x))

                    ))

                    c_mat_main = _col(df, 'Matricula')

                    if c_mat_main:

                        df['Jefe_Directo'] = df[c_mat_main].apply(

                            lambda m: jefe_map.get(_mat_key(m), '')

                        )

                        print(f'[DATA-PRUEBA] Jefe_Directo para {df["Jefe_Directo"].ne("").sum()} personas')

            except Exception as _e:

                print(f'[WARN] Jefe_Directo no cargado: {_e}')



            # Poblar Jefe_Directo por mapa de Seccion/Area desde pa_config (fuente real del org chart)

            _sec_sup_map = _PA_CONFIG.get('vacaciones_seccion_supervisor') or {}

            _area_sup_map = _PA_CONFIG.get('vacaciones_area_supervisor') or {}

            if _sec_sup_map or _area_sup_map:

                c_sec = _col(df, 'Seccion')

                c_area_df = _col(df, 'Area', 'Nombre Area')

                if 'Jefe_Directo' not in df.columns:

                    df['Jefe_Directo'] = ''

                def _jefe_por_seccion(row):

                    existing = _safe(row.get('Jefe_Directo', ''))

                    if existing:

                        return existing

                    sec = _safe(row.get(c_sec, '')).upper() if c_sec else ''

                    area = _safe(row.get(c_area_df, '')).upper() if c_area_df else ''

                    for k, v in _sec_sup_map.items():

                        if k.upper() in sec:

                            return v

                    for k, v in _area_sup_map.items():

                        if k.upper() in area:

                            return v

                    return existing

                df['Jefe_Directo'] = df.apply(_jefe_por_seccion, axis=1)

                n_jefes = df['Jefe_Directo'].ne('').sum()

                n_uniq = df['Jefe_Directo'][df['Jefe_Directo'].ne('')].nunique()

                print(f'[JEFE-MAP] {n_jefes} personas con jefe directo ({n_uniq} jefes distintos)')

    elif prueba_file:

        try:

            _hoja_pd2 = _resolver_hoja_excel(prueba_file, modo='datos')
            _hrow_pd2 = _detectar_header_row(prueba_file, sheet_name=_hoja_pd2)
            df_pd = _safe_read_excel(prueba_file, sheet_name=_hoja_pd2, header=_hrow_pd2, dtype=str)

            df_pd.columns = [_norm(c) for c in df_pd.columns]

            c_mat_pd = _col(df_pd, 'Matricula')

            c_jefe_pd = _col(df_pd, 'Jefe_Directo', 'Jefe Directo')

            c_mat_main = _col(df, 'Matricula')

            if c_mat_pd and c_jefe_pd and c_mat_main:

                def _mat_key2(x):

                    return _norm_id(str(x)).lstrip('0') or '0'

                jefe_map = dict(zip(

                    df_pd[c_mat_pd].apply(_mat_key2),

                    df_pd[c_jefe_pd].apply(lambda x: _safe(x))

                ))

                if 'Jefe_Directo' not in df.columns:

                    df['Jefe_Directo'] = ''

                df['Jefe_Directo'] = df.apply(

                    lambda r: jefe_map.get(_mat_key2(r.get(c_mat_main, '')), _safe(r.get('Jefe_Directo', ''))),

                    axis=1

                )

        except Exception as _e:

            print(f'[WARN] Jefe_Directo no cruzado en modo universo completo: {_e}')

    df = _fusionar_universo_maestro(df)

    # Parsear fechas de ingreso para las filas nuevas del maestro (no tenian _fecha_ingreso)

    if '_fecha_ingreso' not in df.columns:

        df['_fecha_ingreso'] = pd.NaT

    # Detectar filas sin fecha válida (NaN, NaT, string vacío, string que no es Timestamp)

    mask_sin_fecha = ~df['_fecha_ingreso'].apply(lambda x: isinstance(x, pd.Timestamp))

    if mask_sin_fecha.any():

        _fc2 = None

        for _pat in ('Ingreso Compan', 'Ingreso Compa', 'Fecha Ingreso'):

            _fc2 = _col(df, _pat)

            if _fc2:

                break

        if not _fc2:

            _cands2 = [c for c in df.columns if 'Ingreso' in c and 'Afp' not in c and 'AFP' not in c]

            _fc2 = _cands2[0] if _cands2 else None

        if _fc2:

            df.loc[mask_sin_fecha, '_fecha_ingreso'] = pd.to_datetime(

                df.loc[mask_sin_fecha, _fc2], dayfirst=True, errors='coerce'

            )

        if '_meses' not in df.columns:

            df['_meses'] = None

        if '_dias_aniv' not in df.columns:

            df['_dias_aniv'] = None

        if '_elegibilidad' not in df.columns:

            df['_elegibilidad'] = ''

        if '_prox_periodo' not in df.columns:

            df['_prox_periodo'] = None

        mask_sin_fecha2 = df['_fecha_ingreso'].isna() == False

        new_mask = mask_sin_fecha & mask_sin_fecha2

        if new_mask.any():

            df.loc[new_mask, '_meses'] = df.loc[new_mask, '_fecha_ingreso'].apply(

                lambda f: _meses(f, hoy) if pd.notna(f) else None)

            df.loc[new_mask, '_dias_aniv'] = df.loc[new_mask, '_fecha_ingreso'].apply(

                lambda f: _dias_aniv(f, hoy) if pd.notna(f) else None)

            df.loc[new_mask, '_elegibilidad'] = df.loc[new_mask, '_meses'].apply(_elegib)

            df.loc[new_mask, '_prox_periodo'] = df.loc[new_mask, '_fecha_ingreso'].apply(

                lambda f: _prox_periodo(f, hoy) if pd.notna(f) else None)

    df = _aplicar_overrides_colaboradores(df, modo='datos')

    df = _inyectar_registros_demo(df)

    print(f'[OK] {len(df)} activos (incluye universo maestro TyC)')

    if _mt: 
        _df_cache.update({'df': df, 'mtime': _mt, 'ruta': ruta, 'fecha': hoy})
        try:
            with open(os.path.join(DATAS_DIR, '__cache__', 'df_final_v2.pkl'), 'wb') as fp:
                pickle.dump(_df_cache, fp)
        except: pass

    return df, None



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# CARGA EXCEL OBJETIVOS

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _encontrar_objetivo():

    """Usa solo el archivo fijo de vacaciones configurado para el sistema."""

    ruta = _archivo_data_prueba_unico()

    if ruta:

        print(f'[FUENTE-FIJA] OBJETIVO: {os.path.basename(ruta)}')

    return ruta



@medir_tiempo

def cargar_objetivos():

    """Carga y procesa el Excel de objetivos de vacaciones."""

    ruta = _encontrar_objetivo()

    if not ruta: return None, "No se encontro archivo de Objetivos"



    _mt = os.path.getmtime(ruta)

    _pkl_obj = os.path.join(DATAS_DIR, '__cache__', f'{_OBJ_CACHE_VERSION}.pkl')
    try:
        os.makedirs(os.path.dirname(_pkl_obj), exist_ok=True)
        if os.path.exists(_pkl_obj) and os.path.getmtime(_pkl_obj) >= _mt:
            with open(_pkl_obj, 'rb') as fp:
                c_data = pickle.load(fp)
                if c_data.get('mtime') == _mt and c_data.get('ruta') == ruta and c_data.get('version') == _OBJ_CACHE_VERSION:
                    _obj_cache.update(c_data)
                    print('[CACHE] obj_final cargado de pkl')
                    return _obj_cache['df'], None
    except Exception as e: print("Cache Pkl err obj:", e)
    if _obj_cache['df'] is not None and _obj_cache['ruta'] == ruta and _obj_cache['mtime'] == _mt:

        return _obj_cache['df'], None



    hoja_obj = _resolver_hoja_excel(ruta, modo='objetivos')

    print(f'[LOAD-OBJ] {os.path.basename(ruta)}  (hoja sugerida: {hoja_obj})')

    df = None

    header_idx = 0

    hojas_candidatas = [hoja_obj]

    try:

        xls_obj = pd.ExcelFile(ruta)

        for sn in xls_obj.sheet_names:

            if sn not in hojas_candidatas:

                hojas_candidatas.append(sn)

    except Exception:

        pass

    ultimo_error = None

    for hoja_candidata in hojas_candidatas:

        try:

            df_raw = _safe_read_excel(ruta, sheet_name=hoja_candidata, dtype=str, header=None)

            header_idx = _detectar_header_row(ruta, sheet_name=hoja_candidata, modo='objetivos')

            if header_idx >= len(df_raw):

                ultimo_error = f'Header fuera de rango en hoja {hoja_candidata}'

                continue

            new_cols = []

            for j, v in enumerate(df_raw.iloc[header_idx]):

                if pd.notna(v) and str(v).strip():

                    new_cols.append(_norm(str(v).strip()))

                else:

                    new_cols.append(f'_col{j}')

            df_try = df_raw.iloc[header_idx+1:].copy()

            df_try.columns = new_cols

            df_try = df_try.reset_index(drop=True)

            tiene_matricula = any('matricula' in c.lower() for c in df_try.columns)

            tiene_campos_vac = any(

                any(tok in c.lower() for tok in ('objetivo', 'meta', 'gozado', 'cumpl', 'pend', 'trunc', 'venc', 'restante', 'registro'))

                for c in df_try.columns

            )

            if tiene_matricula and tiene_campos_vac:

                hoja_obj = hoja_candidata

                df = df_try

                print(f'[LOAD-OBJ] Hoja valida detectada: {hoja_obj} (header fila {header_idx})')

                break

            ultimo_error = f'Hoja {hoja_candidata} sin columnas clave de objetivos'

        except Exception as e:

            ultimo_error = f'Hoja {hoja_candidata}: {e}'

    if df is None:

        return None, ultimo_error or 'No se encontro una hoja valida con matricula y columnas de vacaciones'



    # Limpiar: excluir filas sin matrÃ­cula

    mat_col = next((c for c in df.columns if 'matricula' in c.lower()), None)

    if mat_col:

        df = df[df[mat_col].notna() & (df[mat_col].str.strip() != '')].reset_index(drop=True)



    # Normalizar columnas numÃ©ricas

    def _to_num(col_name):

        c = next((k for k in df.columns if col_name.lower() in k.lower()), None)

        if c: df[c] = pd.to_numeric(df[c], errors='coerce')

        return c



    c_vencidas   = _to_num('Vac. Vencidas')  or _to_num('Vencidas')

    c_pendientes = _to_num('Vac Pendiente')   or _to_num('Pendiente')

    c_truncos    = _to_num('Vac. Trunco')     or _to_num('Trunco')

    c_total      = _to_num('Suma de Dias')    or _to_num('Dias Total')

    c_objetivo   = _to_num('Objetivo')        or _to_num('Meta')

    c_restantes  = _to_num('Dias Restante')

    c_gozados    = _to_num('Dias Gozado')     or _to_num('Registrad')

    c_cumpl      = _to_num('Cumplimiento')    or _to_num('Meta%')

    c_obj_hrbp = _col(df, 'HRBP', 'Business Partner', 'BP')
    c_obj_dep = _col(df, 'Departamento', 'Nombre Departamento', 'Division')
    c_obj_area = _col(df, 'Area', 'Nombre Area')
    c_obj_puesto = _col(df, 'Puesto', 'Nombre Puesto')

    if c_obj_hrbp:
        def _normalizar_hrbp_obj(row):
            hrbp_val = _resolver_hrbp_tyc(
                row.get(c_obj_hrbp, ''),
                area=row.get(c_obj_area, '') if c_obj_area else '',
                departamento=row.get(c_obj_dep, '') if c_obj_dep else '',
                puesto=row.get(c_obj_puesto, '') if c_obj_puesto else '',
            )
            n = _norm(hrbp_val).upper()
            if 'CARLOS' in n and 'JARA' in n:
                return 'Carlos Jara'
            if 'CESAR' in n and 'REYES' in n:
                return 'César Reyes'
            if 'FATIMA' in n and 'SALAZAR' in n:
                return 'Fatima Salazar'
            if 'LESL' in n and 'REYES' in n:
                return 'Lesley Reyes'
            if 'MELISSA' in n and 'HIGA' in n:
                return 'Melissa Higa'
            return _safe(hrbp_val)

        df[c_obj_hrbp] = df.apply(_normalizar_hrbp_obj, axis=1)



    limitar_tyc = _cfg_bool(_PA_CONFIG.get('vacaciones_limitar_tyc', True), True)

    if limitar_tyc:

        masks = []

        for c in (c_obj_dep, c_obj_area, c_obj_puesto):

            if c:

                masks.append(df[c].fillna('').astype(str).str.upper().str.contains('TALENTO|CULTURA|GESTION HUMANA|ATRACCION', na=False))

        if masks:

            mask_ok = masks[0]

            for m in masks[1:]:

                mask_ok = (mask_ok | m)

            df = df[mask_ok].reset_index(drop=True)

            print(f'[FILTRO-TYC-OBJ] {len(df)} registros de objetivos Talento y Cultura')



    print(f'[OK-OBJ] {len(df)} registros de objetivos')

    df = _aplicar_overrides_colaboradores(df, modo='objetivos')

    df = _inyectar_registros_demo(df)

    try:
        df_maestro, err_m = _cargar_maestro_universo()
        if df_maestro is not None and not df_maestro.empty:
            print("[JOIN] Iniciando cruce de Objetivos con Personal Maestro para traer Jefes...")
            llaves = _detectar_llave_union(df, df_maestro)
            col_obj = llaves.get('col_obj')
            col_maestro = llaves.get('col_maestro')
            
            if col_obj and col_maestro:
                c_supervisor = next((c for c in df_maestro.columns if 'supervisor' in c.lower() or 'jefe' in c.lower() or 'lider' in c.lower()), None)
                if not c_supervisor:
                    c_supervisor = next((c for c in df_maestro.columns if 'reporta a' in c.lower() or 'apellidos y nombres supervisor' in c.lower() or 'jefatura' in c.lower()), None)
                
                if c_supervisor:
                    df_maestro_norm = df_maestro.copy()
                    df_maestro_norm[col_maestro] = df_maestro_norm[col_maestro].astype(str).str.strip().str.upper()
                    
                    df_obj_norm = df[col_obj].astype(str).str.strip().str.upper()
                    
                    mapa_jefes = dict(zip(df_maestro_norm[col_maestro], df_maestro_norm[c_supervisor]))
                    
                    c_jefe_existente = next((c for c in df.columns if c.lower() == 'supervisor' or 'jefe_directo' in c.lower()), None)
                    if not c_jefe_existente:
                        c_jefe_existente = 'supervisor'
                        df[c_jefe_existente] = None
                        
                    df[c_jefe_existente] = df_obj_norm.map(mapa_jefes).combine_first(df[c_jefe_existente])
                    
                    print(f"[JOIN-OK] Cruce exitoso. Columna supervisor mapeada usando {col_obj} y {col_maestro}.")
                else:
                    print("[JOIN-WARN] No se encontró columna de Supervisor en Personal Maestro.")
            else:
                print("[JOIN-WARN] No se detectaron llaves de unión válidas entre Objetivos y Maestro.")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[JOIN-ERR] Error al cruzar con Maestro: {e}")

    _obj_cache.update({'df': df, 'mtime': _mt, 'ruta': ruta, 'version': _OBJ_CACHE_VERSION})
    # Guardar pickle cache de objetivos
    try:
        _pkl_obj_save = os.path.join(DATAS_DIR, '__cache__', f'{_OBJ_CACHE_VERSION}.pkl')
        os.makedirs(os.path.dirname(_pkl_obj_save), exist_ok=True)
        with open(_pkl_obj_save, 'wb') as fp:
            pickle.dump(_obj_cache, fp)
        print('[CACHE] obj_final guardado en pkl')
    except Exception as _e:
        print(f'[CACHE] Error guardando obj pkl: {_e}')

    return df, None



def _columnas_objetivos(dfo):

    if dfo is None:

        return {}

    c_nombre_full = next((c for c in dfo.columns if 'apellidos y nombres' in c.lower() or 'nombre completo' in c.lower() or 'colaborador' in c.lower()), None)

    c_ap = next((c for c in dfo.columns if 'apellido paterno' in c.lower()), None)

    c_am = next((c for c in dfo.columns if 'apellido materno' in c.lower()), None)

    c_nom = next((c for c in dfo.columns if c.lower() == 'nombre' or (' nombre' in c.lower()) or ('nombre ' in c.lower())), None)

    return {

        'matricula': next((c for c in dfo.columns if 'matricula' in c.lower()), None),

        # Prioridad: nombre completo > apellidos+nombres separados > fallback por columna de nombre.

        'nombre': c_nombre_full or next((c for c in dfo.columns if 'apellidos' in c.lower() or 'nombre' in c.lower()), None),

        'nombre_full': c_nombre_full,

        'apellido_paterno': c_ap,

        'apellido_materno': c_am,

        'nombres': c_nom,

        'activo': next((c for c in dfo.columns if 'activo' in c.lower() or 'cesado' in c.lower()), None),

        'hrbp': next((c for c in dfo.columns if 'hrbp' in c.lower() or 'bp' in c.lower()), None),

        'tipo': next((c for c in dfo.columns if 'tipo' in c.lower() and 'trab' in c.lower()), None),

        'puesto': next((c for c in dfo.columns if 'puesto' in c.lower()), None),

        'departamento': next((c for c in dfo.columns if 'departamento' in c.lower()), None),

        'area': next((c for c in dfo.columns if c.lower() == 'area' or 'area' in c.lower()), None),

        'seccion': next((c for c in dfo.columns if 'seccion' in c.lower()), None),

        'fecha_ingreso': next((c for c in dfo.columns if 'fecha' in c.lower() and 'ingreso' in c.lower()), None),

        'objetivo': next((c for c in dfo.columns if 'objetivo' in c.lower() or 'meta' in c.lower()), None),

        'gozados': next((c for c in dfo.columns if 'gozado' in c.lower() or 'registrad' in c.lower()), None),

        'cumplimiento': next((c for c in dfo.columns if 'cumplimiento' in c.lower() or 'meta%' in c.lower() or 'meta %' in c.lower()), None),

        'truncas': next((c for c in dfo.columns if 'trunco' in c.lower()), None),

        'pendientes': next((c for c in dfo.columns if 'pendiente' in c.lower()), None),

        'vencidas': next((c for c in dfo.columns if 'vencida' in c.lower()), None),

        'total': next((c for c in dfo.columns if 'suma' in c.lower() and 'dia' in c.lower()), None),

        'comentario': next((c for c in dfo.columns if 'comentario' in c.lower()), None),

        'supervisor': next((c for c in dfo.columns if 'jefe_directo' in c.lower() or c.lower() == 'supervisor' or 'supervisor' in c.lower() or c.lower() == 'jefe' or 'jefe' in c.lower() or 'business_partner' in c.lower() or 'bp' in c.lower()), None),

    }



def _objetivos_activos(dfo):

    cols = _columnas_objetivos(dfo)

    if dfo is None:

        return None, cols

    df_act = dfo.copy()

    c_act = cols.get('activo')

    c_tipo = cols.get('tipo')

    if c_act:

        df_act = df_act[df_act[c_act].astype(str).str.upper().str.contains('ACTIVO', na=False)]

    if c_tipo:

        df_act = df_act[~df_act[c_tipo].astype(str).str.upper().str.contains('PRACT|TRAINEE|FORMATIV', na=False)]

    return df_act.reset_index(drop=True), cols



def _build_objetivos_lookup(dfo):

    df_act, cols = _objetivos_activos(dfo)

    lookup = {}

    if df_act is None:

        return lookup, df_act, cols



    for _, row in df_act.iterrows():

        key = _id_key(row.get(cols.get('matricula'), '')) if cols.get('matricula') else ''

        if not key:

            continue



        nombre_full = ''

        ap = _safe(row.get(cols.get('apellido_paterno'), '')) if cols.get('apellido_paterno') else ''

        am = _safe(row.get(cols.get('apellido_materno'), '')) if cols.get('apellido_materno') else ''

        no = _safe(row.get(cols.get('nombres'), '')) if cols.get('nombres') else ''

        if ap or am or no:

            nombre_full = ' '.join([x for x in (ap, am, no) if x]).strip()

        if not nombre_full and cols.get('nombre_full'):

            nombre_full = _safe(row.get(cols.get('nombre_full'), ''))

        if not nombre_full and cols.get('nombre'):

            nombre_full = _safe(row.get(cols.get('nombre'), '-'))



        item = {

            'matricula': _norm_id(row.get(cols.get('matricula'), '')) if cols.get('matricula') else '',

            'nombre': nombre_full or '-',

            'hrbp': _safe(row.get(cols.get('hrbp'), '')) if cols.get('hrbp') else '',

            'puesto': _safe(row.get(cols.get('puesto'), '')) if cols.get('puesto') else '',

            'departamento': _safe(row.get(cols.get('departamento'), '')) if cols.get('departamento') else '',

            'area': _safe(row.get(cols.get('area'), '')) if cols.get('area') else '',

            'seccion': _safe(row.get(cols.get('seccion'), '')) if cols.get('seccion') else '',

            'fecha_ingreso': _safe(row.get(cols.get('fecha_ingreso'), '')) if cols.get('fecha_ingreso') else '',

            'objetivo': _to_int0(row.get(cols.get('objetivo'), 0)) if cols.get('objetivo') else 0,

            'gozados': _to_int0(row.get(cols.get('gozados'), 0)) if cols.get('gozados') else 0,

            'cumplimiento': _to_int0(row.get(cols.get('cumplimiento'), 0)) if cols.get('cumplimiento') else 0,

            'truncas': _to_float0(row.get(cols.get('truncas'), 0)) if cols.get('truncas') else 0.0,

            'pendientes': _to_float0(row.get(cols.get('pendientes'), 0)) if cols.get('pendientes') else 0.0,

            'vencidas': _to_float0(row.get(cols.get('vencidas'), 0)) if cols.get('vencidas') else 0.0,

            'total_vacaciones': _to_float0(row.get(cols.get('total'), 0)) if cols.get('total') else 0.0,

            'supervisor': _safe(row.get(cols.get('supervisor'), '')) if cols.get('supervisor') else '',

            'comentario': _safe(row.get(cols.get('comentario'), '')) if cols.get('comentario') else '',

        }

        item['hrbp'] = _resolver_hrbp_tyc(

            item.get('hrbp', ''),

            area=item.get('area', ''),

            departamento=item.get('departamento', ''),

            puesto=item.get('puesto', ''),

        )

        item['cantidad_pendiente'] = round(float(item.get('vencidas', 0) or 0) + float(item.get('pendientes', 0) or 0), 1)

        aviso, recomendacion = _aviso_recomendacion_persona(item)

        item['aviso'] = aviso

        item['recomendacion'] = recomendacion

        lookup[key] = item

    return lookup, df_act, cols



def _enriquecer_alertas_con_objetivos(alertas, obj_lookup):

    out = []

    for p in alertas or []:

        key = _id_key(p.get('matricula', ''))

        obj = obj_lookup.get(key)

        if not obj:

            continue

        item = dict(p)

        for k in ('nombre', 'area', 'puesto', 'hrbp', 'supervisor', 'fecha_ingreso', 'truncas', 'pendientes', 'vencidas', 'total_vacaciones', 'cantidad_pendiente', 'objetivo', 'gozados', 'comentario'):

            if obj.get(k) not in (None, ''):

                item[k] = obj.get(k)

        aviso, recomendacion = _aviso_recomendacion_persona(item)

        item['aviso'] = aviso

        item['recomendacion'] = recomendacion

        out.append(item)

    return out



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# CONSTRUIR REGISTRO

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _find_key(row, *keys):

    """Buscar clave en un Series de pandas."""

    cols = list(row.index) if hasattr(row, 'index') else []

    for k in keys:

        for c in cols:

            if k.lower() in str(c).lower():

                return c

    return None



def _persona(row, hoy):

    cp = _find_key(row, 'Apellido Paterno')

    cm = _find_key(row, 'Apellido Materno')

    cn = _find_key(row, 'Nombre')

    cmat = _find_key(row, 'Matricula')

    partes_nombre = []

    if cp:

        partes_nombre.append(_safe(row.get(cp, '')))

    if cm:

        partes_nombre.append(_safe(row.get(cm, '')))

    if cn:

        partes_nombre.append(_safe(row.get(cn, '')))

    nombre = ' '.join([p for p in partes_nombre if p]).strip()

    if not nombre:

        c_comb = _find_key(

            row,

            'Apellidos y Nombres',

            'Apellidos_Nombres',

            'Apellidos Nombres',

            'Nombres y Apellidos',

            'Nombre Completo',

            'Colaborador',

            'Trabajador'

        )

        if c_comb:

            nombre = _safe(row.get(c_comb, ''))

    if not nombre and cn:

        nombre = _safe(row.get(cn, ''))



    ca  = _find_key(row, 'Nombre Area', 'Area')

    cpu = _find_key(row, 'Nombre Puesto', 'Puesto')

    cd  = _find_key(row, 'Nombre Departamento', 'Division', 'Departamento')

    ces = _find_key(row, 'Esquema Vacac')

    csup = _find_key(row, 'Jefe_Directo', 'Jefe Directo', 'Supervisor', 'Jefe', 'Business Partner', 'HRBP')

    chrbp = _find_key(row, 'HRBP', 'Business Partner', 'BP', 'Jefe')

    cmat_sup = _find_key(row, 'Matricula Supervisor')

    cem = _find_key(

        row,

        'Email Institucional',

        'Email Trabajo',

        'Correo Institucional',

        'Correo Corporativo',

        'Correo Electronico',

        'Correo',

        'Email',

        'E-mail',

        'Mail'

    )



    esq = _safe(row.get(ces,'')) if ces else ''

    fi = row.get('_fecha_ingreso')

    truncas, pend, venc, total = _calc_vac(fi, hoy, esq)



    mv = row.get('_meses')

    dv = row.get('_dias_aniv')

    m_int = int(mv) if pd.notna(mv) else 0

    d_int = int(dv) if pd.notna(dv) else None



    anios_v  = m_int // 12

    meses_r  = m_int % 12

    elegib_v = row.get('_elegibilidad', 'Sin dato')

    prox_v   = row.get('_prox_periodo', '')

    fi_txt = ''

    if pd.notna(fi):

        try:

            fi_txt = fi.strftime('%d/%m/%Y') if hasattr(fi, 'strftime') else str(fi)

        except Exception:

            fi_txt = str(fi)

    hrbp_base = _safe(row.get(chrbp,'')) if chrbp else (_safe(row.get(csup,'')) if csup else '')

    area_v = _safe(row.get(ca,'')) if ca else ''

    div_v = _safe(row.get(cd,'')) if cd else ''

    puesto_v = _safe(row.get(cpu,'')) if cpu else ''



    return {

        'matricula': _norm_id(row.get(cmat, '')) if cmat else '',

        'nombre': nombre,

        'puesto': puesto_v,

        'area': area_v,

        'division': div_v,

        'meses': m_int,

        'anios': anios_v,

        'meses_resto': meses_r,

        'antiguedad': f"{anios_v}a {meses_r}m",

        'dias_para_cumple': abs(d_int) if d_int is not None else None,

        'dias_para_aniversario': d_int,

        'dias': d_int,

        'en_retraso': d_int is not None and d_int < 0,

        'elegibilidad': elegib_v,

        'estado': elegib_v,

        'proximo_periodo': prox_v,

        'proximo': prox_v,

        'fecha_ingreso': fi_txt,

        'esquema': esq,

        'truncas': truncas, 'pendientes': pend, 'vencidas': venc, 'total_vacaciones': total,

        'fecha_aniversario': prox_v,

        'supervisor': _safe(row.get(csup,'')) if csup else '',

        'hrbp': _resolver_hrbp_tyc(hrbp_base, area=area_v, departamento=div_v, division=div_v, puesto=puesto_v),

        'supervisor_matricula': _norm_id(row.get(cmat_sup, '')) if cmat_sup else '',

        'correo': _norm_email(row.get(cem, '')) if cem else ''

    }



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# ENDPOINTS

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/api/ping')

def api_ping():

    """Endpoint ultra-liviano para health-check de Electron. No toca el Excel."""

    return jsonify({'ok': True, 'status': 'alive'})



@app.route('/')
def index():
    response = send_from_directory(SCRIPT_DIR, 'index_vacaciones.html')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/api/colaborador/<identificador>/cadena', methods=['GET'])

def api_get_cadena_mando(identificador):

    """Retorna la cadena de mando completa (hacia arriba) de un colaborador."""

    cadena = _obtener_cadena_mando_recursiva(identificador)

    return jsonify({

        'ok': True,

        'identificador': identificador,

        'cadena': cadena,

        'niveles': len(cadena)

    })



@app.route('/api/jefe/<identificador>/equipo_recursivo', methods=['GET'])

def api_get_equipo_recursivo(identificador):

    """Retorna todo el equipo (directos e indirectos) de un jefe."""

    equipo = _obtener_equipo_recursivo(identificador)

    return jsonify({

        'ok': True,

        'jefe': identificador,

        'equipo': equipo,

        'total': len(equipo)

    })


@app.route('/api/colaborador/buscar', methods=['GET'])

def api_buscar_colaborador():

    """Busca colaboradores por nombre o matricula en el universo completo."""

    query = request.args.get('q', '').strip().upper()

    if len(query) < 2:

        return jsonify({'ok': True, 'resultados': [], 'mensaje': 'Query muy corto'})

    

    _, mat_a_info, err = _cargar_tabla_maestra_jefes()

    if err: return jsonify({'ok': False, 'error': err}), 500

    

    resultados = []

    for mat, info in mat_a_info.items():

        if query in mat or query in info['nombre'].upper():

            resultados.append(info)

            if len(resultados) >= 50: break

            

    return jsonify({

        'ok': True,

        'query': query,

        'resultados': resultados,

        'total': len(resultados)

    })


@app.route('/api/jefe/resumen_organizacional', methods=['GET'])

def api_get_resumen_org():

    """Retorna metricas clave de la estructura jerarquica."""

    colab_a_jefe, mat_a_info, err = _cargar_tabla_maestra_jefes()

    if err: return jsonify({'ok': False, 'error': err}), 500

    

    jefes_unicos = set(v['mat_jefe'] for v in colab_a_jefe.values() if v['mat_jefe'])

    total_colab = len(mat_a_info)

    total_jefes = len(jefes_unicos)

    

    return jsonify({

        'ok': True,

        'total_colaboradores': total_colab,

        'total_jefes': total_jefes,

        'ratio_jefatura': round(total_colab / total_jefes, 2) if total_jefes > 0 else 0,

        'fuente': os.path.basename(_MAESTRO_FULL_PATH or 'Desconocida')

    })


@app.route('/api/diagnostico', methods=['GET'])
def api_diagnostico():
    """Diagnóstico rápido del estado de carga de datos."""
    candidatos = _candidatos_vacaciones()
    df, err = cargar_datos()
    df_m, err_m = _cargar_maestro_universo()

    c_mat = _col(df, 'Matricula') if df is not None else None
    c_jefe = _col(df, 'Jefe_Directo', 'Jefe Directo', 'Supervisor') if df is not None else None
    c_sup_mat = _col(df, 'Matricula Supervisor') if df is not None else None
    c_mat_m = _col(df_m, 'Matricula') if df_m is not None else None
    c_sup_m = _col(df_m, 'Matricula Supervisor') if df_m is not None else None

    return jsonify({
        'candidatos_vacaciones': [c['archivo'] for c in candidatos],
        'vacaciones_cargado': df is not None,
        'vacaciones_error': err,
        'vacaciones_filas': len(df) if df is not None else 0,
        'vacaciones_columnas': list(df.columns[:20]) if df is not None else [],
        'col_matricula_vac': c_mat,
        'col_jefe_directo': c_jefe,
        'col_matricula_sup': c_sup_mat,
        'maestro_archivo': os.path.basename(_MAESTRO_FULL_PATH or 'NO ENCONTRADO'),
        'maestro_cargado': df_m is not None,
        'maestro_error': err_m,
        'maestro_filas': len(df_m) if df_m is not None else 0,
        'col_matricula_maestro': c_mat_m,
        'col_sup_maestro': c_sup_m,
    })


@app.route('/api/setup/verificar', methods=['GET'])

def api_setup_verificar():

    """Verifica la integridad del entorno (carpetas, config, etc.)"""

    return jsonify({

        'ok': True,

        'config_encontrada': bool(_PA_CONFIG_PATH),

        'cola_dir': COLA_DIR,

        'cola_dir_valido': os.path.isdir(COLA_DIR),

        'maestro_full': os.path.basename(_MAESTRO_FULL_PATH or 'NO ENCONTRADO'),

        'universo_habilitado': _UNIVERSO_COMPLETO_ENABLED,

        'delay_pa': _PA_DELAY_SECONDS,

        'usuario': _usuario_session()

    })

@app.route('/api/auth/me', methods=['GET'])
def api_auth_me():
    email = session.get('usuario_email', '')
    if email:
        return jsonify({'ok': True, 'autenticado': True, 'usuario': _usuario_session()})
    return jsonify({'ok': True, 'autenticado': False})





@app.route('/api/auth/validar_email', methods=['POST'])

def api_auth_validar_email():

    payload = request.get_json(silent=True) or {}

    email = _norm_email(payload.get('email', ''))

    if not email or '@' not in email:

        return jsonify({'ok': False, 'error': 'Ingresa un correo valido'}), 400



    emails, err = _emails_autorizados()

    if err:

        return jsonify({'ok': False, 'error': err}), 500



    return jsonify({

        'ok': True,

        'autorizado': email in emails,

        'email': email,

        'total_autorizados': len(emails)

    })





@app.route('/api/auth/login', methods=['POST'])

def api_auth_login():

    payload = request.get_json(silent=True) or {}

    email = _norm_email(payload.get('email', ''))

    if not email or '@' not in email:

        return jsonify({'ok': False, 'error': 'Ingresa un correo valido'}), 400



    emails, err = _emails_autorizados()

    if err:

        return jsonify({'ok': False, 'error': err}), 500



    if email not in emails:

        return jsonify({'ok': False, 'error': 'Correo no habilitado en tabla maestra'}), 403



    session['usuario_email'] = email

    session['usuario_nombre'] = payload.get('nombre', '') or email.split('@')[0]



    return jsonify({

        'ok': True,

        'autenticado': True,

        'usuario': _usuario_session(),

        'total_autorizados': len(emails)

    })





@app.route('/api/auth/logout', methods=['POST'])

def api_auth_logout():

    session.clear()

    return jsonify({'ok': True})





@app.route('/api/supervisores_emails', methods=['GET'])

def api_get_supervisores_emails():

    auto_mode = str(request.args.get('auto', '1')).strip().lower() not in ('0', 'false', 'no')

    if auto_mode:

        items, err = _supervisores_automaticos_desde_maestra()

        if err:

            return jsonify({'ok': False, 'error': err, 'items': [], 'total': 0}), 400

        webhook_personal = _PA_CONFIG.get('teams_webhook_personal_url', '') or ''

        return jsonify({'ok': True, 'items': items, 'total': len(items), 'teams_webhook_personal_url': webhook_personal})



    data = _supervisores_cfg_map()

    items = []

    for nombre, cfg in data.items():

        n = str(nombre or '').strip()

        e = _norm_email((cfg or {}).get('email', ''))

        if n and e:

            items.append({

                'nombre': n,

                'email': e,

                'mensaje': str((cfg or {}).get('mensaje', '') or '').strip(),

                'aviso': str((cfg or {}).get('aviso', '') or '').strip(),

                'recomendacion': str((cfg or {}).get('recomendacion', '') or '').strip(),

            })

    items.sort(key=lambda x: x['nombre'].upper())

    webhook_personal = _PA_CONFIG.get('teams_webhook_personal_url', '') or ''

    return jsonify({'ok': True, 'items': items, 'total': len(items), 'teams_webhook_personal_url': webhook_personal})





_SUPERVISORES_CACHE = {'key': None, 'data': None}
_SUPERVISORES_COMPUTE_LOCK = threading.Lock()  # solo un calculo a la vez

def _supervisores_refresh_bg():
    """Recalcula el cache de supervisores en un hilo de fondo (sin bloquear requests)."""
    if not _SUPERVISORES_COMPUTE_LOCK.acquire(blocking=False):
        return  # ya hay un calculo en curso, no lanzar otro
    try:
        _supervisores_automaticos_desde_maestra(force=True)
    except Exception as e:
        print('[SUP-BG] Error en refresco de fondo:', e)
    finally:
        _SUPERVISORES_COMPUTE_LOCK.release()



def _supervisores_cache_key():

    """Genera una clave de caché basada en mtime de los archivos fuente."""

    try:

        mt_vac  = os.path.getmtime(VACACIONES_DATA_FILE) if os.path.isfile(VACACIONES_DATA_FILE) else 0

        mt_mae  = os.path.getmtime(_MAESTRO_FULL_PATH)   if _MAESTRO_FULL_PATH and os.path.isfile(_MAESTRO_FULL_PATH) else 0

        ruta_obj = _encontrar_objetivo()

        mt_obj = os.path.getmtime(ruta_obj) if ruta_obj and os.path.isfile(ruta_obj) else 0

        return ('objetivos_v2', mt_vac, mt_mae, mt_obj)

    except Exception:

        return None



def _supervisores_automaticos_desde_maestra(force=False):
    """Wrapper con cache stale-while-revalidate + lock para serializar el calculo pesado.

    - Si hay cache fresco: lo devuelve al instante.
    - Si hay cache obsoleto: lo devuelve al instante y refresca en fondo.
    - Si no hay cache: calcula bajo lock (max 1 calculo a la vez) para no saturar
      los threads del servidor con multiples lecturas de Excel en paralelo.
    """
    if not force and _SUPERVISORES_CACHE.get('data') is not None:
        _cache_key_now = _supervisores_cache_key()
        if _cache_key_now and _cache_key_now == _SUPERVISORES_CACHE.get('key'):
            return _SUPERVISORES_CACHE['data'], None  # cache fresco
        # Cache obsoleto: devolver igual y refrescar en fondo
        threading.Thread(target=_supervisores_refresh_bg, daemon=True).start()
        return _SUPERVISORES_CACHE['data'], None

    # No hay cache (o force): calcular bajo lock para no saturar threads
    acquired = _SUPERVISORES_COMPUTE_LOCK.acquire(timeout=25)
    if not acquired:
        # No conseguimos el lock en 25s: si entretanto otro hilo lleno el cache, usarlo
        if _SUPERVISORES_CACHE.get('data') is not None:
            return _SUPERVISORES_CACHE['data'], None
        return [], 'Servidor ocupado calculando supervisores, reintenta en unos segundos'
    try:
        # Doble-check: otro hilo pudo completar el calculo mientras esperabamos el lock
        if not force and _SUPERVISORES_CACHE.get('data') is not None:
            _ck = _supervisores_cache_key()
            if _ck and _ck == _SUPERVISORES_CACHE.get('key'):
                return _SUPERVISORES_CACHE['data'], None
        return _supervisores_compute_impl()
    finally:
        _SUPERVISORES_COMPUTE_LOCK.release()


def _supervisores_compute_impl():
    """Calculo real (pesado): lee Excel de objetivos + maestra y arma el arbol de jefes."""
    df, err = cargar_objetivos()

    if err or df is None:

        return [], err or 'No se pudo cargar objetivos de vacaciones'

    c_act = _col(df, 'Activo', 'Activo/Cesado', 'Cesado')
    if c_act:
        df = df[df[c_act].fillna('').astype(str).str.upper().str.contains('ACTIVO', na=False)].copy()

    c_goz = _col(df, 'Dias Gozado', 'Dias Gozados', 'Días Gozado', 'Días Gozados')

    c_obj = _col(df, 'Objetivo')
    if c_obj:
        meta_vals = pd.to_numeric(df[c_obj], errors='coerce').fillna(0)
        if float(meta_vals.sum()) > 0:
            if c_goz:
                goz_vals = pd.to_numeric(df[c_goz], errors='coerce').fillna(0)
                df = df[(meta_vals - goz_vals) > 0].copy()
            else:
                df = df[meta_vals > 0].copy()

    if c_goz:
        dias_gozados = pd.to_numeric(df[c_goz], errors='coerce').fillna(0)
        df = df[dias_gozados > 0].copy()

    if df.empty:

        return [], 'No se encontraron personas activas con vacaciones registradas para el trimestre'



    c_sup = _col(df, 'Jefe_Directo', 'Jefe Directo', 'Supervisor', 'Jefe')

    if not c_sup:

        c_sup = _col(df, 'Business Partner', 'HRBP')



    c_mat_df  = _col(df, 'Matricula')

    c_sup_mat = _col(df, 'Matricula Supervisor')

    c_hrbp = _col(df, 'HRBP', 'Business Partner', 'BP')

    c_ger    = _col(df, 'Gerencia', 'Direccion', 'Vicepresidencia', 'Departamento')

    c_subger = _col(df, 'Subgerencia', 'Nombre Departamento', 'Division', 'Area')

    c_area   = _col(df, 'Nombre Area', 'Seccion')

    c_puesto = _col(df, 'Nombre Puesto', 'Puesto')

    c_correo = _col_correo(df)

    c_ap  = _col(df, 'Apellido Paterno')

    c_am  = _col(df, 'Apellido Materno')

    c_nom = _col(df, 'Nombre')

    c_comb = _col(df, 'Apellidos y Nombres', 'Apellidos_Nombres') if not c_ap else None



    # === NUEVO: Cargar tabla maestra para resolver emails de jefes ===

    # _cargar_tabla_maestra_jefes usa PersonalMaestroReporte y Matricula Supervisor

    # para mapear cada colaborador a su jefe directo (con email real).

    colab_a_jefe_mae, mat_a_info_mae, _mae_err = _cargar_tabla_maestra_jefes()

    if _mae_err:

        print(f'[WARN-JEFES] No se pudo cargar mapa de jefes desde maestra: {_mae_err}')



    # Construir indice de personas desde la tabla maestra completa (tiene emails)

    # para que _resolver_supervisor_identidad pueda encontrar emails por nombre/matricula

    df_mae_univ, _univ_err = _cargar_maestro_universo()

    idx = _build_personal_index(df_mae_univ) if df_mae_univ is not None else _build_personal_index(df)



    def _norm_mat_key(v):

        """Normaliza matricula a int-string sin ceros iniciales para comparacion."""

        try:

            s = str(v).strip()

            if s in ('', 'nan', 'None'):

                return ''

            return str(int(float(s))).zfill(10)

        except Exception:

            return _norm_id(str(v))



    cfg_map = _supervisores_cfg_map()

    sup_stats = {}

    if c_sup or c_mat_df:

        for _, row in df.iterrows():

            # === Estrategia 1: cruce por Matrícula en tabla maestra (método preciso) ===

            jefe_por_mat = None

            if c_mat_df and colab_a_jefe_mae:

                mat_colab = _norm_mat_key(row.get(c_mat_df, ''))

                if mat_colab:

                    jefe_por_mat = colab_a_jefe_mae.get(mat_colab)



            if jefe_por_mat:

                # Email y nombre del jefe obtenidos directamente de la tabla maestra

                sup = jefe_por_mat.get('nombre_jefe', '').strip()

                email_jefe = jefe_por_mat.get('email_jefe', '')

                mat_jefe = jefe_por_mat.get('mat_jefe', '')

                sup_ref = {'nombre': sup, 'email': email_jefe, 'matricula': mat_jefe}

            elif c_sup:

                # === Estrategia 2: fallback por nombre textual del Jefe Directo ===

                sup_mat_val = row.get(c_sup_mat, '') if c_sup_mat else ''

                sup_ref = _resolver_supervisor_identidad(

                    row.get(c_sup, ''),

                    sup_mat_val,

                    idx,

                )

                sup = sup_ref.get('nombre', '').strip()

            else:

                continue



            sup = sup_ref.get('nombre', '').strip()

            if not sup or sup == 'Sin Supervisor':

                continue

            key = sup.strip()

            if key not in sup_stats:

                sup_stats[key] = {

                    'total': 0,

                    'email_maestra': sup_ref.get('email', ''),

                    'matricula': sup_ref.get('matricula', ''),

                    'hrbps': set(),

                    'gerencias': set(),

                    'subgerencias': set(),

                    'areas': set(),

                    'colaboradores': set(),

                }

            sup_stats[key]['total'] += 1

            if not sup_stats[key].get('email_maestra') and sup_ref.get('email'):

                sup_stats[key]['email_maestra'] = sup_ref.get('email')

            hrbp = _safe(row.get(c_hrbp, '')) if c_hrbp else ''

            if hrbp:

                sup_stats[key]['hrbps'].add(hrbp)



            ger = _safe(row.get(c_ger, '')) if c_ger else ''

            subger = _safe(row.get(c_subger, '')) if c_subger else ''

            area = _safe(row.get(c_area, '')) if c_area else ''

            nombre_col = f"{_safe(row.get(c_ap, ''))} {_safe(row.get(c_am, ''))} {_safe(row.get(c_nom, ''))}".strip()

            if not nombre_col and c_comb:

                nombre_col = _safe(row.get(c_comb, ''))

            if ger:

                sup_stats[key]['gerencias'].add(ger)

            if subger:

                sup_stats[key]['subgerencias'].add(subger)

            if area:

                sup_stats[key]['areas'].add(area)

            if nombre_col:

                sup_stats[key]['colaboradores'].add(nombre_col)

    else:

        # Fallback para archivos de Talento y Cultura sin columna "Supervisor":

        # infiere un jefe por area usando el puesto y correo institucional.

        if not c_area or not c_puesto or not c_correo:

            return [], 'No se pudo inferir jefes por area (faltan columnas Area/Puesto/Correo)'



        role_order = ['GERENTE', 'SUBGERENTE', 'DIRECTOR', 'JEFE', 'COORDINADOR', 'BUSINESS PARTNER']



        def _rank_puesto(txt):

            t = _safe(txt).upper()

            for i, rk in enumerate(role_order):

                if rk in t:

                    return i

            return 99



        for area_val, g in df.groupby(c_area):

            area_txt = _safe(area_val)

            if not area_txt:

                continue



            g2 = g[g[c_correo].fillna('').astype(str).str.contains('@', na=False)].copy()

            if g2.empty:

                continue



            g2['__r'] = g2[c_puesto].apply(_rank_puesto)

            cand = g2[g2['__r'] < 99].sort_values(['__r'])

            if cand.empty:

                continue



            row = cand.iloc[0]

            sup = f"{_safe(row.get(c_ap, ''))} {_safe(row.get(c_am, ''))} {_safe(row.get(c_nom, ''))}".strip()

            if not sup and c_comb:

                sup = _safe(row.get(c_comb, ''))

            if not sup:

                continue



            key = sup.strip()

            if key not in sup_stats:

                sup_stats[key] = {

                    'total': 0,

                    'email_maestra': _norm_email(row.get(c_correo, '')),

                    'matricula': _norm_id(row.get(_col(df, 'Matricula') or '', '')),

                    'hrbps': set(),

                    'gerencias': set(),

                    'subgerencias': set(),

                    'areas': set(),

                    'colaboradores': set(),

                }



            sup_stats[key]['total'] += int(len(g))

            if not sup_stats[key].get('email_maestra'):

                sup_stats[key]['email_maestra'] = _norm_email(row.get(c_correo, ''))



            ger = _safe(row.get(c_ger, '')) if c_ger else ''

            subger = _safe(row.get(c_subger, '')) if c_subger else ''

            hrbp = _safe(row.get(c_hrbp, '')) if c_hrbp else ''

            if hrbp:

                sup_stats[key]['hrbps'].add(hrbp)

            if ger:

                sup_stats[key]['gerencias'].add(ger)

            if subger:

                sup_stats[key]['subgerencias'].add(subger)

            sup_stats[key]['areas'].add(area_txt)



            for _, r2 in g.iterrows():

                nombre_col = f"{_safe(r2.get(c_ap, ''))} {_safe(r2.get(c_am, ''))} {_safe(r2.get(c_nom, ''))}".strip()

                if not nombre_col and c_comb:

                    nombre_col = _safe(r2.get(c_comb, ''))

                if nombre_col:

                    sup_stats[key]['colaboradores'].add(nombre_col)


    items = []

    default_em = _resolver_default_email()

    for sup, meta in sup_stats.items():

        cfg = cfg_map.get(sup, {}) if isinstance(cfg_map.get(sup, {}), dict) else {}

        em_cfg = _norm_email(cfg.get('email', ''))

        em = em_cfg or _norm_email(meta.get('email_maestra', ''))

        if not em and _DEMO_JEFE_ENABLED and _nombre_cmp_key(_nombre_supervisor_canonico(sup)) == _nombre_cmp_key(_nombre_supervisor_canonico(_DEMO_JEFE_NOMBRE)):

            em = _norm_email(_DEMO_JEFE_EMAIL)

        origen = 'config' if em_cfg else ('maestra' if em else 'default')

        if not em:

            em = default_em

        items.append({

            'nombre': sup,

            'email': em,

            'hrbp': next(iter(sorted(list(meta.get('hrbps', set())), key=lambda x: x.upper())), '') if meta.get('hrbps') else '',

            'hrbps': sorted(list(meta.get('hrbps', set())), key=lambda x: x.upper()),

            'es_vista_ejecutiva': bool(meta.get('es_vista_ejecutiva', False)),

            'mensaje': str(cfg.get('mensaje', '') or '').strip(),

            'aviso': str(cfg.get('aviso', '') or '').strip(),

            'recomendacion': str(cfg.get('recomendacion', '') or '').strip(),

            'total_colaboradores': int(meta.get('total', 0)),

            'origen_email': origen,

            'matricula_supervisor': meta.get('matricula', ''),

            'gerencias': _filtrar_gerencias_supervisor(sup, meta.get('gerencias', set())),

            'subgerencias': sorted(list(meta.get('subgerencias', set())), key=lambda x: x.upper()),

            'areas': sorted(list(meta.get('areas', set())), key=lambda x: x.upper()),

            'colaboradores': sorted(list(meta.get('colaboradores', set())), key=lambda x: x.upper()),

        })



    # Deduplicar variantes del mismo nombre (ej: "LOLI PICON CARMEN" vs "LOLI PICON, CARMEN").

    merged = {}

    for it in items:

        k = _nombre_cmp_key(it.get('nombre', ''))

        if not k:

            continue

        cur = merged.get(k)

        if not cur:

            merged[k] = dict(it)

            continue

        cur['total_colaboradores'] = int(cur.get('total_colaboradores', 0)) + int(it.get('total_colaboradores', 0))

        if ',' in str(it.get('nombre', '')) and ',' not in str(cur.get('nombre', '')):

            cur['nombre'] = it.get('nombre', cur.get('nombre', ''))

        if (not cur.get('email')) and it.get('email'):

            cur['email'] = it.get('email')

            cur['origen_email'] = it.get('origen_email', cur.get('origen_email', ''))

        cur['hrbps'] = sorted(set((cur.get('hrbps') or []) + (it.get('hrbps') or [])), key=lambda x: x.upper())

        if (not cur.get('hrbp')) and it.get('hrbp'):

            cur['hrbp'] = it.get('hrbp')

        cur['es_vista_ejecutiva'] = bool(cur.get('es_vista_ejecutiva', False) or it.get('es_vista_ejecutiva', False))

        cur['gerencias'] = sorted(set((cur.get('gerencias') or []) + (it.get('gerencias') or [])), key=lambda x: x.upper())

        cur['subgerencias'] = sorted(set((cur.get('subgerencias') or []) + (it.get('subgerencias') or [])), key=lambda x: x.upper())

        cur['areas'] = sorted(set((cur.get('areas') or []) + (it.get('areas') or [])), key=lambda x: x.upper())

        cur['colaboradores'] = sorted(set((cur.get('colaboradores') or []) + (it.get('colaboradores') or [])), key=lambda x: x.upper())



    items = list(merged.values())

    items.sort(key=lambda x: (-x.get('total_colaboradores', 0), x.get('nombre', '').upper()))



    # Segunda pasada: fusionar variantes informales de nombre (ej: "Gabriel Chag" es subconjunto

    # de tokens de "CHANG CHANG GABRIEL ANDRES"). Conservamos la forma canónica (mas tokens)

    # y sumamos colaboradores.

    def _nombre_tokens(nombre):

        return set(re.sub(r'[^A-Z0-9 ]', '', _norm(nombre).upper()).split())



    # Ordenar por longitud de tokens descendente para que el canónico esté primero

    items.sort(key=lambda x: (-len(_nombre_tokens(x.get('nombre', ''))), -x.get('total_colaboradores', 0)))

    merged2 = []

    used = set()

    for i, a in enumerate(items):

        if i in used:

            continue

        tok_a = _nombre_tokens(a.get('nombre', ''))

        if len(tok_a) < 2:

            merged2.append(a)

            continue

        combined = dict(a)

        for j, b in enumerate(items):

            if j <= i or j in used:

                continue

            tok_b = _nombre_tokens(b.get('nombre', ''))

            # b es variante informal si sus tokens son subconjunto de a

            if len(tok_b) >= 2 and tok_b.issubset(tok_a) and tok_b != tok_a:

                combined['total_colaboradores'] = int(combined.get('total_colaboradores', 0)) + int(b.get('total_colaboradores', 0))

                if not combined.get('email') and b.get('email'):

                    combined['email'] = b['email']

                combined['areas'] = sorted(set((combined.get('areas') or []) + (b.get('areas') or [])), key=lambda x: x.upper())

                combined['colaboradores'] = sorted(set((combined.get('colaboradores') or []) + (b.get('colaboradores') or [])), key=lambda x: x.upper())

                used.add(j)

        merged2.append(combined)

        used.add(i)



    items = merged2

    items.sort(key=lambda x: (-x.get('total_colaboradores', 0), x.get('nombre', '').upper()))

    # Guardar en caché para próximas llamadas (recalcular la clave aquí, ya que el
    # cálculo es la parte lenta; la clave por mtime es instantánea)
    _cache_key = _supervisores_cache_key()
    if _cache_key:

        _SUPERVISORES_CACHE['key']  = _cache_key

        _SUPERVISORES_CACHE['data'] = items

    return items, None





@app.route('/api/supervisores_auto', methods=['GET'])

def api_get_supervisores_auto():

    items, err = _supervisores_automaticos_desde_maestra()

    if err:

        return jsonify({'ok': False, 'error': err, 'items': [], 'total': 0}), 400

    return jsonify({'ok': True, 'items': items, 'total': len(items)})





@app.route('/api/jefes_equipo_arbol', methods=['GET'])

def api_jefes_equipo_arbol():

    """Retorna árbol de jefes -> áreas -> conteo de personas.

    RAPIDO: usa _supervisores_automaticos_desde_maestra() directamente,

    sin llamar a _armar_contenido_supervisor() por cada jefe (que era O(n*m) lento).

    """

    items, err = _supervisores_automaticos_desde_maestra()

    if err:

        return jsonify({'ok': False, 'error': err, 'arbol': []}), 400



    arbol = []

    for jefe_info in items:

        jefe_nombre = jefe_info.get('nombre', '')

        jefe_email  = jefe_info.get('email', '')

        jefe_total  = jefe_info.get('total_colaboradores', 0)

        jefe_hrbp   = jefe_info.get('hrbp', '')

        jefe_hrbps  = jefe_info.get('hrbps', []) or ([] if not jefe_hrbp else [jefe_hrbp])



        # Construir nodos de area desde los datos ya computados

        areas_raw = jefe_info.get('areas', [])

        area_nodes = []

        for area_nombre in sorted(areas_raw):

            area_nodes.append({

                'nombre': area_nombre,

                'total': 0,     # no tenemos conteo por area en este nivel

                'personas': []  # se carga on-demand via /api/preview-supervisor

            })



        jefe_node = {

            'nombre':             jefe_nombre,

            'email':              jefe_email,

            'hrbp':               jefe_hrbp,

            'hrbps':              jefe_hrbps,

            'es_vista_ejecutiva': bool(jefe_info.get('es_vista_ejecutiva', False)),

            'total_colaboradores': jefe_total,

            'gerencias':          jefe_info.get('gerencias', []),

            'subgerencias':       jefe_info.get('subgerencias', []),

            'areas':              area_nodes,

            'matricula_supervisor': jefe_info.get('matricula_supervisor', ''),

            'origen_email':       jefe_info.get('origen_email', ''),

        }

        arbol.append(jefe_node)



    return jsonify({'ok': True, 'arbol': arbol, 'total_jefes': len(arbol)})







@app.route('/api/personas_autocomplete', methods=['GET'])

def api_personas_autocomplete():

    """Lista nombres + correo desde la data propia de vacaciones para autocompletar y autollenar email."""

    q = str(request.args.get('q', '') or '').strip()

    try:

        limite = int(request.args.get('limite', '1200') or 1200)

    except Exception:

        limite = 1200

    limite = max(50, min(5000, limite))



    df, err = cargar_datos()

    if err or df is None:

        return jsonify({'ok': False, 'error': err or 'No se pudo cargar datos', 'items': [], 'total': 0}), 400



    idx = _build_personal_index(df)

    vals = list((idx.get('by_name') or {}).values())

    manuales = _overrides_a_items_busqueda()

    if not vals and not manuales:

        return jsonify({'ok': True, 'items': [], 'total': 0})



    q_up = q.upper()

    out = []

    seen = set()

    for data in vals:

        nombre = str((data or {}).get('nombre', '') or '').strip()

        email = _norm_email((data or {}).get('email', '') or '')

        matricula = str((data or {}).get('matricula', '') or '').strip()

        if not nombre:

            continue

        if q_up and q_up not in nombre.upper() and q_up not in matricula.upper():

            continue

        key = f"{nombre.upper()}|{matricula}"

        if key in seen:

            continue

        seen.add(key)

        if email:

            seen.add(f"{nombre.upper()}|{email}")

        out.append({'nombre': nombre, 'email': email, 'matricula': matricula})



    # Incluir overrides manuales (p. ej., colaboradores agregados por People Analytics).

    for it in manuales:

        nombre = str(it.get('nombre', '') or '').strip()

        email = _norm_email(it.get('email', '') or '')

        matricula = str(it.get('matricula', '') or '').strip()

        if not nombre:

            continue

        if q_up and q_up not in nombre.upper() and q_up not in matricula.upper() and q_up not in email.upper():

            continue

        # Dedup por email si ya existe un registro con mismo email+nombre

        email_key = f"{nombre.upper()}|{email}" if email else f"{nombre.upper()}|{matricula}"

        if email_key in seen:

            continue

        seen.add(email_key)

        seen.add(f"{nombre.upper()}|{matricula}")

        out.append({'nombre': nombre, 'email': email, 'matricula': matricula})



    out.sort(key=lambda x: (x.get('nombre', '').upper(), x.get('matricula', '')))

    if len(out) > limite:

        out = out[:limite]

    return jsonify({'ok': True, 'items': out, 'total': len(out)})





@app.route('/api/hrbp/config', methods=['GET'])
def api_hrbp_config_get():
    """Devuelve HRBPs disponibles, asignación por área y áreas detectadas del Excel."""
    hrbp_disponibles = list(_PA_CONFIG.get('hrbp_disponibles', []))
    hrbp_asignacion  = dict(_PA_CONFIG.get('hrbp_asignacion', {}))

    # Auto-detectar HRBPs únicos desde la tabla maestra (columna BP/Business Partner/HRBP)
    areas_cfg = sorted(set(
        list(_PA_CONFIG.get('vacaciones_seccion_supervisor', {}).keys()) +
        list(_PA_CONFIG.get('vacaciones_area_supervisor', {}).keys())
    ))
    try:
        df, _ = cargar_datos()
        if df is not None and not df.empty:
            # Detectar áreas
            col_area = next((c for c in df.columns if c.lower() in ('area', 'área', 'division', 'división', 'gerencia')), None)
            if not col_area:
                col_area = next((c for c in df.columns if 'area' in c.lower()), None)
            if col_area:
                areas_dato = [str(v).strip().upper() for v in df[col_area].dropna().unique() if str(v).strip()]
                areas_cfg = sorted(set(areas_cfg + areas_dato))
    except Exception:
        pass

    # Incluir BPs canónicos del sistema; excluir a Gabriel Chang (no es del equipo TyC activo)
    # y evitar duplicar a Carlos Jara (ya está como JARA ORTIZ, CARLOS HUMBERTO)
    _EXCLUIR_BP = {'GABRIEL CHANG'}
    try:
        for _toks, _nombre in _BP_CANON:
            if not _nombre:
                continue
            if _nombre.upper() in _EXCLUIR_BP:
                continue
            if _nombre not in hrbp_disponibles:
                hrbp_disponibles.append(_nombre)
    except Exception:
        pass

    return jsonify({'ok': True, 'hrbp_disponibles': hrbp_disponibles,
                    'hrbp_asignacion': hrbp_asignacion, 'areas': areas_cfg})


@app.route('/api/hrbp/config', methods=['POST'])
def api_hrbp_config_save():
    """Guarda la asignación de HRBPs por área en pa_config."""
    payload = request.get_json(silent=True) or {}
    hrbp_disponibles = [str(h).strip() for h in payload.get('hrbp_disponibles', []) if str(h).strip()]
    hrbp_asignacion  = {str(k).strip(): str(v).strip()
                        for k, v in (payload.get('hrbp_asignacion') or {}).items()
                        if str(k).strip() and str(v).strip()}
    _guardar_pa_config({'hrbp_disponibles': hrbp_disponibles, 'hrbp_asignacion': hrbp_asignacion})
    return jsonify({'ok': True, 'guardado': len(hrbp_asignacion)})


@app.route('/api/smtp/test', methods=['POST'])
def api_smtp_test():
    """Prueba las credenciales SMTP y opcionalmente las actualiza en pa_config.json."""
    payload = request.get_json(silent=True) or {}
    email_nuevo   = str(payload.get('smtp_email',    '') or '').strip()
    pass_nueva    = str(payload.get('smtp_password', '') or '').strip()
    solo_prueba   = bool(payload.get('solo_prueba', False))

    # Si se enviaron credenciales nuevas, actualizar config en memoria y disco
    if email_nuevo and pass_nueva and not solo_prueba:
        ok_cfg, err_cfg = _guardar_pa_config({'smtp_email': email_nuevo, 'smtp_password': pass_nueva})
        if not ok_cfg:
            return jsonify({'ok': False, 'error': f'No se pudo guardar config: {err_cfg}'}), 500

    # Probar con las credenciales actuales (o las recien guardadas)
    import smtplib
    smtp_user = _PA_CONFIG.get('smtp_email', '') or SMTP_EMAIL
    smtp_pass = _PA_CONFIG.get('smtp_password', '') or SMTP_PASSWORD
    if not smtp_user or not smtp_pass:
        return jsonify({'ok': False, 'error': 'SMTP no configurado'}), 400
    try:
        with smtplib.SMTP('smtp.office365.com', 587, timeout=15) as srv:
            srv.ehlo(); srv.starttls(); srv.ehlo()
            srv.login(smtp_user, smtp_pass)
        return jsonify({'ok': True, 'mensaje': f'SMTP OK — credenciales validas para {smtp_user}',
                        'actualizado': bool(email_nuevo and not solo_prueba)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e),
                        'ayuda': 'Genera una nueva contrasena de app en https://aka.ms/mysecurityinfo'})


@app.route('/api/cola-pa/estado-in', methods=['GET'])
def api_cola_estado_in():
    """Devuelve cuántos archivos hay en in/ y su estado."""
    archivos = []
    for nombre in os.listdir(COLA_IN_DIR):
        if not nombre.lower().endswith('.json'):
            continue
        p = os.path.join(COLA_IN_DIR, nombre)
        if not os.path.isfile(p):
            continue
        import stat as _stat
        mt = os.path.getmtime(p)
        archivos.append({
            'nombre': nombre,
            'tamano_kb': round(os.path.getsize(p) / 1024, 1),
            'creado': datetime.fromtimestamp(mt).strftime('%d/%m/%Y %H:%M:%S'),
            'minutos_esperando': round((time.time() - mt) / 60, 1),
        })
    archivos.sort(key=lambda x: x['creado'])
    smtp_ok = bool(_PA_CONFIG.get('smtp_email') and _PA_CONFIG.get('smtp_password'))
    return jsonify({'ok': True, 'en_cola_in': len(archivos), 'archivos': archivos,
                    'smtp_configurado': smtp_ok,
                    'smtp_email': _PA_CONFIG.get('smtp_email', SMTP_EMAIL)})


@app.route('/api/tester-config', methods=['GET', 'POST'])

def api_tester_config():

    global _TEST_EMAIL_DESTINO

    if request.method == 'GET':

        return jsonify({

            'ok': True,

            'email_tester': _PA_CONFIG.get('vacaciones_test_email', ''),

            'email_default': _resolver_default_email(),

        })



    payload = request.get_json(silent=True) or {}

    email = _norm_email(payload.get('email_tester', payload.get('email', '')))

    if email and '@' not in email:

        return jsonify({'ok': False, 'error': 'Correo tester invalido'}), 400

    ok, err = _guardar_pa_config({'vacaciones_test_email': email})

    if not ok:

        return jsonify({'ok': False, 'error': err or 'No se pudo guardar la configuracion'}), 500

    _TEST_EMAIL_DESTINO = email

    _PA_CONFIG['vacaciones_test_email'] = email

    return jsonify({'ok': True, 'email_tester': email})





@app.route('/api/colaboradores_editor', methods=['GET'])

def api_colaboradores_editor():

    q = str(request.args.get('q', '') or '').strip().upper()

    area_f = str(request.args.get('area', '') or '').strip().upper()

    jefe_f = str(request.args.get('jefe', '') or '').strip().upper()

    try:

        limite = int(request.args.get('limite', '500') or 500)

    except Exception:

        limite = 500

    limite = max(20, min(2000, limite))



    df, err = cargar_datos()

    if err or df is None:

        return jsonify({'ok': False, 'error': err or 'No se pudo cargar datos'}), 400



    dfo, _ = cargar_objetivos()

    obj_map = {}

    if dfo is not None and not dfo.empty:

        c_obj_mat = _col(dfo, 'Matricula')

        c_obj_goz = _col(dfo, 'Dias Gozado')

        c_obj_total = _col(dfo, 'Suma de Dias', 'Dias Total')

        c_obj_pend = _col(dfo, 'Vac Pendiente', 'Pendiente')

        c_obj_venc = _col(dfo, 'Vac. Vencidas', 'Vencidas')

        c_obj_trun = _col(dfo, 'Vac. Trunco', 'Trunco')

        c_obj_sup = _col(dfo, 'Supervisor', 'Jefe_Directo', 'Jefe Directo', 'Jefe', 'Business Partner', 'HRBP')

        if c_obj_mat:

            for _, r_obj in dfo.iterrows():

                mk = _id_key(r_obj.get(c_obj_mat, ''))

                if not mk:

                    continue

                obj_map[mk] = {

                    'dias_gozados_q1': _to_float0(r_obj.get(c_obj_goz, 0)) if c_obj_goz else 0.0,

                    'saldo_total_dias': _to_float0(r_obj.get(c_obj_total, 0)) if c_obj_total else 0.0,

                    'vac_pendiente': _to_float0(r_obj.get(c_obj_pend, 0)) if c_obj_pend else 0.0,

                    'vac_vencidas': _to_float0(r_obj.get(c_obj_venc, 0)) if c_obj_venc else 0.0,

                    'vac_truncas': _to_float0(r_obj.get(c_obj_trun, 0)) if c_obj_trun else 0.0,

                    'supervisor': _safe(r_obj.get(c_obj_sup, '')) if c_obj_sup else '',

                }



    c_mat = _col(df, 'Matricula')

    c_nom = _col(df, 'Apellidos y Nombres', 'Nombre Colaborador', 'Nombre')

    c_mail = _col_correo(df)

    c_dep = _col(df, 'Nombre Departamento', 'Departamento', 'Division', 'Subgerencia')

    c_area = _col(df, 'Nombre Area', 'Area', 'Seccion')

    c_pue = _col(df, 'Nombre Puesto', 'Puesto', 'Cargo')

    c_sup = _col(df, 'Jefe_Directo', 'Jefe Directo', 'Supervisor', 'Jefe', 'Business Partner', 'HRBP')



    # Fallback de jefatura para archivos sin columna supervisor: inferir por area.

    jefe_por_area = {}

    try:

        sup_items, _sup_err = _supervisores_automaticos_desde_maestra()

        if isinstance(sup_items, list):

            for it in sup_items:

                sup_name = str((it or {}).get('nombre', '') or '').strip()

                if not sup_name:

                    continue

                for area_name in (it or {}).get('areas', []) or []:

                    ak = _norm(area_name).upper()

                    if ak and ak not in jefe_por_area:

                        jefe_por_area[ak] = sup_name

    except Exception:

        jefe_por_area = {}



    rows = []

    overrides_map = _leer_colaboradores_overrides()

    for _, row in df.iterrows():

        mat_key = _id_key(row.get(c_mat, '')) if c_mat else ''

        obj = obj_map.get(mat_key, {})

        area_val = _safe(row.get(c_area, '')) if c_area else ''

        dep_val = _safe(row.get(c_dep, '')) if c_dep else ''

        jefe_val = _safe(row.get(c_sup, '')) if c_sup else ''

        if not jefe_val and obj.get('supervisor'):

            jefe_val = _safe(obj.get('supervisor', ''))

        if not jefe_val:

            area_key = _norm(area_val or dep_val).upper()

            jefe_val = jefe_por_area.get(area_key, '') if area_key else ''

        item = {

            'matricula': _safe(row.get(c_mat, '')),

            'nombre': _safe(row.get(c_nom, '')),

            'email': _norm_email(row.get(c_mail, '')) if c_mail else '',

            'departamento': dep_val,

            'area': area_val,

            'puesto': _safe(row.get(c_pue, '')) if c_pue else '',

            'jefe': jefe_val,

            'dias_gozados_q1': float(obj.get('dias_gozados_q1', 0.0)),

            'saldo_total_dias': float(obj.get('saldo_total_dias', 0.0)),

            'vac_pendiente': float(obj.get('vac_pendiente', 0.0)),

            'vac_vencidas': float(obj.get('vac_vencidas', 0.0)),

            'vac_truncas': float(obj.get('vac_truncas', 0.0)),

        }

        

        # Aplicar overrides de vacaciones si existen

        if mat_key and isinstance(overrides_map.get(mat_key), dict):

            ov = overrides_map.get(mat_key)

            for vac_field in ['dias_gozados_q1', 'vac_pendiente', 'vac_vencidas', 'vac_truncas', 'saldo_total_dias']:

                if vac_field in ov and ov.get(vac_field) is not None:

                    item[vac_field] = float(ov.get(vac_field, 0))

        

        haystack = ' '.join([item['matricula'], item['nombre'], item['email'], item['departamento'], item['area'], item['puesto'], item['jefe']]).upper()

        if q and q not in haystack:

            continue

        if area_f and area_f not in item['area'].upper() and area_f not in item['departamento'].upper():

            continue

        if jefe_f and jefe_f not in item['jefe'].upper():

            continue

        rows.append(item)



    rows.sort(key=lambda x: (x.get('nombre', '').upper(), x.get('matricula', '')))

    return jsonify({'ok': True, 'items': rows[:limite], 'total': len(rows)})





@app.route('/api/colaboradores_editor', methods=['POST'])

def api_colaboradores_editor_save():

    payload = request.get_json(silent=True) or {}

    matricula = _id_key(payload.get('matricula', ''))

    if not matricula:

        return jsonify({'ok': False, 'error': 'Matricula requerida'}), 400



    allowed = {

        'nombre': str(payload.get('nombre', '') or '').strip(),

        'email': _norm_email(payload.get('email', '')),

        'departamento': str(payload.get('departamento', '') or '').strip(),

        'area': str(payload.get('area', '') or '').strip(),

        'puesto': str(payload.get('puesto', '') or '').strip(),

        'supervisor': str(payload.get('jefe', '') or '').strip(),

    }

    

    # Agregar campos de vacaciones si se proporcionan

    for vac_field in ['dias_gozados_q1', 'vac_pendiente', 'vac_vencidas', 'vac_truncas', 'saldo_total_dias']:

        try:

            val = float(payload.get(vac_field, 0))

            if val >= 0:

                allowed[vac_field] = val

        except (ValueError, TypeError):

            pass

    

    if allowed['email'] and '@' not in allowed['email']:

        return jsonify({'ok': False, 'error': 'Correo invalido'}), 400



    with _overrides_lock:

        data = _leer_colaboradores_overrides()

        prev = data.get(matricula, {}) if isinstance(data.get(matricula, {}), dict) else {}

        merged = dict(prev)

        merged.update({k: v for k, v in allowed.items() if v not in (None, '')})

        data[matricula] = merged

        _guardar_colaboradores_overrides(data)



    _invalidate_data_caches()

    return jsonify({'ok': True, 'matricula': matricula, 'override': merged})





@app.route('/api/supervisores_emails', methods=['POST'])

def api_set_supervisores_emails():

    payload = request.get_json(silent=True) or {}

    raw = payload.get('items', [])

    if not isinstance(raw, list):

        return jsonify({'ok': False, 'error': 'Formato invalido'}), 400



    cfg_map = {}

    mapping_legacy = {}

    for it in raw:

        if not isinstance(it, dict):

            continue

        nombre = str(it.get('nombre', '')).strip()

        email = _norm_email(it.get('email', ''))

        if not nombre or not email or '@' not in email:

            continue

        cfg_map[nombre] = {

            'email': email,

            'mensaje': str(it.get('mensaje', '') or '').strip(),

            'aviso': str(it.get('aviso', '') or '').strip(),

            'recomendacion': str(it.get('recomendacion', '') or '').strip(),

        }

        mapping_legacy[nombre] = email



    ok, err = _guardar_pa_config({

        'supervisores_config': cfg_map,

        'supervisores_emails': mapping_legacy,

    })

    if not ok:

        return jsonify({'ok': False, 'error': err or 'No se pudo guardar pa_config'}), 500



    # NotificaciÃ³n Teams personal al guardar correos destino

    total_guardados = len(cfg_map)

    def _notif():

        items_card = [{'titulo': 'Total configurados:', 'valor': f'{total_guardados} correo(s)'}]

        for nm, em in list(mapping_legacy.items())[:10]:

            items_card.append({'titulo': nm[:40], 'valor': em})

        if total_guardados > 10:

            items_card.append({'titulo': '...', 'valor': f'y {total_guardados - 10} mas'})

        _notificar_teams_personal(

            titulo='Correos destino actualizados - Sistema Vacaciones',

            detalle_items=items_card,

            color='Good' if total_guardados > 0 else 'Warning'

        )

    threading.Thread(target=_notif, daemon=True).start()



    return jsonify({'ok': True, 'total': total_guardados})



@app.route('/api/resumen')

def api_resumen():

    df, err = cargar_datos()

    if err: return jsonify({'ok': False, 'error': err})

    e = '_elegibilidad'

    return jsonify({

        'ok': True, 'total': len(df),

        'elegibles': int((df[e]=='Elegible').sum()) if e in df.columns else 0,

        'proximos':  int((df[e]=='Proximo (< 1 mes)').sum()) if e in df.columns else 0,

        'alerta':    int((df[e]=='En alerta (1-2 meses)').sum()) if e in df.columns else 0,

        'no_aplica': int((df[e]=='No aplica').sum()) if e in df.columns else 0,

        'retraso':   int((df[e]=='En retraso').sum()) if e in df.columns else 0,

    })



@app.route('/api/ranking')

def api_ranking():

    df, err = cargar_datos()

    if err: return jsonify({'ok': False, 'error': err})



    fa, fd, fs = request.args.get('area',''), request.args.get('division',''), request.args.get('supervisor','')

    fe, fp, fb = request.args.get('elegibilidad',''), request.args.get('puesto',''), request.args.get('buscar','')

    lim = int(request.args.get('limite', 2000))



    ca = _col(df,'Nombre Area','Area')

    cd = _col(df,'Nombre Departamento','Division')

    cs = _col(df,'Supervisor', 'Jefe_Directo', 'Jefe Directo', 'Jefe', 'Business Partner', 'HRBP')

    cp = _col(df,'Nombre Puesto')



    if fa and ca: df = df[df[ca]==fa]

    if fd and cd: df = df[df[cd]==fd]

    if fs and cs: df = df[df[cs].str.upper().str.contains(fs.upper(), na=False)]

    if fe and fe != 'todos' and '_elegibilidad' in df.columns:

        _mapa_fe = {

            'elegibles':  'Elegible',

            'proximos':   'Proximo',

            'alerta':     'En alerta',

            'no_aplica':  'No aplica',

            'retraso':    'En retraso',

            'Con retraso':'En retraso',

            'Proximo':    'Proximo',

            'En alerta':  'En alerta',

        }

        fe_norm = _mapa_fe.get(fe, fe)

        df = df[df['_elegibilidad'].str.startswith(fe_norm, na=False)]

    if fp and cp: df = df[df[cp].str.upper().str.contains(fp.upper(), na=False)]



    if '_dias_aniv' in df.columns:

        df = df.sort_values('_dias_aniv', ascending=True, na_position='last')

    elif '_meses' in df.columns:

        df = df.sort_values('_meses', ascending=False, na_position='last')



    df = df.head(lim)

    hoy = date.today()

    regs = [_persona(row, hoy) for _, row in df.iterrows()]

    # ── Cargar objetivos para los KPIs de cumplimiento ──────────────────────
    dfo, _oerr = cargar_objetivos()
    c_obj = c_goz = c_cum = c_com = c_hrbp = c_act = None
    df_act = pd.DataFrame()
    total_obj = dias_objetivo_total = dias_gozados_real = cumplieron = 0
    personas_pendientes = dias_pendientes = obligatorios = 0
    pct_cumpl = 0.0

    if dfo is not None:
        c_obj  = next((c for c in dfo.columns if 'objetivo' in c.lower()), None)
        c_goz  = next((c for c in dfo.columns if 'gozado' in c.lower() or 'registrad' in c.lower()), None)
        c_cum  = next((c for c in dfo.columns if 'cumplimiento' in c.lower()), None)
        c_com  = next((c for c in dfo.columns if 'comentario' in c.lower()), None)
        c_hrbp = next((c for c in dfo.columns if 'hrbp' in c.lower() or 'bp' in c.lower()), None)
        c_act  = next((c for c in dfo.columns if 'activo' in c.lower() or 'cesado' in c.lower()), None)
        df_act = dfo[dfo[c_act].str.upper().str.contains('ACTIVO', na=False)] if c_act else dfo
        total_obj = len(df_act)
        dias_objetivo_total = float(pd.to_numeric(df_act[c_obj], errors='coerce').fillna(0).sum()) if c_obj else 0.0
        if c_cum:
            deficit_dias = int(df_act[c_cum][df_act[c_cum] < 0].sum())
            dias_gozados_real = dias_objetivo_total + deficit_dias
            cumplieron = int((df_act[c_cum] >= 0).sum())
            personas_pendientes = int((df_act[c_cum] < 0).sum())
            dias_pendientes = abs(deficit_dias)
        else:
            dias_gozados_real = float(pd.to_numeric(df_act[c_goz], errors='coerce').fillna(0).sum()) if c_goz else 0.0
        pct_cumpl = round(float(dias_gozados_real) / float(dias_objetivo_total) * 100, 1) if dias_objetivo_total > 0 else 0
        if c_com:
            obligatorios = int(df_act[c_com].str.upper().str.contains('OBLIGATOR', na=False).sum())
    # ────────────────────────────────────────────────────────────────────────

    obj_kpis = {

            'colaboradores': int(total_obj),

            'dias_objetivo': dias_objetivo_total,

            'dias_gozados': dias_gozados_real,

            'cumplimiento': pct_cumpl,

            'cumplieron': cumplieron,

            'pendiente_meta': dias_pendientes,

            'personas_pendientes': personas_pendientes,

            'obligatorios': obligatorios,

        }

    if c_hrbp:

        hrbps_raw = df_act[c_hrbp].dropna().unique().tolist()

        hrbps_clean = set()

        for name in hrbps_raw:

            n = _norm(name).upper()

            if 'GABRIEL CHANG' in n: continue

            if 'CARLOS' in n and 'JARA' in n:

                hrbps_clean.add('JARA ORTIZ, CARLOS HUMBERTO')

            else:

                hrbps_clean.add(name.strip())

        hrbps = sorted(list(hrbps_clean))



        # === NUEVAS MÉTRICAS ===

        c_reg = next((c for c in dfo.columns if 'registrada' in c.lower()), None)

        con_registradas = int((pd.to_numeric(df_act[c_reg], errors='coerce').fillna(0) > 0).sum()) if c_reg else 0

        obj_kpis['con_registradas'] = con_registradas



        c_venc = next((c for c in dfo.columns if 'vencida' in c.lower()), None)

        obj_kpis['total_vencidas_dias'] = int(pd.to_numeric(df_act[c_venc], errors='coerce').fillna(0).sum()) if c_venc else 0

        obj_kpis['personas_con_vencidas'] = int((pd.to_numeric(df_act[c_venc], errors='coerce').fillna(0) > 0).sum()) if c_venc else 0



        c_dept_o = next((c for c in dfo.columns if 'departamento' in c.lower()), None)

        top_vencidas_dept = []

        if c_dept_o and c_venc:

            _g = df_act.copy()

            _g['_v'] = pd.to_numeric(_g[c_venc], errors='coerce').fillna(0)

            _dg = _g.groupby(c_dept_o)['_v'].agg(['sum', 'count']).reset_index()

            _dg = _dg[_dg['sum'] > 0].sort_values('sum', ascending=False).head(8)

            for _, row in _dg.iterrows():

                top_vencidas_dept.append({'dept': str(row[c_dept_o])[:35], 'dias': int(row['sum']), 'personas': int(row['count'])})



        objetivo_por_hrbp = []

        if c_hrbp and c_obj and c_reg:

            _g2 = df_act.copy()

            _g2['_o'] = pd.to_numeric(_g2[c_obj], errors='coerce').fillna(0)

            _g2['_r'] = pd.to_numeric(_g2[c_reg], errors='coerce').fillna(0)

            _hg = _g2.groupby(c_hrbp).agg(objetivo=('_o','sum'), registrado=('_r','sum'), personas=(c_hrbp,'count')).reset_index()

            _hg = _hg[_hg['objetivo'] > 0].sort_values('objetivo', ascending=False)

            for _, row in _hg.iterrows():

                bp = str(row[c_hrbp]).strip()

                if not bp or bp.lower() in ('nan','none'): continue

                ov, rv = int(row['objetivo']), int(row['registrado'])

                objetivo_por_hrbp.append({'hrbp': bp, 'objetivo': ov, 'registrado': rv, 'pct': round(rv/ov*100,1) if ov>0 else 0, 'personas': int(row['personas'])})



    else:

        top_vencidas_dept = []

        objetivo_por_hrbp = []



    fuente = _resumen_fuente_datos()

    return jsonify({'ok': True,

        'resumen': {'total':len(df),

            'elegibles':int((df['_elegibilidad']=='Elegible').sum()) if '_elegibilidad' in df.columns else 0,

            'proximos':int((df['_elegibilidad']=='Proximo (< 1 mes)').sum()) if '_elegibilidad' in df.columns else 0,

            'alerta':int((df['_elegibilidad']=='En alerta (1-2 meses)').sum()) if '_elegibilidad' in df.columns else 0,

            'no_aplica':int((df['_elegibilidad']=='No aplica').sum()) if '_elegibilidad' in df.columns else 0,

            'retraso':int((df['_elegibilidad']=='En retraso').sum()) if '_elegibilidad' in df.columns else 0},

        'obj_kpis': obj_kpis,

        'fuente_datos': fuente
    })

@app.route('/api/filtros')
def api_filtros():
    df, err = cargar_datos()

    if err: return jsonify({'ok': False, 'error': err})

    ca, cd, cp = _col(df,'Nombre Area','Area'), _col(df,'Nombre Departamento','Division'), _col(df,'Nombre Puesto')

    return jsonify({'ok': True,

        'areas':sorted(df[ca].dropna().unique().tolist()) if ca else [],

        'divisiones':sorted(df[cd].dropna().unique().tolist()) if cd else [],

        'puestos':sorted(df[cp].dropna().unique().tolist()) if cp else []})









@app.route('/api/cola-pa/liberar-todo', methods=['POST'])

def api_cola_liberar_todo():

    num = _liberar_todo_pa_ahora()

    return jsonify({'ok': True, 'total': num})


@app.route('/api/cola-pa/enviar-smtp-ahora', methods=['POST'])
def api_cola_enviar_smtp_ahora():
    """Procesa inmediatamente todos los archivos en in/ via SMTP.
    Útil cuando Power Automate no funciona."""
    import threading
    def _run():
        try:
            _smtp_procesar_in.__globals__['time'] = time  # asegurar import
        except Exception:
            pass
        _smtp_procesar_in._espera_override = 0  # sin espera minima
        _smtp_procesar_in()
    # Forzar sin espera: copiar la logica directamente para no modificar la funcion
    procesados_dir = os.path.join(COLA_DIR, 'procesados')
    errores_dir    = os.path.join(COLA_DIR, 'errores')
    os.makedirs(procesados_dir, exist_ok=True)
    os.makedirs(errores_dir, exist_ok=True)
    resultados = []
    for nombre in list(os.listdir(COLA_IN_DIR)):
        if not nombre.lower().endswith('.json'):
            continue
        src = os.path.join(COLA_IN_DIR, nombre)
        if not os.path.isfile(src):
            continue
        try:
            with open(src, 'r', encoding='utf-8') as f:
                entradas = json.load(f)
            if not isinstance(entradas, list):
                entradas = [entradas]
            enviados, errores_item = 0, []
            for entrada in entradas:
                email  = (entrada.get('email_jefe') or entrada.get('email_destino_real') or '').strip()
                asunto = entrada.get('asunto', 'Alertas Vacaciones USIL')
                html   = entrada.get('mensaje_html', '')
                nombre_dest = entrada.get('nombre_jefe', '')
                if not email or '@' not in email or not html:
                    continue
                ok, err = _enviar_correo_smtp(email, nombre_dest, asunto, html)
                if ok:
                    enviados += 1
                else:
                    errores_item.append(err)
            destino = procesados_dir if not errores_item else errores_dir
            shutil.move(src, os.path.join(destino, nombre))
            resultados.append({'archivo': nombre, 'enviados': enviados, 'errores': errores_item})
        except Exception as e:
            resultados.append({'archivo': nombre, 'error': str(e)})
    total_ok = sum(r.get('enviados', 0) for r in resultados)
    total_err = sum(len(r.get('errores', [])) for r in resultados)
    return jsonify({'ok': True, 'archivos': len(resultados), 'enviados': total_ok, 'errores': total_err, 'detalle': resultados})


@app.route('/api/cola-pa/enviar-outlook-ahora', methods=['POST'])
def api_cola_enviar_outlook_ahora():
    """Procesa in/ y errores/ via Outlook COM (no requiere SMTP AUTH).
    Requiere Outlook de escritorio abierto con sesión activa."""
    import subprocess, re as _re
    script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'enviar_cola_outlook.py')
    if not os.path.isfile(script):
        return jsonify({'ok': False, 'error': 'enviar_cola_outlook.py no encontrado junto a servidor.py'})
    try:
        result = subprocess.run(
            [sys.executable, script],
            capture_output=True, text=True, timeout=120, encoding='utf-8', errors='replace'
        )
        stdout = result.stdout or ''
        stderr = result.stderr or ''
        enviados, errores_n = 0, 0
        for linea in stdout.splitlines():
            m = _re.search(r'(\d+) enviados?,\s*(\d+) error', linea)
            if m:
                enviados, errores_n = int(m.group(1)), int(m.group(2))
                break
        ok = enviados > 0 or (result.returncode == 0 and errores_n == 0)
        return jsonify({
            'ok': ok,
            'enviados': enviados,
            'errores': errores_n,
            'log': stdout[-3000:],
            'stderr': stderr[-500:] if stderr else ''
        })
    except subprocess.TimeoutExpired:
        return jsonify({'ok': False, 'error': 'Timeout: Outlook tardó más de 2 minutos. ¿Está Outlook abierto?'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/init')

def api_init():

    df, err = cargar_datos()

    if err: return jsonify({'ok': False, 'error': err})

    e = '_elegibilidad'

    ca, cd, cp = _col(df,'Nombre Area','Area'), _col(df,'Nombre Departamento','Division'), _col(df,'Nombre Puesto')



    # Cargar objetivos para KPIs de cumplimiento y HRBPs

    obj_kpis = {}

    hrbps = []

    dfo, oerr = cargar_objetivos()

    if dfo is not None:

        c_obj  = next((c for c in dfo.columns if 'objetivo' in c.lower()), None)

        c_goz  = next((c for c in dfo.columns if 'gozado' in c.lower() or 'registrad' in c.lower()), None)

        c_cum  = next((c for c in dfo.columns if 'cumplimiento' in c.lower()), None)

        c_tot  = next((c for c in dfo.columns if 'suma' in c.lower() and 'dia' in c.lower()), None)

        c_com  = next((c for c in dfo.columns if 'comentario' in c.lower()), None)

        c_hrbp = next((c for c in dfo.columns if 'hrbp' in c.lower() or 'bp' in c.lower()), None)

        c_act  = next((c for c in dfo.columns if 'activo' in c.lower() or 'cesado' in c.lower()), None)



        # Solo activos

        df_act = dfo

        if c_act:

            df_act = dfo[dfo[c_act].str.upper().str.contains('ACTIVO', na=False)]



        total_obj = len(df_act)

        dias_objetivo_total = float(pd.to_numeric(df_act[c_obj], errors='coerce').fillna(0).sum()) if c_obj else 0.0

        # DÃ­as gozados HACIA el objetivo: objetivo_total + suma(cumplimientos_negativos)

        # Equivale a sum(min(DiasGozados, Objetivo)) por persona â€” evita inflar con sobre-logros

        if c_cum:

            deficit_dias = int(df_act[c_cum][df_act[c_cum] < 0].sum())  # negativo

            dias_gozados_real = dias_objetivo_total + deficit_dias        # 1873 + (-945) = 928

            cumplieron = int((df_act[c_cum] >= 0).sum())

            personas_pendientes = int((df_act[c_cum] < 0).sum())

        else:

            deficit_dias = 0

            dias_gozados_real = float(pd.to_numeric(df_act[c_goz], errors='coerce').fillna(0).sum()) if c_goz else 0.0

            cumplieron = 0

            personas_pendientes = 0

        # Cumplimiento = % de dÃ­as logrados vs. objetivo (no % de personas)

        pct_cumpl = round(float(dias_gozados_real) / float(dias_objetivo_total) * 100, 1) if dias_objetivo_total > 0 else 0

        dias_pendientes = abs(deficit_dias)  # DÃAS que faltan para completar objetivo

        obligatorios = 0

        if c_com:

            obligatorios = int(df_act[c_com].str.upper().str.contains('OBLIGATOR', na=False).sum())

        obj_kpis = {

            'colaboradores': int(total_obj),

            'dias_objetivo': dias_objetivo_total,

            'dias_gozados': dias_gozados_real,

            'cumplimiento': pct_cumpl,

            'cumplieron': cumplieron,

            'pendiente_meta': dias_pendientes,       # DÃAS pendientes

            'personas_pendientes': personas_pendientes,  # PERSONAS sin cumplir

            'obligatorios': obligatorios,

        }

        if c_hrbp:

            # Normalizar nombres de HRBP para evitar duplicados como "Carlos Jara" vs "Jara Ortiz..."

            hrbps_raw = df_act[c_hrbp].dropna().unique().tolist()

            hrbps_clean = set()

            for name in hrbps_raw:

                n = _norm(name).upper()

                if 'GABRIEL CHANG' in n: continue 

                if 'CARLOS' in n and 'JARA' in n:

                    hrbps_clean.add('Carlos Jara')

                elif 'CESAR' in n and 'REYES' in n:

                    hrbps_clean.add('César Reyes')

                else:

                    hrbps_clean.add(name.strip())



            hrbps = sorted(list(hrbps_clean))





    # Calcular KPIs para el frontend basados en vacaciones gozadas y saldo
    dfo_kpis, _ = cargar_objetivos()
    con_vacaciones = 0
    con_saldo = 0
    sin_vacaciones = 0

    if dfo_kpis is not None:
        c_goz_kpi = next((c for c in dfo_kpis.columns if 'gozado' in c.lower() or 'registrad' in c.lower()), None)
        c_total_kpi = next((c for c in dfo_kpis.columns if 'suma' in c.lower() and 'dia' in c.lower()), None)
        c_act_kpi = next((c for c in dfo_kpis.columns if 'activo' in c.lower() or 'cesado' in c.lower()), None)

        df_kpis = dfo_kpis
        if c_act_kpi:
            df_kpis = dfo_kpis[dfo_kpis[c_act_kpi].str.upper().str.contains('ACTIVO', na=False)]

        if c_goz_kpi:
            con_vacaciones = int(pd.to_numeric(df_kpis[c_goz_kpi], errors='coerce').fillna(0).gt(0).sum())
        if c_total_kpi:
            con_saldo = int(pd.to_numeric(df_kpis[c_total_kpi], errors='coerce').fillna(0).gt(0).sum())
        if c_goz_kpi and c_total_kpi:
            sin_vac_mask = (pd.to_numeric(df_kpis[c_goz_kpi], errors='coerce').fillna(0) == 0) & (pd.to_numeric(df_kpis[c_total_kpi], errors='coerce').fillna(0) > 0)
            sin_vacaciones = int(sin_vac_mask.sum())

    fuente = _resumen_fuente_datos()

    return jsonify({'ok': True,

        'resumen': {'total':len(df),

            'elegibles':int((df[e]=='Elegible').sum()) if e in df.columns else 0,

            'proximos':int((df[e]=='Proximo (< 1 mes)').sum()) if e in df.columns else 0,

            'alerta':int((df[e]=='En alerta (1-2 meses)').sum()) if e in df.columns else 0,

            'no_aplica':int((df[e]=='No aplica').sum()) if e in df.columns else 0,

            'retraso':int((df[e]=='En retraso').sum()) if e in df.columns else 0,

            'con_vacaciones': con_vacaciones,

            'con_saldo': con_saldo,

            'sin_vacaciones': sin_vacaciones},

        'obj_kpis': obj_kpis,

        'fuente_datos': fuente,

        'trimestre_control': fuente.get('trimestre_control', {}),

        'filtros': {

            'areas':sorted(df[ca].dropna().unique().tolist()) if ca else [],

            'divisiones':sorted(df[cd].dropna().unique().tolist()) if cd else [],

            'puestos':sorted(df[cp].dropna().unique().tolist()) if cp else [],

            'hrbps': hrbps}

    })





@app.route('/api/fuente_datos', methods=['GET'])

def api_fuente_datos():

    return jsonify({'ok': True, 'fuente_datos': _resumen_fuente_datos()})


@app.route('/api/diagnostico_datos', methods=['GET'])
def api_diagnostico_datos():
    df_base, err_base = cargar_datos()
    df_obj, err_obj = cargar_objetivos()
    df_m, err_m = _cargar_maestro_universo()

    llave = _detectar_llave_union(df_obj, df_m) if (df_obj is not None and df_m is not None) else {
        'llave': None, 'col_obj': None, 'col_maestro': None, 'overlap': 0
    }

    return jsonify({
        'ok': True,
        'fuente_datos': _resumen_fuente_datos(),
        'backend': {
            'datos_base_ok': err_base is None,
            'objetivos_ok': err_obj is None,
            'maestro_ok': err_m is None,
            'error_datos_base': err_base,
            'error_objetivos': err_obj,
            'error_maestro': err_m,
            'filas_base': int(len(df_base)) if df_base is not None else 0,
            'filas_objetivos': int(len(df_obj)) if df_obj is not None else 0,
            'filas_maestro': int(len(df_m)) if df_m is not None else 0,
            'columnas_objetivos': [str(c) for c in list(df_obj.columns)] if df_obj is not None else [],
            'columnas_maestro': [str(c) for c in list(df_m.columns)] if df_m is not None else [],
            'llave_union_detectada': llave,
        }
    })





@app.route('/api/trimestre/estado', methods=['GET'])

def api_trimestre_estado():

    fuente = _resumen_fuente_datos()

    return jsonify({

        'ok': True,

        'trimestre_control': fuente.get('trimestre_control', {}),

        'fuente_datos': fuente,

        'manual_configurado': str(_PA_CONFIG.get('vacaciones_trimestre_actual', '') or '').strip()

    })





@app.route('/api/trimestre/configurar', methods=['POST'])

def api_trimestre_configurar():

    payload = request.get_json(silent=True) or {}

    modo = str(payload.get('modo', 'manual') or 'manual').strip().lower()

    trimestre_raw = str(payload.get('trimestre', '') or '').strip()



    if modo not in {'manual', 'auto'}:

        return jsonify({'ok': False, 'error': 'Modo invalido. Usa manual o auto'}), 400



    if modo == 'auto':

        ok, err = _guardar_pa_config({'vacaciones_trimestre_actual': ''})

        if not ok:

            return jsonify({'ok': False, 'error': err or 'No se pudo guardar pa_config'}), 500

        _PA_CONFIG['vacaciones_trimestre_actual'] = ''

    else:

        tri = _normalizar_trimestre_txt(trimestre_raw)

        if not tri:

            return jsonify({'ok': False, 'error': 'Formato de trimestre invalido. Usa Q1-2026, Q2-2026, Q3-2026 o Q4-2026'}), 400

        candidatos = _candidatos_vacaciones()
        tri_label = tri.get('label', '')
        existe_trimestre = any(((_extraer_trimestre_de_nombre(item.get('archivo', '')) or {}).get('label', '') == tri_label) for item in candidatos)
        if not existe_trimestre:
            return jsonify({'ok': False, 'error': f'No se encontró archivo de objetivos para {tri_label}. Carga primero un archivo de ese trimestre.'}), 400

        ok, err = _guardar_pa_config({'vacaciones_trimestre_actual': tri.get('label', '')})

        if not ok:

            return jsonify({'ok': False, 'error': err or 'No se pudo guardar pa_config'}), 500

        _PA_CONFIG['vacaciones_trimestre_actual'] = tri.get('label', '')



    fuente = _resumen_fuente_datos()

    return jsonify({'ok': True, 'trimestre_control': fuente.get('trimestre_control', {}), 'fuente_datos': fuente})



@app.route('/api/config/cola_dir', methods=['GET', 'POST'])

def api_config_cola_dir():

    if request.method == 'GET':

        return jsonify({

            'ok': True,

            'dir': COLA_DIR,

            'exists': os.path.isdir(COLA_DIR),

            'historial_dir': COLA_FUENTES_DIR,

            'subcarpetas': {

                'in': COLA_IN_DIR,

                'pendientes': COLA_PENDIENTES_DIR,

                'archivados': COLA_ARCHIVADOS_DIR,

                'cancelados': COLA_CANCELADOS_DIR,

                'maestra': COLA_FUENTES_MAESTRA_DIR,

                'objetivos': COLA_FUENTES_OBJETIVOS_DIR,

            }

        })



    payload = request.get_json(silent=True) or {}

    dir_raw = str(payload.get('dir', '') or '').strip()

    if not dir_raw:

        return jsonify({'ok': False, 'error': 'Ingresa una ruta de carpeta compartida'}), 400



    dir_abs = os.path.abspath(os.path.expanduser(os.path.expandvars(dir_raw)))

    if not os.path.isabs(dir_abs):

        return jsonify({'ok': False, 'error': 'La ruta debe ser absoluta'}), 400



    try:

        _aplicar_rutas_cola(dir_abs)

    except Exception as e:

        return jsonify({'ok': False, 'error': f'No se pudo preparar la carpeta compartida: {e}'}), 500



    ok, err = _guardar_pa_config({'alertas_cola_dir': dir_abs})

    if not ok:

        return jsonify({'ok': False, 'error': err or 'No se pudo guardar pa_config'}), 500



    _PA_CONFIG['alertas_cola_dir'] = dir_abs

    return jsonify({

        'ok': True,

        'dir': COLA_DIR,

        'exists': os.path.isdir(COLA_DIR),

        'historial_dir': COLA_FUENTES_DIR,

        'subcarpetas': {

            'in': COLA_IN_DIR,

            'pendientes': COLA_PENDIENTES_DIR,

            'archivados': COLA_ARCHIVADOS_DIR,

            'cancelados': COLA_CANCELADOS_DIR,

            'maestra': COLA_FUENTES_MAESTRA_DIR,

            'objetivos': COLA_FUENTES_OBJETIVOS_DIR,

        }

    })


@app.route('/api/config/cola_dir/open', methods=['POST'])

def api_open_config_cola_dir():

    payload = request.get_json(silent=True) or {}

    dir_raw = str(payload.get('dir', '') or COLA_DIR).strip()

    dir_abs = os.path.abspath(os.path.expanduser(os.path.expandvars(dir_raw)))

    if not dir_abs:

        return jsonify({'ok': False, 'error': 'No hay carpeta configurada para abrir'}), 400

    if not os.path.isdir(dir_abs):

        return jsonify({'ok': False, 'error': 'La carpeta configurada no existe'}), 400

    try:

        os.startfile(dir_abs)

    except Exception as e:

        return jsonify({'ok': False, 'error': f'No se pudo abrir la carpeta: {e}'}), 500

    return jsonify({'ok': True, 'dir': dir_abs})


@app.route('/api/config/cola_dir/pick', methods=['POST'])

def api_pick_config_cola_dir():

    payload = request.get_json(silent=True) or {}

    dir_raw = str(payload.get('dir', '') or COLA_DIR).strip()

    dir_abs = os.path.abspath(os.path.expanduser(os.path.expandvars(dir_raw))) if dir_raw else COLA_DIR

    initial_dir = dir_abs if os.path.isdir(dir_abs) else os.path.dirname(dir_abs)

    root = None
    try:

        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        root.update_idletasks()
        root.lift()
        root.focus_force()
        selected = filedialog.askdirectory(
            parent=root,
            title='Selecciona la carpeta compartida para alertas',
            initialdir=initial_dir if os.path.isdir(initial_dir) else COLA_DIR,
            mustexist=False,
        )

    except Exception as e:

        return jsonify({'ok': False, 'error': f'No se pudo abrir el selector de carpetas: {e}'}), 500

    finally:

        if root is not None:

            try:

                root.destroy()

            except Exception:

                pass

    if not selected:

        return jsonify({'ok': False, 'cancelled': True})

    return jsonify({'ok': True, 'dir': os.path.abspath(selected), 'exists': os.path.isdir(selected)})



@app.route('/api/objetivos')

def api_objetivos():

    """Endpoint para datos del reporte de objetivos de vacaciones."""

    dfo, err = cargar_objetivos()

    if err:
        return jsonify({'ok': False, 'error': err})



    f_hrbp  = request.args.get('hrbp', '')

    f_area  = request.args.get('area', '')

    f_buscar = request.args.get('buscar', '')

    lim = int(request.args.get('limite', 2000))



    c_mat  = next((c for c in dfo.columns if 'matricula' in c.lower()), None)

    c_nom  = next((c for c in dfo.columns if 'apellidos y nombres' in c.lower() or 'apellidos' in c.lower()), None)

    c_ap   = next((c for c in dfo.columns if 'apellido paterno' in c.lower()), None)

    c_am   = next((c for c in dfo.columns if 'apellido materno' in c.lower()), None)

    c_no   = next((c for c in dfo.columns if c.lower() == 'nombre'), None)

    c_act  = next((c for c in dfo.columns if 'activo' in c.lower() or 'cesado' in c.lower()), None)

    c_suc  = next((c for c in dfo.columns if 'sucursal' in c.lower()), None)

    c_hrbp = next((c for c in dfo.columns if 'hrbp' in c.lower() or 'bp' in c.lower()), None)

    c_tipo = next((c for c in dfo.columns if 'tipo' in c.lower() and 'trabajador' in c.lower()), None)

    c_pues = next((c for c in dfo.columns if 'puesto' in c.lower()), None)

    c_dept = next((c for c in dfo.columns if 'departamento' in c.lower()), None)

    c_area = next((c for c in dfo.columns if c.lower() == 'area' or 'area' in c.lower()), None)

    c_sup = next((c for c in dfo.columns if 'jefe_directo' in c.lower() or 'jefe directo' in c.lower() or c.lower() == 'supervisor' or 'supervisor' in c.lower() or c.lower() == 'jefe' or 'jefe' in c.lower()), None)

    c_secc = next((c for c in dfo.columns if 'seccion' in c.lower()), None)

    c_fing = next((c for c in dfo.columns if 'fecha' in c.lower() and 'ingreso' in c.lower()), None)

    c_mven = next((c for c in dfo.columns if 'mes' in c.lower() and 'venc' in c.lower()), None)

    c_venc = next((c for c in dfo.columns if 'vencida' in c.lower()), None)

    c_pend = next((c for c in dfo.columns if 'pendiente' in c.lower()), None)

    c_trunc = next((c for c in dfo.columns if 'trunco' in c.lower()), None)

    c_total = next((c for c in dfo.columns if 'suma' in c.lower() and 'dia' in c.lower()), None)

    c_obj  = next((c for c in dfo.columns if 'objetivo' in c.lower()), None)

    c_rest = next((c for c in dfo.columns if 'restante' in c.lower()), None)

    c_goz  = next((c for c in dfo.columns if 'gozado' in c.lower() or 'registrad' in c.lower()), None)

    c_cum  = next((c for c in dfo.columns if 'cumplimiento' in c.lower()), None)

    c_com  = next((c for c in dfo.columns if 'comentario' in c.lower()), None)



    columnas_detectadas = [str(c) for c in dfo.columns]

    # Validación explícita para evitar respuestas vacías silenciosas cuando el Excel trae una hoja resumen.
    metric_cols = [c_obj, c_rest, c_goz, c_cum, c_total, c_pend, c_venc, c_trunc]
    if not c_mat:
        return jsonify({
            'ok': False,
            'error': 'No se encontró columna de matrícula en el archivo de objetivos. Verifique la hoja activa del Excel.',
            'debug': {
                'columnas_detectadas': columnas_detectadas[:80]
            }
        })
    if not any(metric_cols):
        return jsonify({
            'ok': False,
            'error': 'No se detectaron columnas métricas de vacaciones (objetivo/gozado/saldo/pendientes).',
            'debug': {
                'columnas_detectadas': columnas_detectadas[:80]
            }
        })

    print(f"[API-OBJ] Columnas clave -> matricula={c_mat}, objetivo={c_obj}, gozados={c_goz}, total={c_total}, pendientes={c_pend}, vencidas={c_venc}, truncos={c_trunc}")

    df = dfo.copy()

    # Solo activos (si la columna existe en el archivo de objetivos)
    if c_act:
        df = df[df[c_act].str.upper().str.contains('ACTIVO', na=False)]

    # Excluir cesados: cruzar contra el maestro y eliminar matriculas que ya no existen.
    # Esto elimina "fantasmas" — personas que estaban en el Excel de objetivos pero
    # ya fueron retiradas del PersonalMaestroReporte (fin de contrato, cese, etc.)
    try:
        dfm_act, _ = _cargar_maestro_universo()
        if dfm_act is not None and c_mat:
            c_mat_m = _col(dfm_act, 'Matricula')
            if c_mat_m:
                matriculas_activas = set(dfm_act[c_mat_m].dropna().apply(_norm_id))
                antes = len(df)
                df = df[df[c_mat].apply(_norm_id).isin(matriculas_activas)]
                eliminados = antes - len(df)
                if eliminados > 0:
                    print(f'[API-OBJ] {eliminados} registros excluidos por no estar en el maestro (cesados/fantasmas)')
    except Exception as e:
        print(f'[API-OBJ] Advertencia al cruzar con maestro: {e}')

    # Filtros

    if f_hrbp and c_hrbp:

        df = df[df[c_hrbp] == f_hrbp]

    if f_area and c_area:

        df = df[df[c_area] == f_area]

    if f_buscar and c_nom:

        df = df[df[c_nom].str.upper().str.contains(f_buscar.upper(), na=False)]



    # Solo filtrar por meta > 0 si realmente existe una meta distinta de cero.

    if c_obj:

        meta_vals = pd.to_numeric(df[c_obj], errors='coerce').fillna(0)

        if float(meta_vals.sum()) > 0:

            df = df[meta_vals > 0]



    # Ordenar por cumplimiento (los mÃ¡s negativos primero = mÃ¡s urgente)

    if c_cum:

        df = df.sort_values(c_cum, ascending=True, na_position='last')



    df = df.head(lim)

    maestro_org = {}
    dfm, errm = _cargar_maestro_universo()
    if dfm is not None and not errm:
        c_mat_m = _col(dfm, 'Matricula')
        c_ger_m = _col(dfm, 'Gerencia', 'Division', 'Direccion', 'Vicepresidencia')
        c_subger_m = _col(dfm, 'Subgerencia', 'Nombre Departamento', 'Departamento')
        c_area_m = _col(dfm, 'Nombre Area', 'Area')
        c_sup_m = _col(dfm, 'Supervisor', 'Jefe Directo', 'Jefe_Directo', 'Jefe')
        if c_mat_m:
            # zip de columnas en lugar de iterrows(): ~10-50x mas rapido sobre el maestro completo
            n = len(dfm)
            vacio = [''] * n
            mats = dfm[c_mat_m].tolist()
            gers = dfm[c_ger_m].tolist() if c_ger_m else vacio
            subs = dfm[c_subger_m].tolist() if c_subger_m else vacio
            ars  = dfm[c_area_m].tolist() if c_area_m else vacio
            sups = dfm[c_sup_m].tolist() if c_sup_m else vacio
            for mat_raw, g, s, a, sup in zip(mats, gers, subs, ars, sups):
                mat_m = _norm_id(mat_raw)
                if not mat_m:
                    continue
                maestro_org[mat_m] = {
                    'gerencia': _safe(g),
                    'subgerencia': _safe(s),
                    'area': _safe(a),
                    'supervisor': _safe(sup),
                }



    registros = []

    for _, row in df.iterrows():

        def _n(col):

            if not col: return None

            v = row.get(col)

            if pd.notna(v):

                try: return float(v)

                except: return None

            return None



        nombre_full = _safe(row.get(c_nom, '')) if c_nom else ''

        if not nombre_full:

            nombre_full = f"{_safe(row.get(c_ap, ''))} {_safe(row.get(c_am, ''))} {_safe(row.get(c_no, ''))}".strip()



        mat_val = _safe(row.get(c_mat, ''))

        org_meta = maestro_org.get(_norm_id(mat_val), {}) if mat_val else {}

        area_val = _safe(row.get(c_area, '')) if c_area else ''

        dept_val = _safe(row.get(c_dept, '')) if c_dept else ''

        secc_val = _safe(row.get(c_secc, '')) if c_secc else ''

        gerencia_val = org_meta.get('gerencia', '') or dept_val

        subgerencia_val = org_meta.get('subgerencia', '') or dept_val

        area_resuelta = area_val or org_meta.get('area', '') or subgerencia_val or dept_val

        supervisor_val = (_safe(row.get(c_sup, '')) if c_sup else '') or org_meta.get('supervisor', '')

        puesto_val = _safe(row.get(c_pues, '')) if c_pues else ''

        total_dias = _n(c_total) or 0

        dias_gozados = _n(c_goz) or 0

        pendientes = _n(c_pend) or 0

        vencidas = _n(c_venc) or 0

        truncos = _n(c_trunc) or 0

        objetivo = _n(c_obj) or 0

        por_programar_real = max(pendientes + vencidas + truncos, 0)

        por_programar_meta = max(objetivo - dias_gozados, 0)

        por_programar = por_programar_real if por_programar_real > 0 else por_programar_meta

        tiene_vacaciones_q1 = dias_gozados > 0

        tiene_saldo = total_dias > 0

        estado = 'Con vacaciones Q1' if tiene_vacaciones_q1 else ('Con saldo vacacional' if tiene_saldo else 'Sin saldo')



        registros.append({

            'matricula': mat_val,

            'nombre': nombre_full,

            'hrbp': _safe(row.get(c_hrbp, '')) if c_hrbp else '',

            'tipo': _safe(row.get(c_tipo, '')) if c_tipo else '',

            'puesto': puesto_val,

            'departamento': dept_val,

            'gerencia': gerencia_val,

            'subgerencia': subgerencia_val,

            'supervisor': supervisor_val,

            'area': area_resuelta,

            'seccion': secc_val or area_resuelta or subgerencia_val or dept_val,

            'fecha_ingreso': _safe(row.get(c_fing, ''))[:10] if c_fing else '',

            'mes_vencimiento': _safe(row.get(c_mven, '')) if c_mven else '',

            'vencidas': vencidas,

            'pendientes': pendientes,

            'truncos': truncos,

            'total_dias': total_dias,

            'objetivo': objetivo,

            'dias_restantes': _n(c_rest),

            'dias_gozados': dias_gozados,

            'cumplimiento': _n(c_cum),

            'por_programar': por_programar,

            'tiene_vacaciones_q1': tiene_vacaciones_q1,

            'tiene_saldo_vacacional': tiene_saldo,

            'estado': estado,

            'persona_objetivo': 'Si',

            'comentario': _safe(row.get(c_com, '')) if c_com else '',

        })



    print(f"[API-OBJ] Registros devolviendo: {len(registros)}")
    return jsonify({'ok': True, 'total': len(registros), 'registros': registros})



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# HELPERS: EMAIL POR JEFE (Power Automate JSON)

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# Email fijo del administrador (siempre recibe copia si no hay emails en el Excel)

_ADMIN_EMAIL = 'jlopezp@usil.edu.pe'

_ADMIN_NOMBRE = 'Jose Lopez'



_SUPERVISOR_ALIAS = {

}



_MSG_BASE_DEFAULT = (

    'Desde la Gerencia de Talento y Cultura, seguimos impulsando la eficiencia institucional '

    'y el bienestar de nuestros equipos. Para lograrlo, la gestion oportuna del descanso es fundamental.\n\n'

    'Hemos compartido contigo el Reporte de Vacaciones de las personas a tu cargo, donde se detalla la meta '

    'de dias a programar para este trimestre. Para visualizar este reporte, te solicitamos revises tu bandeja '

    'de entrada de tu cuenta de correo institucional.\n\n'

    '<strong>En el reporte encontraras:</strong>\n'

    '&bull; <strong>Meta de Vacaciones para el {trimestre}:</strong> El reporte indica una meta especifica para cada colaborador. '

    'Si algun miembro de tu equipo no tiene una meta establecida, eso no es impedimento para que pueda '

    'programar dias de descanso si fuera el caso.\n'

    '&bull; <strong>Vacaciones Corporativas:</strong> El periodo del 20 de julio al 02 de agosto ha sido definido como prioridad '

    'para la Institucion, debido a la suspension de actividades academicas.\n'

    '&bull; <strong>Cumplimiento:</strong> Si la meta trimestral de un colaborador es superior a los 14 dias del periodo de '

    'vacaciones corporativas, es necesario programar los dias restantes dentro del presente trimestre '

    'para alcanzar el objetivo asignado.\n\n'

    '<strong>Consejos para la programacion de vacaciones con tu equipo:</strong>\n'

    '1. <strong>Planifica y motiva la participacion:</strong> Reunete con tu equipo (de forma grupal o individual) '

    'para recibir sus propuestas de fechas.\n'

    '2. <strong>Busca acuerdos:</strong> Los periodos deben ser aceptados y acordados previamente contigo. A falta de acuerdo, '

    'el lider tiene la facultad de programar las vacaciones segun las necesidades del area.\n'

    '3. <strong>Registrar en Adryan:</strong> Una vez establecida la fecha, el colaborador debe registrarla en Adryan. '

    'Tu rol final sera la aprobacion en el mismo sistema para que el flujo administrativo sea exitoso.\n\n'

    'Tienes plazo hasta el <strong>{fecha_limite}</strong> para gestionar la programacion con tu equipo.\n\n'

    'Atentamente,<br>'

    '<strong>{hrbp}</strong><br>'

    'HRBP de la Subgerencia de Talento y Cultura<br>'

    '<em>"Formar personas con valores para que dejen su huella en el mundo"</em>'

)

_AVISO_DEFAULT = (

    'Atencion especial a quienes tienen dias Vencidos o cumplimiento negativo: '

    'estas personas deben programar sus vacaciones con urgencia para evitar mayor acumulacion.'

)

_RECOM_DEFAULT = (

    'Coordina con cada colaborador las fechas de descanso y confirmalas en el sistema este mes. '

    'Recuerda que las vacaciones Vencidas tienen prioridad sobre las Pendientes.'

)





def _resolver_default_email():

    """Devuelve el email por defecto en orden de prioridad."""

    return (_PA_CONFIG.get('teams_sender_email') or

            _PA_CONFIG.get('smtp_email') or

            _ADMIN_EMAIL)



def _resolver_email_supervisor_maestro(nombre):
    """Busca el email de un supervisor en la tabla maestra por nombre.
    Normaliza comas y acentos para cruzar 'VERA, NANCY' (vacaciones) con 'VERA NANCY' (maestro)."""
    if not nombre:
        return ''
    _, mat_a_info, _ = _cargar_tabla_maestra_jefes()
    if not mat_a_info:
        return ''
    target_key = _nombre_cmp_key(nombre)
    for info in mat_a_info.values():
        nom = info.get('nombre', '')
        if nom and _nombre_cmp_key(nom) == target_key:
            return _norm_email(info.get('email', ''))
    return ''





def _nombre_supervisor_canonico(nombre_raw):

    nombre = str(nombre_raw or '').strip()

    if not nombre:

        return nombre

    n_up = _nombre_cmp_key(nombre)

    for alias, canon in _SUPERVISOR_ALIAS.items():

        if _nombre_cmp_key(alias) in n_up:

            return canon

    return nombre





def _contexto_campania(fecha_base=None):

    ref = fecha_base or date.today()

    tri = _trimestre_vigente_info(_resumen_fuente_datos().get('archivo', ''))

    q = int(tri.get('q') or (((ref.month - 1) // 3) + 1))

    anio = int(tri.get('anio') or ref.year)

    q_start_month = ((q - 1) * 3) + 1

    q_start = date(anio, q_start_month, 1)

    semana_trimestre = ((ref - q_start).days // 7) + 1

    if semana_trimestre < 1:

        semana_trimestre = 1

    return {

        'frecuencia': 'semanal',

        'meta_periodo': 'trimestral',

        'trimestre': tri.get('label', f'Q{q}-{anio}'),

        'semana_trimestre': semana_trimestre,

        'fecha': ref.strftime('%d/%m/%Y')

    }


def _buscar_hrbp_destinatario(equipo_full, nombre_objetivo='', fallback=''):

    nombre_key = _nombre_cmp_key(nombre_objetivo)
    if nombre_key:
        for persona in (equipo_full or []):
            if _nombre_cmp_key(persona.get('nombre', '')) == nombre_key:
                hrbp = str(persona.get('hrbp', '') or '').strip()
                if hrbp:
                    return hrbp

    conteo = {}
    for persona in (equipo_full or []):
        hrbp = str((persona or {}).get('hrbp', '') or '').strip()
        if not hrbp:
            continue
        key = _nombre_cmp_key(hrbp)
        if not key:
            continue
        data = conteo.setdefault(key, {'nombre': hrbp, 'total': 0})
        data['total'] += 1

    if conteo:
        ganador = max(conteo.values(), key=lambda item: (item.get('total', 0), item.get('nombre', '')))
        if str(ganador.get('nombre', '')).strip():
            return str(ganador.get('nombre', '')).strip()

    return str(fallback or '').strip()


def _resolver_hrbp_area_override(equipo_full, hrbp_fallback=''):
    """Retorna el HRBP configurado manualmente para el área dominante del equipo.
    Si no hay override configurado, devuelve hrbp_fallback."""
    asignacion = _PA_CONFIG.get('hrbp_asignacion', {})
    if not asignacion:
        return hrbp_fallback
    area_count = {}
    for p in (equipo_full or []):
        area = str(p.get('area', '') or p.get('division', '') or '').strip().upper()
        if area:
            area_count[area] = area_count.get(area, 0) + 1
    if not area_count:
        return hrbp_fallback
    area_principal = max(area_count, key=area_count.get)
    for k, v in asignacion.items():
        if k.strip().upper() == area_principal:
            return str(v).strip() if v else hrbp_fallback
    return hrbp_fallback




def _seleccionar_candidatos_adicionales(equipo_list, ret_list, prox_list, sc_list, limite=5):

    usados = set()

    for arr in (ret_list or [], prox_list or [], sc_list or []):

        for p in arr:

            key = _id_key(p.get('matricula', ''))

            if key:

                usados.add(key)



    candidatos = []

    for p in (equipo_list or []):

        key = _id_key(p.get('matricula', ''))

        if key and key in usados:

            continue

        objetivo = _to_int0(p.get('objetivo', 0))

        cant_p = _to_float0(p.get('cantidad_pendiente', 0))

        total_v = _to_float0(p.get('total_vacaciones', 0))

        dias_aniv = p.get('dias_para_aniversario')

        if objetivo <= 0 and cant_p <= 0:

            continue

        score = cant_p + (total_v * 0.1)

        if dias_aniv is not None and 0 <= int(dias_aniv) <= 90:

            score += 3

        if dias_aniv is not None and int(dias_aniv) < 0:

            score += 5

        aviso, reco = _aviso_recomendacion_persona(p)

        item = dict(p)

        item['aviso'] = item.get('aviso') or aviso

        item['recomendacion'] = item.get('recomendacion') or reco

        item['_score_extra'] = score

        candidatos.append(item)



    candidatos.sort(key=lambda x: (x.get('_score_extra', 0), x.get('cantidad_pendiente', 0)), reverse=True)

    out = candidatos[:max(0, int(limite))]

    for it in out:

        it.pop('_score_extra', None)

    return out





def _resumen_meta_equipo(equipo_list):

    """Calcula avance de meta trimestral y metricas completas del equipo."""

    dias_objetivo = 0.0
    dias_gozados_hacia_meta = 0.0
    colaboradores_meta = 0
    colaboradores_cumplieron = 0
    dias_saldo_total = 0.0

    total_equipo = len(equipo_list or [])
    colaboradores_sin_meta = 0
    dias_vencidos_equipo = 0.0
    dias_truncos_equipo = 0.0
    dias_pendientes_equipo = 0.0
    colaboradores_con_obligatorias = 0
    colaboradores_al_dia = 0

    for p in (equipo_list or []):
        venc = _to_float0(p.get('vencidas', 0))
        pend = _to_float0(p.get('pendientes', 0))
        trun = _to_float0(p.get('truncas', 0))
        total_v = max(_to_float0(p.get('total_vacaciones', 0)), 0.0)

        dias_saldo_total += total_v
        dias_vencidos_equipo += max(venc, 0.0)
        dias_truncos_equipo += max(trun, 0.0)
        dias_pendientes_equipo += max(venc + pend, 0.0)

        if venc > 0 or pend > 0:
            colaboradores_con_obligatorias += 1

        obj = _to_float0(p.get('objetivo', 0))
        if obj <= 0:
            colaboradores_sin_meta += 1
            if venc <= 0 and pend <= 0:
                colaboradores_al_dia += 1
            continue

        colaboradores_meta += 1
        goz = _to_float0(p.get('gozados', 0))
        dias_objetivo += obj
        dias_gozados_hacia_meta += min(max(goz, 0.0), obj)
        if goz >= obj:
            colaboradores_cumplieron += 1
            colaboradores_al_dia += 1

    dias_obj_i = int(round(dias_objetivo))
    dias_goz_i = int(round(dias_gozados_hacia_meta))
    dias_pend_i = max(dias_obj_i - dias_goz_i, 0)
    pct = round((dias_gozados_hacia_meta / dias_objetivo) * 100, 1) if dias_objetivo > 0 else 0.0
    col_pend = max(colaboradores_meta - colaboradores_cumplieron, 0)
    pct_al_dia = round((colaboradores_al_dia / total_equipo) * 100, 1) if total_equipo > 0 else 0.0

    return {
        'colaboradores_meta': int(colaboradores_meta),
        'colaboradores_cumplieron': int(colaboradores_cumplieron),
        'colaboradores_pendientes': int(col_pend),
        'dias_objetivo': int(dias_obj_i),
        'dias_gozados_hacia_meta': int(dias_goz_i),
        'dias_pendientes_meta': int(dias_pend_i),
        'cumplimiento_meta_pct': float(pct),
        'dias_saldo_total': int(round(dias_saldo_total)),
        'total_equipo': int(total_equipo),
        'colaboradores_sin_meta': int(colaboradores_sin_meta),
        'dias_vencidos_equipo': int(round(dias_vencidos_equipo)),
        'dias_truncos_equipo': int(round(dias_truncos_equipo)),
        'dias_pendientes_equipo': int(round(dias_pendientes_equipo)),
        'colaboradores_con_obligatorias': int(colaboradores_con_obligatorias),
        'colaboradores_al_dia': int(colaboradores_al_dia),
        'pct_equipo_al_dia': float(pct_al_dia),
    }





def _get_maestro_matriculas_activas():
    """Devuelve el set de matrículas activas del maestro más reciente. Si falla, devuelve None (no filtrar)."""
    try:
        dfm, _ = _cargar_maestro_universo()
        if dfm is None:
            return None
        c_mat = _col(dfm, 'Matricula', 'Matrícula', 'matricula')
        if not c_mat:
            return None
        return {_id_key(str(v)) for v in dfm[c_mat].dropna() if str(v).strip()}
    except Exception:
        return None


def _equipo_objetivo_por_supervisor(df_obj_act, cols_obj, obj_lookup):

    out = {}

    if df_obj_act is None:

        return out

    c_sup = cols_obj.get('supervisor') if isinstance(cols_obj, dict) else None

    c_mat = cols_obj.get('matricula') if isinstance(cols_obj, dict) else None

    if not c_mat:

        return out

    # Filtrar colaboradores que ya no están en el maestro (personas que salieron de USIL)
    mats_activas = _get_maestro_matriculas_activas()

    if c_sup:

        for _, row in df_obj_act.iterrows():

            mat_key = _id_key(row.get(c_mat, ''))

            if not mat_key:

                continue

            # Excluir si no está en el maestro actual (ya no trabaja en USIL)
            if mats_activas is not None and mat_key not in mats_activas:
                continue

            item = obj_lookup.get(mat_key)

            if not item:

                continue

            sup = _nombre_supervisor_canonico(_safe(row.get(c_sup, '')) if c_sup else item.get('supervisor', ''))

            if not sup:

                continue

            out.setdefault(sup, {})

            out[sup][mat_key] = dict(item)

        return {k: list(v.values()) for k, v in out.items()}



    # Fallback: sin columna supervisor en objetivos -> inferir lider por area.

    c_area = cols_obj.get('area') if isinstance(cols_obj, dict) else None

    c_puesto = cols_obj.get('puesto') if isinstance(cols_obj, dict) else None

    c_nom = cols_obj.get('nombre') if isinstance(cols_obj, dict) else None

    c_ap = _col(df_obj_act, 'Apellido Paterno')

    c_am = _col(df_obj_act, 'Apellido Materno')

    c_no = _col(df_obj_act, 'Nombre')

    if not c_area or not c_puesto or not c_nom:

        return out



    role_order = ['GERENTE', 'SUBGERENTE', 'DIRECTOR', 'JEFE', 'COORDINADOR', 'BUSINESS PARTNER']



    def _rank_puesto(txt):

        t = _safe(txt).upper()

        for i, rk in enumerate(role_order):

            if rk in t:

                return i

        return 99



    for _, g in df_obj_act.groupby(c_area):

        g2 = g.copy()

        g2['__r'] = g2[c_puesto].apply(_rank_puesto)

        cand = g2[g2['__r'] < 99].sort_values(['__r'])

        if cand.empty:

            continue

        r_lider = cand.iloc[0]

        lider_full = ''

        if c_ap or c_am or c_no:

            lider_full = f"{_safe(r_lider.get(c_ap, ''))} {_safe(r_lider.get(c_am, ''))} {_safe(r_lider.get(c_no, ''))}".strip()

        lider = _nombre_supervisor_canonico(lider_full or _safe(r_lider.get(c_nom, '')))

        if not lider:

            continue



        out.setdefault(lider, {})

        for _, row in g2.iterrows():

            mat_key = _id_key(row.get(c_mat, ''))

            if not mat_key:

                continue

            item = obj_lookup.get(mat_key)

            if not item:

                continue

            out[lider][mat_key] = dict(item)



    # Vista ejecutiva global para el gerente de Talento y Cultura.

    gerentes = df_obj_act[df_obj_act[c_puesto].fillna('').astype(str).str.upper().str.match(r'(?:GERENTE DE TALENTO Y CULTURA).*') & ~df_obj_act[c_puesto].fillna('').astype(str).str.upper().str.contains('SUBGERENTE', na=False)]

    if not gerentes.empty:

        r_ger = gerentes.iloc[0]

        ger_full = ''

        if c_ap or c_am or c_no:

            ger_full = f"{_safe(r_ger.get(c_ap, ''))} {_safe(r_ger.get(c_am, ''))} {_safe(r_ger.get(c_no, ''))}".strip()

        gerente = _nombre_supervisor_canonico(ger_full or _safe(r_ger.get(c_nom, '')))

        if gerente:

            out[gerente] = {k: dict(v) for k, v in obj_lookup.items()}



    return {k: list(v.values()) for k, v in out.items()}





def _build_webhook_payload(contenido):

    mensaje_html = str((contenido or {}).get('mensaje_html', '') or '')

    mensaje_texto = str((contenido or {}).get('mensaje', '') or '')

    def _slim(items, max_items=50):

        out = []

        for p in (items or [])[:max_items]:

            out.append({

                'matricula': p.get('matricula', ''),

                'nombre': p.get('nombre', ''),

                'meta': p.get('objetivo', 0),

                'fecha_ingreso': p.get('fecha_ingreso', ''),

                'puesto': p.get('puesto', ''),

                'departamento': p.get('departamento', ''),

                'area': p.get('area', ''),

                'pendientes': p.get('pendientes', 0),

                'truncos': p.get('truncas', 0),

                'total_dias': p.get('total_vacaciones', 0),

                'gozados': p.get('gozados', 0),

                'por_programar': max(int(p.get('objetivo', 0) or 0) - int(p.get('gozados', 0) or 0), 0),

                'business_partner': p.get('hrbp', ''),

                'cantidad_pendiente': p.get('cantidad_pendiente', 0),

            })

        return out



    return {

        # Compatibilidad Power Automate: mantener campos legacy para flujos existentes.

        'destinatario': contenido.get('email_jefe', ''),

        'jefe': contenido.get('nombre_jefe', ''),

        'email_jefe': contenido.get('email_jefe', ''),

        'nombre_jefe': contenido.get('nombre_jefe', ''),

        'asunto': contenido.get('asunto', ''),

        'mensaje': mensaje_texto,

        # Campos HTML recomendados para Power Automate / Outlook.

        'mensaje_html': mensaje_html,

        'body_html': mensaje_html,

        'html_body': mensaje_html,

        'body_content_type': 'html' if mensaje_html else 'text',

        'is_html': bool(mensaje_html),

        'aviso': contenido.get('aviso', ''),

        'recomendacion': contenido.get('recomendacion', ''),

        'mensaje_teams': contenido.get('mensaje_teams', ''),

        'campania': contenido.get('campania', {}),

        'resumen': {

            'meta': contenido.get('meta_resumen', {}),

            'en_retraso': contenido.get('en_retraso', 0),

            'proximos': contenido.get('proximos', 0),

            'sin_cumplir': contenido.get('sin_cumplir', 0),

            'adicionales': contenido.get('adicionales', 0),

            'total_colaboradores': contenido.get('total_colaboradores', len(contenido.get('detalle_equipo_full', []))),

        },

        'detalle': {

            'retraso': _slim(contenido.get('detalle_retraso', [])),

            'proximos': _slim(contenido.get('detalle_proximos', [])),

            'sin_cumplir': _slim(contenido.get('detalle_sin_cumplir', [])),

            'adicionales': _slim(contenido.get('detalle_adicionales', [])),

            'equipo_full': _slim(contenido.get('detalle_equipo_full', [])),

        },

    }





def _build_personal_index(df_pm):

    idx = {'by_mat': {}, 'by_name': {}, 'by_short': {}}

    if df_pm is None or len(df_pm) == 0:

        return idx



    c_mat = _col(df_pm, 'Matricula')

    c_em = _col_correo(df_pm)

    c_ap = _col(df_pm, 'Apellido Paterno')

    c_am = _col(df_pm, 'Apellido Materno')

    c_no = _col(df_pm, 'Nombre')

    c_comb = _col(

        df_pm,

        'Apellidos y Nombres',

        'Apellidos_Nombres',

        'Nombres y Apellidos',

        'Nombre Completo',

        'Colaborador',

        'Trabajador'

    )



    for _, row in df_pm.iterrows():

        full = f"{_safe(row.get(c_ap,''))} {_safe(row.get(c_am,''))} {_safe(row.get(c_no,''))}".strip()

        if not full and c_comb:

            full = _safe(row.get(c_comb, ''))

        short = f"{_safe(row.get(c_ap,''))} {_safe(row.get(c_no,''))}".strip()

        if not short and full:

            toks = [t for t in str(full).split() if t]

            short = ' '.join([toks[0], toks[-1]]) if len(toks) >= 2 else full

        email = _norm_email(row.get(c_em, '')) if c_em else ''

        mat = _norm_id(row.get(c_mat, '')) if c_mat else ''

        data = {'nombre': full or short or mat, 'email': email, 'matricula': mat}



        if mat:

            idx['by_mat'][mat] = data

        if full:

            idx['by_name'][full.upper()] = data

            idx['by_name'][_norm(full).upper()] = data

        if short and short.upper() not in idx['by_short']:

            idx['by_short'][short.upper()] = data



    return idx





def _resolver_supervisor_identidad(nombre_raw='', matricula_raw='', idx=None):

    idx = idx or {'by_mat': {}, 'by_name': {}, 'by_short': {}}

    nombre_raw = _nombre_supervisor_canonico(nombre_raw)

    matricula = _norm_id(matricula_raw)

    nombre_norm = nombre_raw.upper()



    data = None

    if matricula:

        data = idx.get('by_mat', {}).get(matricula)

    if data is None and nombre_norm:

        data = idx.get('by_name', {}).get(nombre_norm)

    if data is None and nombre_norm:

        data = idx.get('by_short', {}).get(nombre_norm)

    if data is None and _norm_id(nombre_raw):

        data = idx.get('by_mat', {}).get(_norm_id(nombre_raw))

    # Lookup normalizado: elimina comas y espacios extra para cruzar
    # "SUAREZ VERA, NANCY" (vacaciones) con "SUAREZ VERA NANCY" (maestro)
    if data is None and nombre_norm:

        data = idx.get('by_name', {}).get(_nombre_cmp_key(nombre_raw))



    if data:

        return {

            'nombre': data.get('nombre') or nombre_raw or matricula or 'Sin Supervisor',

            'email': _norm_email(data.get('email', '')),

            'matricula': data.get('matricula', '') or matricula,

        }



    return {

        'nombre': nombre_raw or matricula or 'Sin Supervisor',

        'email': '',

        'matricula': matricula or _norm_id(nombre_raw),

    }





def _supervisores_cfg_map():

    """Retorna configuracion avanzada por supervisor.

    Formato: {"SUPERVISOR": {email, mensaje, aviso, recomendacion}}

    """

    cfg = _PA_CONFIG.get('supervisores_config', {}) or {}

    if not isinstance(cfg, dict) or not cfg:

        legacy = _PA_CONFIG.get('supervisores_emails', {}) or {}

        if isinstance(legacy, dict):

            migrado = {}

            for nombre, email in legacy.items():

                n = str(nombre or '').strip()

                e = _norm_email(email)

                if n and e:

                    migrado[n] = {

                        'email': e,

                        'mensaje': '',

                        'aviso': '',

                        'recomendacion': ''

                    }

            return migrado

        return {}



    out = {}

    for nombre, data in cfg.items():

        n = str(nombre or '').strip()

        if not n:

            continue

        if isinstance(data, str):

            e = _norm_email(data)

            if e:

                out[n] = {'email': e, 'mensaje': '', 'aviso': '', 'recomendacion': ''}

            continue

        if not isinstance(data, dict):

            continue

        e = _norm_email(data.get('email', ''))

        if not e:

            continue

        out[n] = {

            'email': e,

            'mensaje': str(data.get('mensaje', '') or '').strip(),

            'aviso': str(data.get('aviso', '') or '').strip(),

            'recomendacion': str(data.get('recomendacion', '') or '').strip(),

        }

    return out





def _resolver_supervisor_cfg(nombre):

    cfg = _supervisores_cfg_map().get(str(nombre or '').strip(), {})

    if not isinstance(cfg, dict):

        cfg = {}

    return {

        'email': _norm_email(cfg.get('email', '')),

        'mensaje': str(cfg.get('mensaje', '') or '').strip(),

        'aviso': str(cfg.get('aviso', '') or '').strip(),

        'recomendacion': str(cfg.get('recomendacion', '') or '').strip(),

    }





def _tpl(texto, ctx):

    txt = str(texto or '')

    for k, v in (ctx or {}).items():

        txt = txt.replace('{' + str(k) + '}', str(v or ''))

    return txt


def _normalizar_texto_alerta_supervisor(texto, nombre_destinatario='', trimestre='', fecha_limite=''):

    txt = str(texto or '')

    if not txt.strip():

        return txt

    tri = str(trimestre or '').strip() or str(_resumen_fuente_datos().get('trimestre_vigente', '') or '').strip() or 'trimestre vigente'

    dest = str(nombre_destinatario or '').strip()

    fecha_ref = str(fecha_limite or '').strip() or 'la fecha acordada'

    txt = re.sub(r'Q\?-\?\?\?\?', tri, txt, flags=re.IGNORECASE)

    txt = re.sub(r'(?i)(a\s+m[aá]s\s+tardar\s+el)\s+(la\s+fecha\s+acordada)', r'a más tardar \2', txt)

    if dest:

        txt = re.sub(r'(?im)^\s*hola+a*\s*,?\s*l[ií]der\s*:?\s*', f'Hola, <strong>{dest}</strong>:\n\n', txt, count=1)

        txt = re.sub(r'(?im)^\s*hola+a*\s*,?\s*jefe\s*:?\s*', f'Hola, <strong>{dest}</strong>:\n\n', txt, count=1)

        # Si el nombre real ya está en el saludo (después de _tpl), también ponerlo en negrita
        txt = re.sub(r'(?im)^(hola,\s*)(' + re.escape(dest) + r')(\s*:)', r'\1<strong>\2</strong>\3', txt, count=1)

    return txt


def _render_texto_editable_html(texto):

    txt = str(texto or '')

    if not txt.strip():

        return ''

    txt_norm = txt.replace('\r\n', '\n').replace('\r', '\n')

    txt_lower = txt_norm.lower()

    if any(tag in txt_lower for tag in ('<p', '<ul', '<ol', '<li', '<div', '<table', '<tr', '<td', '<th', '<h1', '<h2', '<h3', '<h4', '<h5', '<h6')):

        return txt_norm

    html_txt = txt_norm

    html_txt = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', html_txt, flags=re.S)

    html_txt = re.sub(r'__(.+?)__', r'<strong>\1</strong>', html_txt, flags=re.S)

    bloques = []
    lineas = [ln.rstrip() for ln in html_txt.split('\n')]
    i = 0

    while i < len(lineas):
        actual = lineas[i].strip()
        if not actual:
            i += 1
            continue

        if re.match(r'^(?:&bull;|•|\-)\s+', actual):
            items = []
            while i < len(lineas):
                linea = lineas[i].strip()
                if not linea or not re.match(r'^(?:&bull;|•|\-)\s+', linea):
                    break
                items.append(re.sub(r'^(?:&bull;|•|\-)\s+', '', linea, count=1).strip())
                i += 1
            bloques.append(
                '<ul style="margin:0 0 14px 18px;padding:0;line-height:1.7">'
                + ''.join(f'<li style="margin:0 0 8px 0">{item}</li>' for item in items)
                + '</ul>'
            )
            continue

        if re.match(r'^\d+\.\s+', actual):
            items = []
            while i < len(lineas):
                linea = lineas[i].strip()
                if not linea or not re.match(r'^\d+\.\s+', linea):
                    break
                items.append(re.sub(r'^\d+\.\s+', '', linea, count=1).strip())
                i += 1
            bloques.append(
                '<ol style="margin:0 0 14px 22px;padding:0;line-height:1.7">'
                + ''.join(f'<li style="margin:0 0 8px 0">{item}</li>' for item in items)
                + '</ol>'
            )
            continue

        parrafo = []
        while i < len(lineas):
            linea = lineas[i].strip()
            if not linea:
                break
            if re.match(r'^(?:&bull;|•|\-)\s+', linea) or re.match(r'^\d+\.\s+', linea):
                break
            parrafo.append(linea)
            i += 1

        if parrafo:
            if len(parrafo) == 1:
                bloques.append(f'<p style="margin:0 0 12px 0">{parrafo[0]}</p>')
            else:
                bloques.append(f'<p style="margin:0 0 12px 0">{"<br>".join(parrafo)}</p>')
            continue

        i += 1

    return ''.join(bloques)


def _mensaje_ya_incluye_saludo(texto):

    base = str(texto or '').strip().lower()

    if not base:

        return False

    base = re.sub(r'^<[^>]+>', '', base).strip()

    return base.startswith('hola') or base.startswith('estimado') or base.startswith('estimada')


def _texto_para_teams(texto):

    txt = str(texto or '')

    if not txt.strip():

        return ''

    txt = txt.replace('\r\n', '\n').replace('\r', '\n')

    txt = re.sub(r'(?is)<\s*br\s*/?\s*>', '\n', txt)

    txt = re.sub(r'(?is)<\s*/\s*p\s*>', '\n\n', txt)

    txt = re.sub(r'(?is)<\s*p\b[^>]*>', '', txt)

    txt = re.sub(r'(?is)<\s*li\b[^>]*>', '• ', txt)

    txt = re.sub(r'(?is)<\s*/\s*li\s*>', '\n', txt)

    txt = re.sub(r'(?is)<\s*/\s*(ul|ol|div)\s*>', '\n', txt)

    txt = re.sub(r'(?is)<\s*(strong|b)\s*>', '**', txt)

    txt = re.sub(r'(?is)<\s*/\s*(strong|b)\s*>', '**', txt)

    txt = re.sub(r'(?is)<\s*(em|i)\s*>', '_', txt)

    txt = re.sub(r'(?is)<\s*/\s*(em|i)\s*>', '_', txt)

    txt = re.sub(r'(?is)<[^>]+>', '', txt)

    txt = html_lib.unescape(txt)

    txt = re.sub(r'\n{3,}', '\n\n', txt)

    return txt.strip()





def _aviso_recomendacion_persona(p):

    d = p.get('dias_para_aniversario')

    venc = float(p.get('vencidas') or 0)

    tot = float(p.get('total_vacaciones') or 0)

    if d is not None and d < 0:

        return (

            'Aniversario laboral vencido con atraso vacacional.',

            'Coordinar goce inmediato de vacaciones en un plazo maximo de 7 dias.'

        )

    if venc > 0:

        return (

            'Colaborador con vacaciones vencidas acumuladas.',

            'Programar primero las vacaciones vencidas antes de nuevas acumulaciones.'

        )

    if d is not None and 0 <= d <= 30:

        return (

            'Aniversario proximo en menos de 30 dias.',

            'Definir fechas de descanso y validar cobertura operativa del equipo.'

        )

    if tot >= 45:

        return (

            'Saldo total de vacaciones alto.',

            'Revisar plan trimestral para reducir saldo y evitar vencimientos.'

        )

    return (

        'Seguimiento preventivo de vacaciones.',

        'Mantener programacion vacacional trimestral con revision mensual.'

    )





def _enviar_correo_smtp(to_email, to_nombre, asunto, html_body):

    """EnvÃ­a correo directo via SMTP Office 365. Retorna (ok, error_str).

    No depende de Power Automate â€” funciona aunque el flujo estÃ© roto."""

    import smtplib

    from email.mime.multipart import MIMEMultipart

    from email.mime.text import MIMEText

    # Leer siempre en vivo para capturar cambios en pa_config.json sin reiniciar
    smtp_user = _PA_CONFIG.get('smtp_email', '') or SMTP_EMAIL
    smtp_pass = _PA_CONFIG.get('smtp_password', '') or SMTP_PASSWORD

    if not smtp_user or not smtp_pass:

        return False, 'SMTP no configurado: falta smtp_email o smtp_password en pa_config.json'

    if not to_email or '@' not in to_email:

        return False, f'Email destino invalido: {to_email!r}'

    try:

        msg = MIMEMultipart('alternative')

        msg['Subject'] = asunto

        msg['From']    = f'People Analytics USIL <{smtp_user}>'

        msg['To']      = to_email

        msg['X-USIL-PA'] = 'Sistema-Vacaciones'

        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        with smtplib.SMTP('smtp.office365.com', 587, timeout=25) as srv:

            srv.ehlo()

            srv.starttls()

            srv.ehlo()

            srv.login(smtp_user, smtp_pass)

            srv.sendmail(smtp_user, [to_email], msg.as_bytes())

        print(f'[SMTP] OK -> {to_email} | {asunto[:60]}')

        return True, None

    except Exception as e:

        print(f'[SMTP] Error -> {to_email}: {e}')

        return False, str(e)





def _agrupar_por_supervisor(ret, prox, sin_cumplir, df_pm):

    """Agrupa aniversarios y sin_cumplir por supervisor y resuelve su email.

    Siempre garantiza al menos una entrada aunque no haya emails en el Excel."""

    idx = _build_personal_index(df_pm)

    # Enriquecer el índice con emails del maestro:
    # el dataframe de vacaciones no tiene columnas de correo, así que sin este paso
    # todos los emails quedan vacíos y se cae en el fallback (jlopezp).
    _, mat_a_info_mae, _ = _cargar_tabla_maestra_jefes()
    if mat_a_info_mae:
        for _mat, _info in mat_a_info_mae.items():
            _em = _info.get('email', '')
            _nom = _info.get('nombre', '')
            _entry = {'nombre': _nom, 'email': _em, 'matricula': _mat}
            if _mat:
                existing = idx['by_mat'].get(_mat)
                if not existing:
                    idx['by_mat'][_mat] = _entry
                elif not existing.get('email') and _em:
                    existing['email'] = _em
            if _nom:
                for _key in (_nom.upper(), _norm(_nom).upper()):
                    existing = idx['by_name'].get(_key)
                    if not existing:
                        idx['by_name'][_key] = _entry
                    elif not existing.get('email') and _em:
                        existing['email'] = _em

    default  = _resolver_default_email()

    cfg_sup = _supervisores_cfg_map()

    for nm, cfg in cfg_sup.items():

        em = _norm_email((cfg or {}).get('email', ''))

        if em and '@' in em:

            idx['by_name'][str(nm).upper()] = {

                'nombre': str(nm),

                'email': em,

                'matricula': ''

            }



    jefes = {}

    for p in list(ret) + list(prox):

        sup_info = _resolver_supervisor_identidad(p.get('supervisor', ''), p.get('supervisor_matricula', ''), idx)

        sup = sup_info.get('nombre', 'Sin Supervisor')

        scfg = _resolver_supervisor_cfg(sup)

        em  = sup_info.get('email', '') or scfg.get('email', '') or default

        jefes.setdefault(sup, {

            'nombre': sup,

            'email': em,

            'matricula': sup_info.get('matricula', ''),

            'mensaje': scfg.get('mensaje', ''),

            'aviso': scfg.get('aviso', ''),

            'recomendacion': scfg.get('recomendacion', ''),

            'ret': [], 'prox': [], 'sc': []

        })

        jefes[sup]['ret' if p.get('en_retraso') else 'prox'].append(p)

    for p in sin_cumplir:

        sup_info = _resolver_supervisor_identidad(p.get('supervisor', ''), p.get('supervisor_matricula', ''), idx)

        sup = sup_info.get('nombre', 'Sin Supervisor')

        scfg = _resolver_supervisor_cfg(sup)

        em  = sup_info.get('email', '') or scfg.get('email', '') or default

        jefes.setdefault(sup, {

            'nombre': sup,

            'email': em,

            'matricula': sup_info.get('matricula', ''),

            'mensaje': scfg.get('mensaje', ''),

            'aviso': scfg.get('aviso', ''),

            'recomendacion': scfg.get('recomendacion', ''),

            'ret': [], 'prox': [], 'sc': []

        })

        jefes[sup]['sc'].append(p)

    return jefes





def _build_html_jefe(nombre_jefe, ret_list, prox_list, sc_list, fecha_str, plantilla_msg='', aviso_txt='', recomendacion_txt='', adicionales_list=None, campania=None, meta_resumen=None, hrbp_nombre='', fecha_limite='', equipo_full=None, confirmaciones=None, nombre_objetivo=''):

    """Construye el HTML del correo para un jefe con su lista de colaboradores."""

    adicionales_list = adicionales_list or []

    campania = campania or _contexto_campania()

    meta_resumen = meta_resumen or {}



    def _fnum(v):

        try:

            f = round(float(v or 0), 1)

            return int(f) if f == int(f) else f

        except Exception:

            return 0



    def _row(p):

        goz    = _fnum(p.get('gozados', 0))

        obj    = int(p.get('objetivo') or 0)

        total  = _fnum(p.get('total_vacaciones'))

        por_prog = max(obj - int(goz), 0)

        # Saldo Final = días pendientes para cumplir la meta (no el saldo total de vacaciones)
        saldo    = por_prog

        return (f'<tr style="border-bottom:1px solid #e9ecef">'

                f'<td style="padding:8px 7px;text-align:center">{p.get("matricula") or "-"}</td>'

                f'<td style="padding:8px 9px"><strong>{p.get("nombre") or "-"}</strong></td>'

                f'<td style="padding:8px 7px;text-align:center">{p.get("fecha_ingreso") or "-"}</td>'

                f'<td style="padding:8px 9px">{(p.get("puesto") or "-")[:45]}</td>'

                f'<td style="padding:8px 9px">{(p.get("departamento") or "-")[:45]}</td>'

                f'<td style="padding:8px 9px">{(p.get("area") or "-")[:35]}</td>'

                f'<td style="padding:8px 7px;text-align:center">{_fnum(p.get("pendientes"))}</td>'

                f'<td style="padding:8px 7px;text-align:center">{_fnum(p.get("truncas"))}</td>'

                f'<td style="padding:8px 7px;text-align:center">{total}</td>'

                f'<td style="padding:8px 7px;text-align:center">{obj}</td>'

                f'<td style="padding:8px 7px;text-align:center">{goz}</td>'

                f'<td style="padding:8px 7px;text-align:center;font-weight:700">{por_prog}</td>'

                f'<td style="padding:8px 7px;text-align:center;background:#fff3bf;font-weight:700">{saldo}</td>'

                f'</tr>')



    _TH = (

        '<th style="padding:8px 7px;text-align:center;background:#1f26ff;color:#fff">Matricula</th>'

        '<th style="padding:8px 9px;text-align:left;background:#1f26ff;color:#fff">Apellidos y Nombres</th>'

        '<th style="padding:8px 7px;text-align:center;background:#1f26ff;color:#fff">Fecha Ingreso Compania</th>'

        '<th style="padding:8px 9px;text-align:left;background:#1f26ff;color:#fff">Nombre Puesto</th>'

        '<th style="padding:8px 9px;text-align:left;background:#1f26ff;color:#fff">Nombre Departamento</th>'

        '<th style="padding:8px 9px;text-align:left;background:#1f26ff;color:#fff">Nombre Area</th>'

        '<th style="padding:8px 7px;text-align:center;background:#1f26ff;color:#fff">Cant Vac Pendiente</th>'

        '<th style="padding:8px 7px;text-align:center;background:#1f26ff;color:#fff">Cant Vac Truncos</th>'

        '<th style="padding:8px 7px;text-align:center;background:#1f26ff;color:#fff">Suma de Dias Total</th>'

        '<th style="padding:8px 7px;text-align:center;background:#d40000;color:#fff">Meta (Abril - Jul 2026)</th>'

        '<th style="padding:8px 7px;text-align:center;background:#d40000;color:#fff">Avance</th>'

        '<th style="padding:8px 7px;text-align:center;background:#d40000;color:#fff">Por programar</th>'

        '<th style="padding:8px 7px;text-align:center;background:#d40000;color:#fff">Saldo Final</th>'

    )



    def _tabla(_color, filas_html):

        return (f'<div style="border-radius:8px;overflow-x:auto;border:1px solid #e9ecef">'

                f'<table style="width:100%;border-collapse:collapse;font-size:13px">'

                f'<thead><tr>{_TH}</tr></thead>'

                f'<tbody>{filas_html}</tbody></table></div>')



    # Tabla completa del equipo (todos los colaboradores bajo este jefe)

    todos = equipo_full if equipo_full else (ret_list + prox_list + sc_list + adicionales_list)

    secs = ''

    if todos:

        filas_todos = ''.join([_row(p) for p in todos])

        secs += ('<h3 style="color:#1e40af;margin:28px 0 8px;font-size:15px">Reporte Completo del Equipo</h3>'

                 f'<p style="margin:0 0 8px 0;color:#1e3a8a;font-size:12px">Total de {len(todos)} colaboradores bajo tu cargo.</p>'

                 + _tabla('#1e40af', filas_todos))



    msg_jefe = (plantilla_msg or _MSG_BASE_DEFAULT)
    # Eliminar placeholder literal ( TABLA ) que algunos templates guardados tienen
    for _placeholder in ('( TABLA )', '(TABLA)', '{tabla}', '[ TABLA ]'):
        msg_jefe = msg_jefe.replace(_placeholder, '')

    aviso_final = aviso_txt or _AVISO_DEFAULT

    reco_final = recomendacion_txt or _RECOM_DEFAULT

    resumen_txt = f'{len(ret_list)} en retraso, {len(prox_list)} proximos aniversario, {len(sc_list)} sin cumplir objetivo y {len(adicionales_list)} adicionales sugeridos'

    meta_obj = int(meta_resumen.get('dias_objetivo', 0) or 0)

    meta_goz = int(meta_resumen.get('dias_gozados_hacia_meta', 0) or 0)

    meta_pend = int(meta_resumen.get('dias_pendientes_meta', 0) or 0)

    meta_pct = float(meta_resumen.get('cumplimiento_meta_pct', 0.0) or 0.0)

    meta_cols = int(meta_resumen.get('colaboradores_meta', 0) or 0)

    meta_cols_ok = int(meta_resumen.get('colaboradores_cumplieron', 0) or 0)

    meta_cols_pen = int(meta_resumen.get('colaboradores_pendientes', 0) or 0)
    hrbp_display = str(hrbp_nombre or '').strip() or _HRBP_TYC_DEFAULT
    # Normalizar a nombre formal del pa_config si el apellido principal coincide
    if hrbp_display and hrbp_display != _HRBP_TYC_DEFAULT:
        _key_display = _nombre_cmp_key(hrbp_display)
        _key_default = _nombre_cmp_key(_HRBP_TYC_DEFAULT)
        # Si todos los tokens del nombre informal están contenidos en el nombre formal, usar el formal
        if all(tok in _key_default for tok in _key_display.split() if len(tok) > 2):
            hrbp_display = _HRBP_TYC_DEFAULT
    # Simplificar firma: "JARA ORTIZ, CARLOS HUMBERTO" → "Carlos Jara"
    hrbp_display = _nombre_firma_usil(hrbp_display) or hrbp_display

    # El saludo siempre va dirigido al JEFE, solo usamos su primer nombre
    nombre_destinatario = _primer_nombre_usil(nombre_jefe) or str(nombre_jefe or '').strip()

    ctx = {

        'jefe': nombre_jefe,

        'nombre': nombre_destinatario,

        'destinatario': nombre_destinatario,

        'fecha': fecha_str,

        'frecuencia': campania.get('frecuencia', 'semanal'),

        'trimestre': campania.get('trimestre', '2T'),

        'semana_trimestre': campania.get('semana_trimestre', ''),

        'resumen': resumen_txt,

        'meta_cumplimiento': f'{meta_pct}%',

        'meta_dias_objetivo': meta_obj,

        'meta_dias_gozados': meta_goz,

        'meta_dias_pendientes': meta_pend,

        'meta_colaboradores': meta_cols,

        'meta_colaboradores_cumplieron': meta_cols_ok,

        'meta_colaboradores_pendientes': meta_cols_pen,

        'aviso': aviso_final,

        'recomendacion': reco_final,

        'hrbp': hrbp_display,

        'hrbp_nombre': hrbp_display,

        'fecha_limite': fecha_limite or 'la fecha acordada',

    }

    msg_tpl = _tpl(msg_jefe, ctx)

    # Poner el nombre del destinatario en negrita dentro del saludo "Hola, NOMBRE:"
    if nombre_destinatario:
        msg_tpl = re.sub(
            r'(?im)^(hola,\s*)(' + re.escape(nombre_destinatario) + r')(\s*:)',
            r'\1<strong>\2</strong>\3',
            msg_tpl, count=1
        )

    # Separar cuerpo del mensaje de la firma/importante
    corte_idx = -1
    marcadores = [
        '<span style="color:#b91c1c">',
        '<span style="color: #b91c1c">',
        '<b>Importante:</b>',
        'Importante:',
        'Atentamente,',
        'Atentamente:',
        'atentamente,'
    ]
    for m in marcadores:
        idx = msg_tpl.find(m)
        if idx != -1:
            if corte_idx == -1 or idx < corte_idx:
                corte_idx = idx

    if corte_idx != -1:
        msg_body_tpl = msg_tpl[:corte_idx].strip()
        msg_footer_tpl = msg_tpl[corte_idx:].strip()
    else:
        msg_body_tpl = msg_tpl.strip()
        msg_footer_tpl = ''

    # Asegurar saludo al inicio si no está
    if not _mensaje_ya_incluye_saludo(msg_body_tpl):
        msg_body_tpl = f'Hola, <strong>{nombre_destinatario}</strong>:\n\n' + msg_body_tpl

    msg_body_rendered = _render_texto_editable_html(msg_body_tpl)
    msg_footer_rendered = _render_texto_editable_html(msg_footer_tpl)

    body_content_html = f'<div style="font-size:13px;color:#444;line-height:1.6;margin-bottom:12px">{msg_body_rendered}</div>'
    footer_content_html = f'<div style="font-size:13px;color:#444;line-height:1.6;margin-top:18px">{msg_footer_rendered}</div>' if msg_footer_rendered else ''

    return ('<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>'

            '<body style="margin:0;padding:0;font-family:Segoe UI,Arial,sans-serif;background:#f5f5f5">'

            '<div style="max-width:960px;margin:20px auto;background:#fff;border-radius:10px;box-shadow:0 4px 12px rgba(0,0,0,.1);padding:24px 28px">'

            + body_content_html +

            '<div style="margin:10px 0 0 0;padding:10px;background:#eff6ff;border-radius:8px;border:1px solid #bfdbfe">'

            f'<p style="margin:0;color:#1d4ed8;font-size:12px"><strong>Cadencia:</strong> Seguimiento {campania.get("frecuencia", "semanal")} · <strong>Meta:</strong> {campania.get("meta_periodo", "trimestral")} · <strong>Periodo:</strong> {_periodo_campania_texto(campania)}</p>'

            '</div>'

            '<div style="margin-top:10px;padding:12px;background:#ecfdf5;border-radius:8px;border:1px solid #86efac">'

            f'<p style="margin:0;color:#065f46;font-size:13px"><strong>Meta del equipo:</strong> {meta_pct}% ({meta_goz}/{meta_obj} dias) · <strong>Faltan:</strong> {meta_pend} dias</p>'

            f'<p style="margin:4px 0 0 0;color:#065f46;font-size:12px">Colaboradores con meta: {meta_cols} · Cumplieron: {meta_cols_ok} · Pendientes: {meta_cols_pen}</p>'

            '</div>'

            + secs +

            '<div style="margin-top:14px;padding:12px;background:#fff7ed;border-radius:8px;border:1px solid #fed7aa">'

            f'<p style="margin:0 0 5px 0;color:#9a3412;font-size:12px"><strong>Aviso:</strong> {_tpl(aviso_final, ctx)}</p>'

            f'<p style="margin:0;color:#9a3412;font-size:12px"><strong>Recomendacion:</strong> {_tpl(reco_final, ctx)}</p>'

            '</div>'

            + footer_content_html +

            '<div style="margin-top:22px;padding:14px;background:#f0f9ff;border-radius:8px;border:1px solid #bae6fd">'

            f'<p style="margin:0;color:#0369a1;font-size:12px">Generado automaticamente por People Analytics USIL - {fecha_str}</p>'

            '</div></div></body></html>')





# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# TEAMS WEBHOOK

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _mensajes_teams(retraso, proximos, dias, sin_cumplir, kpis_obj):

    """Devuelve lista de Adaptive Cards: 1 resumen + 1 por HRBP (agrupado por supervisor) + resumen aniversarios."""

    hoy_str = date.today().strftime('%d/%m/%Y')

    MAX_PER_CARD = 25   # max TextBlocks por card para no saturar



    def _card(body):

        return {'type':'message','attachments':[{'contentType':'application/vnd.microsoft.card.adaptive',

            'content':{'$schema':'http://adaptivecards.io/schemas/adaptive-card.json',

                       'type':'AdaptiveCard','version':'1.4','body':body}}]}



    mensajes = []



    # â”€â”€ Mensaje 1: Resumen KPIs â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    facts = []

    if kpis_obj:

        facts = [

            {'title':'Objetivo Q1:',  'value':str(kpis_obj.get('dias_objetivo',0))+' dias'},

            {'title':'Gozados:',      'value':str(kpis_obj.get('dias_gozados',0))+' dias ('+str(kpis_obj.get('cumplimiento',0))+'%)'},

            {'title':'Sin cumplir:',  'value':str(kpis_obj.get('sin_cumplir',0))+' personas \u2014 '+str(kpis_obj.get('dias_pendientes',0))+' dias pendientes'},

            {'title':'Aniversarios:', 'value':f'{len(retraso)} en retraso, {len(proximos)} proximos {dias}d'},

        ]

    resumen = [

        {'type':'TextBlock','size':'Large','weight':'Bolder','color':'Accent',

         'text':'\U0001f3d6 Alerta Vacaciones Q1 \u2014 USIL'},

        {'type':'TextBlock','isSubtle':True,'size':'Small','text':f'People Analytics \u2022 {hoy_str}'},

    ]

    if facts:

        resumen.append({'type':'FactSet','separator':True,'spacing':'Medium','facts':facts})

    mensajes.append(_card(resumen))



    # â”€â”€ Mensajes: Sin Cumplir agrupado por HRBP â†’ supervisor â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    # Primero agrupar por HRBP, luego por supervisor dentro de cada HRBP

    hrbp_map = {}

    for p in sin_cumplir:

        hrbp = (p.get('hrbp') or 'Sin HRBP').strip()

        sup  = (p.get('supervisor') or '').strip()

        hrbp_map.setdefault(hrbp, {})

        hrbp_map[hrbp].setdefault(sup, []).append(p)



    for hrbp_name in sorted(hrbp_map.keys()):

        sup_dict = hrbp_map[hrbp_name]

        total_hrbp = sum(len(v) for v in sup_dict.values())

        body = [

            {'type':'TextBlock','size':'Medium','weight':'Bolder','color':'Attention',

             'text':f'\U0001f534 {hrbp_name} \u2014 {total_hrbp} sin cumplir objetivo'},

        ]

        lines = 0

        for sup_name in sorted(sup_dict.keys()):

            personas = sup_dict[sup_name]

            if lines >= MAX_PER_CARD:

                # Partir en nueva card

                mensajes.append(_card(body))

                body = [{'type':'TextBlock','size':'Medium','weight':'Bolder','color':'Attention',

                         'text':f'\U0001f534 {hrbp_name} (cont.)'}]

                lines = 0

            # Encabezado del supervisor con @

            if sup_name:

                body.append({'type':'TextBlock','weight':'Bolder','size':'Small','spacing':'Medium','wrap':True,

                    'text':f'\U0001f464 @{sup_name} ({len(personas)} colaborador{"es" if len(personas)>1 else ""}):'})

            else:

                body.append({'type':'TextBlock','weight':'Bolder','size':'Small','spacing':'Medium','wrap':True,

                    'text':f'\U0001f464 Sin supervisor ({len(personas)}):'})

            lines += 1

            for p in personas:

                pct = round(p['gozados'] / p['objetivo'] * 100) if p['objetivo'] > 0 else 0

                aviso_sc = p.get('aviso', 'Sin aviso')

                reco_sc = p.get('recomendacion', 'Sin recomendacion')

                cant_p = p.get('cantidad_pendiente')

                if cant_p is None:

                    cant_p = round(float(p.get('vencidas') or 0) + float(p.get('pendientes') or 0), 1)

                body.append({'type':'TextBlock','spacing':'None','size':'Small','wrap':True,

                    'text':(

                        f"  \u2022 {p['matricula']} | {p['nombre']} | Meta:{p['objetivo']} | F.Ing:{p.get('fecha_ingreso') or '-'} "

                        f"| Puesto:{(p.get('puesto') or '-')[:30]} | Pend:{round(float(p.get('pendientes') or 0),1)} "

                        f"| Trun:{round(float(p.get('truncas') or 0),1)} | Total:{round(float(p.get('total_vacaciones') or 0),1)} "

                        f"| BP:{p.get('hrbp') or '-'} | Cant.Pend:{cant_p} "

                        f"| Tomados:{p['gozados']} | Cumpl:{pct}% ({p['cumplimiento']}d) "

                        f"| Aviso: {aviso_sc} | Recom: {reco_sc}"

                    )})

                lines += 1

        mensajes.append(_card(body))



    # â”€â”€ Mensajes: Aniversarios en retraso (top 15 + resumen) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    if retraso:

        TOP = 15

        body = [

            {'type':'TextBlock','size':'Medium','weight':'Bolder','color':'Attention',

             'text':f'\U0001f534 Aniversarios en Retraso \u2014 {len(retraso)} personas'},

        ]

        for a in retraso[:TOP]:

            d = a.get('dias_para_aniversario', 0)

            sup = a.get('supervisor', '')

            sup_txt = f' (@{sup})' if sup else ''

            aviso_a, reco_a = _aviso_recomendacion_persona(a)

            cant_p = a.get('cantidad_pendiente')

            if cant_p is None:

                cant_p = round(float(a.get('vencidas') or 0) + float(a.get('pendientes') or 0), 1)

            body.append({'type':'TextBlock','spacing':'None','size':'Small','wrap':True,

                'text':(

                    f"\u2022 {a.get('matricula') or '-'} | **{a['nombre']}** | Meta:{a.get('objetivo', 0)} "

                    f"| F.Ing:{a.get('fecha_ingreso') or '-'} | Puesto:{(a.get('puesto') or '-')[:30]} "

                    f"| Pend:{round(float(a.get('pendientes') or 0),1)} | Trun:{round(float(a.get('truncas') or 0),1)} "

                    f"| Total:{round(float(a.get('total_vacaciones') or 0),1)} | BP:{a.get('hrbp') or '-'} "

                    f"| Cant.Pend:{cant_p} | {abs(d)}d retraso{sup_txt} "

                    f"| Aviso: {aviso_a} | Recom: {reco_a}"

                )})

        if len(retraso) > TOP:

            body.append({'type':'TextBlock','isSubtle':True,'size':'Small','spacing':'Small',

                'text':f'... y {len(retraso)-TOP} personas mas en retraso.'})

        mensajes.append(_card(body))



    # â”€â”€ Mensajes: PrÃ³ximos N dÃ­as (top 15 + resumen) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    if proximos:

        TOP = 15

        body = [

            {'type':'TextBlock','size':'Medium','weight':'Bolder','color':'Warning',

             'text':f'\u26a0\ufe0f Proximos {dias} Dias \u2014 {len(proximos)} personas'},

        ]

        for a in proximos[:TOP]:

            d = a.get('dias_para_aniversario', 0)

            sup = a.get('supervisor', '')

            sup_txt = f' (@{sup})' if sup else ''

            aviso_a, reco_a = _aviso_recomendacion_persona(a)

            cant_p = a.get('cantidad_pendiente')

            if cant_p is None:

                cant_p = round(float(a.get('vencidas') or 0) + float(a.get('pendientes') or 0), 1)

            body.append({'type':'TextBlock','spacing':'None','size':'Small','wrap':True,

                'text':(

                    f"\u2022 {a.get('matricula') or '-'} | **{a['nombre']}** | Meta:{a.get('objetivo', 0)} "

                    f"| F.Ing:{a.get('fecha_ingreso') or '-'} | Puesto:{(a.get('puesto') or '-')[:30]} "

                    f"| Pend:{round(float(a.get('pendientes') or 0),1)} | Trun:{round(float(a.get('truncas') or 0),1)} "

                    f"| Total:{round(float(a.get('total_vacaciones') or 0),1)} | BP:{a.get('hrbp') or '-'} "

                    f"| Cant.Pend:{cant_p} | {d}d restantes{sup_txt} "

                    f"| Aviso: {aviso_a} | Recom: {reco_a}"

                )})

        if len(proximos) > TOP:

            body.append({'type':'TextBlock','isSubtle':True,'size':'Small','spacing':'Small',

                'text':f'... y {len(proximos)-TOP} personas mas proximas a cumplir aniversario.'})

        mensajes.append(_card(body))



    return mensajes



@app.route('/api/test-notif-teams', methods=['POST'])

def _resolver_teams_webhook():
    """Lee siempre en vivo para capturar cambios en pa_config sin reiniciar."""
    return (
        _PA_CONFIG.get('teams_webhook_url', '') or TEAMS_WEBHOOK_URL or ''
    ).strip()


def _post_teams_webhook(url, payload_card):
    """Envía al webhook de Teams.
    - URL clásica (webhook.office.com): manda la Adaptive Card directo.
    - URL nueva Flujos de trabajo (logic.azure.com / azure-apim.net):
      manda el texto plano en {"text": "..."} porque el flujo de PA
      espera un body simple y lo publica en el canal.
    Devuelve (ok: bool, error: str|None).
    """
    import urllib.request as _ur
    if not url:
        return False, 'URL de webhook vacía'
    try:
        # Detectar tipo de URL
        es_flujos_trabajo = any(d in url for d in ('logic.azure', 'azure-apim', 'prod-', 'azure.com/workflows'))
        if es_flujos_trabajo:
            # Flujos de trabajo acepta JSON libre; extraemos el texto legible
            texto = ''
            body = payload_card
            if isinstance(body, dict):
                # Intentar sacar texto de Adaptive Card
                for blk in (body.get('attachments') or []):
                    content = (blk or {}).get('content') or {}
                    for item in (content.get('body') or []):
                        t = (item or {}).get('text', '')
                        if t: texto += t + '\n'
                if not texto:
                    texto = str(body)
            body_bytes = json.dumps({'text': texto.strip() or 'Alerta Vacaciones USIL'}).encode('utf-8')
        else:
            # Webhook clásico: Adaptive Card completa
            body_bytes = json.dumps(payload_card).encode('utf-8')

        req = _ur.Request(url, data=body_bytes, method='POST')
        req.add_header('Content-Type', 'application/json')
        with _ur.urlopen(req, timeout=15) as resp:
            resp.read()
        return True, None
    except Exception as e:
        return False, str(e)


@app.route('/api/teams/webhook-update', methods=['POST'])
def api_teams_webhook_update():
    """Actualiza la URL del webhook de Teams en vivo y la guarda en pa_config.json."""
    payload = request.get_json(silent=True) or {}
    nueva_url = str(payload.get('url', '') or '').strip()
    if not nueva_url or 'https://' not in nueva_url:
        return jsonify({'ok': False, 'error': 'URL inválida'}), 400

    # Probar antes de guardar
    test_card = {'type': 'message', 'attachments': [{'contentType': 'application/vnd.microsoft.card.adaptive',
        'content': {'$schema': 'http://adaptivecards.io/schemas/adaptive-card.json', 'type': 'AdaptiveCard',
            'version': '1.2', 'body': [{'type': 'TextBlock', 'text': '✅ Webhook actualizado correctamente — Sistema Vacaciones USIL', 'wrap': True}]}}]}
    ok, err = _post_teams_webhook(nueva_url, test_card)
    if not ok:
        return jsonify({'ok': False, 'error': f'La URL no funciona: {err}'}), 400

    _guardar_pa_config({'teams_webhook_url': nueva_url, 'teams_webhook_personal_url': nueva_url})
    return jsonify({'ok': True, 'mensaje': 'Webhook actualizado y verificado. Mensaje de prueba enviado a Teams.'})


@app.route('/api/teams/webhook-test', methods=['POST'])
def api_teams_webhook_test():
    """Prueba el webhook actual sin modificarlo."""
    url = _resolver_teams_webhook()
    if not url:
        return jsonify({'ok': False, 'error': 'No hay URL de webhook configurada en pa_config.json'})
    test_card = {'type': 'message', 'attachments': [{'contentType': 'application/vnd.microsoft.card.adaptive',
        'content': {'$schema': 'http://adaptivecards.io/schemas/adaptive-card.json', 'type': 'AdaptiveCard',
            'version': '1.2', 'body': [{'type': 'TextBlock', 'text': f'🔔 Prueba webhook {__import__("datetime").datetime.now().strftime("%H:%M:%S")} — Sistema Vacaciones USIL', 'wrap': True}]}}]}
    ok, err = _post_teams_webhook(url, test_card)
    return jsonify({'ok': ok, 'url_tipo': 'flujos_de_trabajo' if any(d in url for d in ('logic.azure', 'prod-')) else 'clasico', 'error': err})


def api_test_notif():

    import urllib.request

    payload = request.get_json(silent=True) or {}

    solo_prueba = bool(payload.get('solo_prueba', False))

    dias = int(payload.get('dias_proximos', 30))

    enviar_teams = _cfg_bool(payload.get('enviar_teams', True), True)

    enviar_smtp = _cfg_bool(payload.get('enviar_smtp', False), False)

    encolar_pa = _cfg_bool(payload.get('encolar_pa', _ENCOLAR_ALERTAS_PA), _ENCOLAR_ALERTAS_PA)

    asunto_payload = payload.get('asunto', '').strip()

    tiene_mensaje_override = 'mensaje' in payload

    mensaje_payload = str(payload.get('mensaje', '') or '').strip()

    tiene_aviso_override = 'aviso' in payload

    aviso_payload = str(payload.get('aviso', '') or '').strip()

    tiene_reco_override = 'recomendacion' in payload

    reco_payload = str(payload.get('recomendacion', '') or '').strip()

    selected_supervisores = payload.get('selected_supervisores', [])
    hrbp_filtro = str(payload.get('hrbp_filtro', '') or '').strip()

    selected_keys = set()

    if isinstance(selected_supervisores, list):

        selected_keys = {

            _nombre_cmp_key(_nombre_supervisor_canonico(str(x or ''))) for x in selected_supervisores

            if str(x or '').strip()

        }



    if not enviar_teams and not enviar_smtp and not encolar_pa:

        return jsonify({'ok': False, 'error': 'Selecciona al menos un canal de envio'})



    # --- Aniversarios ---

    df, err = cargar_datos()

    if err: return jsonify({'ok': False, 'error': err})

    hoy = date.today()

    dfo, _ = cargar_objetivos()

    obj_lookup, df_obj_act, cols_obj = _build_objetivos_lookup(dfo)

    excluded_ids = {_id_key(x) for x in _TEAMS_EXCLUIR_MAT}



    # Excluir personas que NUNCA reciben notificaciÃ³n Teams

    _c_mat_df = _col(df, 'Matricula')

    if _c_mat_df:

        df_teams = df[df[_c_mat_df].apply(lambda v: _id_key(v) not in excluded_ids)].copy()

    else:

        df_teams = df



    ret_raw  = [_persona(r, hoy) for _, r in df_teams[df_teams['_dias_aniv'] < 0].iterrows()] \
        if '_dias_aniv' in df_teams.columns else []

    mask_p = (df_teams['_dias_aniv'] > 0) & (df_teams['_dias_aniv'] <= dias) \
        if '_dias_aniv' in df_teams.columns else pd.Series([False]*len(df_teams))

    prox_raw = [_persona(r, hoy) for _, r in df_teams[mask_p].iterrows()]



    ret = _enriquecer_alertas_con_objetivos(ret_raw, obj_lookup)

    prox = _enriquecer_alertas_con_objetivos(prox_raw, obj_lookup)



    persona_lookup = {}

    if _c_mat_df:

        for _, rr in df_teams.iterrows():

            mat_key = _id_key(rr.get(_c_mat_df, ''))

            if not mat_key:

                continue

            pp = _persona(rr, hoy)

            obj = obj_lookup.get(mat_key)

            if not obj:

                continue

            for k in ('nombre', 'area', 'puesto', 'hrbp', 'supervisor', 'fecha_ingreso', 'objetivo', 'gozados', 'truncas', 'pendientes', 'vencidas', 'total_vacaciones', 'cantidad_pendiente', 'comentario'):

                if obj.get(k) not in (None, ''):

                    pp[k] = obj.get(k)

            avp, rcp = _aviso_recomendacion_persona(pp)

            pp['aviso'] = avp

            pp['recomendacion'] = rcp

            persona_lookup[mat_key] = pp



    # --- Objetivo trimestral: colaboradores sin cumplir ---

    sin_cumplir, kpis_obj = [], {}

    # Lookup: matricula â†’ supervisor (desde PersonalMaestroReporte ya cargado)

    c_mat_pm = _col(df, 'Matricula')

    c_sup_pm = _col(df, 'Supervisor', 'Jefe_Directo', 'Jefe Directo', 'Jefe', 'Business Partner', 'HRBP')

    sup_lookup = {}

    if c_mat_pm and c_sup_pm:

        for _, row in df[[c_mat_pm, c_sup_pm]].iterrows():

            mat = _id_key(row[c_mat_pm])

            sup = str(row.get(c_sup_pm, '')).strip()

            if mat and sup and sup.lower() not in ('nan', '', 'no aplica', 'no asignado'):

                sup_lookup[mat] = sup

    c_sup_obj = cols_obj.get('supervisor')

    if df_obj_act is not None and c_sup_obj and cols_obj.get('matricula'):

        for _, row in df_obj_act.iterrows():

            mat = _id_key(row.get(cols_obj.get('matricula'), ''))

            sup = _safe(row.get(c_sup_obj, ''))

            if mat and sup:

                sup_lookup[mat] = sup



    if df_obj_act is not None and cols_obj.get('cumplimiento'):

        c_cum = cols_obj.get('cumplimiento')

        c_obj = cols_obj.get('objetivo')

        df_sin = df_obj_act[df_obj_act[c_cum] < 0].sort_values(c_cum)

        for _, row in df_sin.iterrows():

            mat_key = _id_key(row.get(cols_obj.get('matricula'), '')) if cols_obj.get('matricula') else ''

            if not mat_key or mat_key in excluded_ids:

                continue

            item = dict(obj_lookup.get(mat_key, {}))

            if not item:

                continue

            item['supervisor'] = item.get('supervisor') or sup_lookup.get(mat_key, '')

            if mat_key in persona_lookup:

                item['correo'] = persona_lookup[mat_key].get('correo', '')

            sin_cumplir.append(item)

        dias_obj    = int(df_obj_act[c_obj].sum()) if c_obj else 0

        deficit     = int(df_obj_act[c_cum][df_obj_act[c_cum] < 0].sum())

        dias_goz    = dias_obj + deficit

        pct         = round(dias_goz / dias_obj * 100, 1) if dias_obj > 0 else 0

        kpis_obj    = {'dias_objetivo': dias_obj, 'dias_gozados': dias_goz,

                       'cumplimiento': pct, 'sin_cumplir': len(sin_cumplir),

                       'dias_pendientes': abs(deficit)}



    if not ret and not prox and not sin_cumplir:

        return jsonify({'ok': True, 'en_retraso': 0, 'proximos': 0, 'sin_cumplir': 0, 'mensaje': 'Sin alertas'})

    if solo_prueba:

        return jsonify({'ok': True, 'en_retraso': len(ret), 'proximos': len(prox),

                        'sin_cumplir': len(sin_cumplir), 'mensaje': 'Solo prueba'})



    webhook = _resolver_teams_webhook()

    if enviar_teams and not webhook:

        return jsonify({'ok': False, 'error': 'No hay webhook en pa_config.json'})



    teams_ok, teams_err, msgs_sent = False, None, 0

    if enviar_teams:

        mensajes = _mensajes_teams(ret, prox, dias, sin_cumplir, kpis_obj)

        import time as _time

        for i, card in enumerate(mensajes):

            try:

                ok_w, err_w = _post_teams_webhook(webhook, card)

                if ok_w:
                    msgs_sent += 1

                teams_ok = True

                if i < len(mensajes) - 1:

                    _time.sleep(0.8)

            except Exception as e:

                teams_err = str(e)

                break



    # â”€â”€ CONFIRMO: mensaje final al canal Teams confirmando el envÃ­o â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    if enviar_teams and teams_ok and msgs_sent > 0:

        _ts_confirmo = datetime.now().strftime('%d/%m/%Y %H:%M')

        _confirmo = {'type':'message','attachments':[{

            'contentType':'application/vnd.microsoft.card.adaptive',

            'content':{'$schema':'http://adaptivecards.io/schemas/adaptive-card.json',

                       'type':'AdaptiveCard','version':'1.4','body':[

                {'type':'TextBlock','size':'Large','weight':'Bolder','color':'Good',

                 'text':'\u2705  CONFIRMO \u2014 Alerta enviada correctamente'},

                {'type':'FactSet','separator':True,'spacing':'Medium','facts':[

                    {'title':'Fecha / Hora:',       'value': _ts_confirmo},

                    {'title':'Tarjetas Teams:',     'value': f'{msgs_sent} mensajes enviados al canal'},

                    {'title':'Aniversarios:',       'value': f'{len(ret)} en retraso \u00b7 {len(prox)} pr\u00f3ximos {dias}d'},

                    {'title':'Sin cumplir obj.:',   'value': str(len(sin_cumplir))+' colaboradores'},

                    {'title':'Sistema:',            'value': 'People Analytics USIL'},

                ]},

                {'type':'TextBlock','isSubtle':True,'size':'Small','wrap':True,

                 'text':'Este mensaje confirma que las alertas fueron generadas y enviadas correctamente.'},

            ]}}]}

        try:

            _cb = json.dumps(_confirmo).encode('utf-8')

            _cr = urllib.request.Request(webhook, data=_cb, method='POST')

            _cr.add_header('Content-Type', 'application/json')

            with urllib.request.urlopen(_cr, timeout=15) as _r:

                _r.read()

            msgs_sent += 1

            print('[CONFIRMO] Mensaje de confirmacion enviado al canal Teams')

        except Exception as _e:

            print(f'[CONFIRMO] Error al enviar confirmacion: {_e}')



    # Construir entradas por jefe (multi-destinatario)

    fecha_str  = hoy.strftime('%d/%m/%Y')

    campania = _contexto_campania(hoy)

    equipo_sup_map = _equipo_objetivo_por_supervisor(df_obj_act, cols_obj, obj_lookup)
    equipo_sup_lookup = {
        _nombre_cmp_key(k): v for k, v in (equipo_sup_map or {}).items()
    }

    jefes_data = _agrupar_por_supervisor(ret, prox, sin_cumplir, df)
    personal_idx = _build_personal_index(df)
    default_email = _resolver_default_email()

    cola = []

    for sup, data in jefes_data.items():
        sup_key = _nombre_cmp_key(_nombre_supervisor_canonico(data.get('nombre', '')))
        if selected_keys and sup_key not in selected_keys:
            continue
        if not data['email']:
            continue
        
        r_l = data['ret']
        p_l = data['prox']
        s_l = data['sc']
        
        equipo_sup = equipo_sup_lookup.get(_nombre_cmp_key(_nombre_supervisor_canonico(data.get('nombre', ''))), [])
        meta_resumen = _resumen_meta_equipo(equipo_sup)
        ad_l = _seleccionar_candidatos_adicionales(equipo_sup, r_l, p_l, s_l, limite=5)
        
        partes = []
        if r_l: partes.append(f"{len(r_l)} aniversario{'s' if len(r_l)>1 else ''} en retraso")
        if p_l: partes.append(f"{len(p_l)} próximo{'s' if len(p_l)>1 else ''}")
        if s_l: partes.append(f"{len(s_l)} sin cumplir")
        if ad_l: partes.append(f"{len(ad_l)} adicionales")

        if asunto_payload:

            asunto = _asegurar_asunto_alerta(_tpl(asunto_payload, {
                'nombre': data.get('nombre', ''),
                'jefe': data.get('nombre', ''),
                'trimestre': campania.get('trimestre', ''),
                'fecha': hoy.strftime('%d/%m/%Y'),
                'fecha_limite': _PA_CONFIG.get('vacaciones_fecha_limite_gestion', 'la fecha acordada'),
                'hrbp': hrbp_filtro or _buscar_hrbp_destinatario(equipo_sup, fallback='') or _HRBP_TYC_DEFAULT,
                'hrbp_nombre': hrbp_filtro or _buscar_hrbp_destinatario(equipo_sup, fallback='') or _HRBP_TYC_DEFAULT,
            }))

        else:

            asunto = (

                f'\U0001f3af Meta Vacaciones USIL \u2014 {meta_resumen.get("cumplimiento_meta_pct", 0.0)}% '

                f'({meta_resumen.get("dias_gozados_hacia_meta", 0)}/{meta_resumen.get("dias_objetivo", 0)} dias) '

                f'| Faltan {meta_resumen.get("dias_pendientes_meta", 0)} dias'

            )

            if partes:

                asunto += ' | ' + ' | '.join(partes)

        mensaje_cfg = mensaje_payload if tiene_mensaje_override else str(data.get('mensaje', '') or '').strip()

        aviso_cfg = aviso_payload if tiene_aviso_override else str(data.get('aviso', '') or '').strip()

        reco_cfg = reco_payload if tiene_reco_override else str(data.get('recomendacion', '') or '').strip()

        hrbp_nombre = hrbp_filtro or _buscar_hrbp_destinatario(equipo_sup, fallback='')

        contenido_jefe = {

            'email_jefe': data['email'],
            'email_destino_real': data['email'],

            'nombre_jefe': data['nombre'],

            'matricula_jefe': data.get('matricula', ''),

            'modo_prueba': False,

            'asunto': asunto,

            'mensaje_html': _build_html_jefe(

                data['nombre'],

                r_l,

                p_l,

                s_l,

                fecha_str,

                mensaje_cfg,

                aviso_cfg,

                reco_cfg,

                ad_l,

                campania,

                meta_resumen,

                hrbp_nombre=hrbp_nombre,

                equipo_full=equipo_sup,

            ),

            'mensaje_teams': _texto_teams_supervisor(data['nombre'], r_l, p_l, s_l, ad_l, campania, meta_resumen, equipo_full=equipo_sup),

            'fecha_generacion': datetime.now().isoformat(),

            'meta_resumen': meta_resumen,

            'en_retraso': len(r_l),

            'proximos': len(p_l),

            'sin_cumplir': len(s_l),

            'adicionales': len(ad_l),

            'detalle_retraso': r_l,

            'detalle_proximos': p_l,

            'detalle_sin_cumplir': s_l,

            'detalle_adicionales': ad_l,

            'detalle_equipo_full': equipo_sup,

            'campania': campania,

            'mensaje': mensaje_cfg,

            'aviso': aviso_cfg,

            'recomendacion': reco_cfg,

            'mensaje_cfg': mensaje_cfg,

            'aviso_cfg': aviso_cfg,

            'recomendacion_cfg': reco_cfg,

            'hrbp_nombre': hrbp_nombre,

            'teams_confirmado': teams_ok,

        }

        contenido_jefe['mensaje_webhook'] = json.dumps(_build_webhook_payload(contenido_jefe), ensure_ascii=False, indent=2)
        contenido_jefe['teams_adaptive_card'] = json.dumps(_build_adaptive_card_supervisor(contenido_jefe), ensure_ascii=False)

        cola.append(contenido_jefe)

    cola_keys = {
        _nombre_cmp_key(_nombre_supervisor_canonico((it or {}).get('nombre_jefe', '')))
        for it in cola
        if (it or {}).get('nombre_jefe')
    }

    if selected_keys:

        for selected_key in sorted(selected_keys):

            if selected_key in cola_keys:

                continue

            sup_info = _resolver_supervisor_identidad(selected_key, '', personal_idx)
            nombre_jefe = sup_info.get('nombre') or selected_key.title()
            scfg = _resolver_supervisor_cfg(nombre_jefe)
            sup_canon = _nombre_cmp_key(_nombre_supervisor_canonico(nombre_jefe))
            email_jefe = sup_info.get('email', '') or scfg.get('email', '') or default_email
            equipo_sup = equipo_sup_lookup.get(sup_canon, [])
            meta_resumen = _resumen_meta_equipo(equipo_sup)

            if asunto_payload:

                asunto = _asegurar_asunto_alerta(_tpl(asunto_payload, {
                    'nombre': nombre_jefe,
                    'jefe': nombre_jefe,
                    'trimestre': campania.get('trimestre', ''),
                    'fecha': hoy.strftime('%d/%m/%Y'),
                    'fecha_limite': _PA_CONFIG.get('vacaciones_fecha_limite_gestion', 'la fecha acordada'),
                    'hrbp': hrbp_filtro or _buscar_hrbp_destinatario(equipo_sup, fallback='') or _HRBP_TYC_DEFAULT,
                    'hrbp_nombre': hrbp_filtro or _buscar_hrbp_destinatario(equipo_sup, fallback='') or _HRBP_TYC_DEFAULT,
                }))

            else:

                asunto = (

                    f'\U0001f3af Meta Vacaciones USIL \u2014 {meta_resumen.get("cumplimiento_meta_pct", 0.0)}% '

                    f'({meta_resumen.get("dias_gozados_hacia_meta", 0)}/{meta_resumen.get("dias_objetivo", 0)} dias) '

                    f'| Faltan {meta_resumen.get("dias_pendientes_meta", 0)} dias'

                )

            mensaje_cfg = mensaje_payload if tiene_mensaje_override else str(scfg.get('mensaje', '') or '').strip()

            aviso_cfg = aviso_payload if tiene_aviso_override else str(scfg.get('aviso', '') or '').strip()

            reco_cfg = reco_payload if tiene_reco_override else str(scfg.get('recomendacion', '') or '').strip()

            hrbp_nombre = hrbp_filtro or _buscar_hrbp_destinatario(equipo_sup, fallback='')

            contenido_jefe = {

                'email_jefe': email_jefe,

                'nombre_jefe': nombre_jefe,

                'matricula_jefe': sup_info.get('matricula', ''),

                'asunto': asunto,

                'mensaje_html': _build_html_jefe(

                    nombre_jefe,

                    [],

                    [],

                    [],

                    fecha_str,

                    mensaje_cfg,

                    aviso_cfg,

                    reco_cfg,

                    [],

                    campania,

                    meta_resumen,

                    hrbp_nombre=hrbp_nombre,

                    equipo_full=equipo_sup,

                ),

                'mensaje_teams': _texto_teams_supervisor(nombre_jefe, [], [], [], [], campania, meta_resumen, equipo_full=equipo_sup),

                'fecha_generacion': datetime.now().isoformat(),

                'meta_resumen': meta_resumen,

                'en_retraso': 0,

                'proximos': 0,

                'sin_cumplir': 0,

                'adicionales': 0,

                'detalle_retraso': [],

                'detalle_proximos': [],

                'detalle_sin_cumplir': [],

                'detalle_adicionales': [],

                'detalle_equipo_full': equipo_sup,

                'campania': campania,

                'mensaje': mensaje_cfg,

                'aviso': aviso_cfg,

                'recomendacion': reco_cfg,

                'mensaje_cfg': mensaje_cfg,

                'aviso_cfg': aviso_cfg,

                'recomendacion_cfg': reco_cfg,

                'hrbp_nombre': hrbp_nombre,

                'teams_confirmado': teams_ok,

            }

            contenido_jefe['mensaje_webhook'] = json.dumps(_build_webhook_payload(contenido_jefe), ensure_ascii=False, indent=2)
            contenido_jefe['teams_adaptive_card'] = json.dumps(_build_adaptive_card_supervisor(contenido_jefe), ensure_ascii=False)

            cola.append(contenido_jefe)
            cola_keys.add(selected_key)

    # Fallback: si no hay emails configurados, enviar resumen al email por defecto

    if not cola and not selected_keys:

        _def_em = _resolver_default_email()

        meta_fallback = _resumen_meta_equipo(list(obj_lookup.values()))

        fallback_equipo = list(obj_lookup.values())

        fallback = {'email_jefe': _def_em, 'nombre_jefe': 'Equipo People Analytics',

                 'matricula_jefe': '',

                 'asunto': (

                     f'\U0001f3af Meta Vacaciones USIL \u2014 {meta_fallback.get("cumplimiento_meta_pct", 0.0)}% '

                     f'({meta_fallback.get("dias_gozados_hacia_meta", 0)}/{meta_fallback.get("dias_objetivo", 0)} dias) '

                     f'| Faltan {meta_fallback.get("dias_pendientes_meta", 0)} dias'

                 ),

                 'mensaje_html': _build_html_jefe('Equipo People Analytics', ret, prox, sin_cumplir, fecha_str, _MSG_BASE_DEFAULT, _AVISO_DEFAULT, _RECOM_DEFAULT, [], campania, meta_fallback, equipo_full=fallback_equipo),

                 'mensaje_teams': _texto_teams_supervisor('Equipo People Analytics', ret, prox, sin_cumplir, [], campania, meta_fallback, equipo_full=fallback_equipo),

                 'fecha_generacion': datetime.now().isoformat(),

                 'meta_resumen': meta_fallback,

                 'en_retraso': len(ret), 'proximos': len(prox), 'sin_cumplir': len(sin_cumplir), 'adicionales': 0,

                 'detalle_retraso': ret, 'detalle_proximos': prox, 'detalle_sin_cumplir': sin_cumplir, 'detalle_adicionales': [],

                 'detalle_equipo_full': fallback_equipo,

                 'campania': campania,

                 'teams_confirmado': teams_ok}

        fallback['mensaje_webhook'] = json.dumps(_build_webhook_payload(fallback), ensure_ascii=False, indent=2)
        fallback['teams_adaptive_card'] = json.dumps(_build_adaptive_card_supervisor(fallback), ensure_ascii=False)

        cola = [fallback]

    if selected_keys and not cola:

        return jsonify({'ok': False, 'error': 'No hay jefes seleccionados con data para enviar'}), 400

    json_jefes = len(cola)

    public_base_url = _resolver_public_base_url(request.host_url)
    cola = [_registrar_confirmaciones_contenido(entrada, public_base_url) for entrada in cola]



    # ── JSON para Power Automate: UN archivo por supervisor para que PA procese
    # cada correo de forma independiente (archivos grandes con todos los supervisores
    # juntos pueden superar los límites de PA y fallar silenciosamente).

    json_ok, json_file, json_archivado, json_meta = False, None, None, None

    if encolar_pa:

        try:

            archivos_creados = 0
            for _entrada in cola:
                _jf, _ja, _jm = _guardar_json_cola([_entrada], prefijo='alerta_vacaciones')
                if json_file is None:
                    json_file, json_archivado, json_meta = _jf, _ja, _jm
                archivos_creados += 1

            json_ok = archivos_creados > 0

            print(f'[COLA] {archivos_creados}/{json_jefes} archivos individuales creados en pendientes')

            print(f'[COLA] Liberacion en {_PA_DELAY_SECONDS}s: {COLA_PENDIENTES_DIR}')

        except Exception as e:

            print(f'[COLA] Error: {e}')



    # â”€â”€ SMTP directo: enviar correo a cada jefe sin depender de PA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    smtp_results = []

    if enviar_smtp:

        import time as _t2

        for entrada in cola:

            s_ok, s_err = _enviar_correo_smtp(

                entrada.get('email_jefe', ''),

                entrada.get('nombre_jefe', ''),

                entrada.get('asunto', 'Alerta Vacaciones USIL'),

                entrada.get('mensaje_html', ''),

            )

            smtp_results.append({'email': entrada.get('email_jefe'), 'ok': s_ok, 'error': s_err})

            if len(cola) > 1:

                _t2.sleep(0.5)   # respetar lÃ­mite Office 365

    smtp_ok_total  = sum(1 for r in smtp_results if r['ok'])

    smtp_err_total = [r for r in smtp_results if not r['ok']]



    _registrar_log('teams', 'webhook+smtp', [{

        'nombre':       'Canal Teams',

        'email':        'webhook',

        'ok':            teams_ok,

        'error':         teams_err or '',

        'retraso':       len(ret),

        'proximos':      len(prox),

        'sin_cumplir':   len(sin_cumplir),

        'json_jefes':    json_jefes,

        'archivado':     json_archivado or '',

        'smtp_enviados': smtp_ok_total,

    }])

    return jsonify({

        'ok':              teams_ok or smtp_ok_total > 0 or json_ok,

        'en_retraso':      len(ret),

        'proximos':        len(prox),

        'sin_cumplir':     len(sin_cumplir),

        'teams_enviado':   teams_ok,

        'teams_error':     teams_err,

        'msgs_enviados':   msgs_sent,

        'json_guardado':   json_ok,

        'json_archivo':    json_file,

        'json_archivado':  json_archivado,

        'json_meta':       json_meta,

        'json_jefes':      json_jefes,

        'smtp_enviados':   smtp_ok_total,

        'smtp_errores':    smtp_err_total,

        'confirmo_enviado': teams_ok,

        'canales': {

            'teams_webhook': enviar_teams,

            'smtp_correo': enviar_smtp,

            'flujo_power_automate': encolar_pa,

        }

    })



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# CORREO DE PRUEBA (genera JSON en alertas_cola para que PA lo procese)

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/api/test-correo', methods=['POST'])

def api_test_correo():

    """EnvÃ­a correo de prueba DIRECTO por SMTP (sin depender de Power Automate).

    TambiÃ©n genera el JSON en alertas_cola para que PA lo procese si estÃ¡ activo."""

    payload      = request.get_json(silent=True) or {}

    dest_email   = payload.get('email', _resolver_default_email())

    dest_nombre  = payload.get('nombre', 'People Analytics USIL')

    ahora        = datetime.now()

    fecha_str    = ahora.strftime('%d/%m/%Y %H:%M')

    ts           = ahora.strftime('%Y%m%d_%H%M%S')



    html = (

        '<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"></head>'

        '<body style="margin:0;padding:0;font-family:Segoe UI,Arial,sans-serif;background:#f5f5f5">'

        '<div style="background:linear-gradient(135deg,#0c4a6e,#0369a1);padding:18px 24px;text-align:center">'

        '<h1 style="color:#fff;margin:0;font-size:20px">&#128197; Sistema de Vacaciones &mdash; USIL</h1>'

        '<p style="color:#bae6fd;margin:4px 0 0;font-size:12px">People Analytics &middot; Prueba de conectividad</p>'

        '</div>'

        '<div style="max-width:640px;margin:20px auto;background:#fff;border-radius:10px;'

        'box-shadow:0 4px 12px rgba(0,0,0,.1);padding:24px 28px">'

        f'<p style="font-size:15px;color:#333">Hola, <strong>{dest_nombre}</strong> &#128075;</p>'

        '<div style="background:#f0fdf4;border-radius:8px;padding:16px;border:1px solid #86efac;margin:16px 0">'

        '<h3 style="color:#166534;margin:0 0 8px;font-size:15px">&#9989; SMTP Office 365 funcionando</h3>'

        '<p style="margin:0;font-size:13px;color:#166534">'

        'El sistema envi&oacute; este correo <strong>directamente</strong> desde Python '

        'via SMTP &mdash; sin depender de Power Automate.</p>'

        '</div>'

        '<div style="background:#f0f9ff;border-radius:8px;padding:14px;border:1px solid #bae6fd;margin-top:12px">'

        f'<p style="margin:0;color:#0369a1;font-size:12px">'

        f'&#128276; Generado autom&aacute;ticamente &mdash; {fecha_str} &mdash; People Analytics USIL</p>'

        '</div></div></body></html>'

    )

    asunto = f'\u2705 TEST - Sistema Alertas Vacaciones USIL \u2014 {fecha_str}'



    # 1) Enviar DIRECTO por SMTP (inmediato, independiente de PA)

    smtp_ok, smtp_err = _enviar_correo_smtp(dest_email, dest_nombre, asunto, html)



    # 2) Opcional: guardar JSON para PA si estÃ¡ habilitado

    json_ok, fname_out, json_meta, json_err = False, None, None, None

    if _ENCOLAR_ALERTAS_PA:

        try:

            entrada = {'email_jefe': dest_email, 'nombre_jefe': dest_nombre,

                       'asunto': asunto, 'mensaje_html': html,

                       'fecha_generacion': ahora.isoformat(),

                       'en_retraso': 0, 'proximos': 0, 'sin_cumplir': 0,

                       'smtp_enviado': smtp_ok, 'tipo': 'prueba'}

            fname_out, _, json_meta = _guardar_json_cola([entrada], prefijo='alerta_vacaciones')

            json_ok = True

            print(f'[TEST-CORREO] JSON -> {fname_out}')

        except Exception as e:

            json_err = str(e)



    ok = smtp_ok or json_ok

    return jsonify({

        'ok':            ok,

        'smtp_enviado':  smtp_ok,

        'smtp_error':    smtp_err,

        'json_guardado': json_ok,

        'json_meta':     json_meta,

        'archivo':       fname_out,

        'destinatario':  dest_email,

        'metodo':        'smtp_directo' if smtp_ok else ('json_cola' if json_ok else 'fallido'),

        'error':         smtp_err if not smtp_ok else json_err,

    })





@app.route('/api/teams-ping', methods=['POST'])

def api_teams_ping():

    import urllib.request

    webhook = POWER_AUTOMATE_URL or TEAMS_WEBHOOK_URL

    if not webhook:

        return jsonify({'ok': False, 'error': 'No hay webhook configurado en pa_config.json'})

    hoy_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    card = {'type':'message','attachments':[{'contentType':'application/vnd.microsoft.card.adaptive',

        'content':{'$schema':'http://adaptivecards.io/schemas/adaptive-card.json',

                   'type':'AdaptiveCard','version':'1.4','body':[

            {'type':'TextBlock','size':'Medium','weight':'Bolder','color':'Accent',

             'text':'Prueba de Conectividad'},

            {'type':'TextBlock','text':f'Enviado: {hoy_str}','isSubtle':True,'size':'Small'},

            {'type':'FactSet','facts':[

                {'title':'Sistema:','value':'People Analytics USIL'},

                {'title':'Estado:','value':'\u2705 Operativo'},

                {'title':'Proyecto:','value':'PROYECTO TREINNE'}]},

            {'type':'TextBlock','text':'Si ves este mensaje, el webhook esta funcionando correctamente.',

             'wrap':True,'isSubtle':True}]}}]}

    try:

        body = json.dumps(card).encode('utf-8')

        req = urllib.request.Request(webhook, data=body, method='POST')

        req.add_header('Content-Type', 'application/json')

        with urllib.request.urlopen(req, timeout=15) as resp:

            resp.read()

        return jsonify({'ok': True, 'mensaje': 'Mensaje de prueba enviado correctamente'})

    except Exception as e:

        return jsonify({'ok': False, 'error': str(e)})





def _notificar_teams_personal(titulo, detalle_items=None, color='Accent'):

    """EnvÃ­a una Adaptive Card de confirmaciÃ³n al webhook de Teams configurado."""

    import urllib.request as _ureq

    webhook = TEAMS_WEBHOOK_PERSONAL_URL or TEAMS_WEBHOOK_URL

    if not webhook:

        return

    hoy_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    body_card = [

        {'type': 'TextBlock', 'size': 'Medium', 'weight': 'Bolder', 'color': color, 'text': titulo},

        {'type': 'TextBlock', 'text': f'Hora: {hoy_str}', 'isSubtle': True, 'size': 'Small', 'spacing': 'None'},

    ]

    if detalle_items:

        facts = [{'title': str(it.get('titulo', '')), 'value': str(it.get('valor', ''))} for it in detalle_items]

        body_card.append({'type': 'FactSet', 'separator': True, 'spacing': 'Small', 'facts': facts})

    card = {'type': 'message', 'attachments': [{'contentType': 'application/vnd.microsoft.card.adaptive',

        'content': {'$schema': 'http://adaptivecards.io/schemas/adaptive-card.json',

                    'type': 'AdaptiveCard', 'version': '1.4', 'body': body_card}}]}

    try:

        data = json.dumps(card).encode('utf-8')

        req = _ureq.Request(webhook, data=data, method='POST')

        req.add_header('Content-Type', 'application/json')

        with _ureq.urlopen(req, timeout=10) as r:

            r.read()

    except Exception as e:

        print(f'[TEAMS-PERSONAL] Error al enviar notificacion: {e}')





def _enviar_teams_webhook_supervisor(contenido):

    """Envia una tarjeta resumida al canal Teams/Webhook para un envio individual."""

    import urllib.request as _ureq

    webhook = TEAMS_WEBHOOK_PERSONAL_URL or TEAMS_WEBHOOK_URL

    if not webhook:

        return False, 'No hay webhook configurado en pa_config.json'



    camp = (contenido or {}).get('campania', {}) or {}

    nombre_objetivo = str((contenido or {}).get('nombre_objetivo', (contenido or {}).get('nombre_jefe', '-')) or '-')

    email_objetivo = str((contenido or {}).get('email_objetivo', (contenido or {}).get('email_destino_real', (contenido or {}).get('email_jefe', '-'))) or '-')

    email_entrega = str((contenido or {}).get('email_jefe', '-') or '-')

    modo_prueba = bool((contenido or {}).get('modo_prueba', False))

    resumen = (

        f"Retraso: {(contenido or {}).get('en_retraso', 0)} | "

        f"Proximos: {(contenido or {}).get('proximos', 0)} | "

        f"Sin cumplir: {(contenido or {}).get('sin_cumplir', 0)}"

    )



    body_card = [

        {

            'type': 'TextBlock',

            'size': 'Large',

            'weight': 'Bolder',

            'color': 'Accent',

            'text': 'Alerta Vacaciones USIL',

            'horizontalAlignment': 'Center'

        },

        {

            'type': 'TextBlock',

            'text': nombre_objetivo,

            'size': 'Medium',

            'weight': 'Bolder',

            'color': 'Default',

            'horizontalAlignment': 'Center',

            'spacing': 'Small'

        },

        {

            'type': 'TextBlock',

            'text': f"Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",

            'size': 'Small',

            'isSubtle': True,

            'horizontalAlignment': 'Center'

        },

        {

            'type': 'FactSet',

            'separator': True,

            'spacing': 'Small',

            'facts': [

                {'title': 'Jefe objetivo:', 'value': f'{nombre_objetivo} <{email_objetivo}>'},

                {'title': 'Entrega:', 'value': (f'{email_entrega} (modo prueba)' if modo_prueba else email_entrega)},

            ]

        },

        {

            'type': 'Container',

            'separator': True,

            'spacing': 'Medium',

            'items': [

                {

                    'type': 'ColumnSet',

                    'columns': [

                        {

                            'width': 'stretch',

                            'items': [{

                                'type': 'TextBlock',

                                'text': str((contenido or {}).get('en_retraso', 0)),

                                'size': 'ExtraLarge',

                                'weight': 'Bolder',

                                'color': 'Attention' if (contenido or {}).get('en_retraso', 0) > 0 else 'Good',

                                'horizontalAlignment': 'Center'

                            }],

                        },

                        {

                            'width': 'stretch',

                            'items': [{

                                'type': 'TextBlock',

                                'text': str((contenido or {}).get('proximos', 0)),

                                'size': 'ExtraLarge',

                                'weight': 'Bolder',

                                'color': 'Warning' if (contenido or {}).get('proximos', 0) > 0 else 'Good',

                                'horizontalAlignment': 'Center'

                            }],

                        },

                        {

                            'width': 'stretch',

                            'items': [{

                                'type': 'TextBlock',

                                'text': str((contenido or {}).get('sin_cumplir', 0)),

                                'size': 'ExtraLarge',

                                'weight': 'Bolder',

                                'color': 'Attention' if (contenido or {}).get('sin_cumplir', 0) > 0 else 'Good',

                                'horizontalAlignment': 'Center'

                            }],

                        }

                    ]

                },

                {

                    'type': 'ColumnSet',

                    'columns': [

                        {

                            'width': 'stretch',

                            'items': [{

                                'type': 'TextBlock',

                                'text': 'En Retraso',

                                'size': 'Small',

                                'weight': 'Lighter',

                                'horizontalAlignment': 'Center'

                            }],

                        },

                        {

                            'width': 'stretch',

                            'items': [{

                                'type': 'TextBlock',

                                'text': 'Proximos',

                                'size': 'Small',

                                'weight': 'Lighter',

                                'horizontalAlignment': 'Center'

                            }],

                        },

                        {

                            'width': 'stretch',

                            'items': [{

                                'type': 'TextBlock',

                                'text': 'Sin Cumplir',

                                'size': 'Small',

                                'weight': 'Lighter',

                                'horizontalAlignment': 'Center'

                            }],

                        }

                    ]

                }

            ]

        }

    ]

    # Tabla del equipo completo (igual que el correo)

    equipo_full = (contenido or {}).get('detalle_equipo_full', [])

    if not equipo_full:

        # Fallback: combinar todas las listas de detalle

        seen_mats = set()

        for lst in ['detalle_retraso', 'detalle_proximos', 'detalle_sin_cumplir', 'detalle_adicionales']:

            for p in (contenido or {}).get(lst, []):

                mk = str(p.get('matricula', '')).strip()

                if mk not in seen_mats:

                    seen_mats.add(mk)

                    equipo_full.append(p)



    if equipo_full:

        def _cell(txt, bold=False, color=None):

            tb = {'type': 'TextBlock', 'text': str(txt or '-'), 'wrap': False, 'size': 'Small'}

            if bold:

                tb['weight'] = 'Bolder'

            if color:

                tb['color'] = color

            return {'type': 'TableCell', 'items': [tb]}



        def _hcell(txt):

            return {'type': 'TableCell', 'style': 'accent', 'items': [

                {'type': 'TextBlock', 'text': txt, 'weight': 'Bolder', 'size': 'Small', 'color': 'Light'}

            ]}



        col_widths = [1, 3, 2, 1, 1, 1, 1, 1]

        header_row = {

            'type': 'TableRow',

            'cells': [

                _hcell('Matricula'), _hcell('Nombre'), _hcell('Puesto'),

                _hcell('Meta'), _hcell('Total'), _hcell('Avance'), _hcell('x Prog'), _hcell('Saldo'),

            ]

        }

        data_rows = []

        for p in equipo_full[:50]:

            obj = int(p.get('objetivo') or 0)

            goz = round(float(p.get('gozados') or 0), 1)

            total = round(float(p.get('total_vacaciones') or 0), 1)

            por_prog = max(obj - int(goz), 0)

            saldo = round(max(total - goz, 0), 1)

            data_rows.append({

                'type': 'TableRow',

                'cells': [

                    _cell(p.get('matricula', '-')),

                    _cell((p.get('nombre') or '-')[:35], bold=True),

                    _cell((p.get('puesto') or '-')[:30]),

                    _cell(obj, color='Attention' if obj == 0 else None),

                    _cell(total),

                    _cell(goz),

                    _cell(por_prog, color='Warning' if por_prog > 0 else 'Good'),

                    _cell(saldo, color='Accent' if saldo > 0 else None),

                ]

            })



        body_card.append({

            'type': 'TextBlock',

            'text': f'Reporte del equipo — {len(equipo_full)} colaboradores',

            'size': 'Medium',

            'weight': 'Bolder',

            'separator': True,

            'spacing': 'Medium',

        })

        body_card.append({

            'type': 'Table',

            'columns': [{'width': w} for w in col_widths],

            'rows': [header_row] + data_rows,

            'showGridLines': True,

            'firstRowAsHeaders': True,

            'spacing': 'Small',

        })



    card = {

        'type': 'message',
        'attachments': [{

            'contentType': 'application/vnd.microsoft.card.adaptive',

            'content': {

                '$schema': 'http://adaptivecards.io/schemas/adaptive-card.json',

                'type': 'AdaptiveCard',

                'version': '1.5',

                'body': body_card,

            }

        }]

    }



    try:

        data = json.dumps(card).encode('utf-8')

        req = _ureq.Request(webhook, data=data, method='POST')

        req.add_header('Content-Type', 'application/json')

        with _ureq.urlopen(req, timeout=15) as resp:

            resp.read()

        return True, None

    except Exception as e:

        return False, str(e)





def _texto_teams_supervisor(nombre_jefe, ret_list, prox_list, sc_list, adicionales_list=None, campania=None, meta_resumen=None, equipo_full=None, plantilla_msg=''):

    """Preview del mensaje Teams: encabezado + tabla completa del equipo."""

    adicionales_list = adicionales_list or []

    campania = campania or _contexto_campania()

    meta_resumen = meta_resumen or {}

    hoy_txt = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    meta_obj = int(meta_resumen.get('dias_objetivo', 0) or 0)

    meta_goz = int(meta_resumen.get('dias_gozados_hacia_meta', 0) or 0)

    meta_pend = int(meta_resumen.get('dias_pendientes_meta', 0) or 0)

    meta_pct = float(meta_resumen.get('cumplimiento_meta_pct', 0.0) or 0.0)

    meta_saldo = int(meta_resumen.get('dias_saldo_total', 0) or 0)



    lineas = [

        '=== ALERTA VACACIONES - RESUMEN ===',

        f'Jefe/Area: {nombre_jefe}',

        f'Periodo: {_periodo_campania_texto(campania)}',

        f'Actualizado: {hoy_txt}',

    ]

    if meta_obj > 0:

        lineas.append(f'META: {meta_pct}% ({meta_goz}/{meta_obj} dias) | Faltan: {meta_pend} dias')

    else:

        lineas.append(f'SALDO TOTAL EQUIPO: {meta_saldo} dias | Sin meta asignada')

    lineas += [

        '',

        f'RETRASO: {len(ret_list)}  |  PROXIMOS: {len(prox_list)}  |  SIN CUMPLIR: {len(sc_list)}',

        '',

    ]



    # Tabla completa (encabezado fijo)

    COL = '{:<10} {:<35} {:<5} {:<6} {:<6} {:<6} {:<6}'

    sep  = '-' * 80

    lineas.append(COL.format('Matricula', 'Nombre', 'Meta', 'Total', 'Avance', 'xProg', 'Saldo'))

    lineas.append(sep)



    # Usar equipo_full si está disponible (todos los colaboradores del jefe)

    if equipo_full:

        todos = list(equipo_full)

    else:

        todos = []

        seen = set()

        for lst in [ret_list, prox_list, sc_list, adicionales_list]:

            for p in (lst or []):

                mk = str(p.get('matricula', '')).strip()

                if mk not in seen:

                    seen.add(mk)

                    todos.append(p)



    for p in todos:

        obj  = int(p.get('objetivo') or 0)

        goz  = round(float(p.get('gozados') or 0), 1)

        tot  = round(float(p.get('total_vacaciones') or 0), 1)

        prog = max(obj - int(goz), 0)

        sald = round(max(tot - goz, 0), 1)

        nom  = (p.get('nombre') or '-')[:35]

        mat  = str(p.get('matricula', '-'))[:10]

        lineas.append(COL.format(mat, nom, obj, tot, goz, prog, sald))



    if not todos:

        lineas.append('Sin registros para este periodo.')



    return '\n'.join(lineas).strip()



def _enviar_teams_webhook_supervisor(contenido):
    """Envia mensaje simple a Teams: solo aviso de correo pendiente."""
    import urllib.request as _ureq

    webhook = TEAMS_WEBHOOK_PERSONAL_URL or TEAMS_WEBHOOK_URL
    if not webhook:
        return False, 'No hay webhook configurado en pa_config.json'

    card = _build_adaptive_card_supervisor(contenido)
    try:
        data = json.dumps(card).encode('utf-8')
        req = _ureq.Request(webhook, data=data, method='POST')
        req.add_header('Content-Type', 'application/json')
        with _ureq.urlopen(req, timeout=15) as resp:
            resp.read()
        return True, None
    except Exception as e:
        return False, str(e)


def _enviar_teams_webhook_supervisor_UNUSED(contenido):
    """DESACTIVADA — version anterior con KPIs y tabla compleja."""
    camp = (contenido or {}).get('campania', {}) or {}
    meta = (contenido or {}).get('meta_resumen', {}) or {}
    nombre_objetivo = str((contenido or {}).get('nombre_objetivo', (contenido or {}).get('nombre_jefe', '-')) or '-')
    modo_prueba = bool((contenido or {}).get('modo_prueba', False))
    hoy_txt = datetime.now().strftime('%d/%m/%Y %H:%M')
    trimestre = _periodo_campania_texto(camp)

    total_eq = int(meta.get('total_equipo', contenido.get('total_colaboradores', 0)) or 0)
    col_meta = int(meta.get('colaboradores_meta', 0) or 0)
    col_cumpl = int(meta.get('colaboradores_cumplieron', 0) or 0)
    pct_cumpl = float(meta.get('cumplimiento_meta_pct', 0) or 0)
    dias_obj = int(meta.get('dias_objetivo', 0) or 0)
    dias_goz = int(meta.get('dias_gozados_hacia_meta', 0) or 0)
    dias_pend_meta = int(meta.get('dias_pendientes_meta', 0) or 0)
    dias_saldo = int(meta.get('dias_saldo_total', 0) or 0)
    dias_venc = int(meta.get('dias_vencidos_equipo', 0) or 0)
    dias_trun = int(meta.get('dias_truncos_equipo', 0) or 0)
    dias_pend_eq = int(meta.get('dias_pendientes_equipo', 0) or 0)
    col_oblig = int(meta.get('colaboradores_con_obligatorias', 0) or 0)
    col_al_dia = int(meta.get('colaboradores_al_dia', 0) or 0)
    pct_al_dia = float(meta.get('pct_equipo_al_dia', 0) or 0)
    col_sin_meta = int(meta.get('colaboradores_sin_meta', 0) or 0)
    n_retraso = int(contenido.get('en_retraso', 0) or 0)
    n_prox = int(contenido.get('proximos', 0) or 0)
    n_sin_cumpl = int(contenido.get('sin_cumplir', 0) or 0)

    _barra_pct = lambda v: int(min(max(v, 0), 100))

    body_card = [
        {
            'type': 'ColumnSet',
            'columns': [
                {
                    'type': 'Column', 'width': 'auto',
                    'items': [{'type': 'Image', 'url': 'https://cdn-icons-png.flaticon.com/512/3652/3652267.png', 'size': 'Small', 'altText': 'Vacaciones'}]
                },
                {
                    'type': 'Column', 'width': 'stretch',
                    'items': [
                        {'type': 'TextBlock', 'text': 'Alerta Vacaciones USIL', 'size': 'Large', 'weight': 'Bolder', 'color': 'Accent', 'spacing': 'None'},
                        {'type': 'TextBlock', 'text': f'People Analytics · Gestion Humana', 'size': 'Small', 'isSubtle': True, 'spacing': 'None'},
                    ]
                }
            ]
        },
        {
            'type': 'TextBlock',
            'text': nombre_objetivo,
            'size': 'Medium',
            'weight': 'Bolder',
            'spacing': 'Small',
        },
        {
            'type': 'ColumnSet',
            'spacing': 'Small',
            'columns': [
                {'type': 'Column', 'width': 'stretch', 'items': [
                    {'type': 'TextBlock', 'text': f'{trimestre}', 'size': 'Small', 'isSubtle': True, 'spacing': 'None'},
                ]},
                {'type': 'Column', 'width': 'auto', 'items': [
                    {'type': 'TextBlock', 'text': f'Actualizado: {hoy_txt}', 'size': 'Small', 'isSubtle': True, 'spacing': 'None'},
                ]},
            ]
        },
    ]

    if modo_prueba:
        body_card.append({
            'type': 'TextBlock', 'text': '⚠️ MODO PRUEBA', 'color': 'Warning',
            'weight': 'Bolder', 'size': 'Small', 'horizontalAlignment': 'Center',
        })

    body_card.append({'type': 'TextBlock', 'text': ' ', 'separator': True, 'spacing': 'Small'})

    kpi_cols = [
        {'label': 'Equipo', 'valor': str(total_eq), 'color': 'Default'},
        {'label': 'Con meta', 'valor': str(col_meta), 'color': 'Accent'},
        {'label': 'Cumplieron', 'valor': str(col_cumpl), 'color': 'Good'},
        {'label': 'Al dia', 'valor': f'{pct_al_dia:.0f}%', 'color': 'Good' if pct_al_dia >= 70 else ('Warning' if pct_al_dia >= 40 else 'Attention')},
    ]
    body_card.append({
        'type': 'ColumnSet',
        'columns': [
            {
                'type': 'Column', 'width': 'stretch',
                'items': [
                    {'type': 'TextBlock', 'text': k['valor'], 'size': 'ExtraLarge', 'weight': 'Bolder', 'color': k['color'], 'horizontalAlignment': 'Center', 'spacing': 'None'},
                    {'type': 'TextBlock', 'text': k['label'], 'size': 'Small', 'isSubtle': True, 'horizontalAlignment': 'Center', 'spacing': 'None'},
                ]
            } for k in kpi_cols
        ]
    })

    if dias_obj > 0:
        body_card.append({
            'type': 'FactSet', 'spacing': 'Small',
            'facts': [
                {'title': '🎯 Meta trimestral', 'value': f'{dias_goz}/{dias_obj} dias ({pct_cumpl:.0f}%)'},
                {'title': '📅 Faltan por programar', 'value': f'{dias_pend_meta} dias'},
            ]
        })

    body_card.append({
        'type': 'FactSet', 'spacing': 'Small',
        'facts': [
            {'title': '📊 Dias vencidos equipo', 'value': f'{dias_venc} dias' if dias_venc > 0 else '0 ✔️'},
            {'title': '📊 Dias truncos equipo', 'value': f'{dias_trun} dias' if dias_trun > 0 else '0'},
            {'title': '📊 Dias pendientes total', 'value': f'{dias_pend_eq} dias'},
            {'title': '📊 Saldo total equipo', 'value': f'{dias_saldo} dias'},
        ]
    })

    if n_retraso > 0 or n_sin_cumpl > 0 or n_prox > 0:
        alerta_facts = []
        if n_retraso > 0:
            alerta_facts.append({'title': '🔴 En retraso', 'value': f'{n_retraso} colaborador(es)'})
        if n_sin_cumpl > 0:
            alerta_facts.append({'title': '🟠 Sin cumplir meta', 'value': f'{n_sin_cumpl} colaborador(es)'})
        if n_prox > 0:
            alerta_facts.append({'title': '🟡 Proximos a vencer', 'value': f'{n_prox} colaborador(es)'})
        body_card.append({
            'type': 'Container',
            'style': 'attention',
            'spacing': 'Small',
            'items': [
                {'type': 'TextBlock', 'text': 'Situaciones que requieren atencion', 'weight': 'Bolder', 'size': 'Small', 'spacing': 'None'},
                {'type': 'FactSet', 'facts': alerta_facts, 'spacing': 'Small'},
            ]
        })

    equipo = contenido.get('detalle_equipo_full', []) or []
    if equipo:
        equipo_sorted = sorted(equipo, key=lambda p: (_to_float0(p.get('total_vacaciones', 0))), reverse=True)
        max_filas = min(len(equipo_sorted), 20)
        tabla_header = {
            'type': 'ColumnSet',
            'spacing': 'Small',
            'separator': True,
            'columns': [
                {'type': 'Column', 'width': 'stretch', 'items': [{'type': 'TextBlock', 'text': 'Colaborador', 'weight': 'Bolder', 'size': 'Small', 'spacing': 'None'}]},
                {'type': 'Column', 'width': '60px', 'items': [{'type': 'TextBlock', 'text': 'Meta', 'weight': 'Bolder', 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None'}]},
                {'type': 'Column', 'width': '70px', 'items': [{'type': 'TextBlock', 'text': 'Gozados', 'weight': 'Bolder', 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None'}]},
                {'type': 'Column', 'width': '60px', 'items': [{'type': 'TextBlock', 'text': 'Pend.', 'weight': 'Bolder', 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None'}]},
                {'type': 'Column', 'width': '60px', 'items': [{'type': 'TextBlock', 'text': 'Saldo', 'weight': 'Bolder', 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None'}]},
            ]
        }
        body_card.append({'type': 'TextBlock', 'text': f'Detalle del equipo ({total_eq} colaboradores)', 'weight': 'Bolder', 'spacing': 'Medium', 'separator': True})
        body_card.append(tabla_header)

        for p in equipo_sorted[:max_filas]:
            obj_p = int(_to_float0(p.get('objetivo', 0)))
            goz_p = int(_to_float0(p.get('gozados', 0)))
            pend_p = round(_to_float0(p.get('cantidad_pendiente', 0)), 1)
            saldo_p = round(_to_float0(p.get('total_vacaciones', 0)), 1)
            venc_p = _to_float0(p.get('vencidas', 0))
            nombre_p = str(p.get('nombre', '')).strip()
            if len(nombre_p) > 25:
                nombre_p = nombre_p[:23] + '..'

            if venc_p > 0:
                color_saldo = 'Attention'
                indicador = '🔴'
            elif pend_p > 0:
                color_saldo = 'Warning'
                indicador = '🟡'
            elif obj_p > 0 and goz_p >= obj_p:
                color_saldo = 'Good'
                indicador = '✅'
            else:
                color_saldo = 'Default'
                indicador = '➖'

            body_card.append({
                'type': 'ColumnSet',
                'spacing': 'None',
                'columns': [
                    {'type': 'Column', 'width': 'stretch', 'items': [{'type': 'TextBlock', 'text': f'{indicador} {nombre_p}', 'size': 'Small', 'spacing': 'None', 'wrap': False}]},
                    {'type': 'Column', 'width': '60px', 'items': [{'type': 'TextBlock', 'text': str(obj_p) if obj_p > 0 else '-', 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None'}]},
                    {'type': 'Column', 'width': '70px', 'items': [{'type': 'TextBlock', 'text': str(goz_p), 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None'}]},
                    {'type': 'Column', 'width': '60px', 'items': [{'type': 'TextBlock', 'text': str(pend_p) if pend_p > 0 else '-', 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None', 'color': color_saldo}]},
                    {'type': 'Column', 'width': '60px', 'items': [{'type': 'TextBlock', 'text': str(saldo_p), 'size': 'Small', 'horizontalAlignment': 'Center', 'spacing': 'None', 'weight': 'Bolder', 'color': color_saldo}]},
                ]
            })

        if len(equipo_sorted) > max_filas:
            body_card.append({'type': 'TextBlock', 'text': f'... y {len(equipo_sorted) - max_filas} mas. Ver detalle completo en el correo.', 'size': 'Small', 'isSubtle': True, 'spacing': 'Small'})

    body_card.append({
        'type': 'Container',
        'style': 'emphasis',
        'spacing': 'Medium',
        'items': [
            {'type': 'TextBlock', 'text': 'Revisa tu correo "Alertas Vacaciones USIL" para el detalle completo con recomendaciones por colaborador.', 'wrap': True, 'size': 'Small', 'spacing': 'None'},
        ]
    })

    body_card.append({
        'type': 'TextBlock',
        'text': f'Generado por People Analytics USIL · {hoy_txt}',
        'size': 'Small', 'isSubtle': True, 'horizontalAlignment': 'Center', 'spacing': 'Small',
    })

    card = {
        'type': 'message',
        'attachments': [{
            'contentType': 'application/vnd.microsoft.card.adaptive',
            'content': {
                '$schema': 'http://adaptivecards.io/schemas/adaptive-card.json',
                'type': 'AdaptiveCard',
                'version': '1.5',
                'body': body_card,
            }
        }]
    }

    try:
        data = json.dumps(card).encode('utf-8')
        req = _ureq.Request(webhook, data=data, method='POST')
        req.add_header('Content-Type', 'application/json')
        with _ureq.urlopen(req, timeout=15) as resp:
            resp.read()
        return True, None
    except Exception as e:
        return False, str(e)



def _texto_teams_supervisor(nombre_jefe, ret_list, prox_list, sc_list, adicionales_list=None, campania=None, meta_resumen=None, equipo_full=None, plantilla_msg=''):

    """Resumen con KPIs para Teams con detalle del equipo completo."""

    campania = campania or _contexto_campania()
    meta = meta_resumen or {}
    hoy_txt = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    trimestre = _periodo_campania_texto(campania)

    total_eq = int(meta.get('total_equipo', len(equipo_full or [])) or 0)
    col_meta = int(meta.get('colaboradores_meta', 0) or 0)
    col_cumpl = int(meta.get('colaboradores_cumplieron', 0) or 0)
    pct_cumpl = float(meta.get('cumplimiento_meta_pct', 0) or 0)
    dias_obj = int(meta.get('dias_objetivo', 0) or 0)
    dias_goz = int(meta.get('dias_gozados_hacia_meta', 0) or 0)
    dias_pend_meta = int(meta.get('dias_pendientes_meta', 0) or 0)
    dias_venc = int(meta.get('dias_vencidos_equipo', 0) or 0)
    dias_trun = int(meta.get('dias_truncos_equipo', 0) or 0)
    dias_pend_eq = int(meta.get('dias_pendientes_equipo', 0) or 0)
    dias_saldo = int(meta.get('dias_saldo_total', 0) or 0)
    col_oblig = int(meta.get('colaboradores_con_obligatorias', 0) or 0)
    pct_al_dia = float(meta.get('pct_equipo_al_dia', 0) or 0)
    n_retraso = len(ret_list or [])
    n_prox = len(prox_list or [])
    n_sin_cumpl = len(sc_list or [])

    lineas = [
        f'Alerta Vacaciones USIL - {trimestre}',
        f'Jefe/Area: {nombre_jefe}',
        f'Actualizado: {hoy_txt}',
        '',
        f'EQUIPO: {total_eq} colaboradores | {col_meta} con meta | {col_cumpl} cumplieron | {pct_al_dia:.0f}% al dia',
    ]

    if dias_obj > 0:
        lineas.append(f'META TRIMESTRAL: {dias_goz}/{dias_obj} dias ({pct_cumpl:.0f}%) | Faltan: {dias_pend_meta} dias')

    lineas.append(f'SALDOS: Vencidos {dias_venc}d | Truncos {dias_trun}d | Pendientes {dias_pend_eq}d | Saldo total {dias_saldo}d')

    if n_retraso > 0 or n_sin_cumpl > 0 or n_prox > 0:
        lineas.append('')
        lineas.append('ALERTAS:')
        if n_retraso > 0:
            lineas.append(f'  En retraso: {n_retraso}')
        if n_sin_cumpl > 0:
            lineas.append(f'  Sin cumplir meta: {n_sin_cumpl}')
        if n_prox > 0:
            lineas.append(f'  Proximos a vencer: {n_prox}')

    todos = sorted((equipo_full or []), key=lambda p: _to_float0(p.get('total_vacaciones', 0)), reverse=True)
    if todos:
        lineas.append('')
        lineas.append(f'DETALLE EQUIPO ({len(todos)}):')
        lineas.append(f'{"Nombre":<28} {"Meta":>4} {"Goz.":>5} {"Pend":>5} {"Saldo":>6}')
        lineas.append('-' * 52)
        for p in todos[:30]:
            nom = str(p.get('nombre', ''))[:27]
            obj_p = int(_to_float0(p.get('objetivo', 0)))
            goz_p = int(_to_float0(p.get('gozados', 0)))
            pend_p = round(_to_float0(p.get('cantidad_pendiente', 0)), 1)
            saldo_p = round(_to_float0(p.get('total_vacaciones', 0)), 1)
            meta_txt = str(obj_p) if obj_p > 0 else '-'
            pend_txt = str(pend_p) if pend_p > 0 else '-'
            lineas.append(f'{nom:<28} {meta_txt:>4} {goz_p:>5} {pend_txt:>5} {saldo_p:>6}')
        if len(todos) > 30:
            lineas.append(f'... y {len(todos) - 30} mas en el correo.')

    lineas.append('')
    lineas.append('Ver detalle completo y recomendaciones en el correo "Alertas Vacaciones USIL".')

    return '\n'.join(lineas).strip()





def _build_adaptive_card_supervisor(contenido):
    """Adaptive Card bonita para Teams: aviso de correo pendiente al supervisor."""
    nombre = contenido.get('nombre_jefe', '')
    puesto = ''
    try:
        dfm, _ = _cargar_maestro_universo()
        if dfm is not None:
            c_mat = _col(dfm, 'Matricula')
            c_pue = _col(dfm, 'Puesto', 'Cargo', 'Nombre Puesto')
            mat_jefe = _norm_id(contenido.get('matricula_jefe', ''))
            if c_mat and c_pue and mat_jefe:
                fila = dfm[dfm[c_mat] == mat_jefe]
                if not fila.empty:
                    puesto = _safe(fila.iloc[0].get(c_pue, ''))
    except Exception:
        pass

    fecha_limite = contenido.get('fecha_limite', '') or _PA_CONFIG.get('vacaciones_fecha_limite_gestion', '')
    nombre_display = nombre.title() if nombre else 'Jefe/a de Área'
    saludo = f'Hola, {puesto.title()} {nombre_display} 👋' if puesto else f'Hola, {nombre_display} 👋'

    body = [
        {
            'type': 'Container',
            'style': 'emphasis',
            'bleed': True,
            'items': [
                {
                    'type': 'ColumnSet',
                    'columns': [
                        {
                            'type': 'Column',
                            'width': 'auto',
                            'items': [{'type': 'TextBlock', 'text': '🏖️', 'size': 'ExtraLarge'}]
                        },
                        {
                            'type': 'Column',
                            'width': 'stretch',
                            'items': [
                                {'type': 'TextBlock', 'text': 'Gestión de Vacaciones', 'weight': 'Bolder', 'size': 'Large', 'color': 'Accent'},
                                {'type': 'TextBlock', 'text': 'People Analytics · USIL', 'isSubtle': True, 'size': 'Small', 'spacing': 'None'}
                            ]
                        }
                    ]
                }
            ]
        },
        {'type': 'TextBlock', 'text': saludo, 'weight': 'Bolder', 'size': 'Medium', 'wrap': True, 'spacing': 'Medium'},
        {
            'type': 'TextBlock',
            'text': 'Desde la **Subgerencia de Talento y Cultura** te informamos que tienes pendiente la programación de vacaciones de tu equipo para este trimestre.',
            'wrap': True, 'spacing': 'Small'
        },
        {
            'type': 'TextBlock',
            'text': '📧  Te hemos enviado un reporte detallado a tu **correo institucional**. Por favor, revísalo y gestiona las aprobaciones correspondientes.',
            'wrap': True, 'spacing': 'Medium', 'color': 'Accent', 'weight': 'Bolder'
        },
    ]

    if fecha_limite:
        body.append({
            'type': 'Container',
            'style': 'attention',
            'spacing': 'Medium',
            'items': [{
                'type': 'TextBlock',
                'text': f'⏰  Fecha límite de gestión: **{fecha_limite}**',
                'wrap': True, 'weight': 'Bolder'
            }]
        })

    body.append({
        'type': 'TextBlock',
        'text': 'Agradecemos tu pronta atención. ¡Gracias!',
        'wrap': True, 'isSubtle': True, 'spacing': 'Medium'
    })

    return {
        'type': 'message',
        'attachments': [{
            'contentType': 'application/vnd.microsoft.card.adaptive',
            'content': {
                '$schema': 'http://adaptivecards.io/schemas/adaptive-card.json',
                'type': 'AdaptiveCard',
                'version': '1.4',
                'body': body
            }
        }]
    }


def _armar_contenido_supervisor(nombre, email='', dias=30, plantilla_override=None, nombre_objetivo='', hrbp_objetivo=''):

    """Construye asunto y HTML de alerta para un jefe especifico."""

    nombre = _nombre_supervisor_canonico(nombre)

    if not nombre:

        return None, 'Nombre de jefe/area requerido'



    df, err = cargar_datos()

    if err:

        return None, err

    hoy = date.today()

    c_mat = _col(df, 'Matricula')

    c_sup = _col(df, 'Supervisor', 'Jefe', 'Jefe_Directo', 'Jefe Directo', 'Business Partner', 'HRBP')

    persona_lookup = {}

    ret, prox, equipo_full = [], [], []

    nombre_up = _nombre_cmp_key(nombre)

    dfo, _ = cargar_objetivos()

    obj_lookup, df_obj_act, cols_obj = _build_objetivos_lookup(dfo)

    mats_supervisor = set()



    # Toma el universo del jefe desde objetivos (por supervisor o por area inferida).

    equipo_obj_map = _equipo_objetivo_por_supervisor(df_obj_act, cols_obj, obj_lookup)

    for sup_name, equipo in (equipo_obj_map or {}).items():

        if _nombre_cmp_key(_nombre_supervisor_canonico(sup_name)) != nombre_up:

            continue

        for p in (equipo or []):

            mk = _id_key(p.get('matricula', ''))

            if mk:

                mats_supervisor.add(mk)



    for _, rr in df.iterrows():

        mat_key = _id_key(rr.get(c_mat, '')) if c_mat else ''

        if mats_supervisor:

            if not mat_key or mat_key not in mats_supervisor:

                continue

        else:

            sup = _nombre_supervisor_canonico(str(rr.get(c_sup, '')).strip()) if c_sup else ''

            if not sup or _nombre_cmp_key(sup) != nombre_up:

                continue

        pp = _persona(rr, hoy)

        obj = obj_lookup.get(mat_key)

        if obj:

            for k in ('nombre', 'area', 'puesto', 'hrbp', 'supervisor', 'fecha_ingreso', 'objetivo', 'gozados', 'truncas', 'pendientes', 'vencidas', 'total_vacaciones', 'cantidad_pendiente', 'comentario'):

                if obj.get(k) not in (None, ''):

                    pp[k] = obj.get(k)

        else:

            pp['objetivo'] = int(pp.get('objetivo') or 0)

            pp['truncas'] = float(pp.get('truncas') or 0)

            pp['pendientes'] = float(pp.get('pendientes') or 0)

            pp['vencidas'] = float(pp.get('vencidas') or 0)

            pp['total_vacaciones'] = float(pp.get('total_vacaciones') or 0)

            pp['cantidad_pendiente'] = round(float(pp.get('vencidas') or 0) + float(pp.get('pendientes') or 0), 1)

        avp, rcp = _aviso_recomendacion_persona(pp)

        pp['aviso'] = avp

        pp['recomendacion'] = rcp

        mat = mat_key

        if mat:

            persona_lookup[mat] = pp

        equipo_full.append(pp)

        d = pp.get('dias_para_aniversario')

        if d is None:

            continue

        if d < 0:

            ret.append(pp)

        elif d <= int(dias):
            """Resumen breve para Teams que deriva al correo con el detalle completo."""
            prox.append(pp)



    sin_cumplir = []

    if df_obj_act is not None and cols_obj.get('cumplimiento'):

        c_cum = cols_obj.get('cumplimiento')

        df_sin = df_obj_act[df_obj_act[c_cum] < 0].sort_values(c_cum)

        for _, row in df_sin.iterrows():

            mat_key = _id_key(row.get(cols_obj.get('matricula'), '')) if cols_obj.get('matricula') else ''

            item = obj_lookup.get(mat_key)

            if not item:

                continue

            if mats_supervisor:

                if mat_key not in mats_supervisor:

                    continue

            elif _nombre_cmp_key(_nombre_supervisor_canonico(item.get('supervisor', ''))) != nombre_up:

                continue

            item_out = dict(item)

            if mat_key in persona_lookup:

                item_out['correo'] = persona_lookup[mat_key].get('correo', '')

            sin_cumplir.append(item_out)



    adicionales = _seleccionar_candidatos_adicionales(equipo_full, ret, prox, sin_cumplir, limite=8)

    meta_resumen = _resumen_meta_equipo(equipo_full)

    campania = _contexto_campania(hoy)



    cfg = _resolver_supervisor_cfg(nombre)

    if isinstance(plantilla_override, dict):

        if 'asunto' in plantilla_override and plantilla_override['asunto']:

            cfg['asunto'] = _asegurar_asunto_alerta(plantilla_override.get('asunto', ''))

        if 'mensaje' in plantilla_override:

            cfg['mensaje'] = str(plantilla_override.get('mensaje', '') or '').strip()

        if 'aviso' in plantilla_override:

            cfg['aviso'] = str(plantilla_override.get('aviso', '') or '').strip()

        if 'recomendacion' in plantilla_override:

            cfg['recomendacion'] = str(plantilla_override.get('recomendacion', '') or '').strip()

    modo_prueba = bool((plantilla_override or {}).get('_modo_prueba', False)) if isinstance(plantilla_override, dict) else False
    nombre_destinatario = str(nombre_objetivo or nombre or '').strip()
    hrbp_destinatario = _buscar_hrbp_destinatario(
        equipo_full,
        nombre_objetivo=nombre_destinatario,
        fallback=hrbp_objetivo or next((p.get('hrbp', '') for p in equipo_full if p.get('hrbp', '').strip()), '')
    )
    asunto_ctx = {
        'nombre': nombre_destinatario or nombre,
        'jefe': nombre,
        'trimestre': campania.get('trimestre', ''),
        'fecha': hoy.strftime('%d/%m/%Y'),
        'fecha_limite': _PA_CONFIG.get('vacaciones_fecha_limite_gestion', 'la fecha acordada'),
        'hrbp': hrbp_destinatario or _HRBP_TYC_DEFAULT,
        'hrbp_nombre': hrbp_destinatario or _HRBP_TYC_DEFAULT,
    }

    _dias_obj = int(meta_resumen.get('dias_objetivo', 0) or 0)

    if cfg.get('asunto'):

        asunto = _asegurar_asunto_alerta(_tpl(cfg.get('asunto'), asunto_ctx))

    elif _dias_obj > 0:

        asunto = _asegurar_asunto_alerta(

            f'Meta vacaciones {meta_resumen.get("cumplimiento_meta_pct", 0.0)}% '

            f'({meta_resumen.get("dias_gozados_hacia_meta", 0)}/{_dias_obj} dias) '

            f'| Faltan {meta_resumen.get("dias_pendientes_meta", 0)} dias'

        )

    else:

        _saldo = int(meta_resumen.get('dias_saldo_total', 0) or 0)

        asunto = _asegurar_asunto_alerta(

            f'Saldo equipo: {_saldo} dias | Sin meta asignada este trimestre'

        )

    fecha_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # Extraer el HRBP del destinatario objetivo cuando exista; si no, usar el primero válido del equipo.

    # Prioridad: override manual por área > HRBP del dato > default T&C
    _hrbp_nombre = (_resolver_hrbp_area_override(equipo_full, '') or
                    hrbp_destinatario or
                    _HRBP_TYC_DEFAULT)

    _fecha_limite = _PA_CONFIG.get('vacaciones_fecha_limite_gestion', 'la fecha acordada')

    _trimestre_vigente = str(campania.get('trimestre', '') or '').strip() or str(_resumen_fuente_datos().get('trimestre_vigente', '') or '').strip()

    cfg['mensaje'] = _normalizar_texto_alerta_supervisor(
        cfg.get('mensaje', ''),
        nombre_destinatario=nombre_destinatario,
        trimestre=_trimestre_vigente,
        fecha_limite=_fecha_limite,
    )
    cfg['aviso'] = _normalizar_texto_alerta_supervisor(
        cfg.get('aviso', ''),
        nombre_destinatario=nombre_destinatario,
        trimestre=_trimestre_vigente,
        fecha_limite=_fecha_limite,
    )
    cfg['recomendacion'] = _normalizar_texto_alerta_supervisor(
        cfg.get('recomendacion', ''),
        nombre_destinatario=nombre_destinatario,
        trimestre=_trimestre_vigente,
        fecha_limite=_fecha_limite,
    )

    html = _build_html_jefe(

        nombre,

        ret,

        prox,

        sin_cumplir,

        fecha_str,

        cfg.get('mensaje', ''),

        cfg.get('aviso', ''),

        cfg.get('recomendacion', ''),

        adicionales,

        campania,

        meta_resumen,

        hrbp_nombre=_hrbp_nombre,

        fecha_limite=_fecha_limite,

        equipo_full=equipo_full,

        nombre_objetivo=nombre_destinatario,

    )

    email_destino_real = (_norm_email(email) or cfg.get('email', '') or
                         _resolver_email_supervisor_maestro(nombre) or _resolver_default_email())

    # Si hay vacaciones_test_email configurado globalmente, SIEMPRE redirigir (modo prueba global)

    _test_global = _resolver_test_email_destino()

    email_destino_final = _test_global if _test_global else (email_destino_real if not modo_prueba else _resolver_test_email_destino())



    # Mantener siempre el nombre objetivo del jefe para trazabilidad en Teams/webhook.

    nombre_para_mensaje = nombre



    contenido = {

        'nombre_jefe': nombre_para_mensaje,

        'nombre_objetivo': nombre_destinatario or nombre,

        'email_jefe': email_destino_final,

        'email_destino_real': email_destino_final,

        'email_supervisor_real': email_destino_real,

        'modo_prueba': modo_prueba,

        'asunto': asunto,

        'mensaje': cfg.get('mensaje', ''),

        'aviso': cfg.get('aviso', ''),

        'recomendacion': cfg.get('recomendacion', ''),

        'mensaje_html': html,

        'mensaje_teams': _texto_teams_supervisor(nombre_para_mensaje, ret, prox, sin_cumplir, adicionales, campania, meta_resumen, equipo_full=equipo_full, plantilla_msg=cfg.get('mensaje', '')),

        'meta_resumen': meta_resumen,

        'total_colaboradores': len(equipo_full),

        'en_retraso': len(ret),

        'proximos': len(prox),

        'sin_cumplir': len(sin_cumplir),

        'adicionales': len(adicionales),

        'detalle_retraso': ret,

        'detalle_proximos': prox,

        'detalle_sin_cumplir': sin_cumplir,

        'detalle_adicionales': adicionales,

        'detalle_equipo_full': equipo_full,

        'campania': campania,

        'mensaje_cfg': cfg.get('mensaje', ''),

        'aviso_cfg': cfg.get('aviso', ''),

        'recomendacion_cfg': cfg.get('recomendacion', ''),

        'hrbp_nombre': _hrbp_nombre,

        'fecha_limite_cfg': _fecha_limite,

        'fecha_generacion': fecha_str,

    }

    contenido['mensaje_webhook'] = json.dumps(_build_webhook_payload(contenido), ensure_ascii=False, indent=2)
    contenido['teams_adaptive_card'] = json.dumps(_build_adaptive_card_supervisor(contenido), ensure_ascii=False)

    return contenido, None





# =========================================================================================

# SISTEMA DE COLABORADORES - Estado individual + Plantillas de mensajes

# =========================================================================================

COLAB_PLANTILLAS_FILE = os.path.join(DATAS_DIR, 'colaboradores_plantillas.json')



def _cargar_plantillas_colaboradores():

    """Carga plantillas personalizadas de mensajes para colaboradores."""

    if not os.path.isfile(COLAB_PLANTILLAS_FILE):

        return {}

    try:

        with open(COLAB_PLANTILLAS_FILE, 'r', encoding='utf-8') as f:

            return json.load(f)

    except:

        return {}



def _guardar_plantillas_colaboradores(data):

    """Guarda plantillas personalizadas."""

    try:

        os.makedirs(DATAS_DIR, exist_ok=True)

        with open(COLAB_PLANTILLAS_FILE, 'w', encoding='utf-8') as f:

            json.dump(data, f, indent=2, ensure_ascii=False)

        return True

    except:

        return False



def _obtener_estado_colaborador(query):

    """Obtiene estado detallado de un colaborador. Acepta matrícula exacta o nombre parcial."""

    df, err = cargar_datos()

    if err:

        return None, err



    dfo, _ = cargar_objetivos()

    obj_lookup, _, _ = _build_objetivos_lookup(dfo)



    c_mat = _col(df, 'Matricula')

    if not c_mat:

        return None, 'Columna Matricula no encontrada'



    c_nom = _col(df, 'Apellidos y Nombres', 'Nombre Completo', 'Nombre', 'Colaborador')



    hoy = date.today()

    mat_key = _id_key(query)

    query_norm = _norm(query).lower()



    matched_row = None



    # Intento 1: matrícula exacta

    for _, row in df.iterrows():

        if _id_key(row.get(c_mat, '')) == mat_key:

            matched_row = row

            break



    # Intento 2: búsqueda por nombre si no se encontró matrícula

    if matched_row is None and c_nom and len(query_norm) >= 3:

        tokens = [t for t in query_norm.split() if len(t) >= 2]

        if tokens:

            for _, row in df.iterrows():

                nombre_norm = _norm(row.get(c_nom, '')).lower()

                if all(t in nombre_norm for t in tokens):

                    matched_row = row

                    break



    if matched_row is None:

        # Fallback: colaborador manual guardado en overrides.

        manuales = _overrides_a_items_busqueda()

        if manuales:

            tokens = [t for t in query_norm.split() if len(t) >= 2]

            hit = None

            for it in manuales:

                mk = _id_key(it.get('matricula', ''))

                nm = _norm(it.get('nombre', '')).lower()

                if mat_key and mk == mat_key:

                    hit = it

                    break

                if tokens and all(t in nm for t in tokens):

                    hit = it

                    break

            if hit:

                persona = {

                    'matricula': hit.get('matricula', ''),

                    'nombre': hit.get('nombre', ''),

                    'correo': hit.get('email', ''),

                    'email': hit.get('email', ''),

                    'departamento': hit.get('departamento', ''),

                    'division': hit.get('departamento', ''),

                    'area': hit.get('area', ''),

                    'puesto': hit.get('puesto', ''),

                    'supervisor': hit.get('supervisor', ''),

                    'fecha_ingreso': '',

                    'fecha_aniversario': '',

                    'dias_para_aniversario': None,

                    'estado': 'Seguimiento manual',

                    'en_retraso': False,

                    'objetivo': 0,

                    'gozados': 0.0,

                    'truncas': 0.0,

                    'pendientes': 0.0,

                    'vencidas': 0.0,

                    'total_vacaciones': 0.0,

                    'cantidad_pendiente': 0.0,

                }

                avp, rcp = _aviso_recomendacion_persona(persona)

                persona['hrbp'] = _resolver_hrbp_tyc(

                    '',

                    area=persona.get('area', ''),

                    departamento=persona.get('departamento', ''),

                    division=persona.get('division', ''),

                    puesto=persona.get('puesto', ''),

                )

                persona['aviso'] = avp

                persona['recomendacion'] = rcp

                return persona, None



        return None, 'Colaborador no encontrado'



    row = matched_row

    persona = _persona(row, hoy)

    obj = obj_lookup.get(_id_key(row.get(c_mat, '')))

    if obj:

        for k in ('nombre', 'area', 'puesto', 'hrbp', 'supervisor', 'fecha_ingreso', 'objetivo', 'gozados', 'truncas', 'pendientes', 'vencidas', 'total_vacaciones', 'cantidad_pendiente'):

            if obj.get(k) not in (None, ''):

                persona[k] = obj.get(k)

    else:

        persona['objetivo'] = int(persona.get('objetivo') or 0)

        persona['truncas'] = float(persona.get('truncas') or 0)

        persona['pendientes'] = float(persona.get('pendientes') or 0)

        persona['vencidas'] = float(persona.get('vencidas') or 0)

        persona['total_vacaciones'] = float(persona.get('total_vacaciones') or 0)

        persona['cantidad_pendiente'] = round(float(persona.get('vencidas') or 0) + float(persona.get('pendientes') or 0), 1)



    avp, rcp = _aviso_recomendacion_persona(persona)

    persona['hrbp'] = _resolver_hrbp_tyc(

        persona.get('hrbp', ''),

        area=persona.get('area', ''),

        departamento=persona.get('departamento', ''),

        division=persona.get('division', ''),

        puesto=persona.get('puesto', ''),

    )

    persona['aviso'] = avp

    persona['recomendacion'] = rcp

    return persona, None



def _obtener_plantilla_colaborador(matricula, nombre='', bp=''):

    """Obtiene plantilla personalizada de mensaje para un colaborador."""

    plantillas = _cargar_plantillas_colaboradores()

    key = _id_key(matricula)

    

    if key in plantillas:

        data = plantillas[key] if isinstance(plantillas[key], dict) else {}

        cuerpo = data.get('cuerpo', data.get('mensaje', ''))

        return {

            'matricula': matricula,

            'nombre': data.get('nombre', nombre),

            'bp': data.get('bp', bp),

            'asunto': str(data.get('asunto', f'Estado de Vacaciones - {nombre}') or '').strip(),

            'cuerpo': str(cuerpo or '').strip(),

            'editable': bool(data.get('editable', True))

        }

    

    # Plantilla por defecto

    return {

        'matricula': matricula,

        'nombre': nombre,

        'bp': bp,

        'asunto': f'Estado de Vacaciones - {nombre}',

        'cuerpo': f'Estimado/a {nombre},\n\nPor este medio le comunicamos el estado actualizado de sus vacaciones.\n\nCualquier coordinación sobre fechas o requerimientos, favor dirigirse a su Business Partner: {bp}.\n\nSaludos,\nPeople Analytics USIL',

        'editable': True

    }



def _guardar_plantilla_colaborador(matricula, plantilla):

    """Guarda plantilla personalizada para un colaborador."""

    plantillas = _cargar_plantillas_colaboradores()

    key = _id_key(matricula)

    data = plantilla if isinstance(plantilla, dict) else {}

    plantillas[key] = {

        'matricula': matricula,

        'nombre': str(data.get('nombre', '') or '').strip(),

        'bp': str(data.get('bp', '') or '').strip(),

        'asunto': str(data.get('asunto', '') or '').strip(),

        'cuerpo': str(data.get('cuerpo', data.get('mensaje', '')) or '').strip(),

        'editable': bool(data.get('editable', True))

    }

    return _guardar_plantillas_colaboradores(plantillas)


def _periodo_campania_texto(campania):

    campania = campania or {}

    return str(campania.get('trimestre', '') or '').strip()


def _asegurar_asunto_alerta(asunto):

    base = str(asunto or '').strip()

    if not base:

        return 'Alertas Vacaciones USIL'

    if 'ALERTAS VACACIONES USIL' in base.upper():

        return base

    return f'Alertas Vacaciones USIL | {base}'



@app.route('/api/colaborador/<matricula>', methods=['GET'])

def api_colaborador_detalle(matricula):

    """Obtiene estado completo de un colaborador."""

    persona, err = _obtener_estado_colaborador(matricula)

    if err:

        return jsonify({'ok': False, 'error': err}), 400

    

    plantilla = _obtener_plantilla_colaborador(

        matricula,

        persona.get('nombre', '-'),

        persona.get('hrbp', '-')

    )

    

    campania = _contexto_campania()

    

    return jsonify({

        'ok': True,

        'persona': persona,

        'plantilla': plantilla,

        'campania': campania,

        'fecha_actual': datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    })



@app.route('/api/colaborador/<matricula>/plantilla', methods=['GET', 'POST'])

def api_colaborador_plantilla(matricula):

    """GET: obtiene plantilla | POST: guarda plantilla editada."""

    if request.method == 'GET':

        persona, err = _obtener_estado_colaborador(matricula)

        if err:

            return jsonify({'ok': False, 'error': err}), 400

        

        plantilla = _obtener_plantilla_colaborador(

            matricula,

            persona.get('nombre', '-'),

            persona.get('hrbp', '-')

        )

        return jsonify({'ok': True, 'plantilla': plantilla})

    

    else:  # POST

        payload = request.get_json(silent=True) or {}

        plantilla = payload.get('plantilla', {})

        

        if not _guardar_plantilla_colaborador(matricula, plantilla):

            return jsonify({'ok': False, 'error': 'Error al guardar plantilla'}), 500

        

        return jsonify({'ok': True, 'mensaje': 'Plantilla guardada correctamente'})





@app.route('/api/preview-supervisor', methods=['POST'])

def api_preview_supervisor():

    payload = request.get_json(silent=True) or {}

    nombre = str(payload.get('nombre', '')).strip()

    nombre_objetivo = str(payload.get('nombre_objetivo', '') or '').strip()

    hrbp_objetivo = str(payload.get('hrbp_objetivo', '') or '').strip()

    email = _norm_email(payload.get('email', ''))

    email_objetivo = _norm_email(payload.get('email_objetivo', ''))

    overrides = {

        'asunto': payload.get('asunto', '').strip(),

        'mensaje': payload.get('mensaje', ''),

        'aviso': payload.get('aviso', ''),

        'recomendacion': payload.get('recomendacion', ''),

        '_modo_prueba': _cfg_bool(payload.get('modo_prueba', False), False),

    }

    contenido, err = _armar_contenido_supervisor(nombre, email, plantilla_override=overrides, nombre_objetivo=nombre_objetivo, hrbp_objetivo=hrbp_objetivo)

    if err:

        return jsonify({'ok': False, 'error': err}), 400

    return jsonify({'ok': True, **contenido})





@app.route('/api/enviar-a-colaborador', methods=['POST'])

def api_enviar_a_colaborador():

    """Envía mensaje personalizado a un colaborador via SMTP y/o notifica al canal Teams."""

    payload = request.get_json(silent=True) or {}

    query = str(payload.get('matricula', '') or payload.get('query', '')).strip()



    if not query:

        return jsonify({'ok': False, 'error': 'Matrícula o nombre requerido'}), 400



    persona, err = _obtener_estado_colaborador(query)

    if err or not persona:

        return jsonify({'ok': False, 'error': err or 'Colaborador no encontrado'}), 400



    matricula = persona.get('matricula', query)

    nombre = persona.get('nombre', '-')

    bp_nombre = persona.get('hrbp', '-')



    plantilla = _obtener_plantilla_colaborador(matricula, nombre, bp_nombre)



    if 'asunto' in payload:

        plantilla['asunto'] = payload['asunto']

    if 'cuerpo' in payload:

        plantilla['cuerpo'] = payload['cuerpo']

    if 'mensaje' in payload and 'cuerpo' not in payload:

        plantilla['cuerpo'] = payload['mensaje']

    plantilla['nombre'] = nombre

    plantilla['bp'] = bp_nombre

    _guardar_plantilla_colaborador(matricula, plantilla)



    modo_prueba = _cfg_bool(payload.get('modo_prueba', False), False)

    # Si hay vacaciones_test_email configurado globalmente, forzar modo prueba

    _test_global = _resolver_test_email_destino()

    if _test_global:

        modo_prueba = True



    enviar_teams = _cfg_bool(payload.get('enviar_teams', True), True)

    enviar_email = _cfg_bool(payload.get('enviar_email', False), False)



    email_destino = _norm_email(persona.get('correo', ''))

    if modo_prueba:

        email_destino = _test_global or _resolver_test_email_destino()



    asunto = str(plantilla.get('asunto', f'Estado de Vacaciones - {nombre}') or '').strip()

    cuerpo_txt = str(plantilla.get('cuerpo', plantilla.get('mensaje', '')) or '')

    cuerpo_html = '<html><body style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;color:#1b2a3d">' + \
        cuerpo_txt.replace('\n', '<br>') + \
        f'<hr style="margin-top:20px"><p style="font-size:11px;color:#888">People Analytics USIL &bull; {datetime.now().strftime("%d/%m/%Y %H:%M")}</p></body></html>'



    smtp_ok, smtp_err = False, None

    teams_ok, teams_err = False, None



    # --- Envío SMTP al colaborador ---

    if enviar_email:

        if email_destino and '@' in email_destino:

            smtp_ok, smtp_err = _enviar_correo_smtp(email_destino, nombre, asunto, cuerpo_html)

        else:

            smtp_err = f'Email no disponible para {nombre}'



    # --- Notificación al canal Teams (webhook personal/PA) ---

    if enviar_teams:

        webhook = TEAMS_WEBHOOK_PERSONAL_URL or TEAMS_WEBHOOK_URL

        if webhook:

            try:

                import urllib.request as _ureq, json as _json

                pendientes = float(persona.get('pendientes', 0) or 0)

                vencidas = float(persona.get('vencidas', 0) or 0)

                card = {

                    'type': 'message',

                    'attachments': [{

                        'contentType': 'application/vnd.microsoft.card.adaptive',

                        'content': {

                            '$schema': 'http://adaptivecards.io/schemas/adaptive-card.json',

                            'type': 'AdaptiveCard',

                            'version': '1.4',

                            'body': [

                                {'type': 'TextBlock', 'size': 'Large', 'weight': 'Bolder',

                                 'text': f'📋 Mensaje enviado a colaborador', 'color': 'Accent'},

                                {'type': 'TextBlock', 'text': nombre, 'size': 'Medium', 'weight': 'Bolder'},

                                {'type': 'FactSet', 'facts': [

                                    {'title': 'Matrícula:', 'value': matricula},

                                    {'title': 'BP:', 'value': bp_nombre},

                                    {'title': 'Pendientes:', 'value': f'{pendientes:.1f} días'},

                                    {'title': 'Vencidas:', 'value': f'{vencidas:.1f} días'},

                                    {'title': 'Modo:', 'value': 'Prueba' if modo_prueba else 'Real'},

                                    {'title': 'Enviado:', 'value': datetime.now().strftime('%d/%m/%Y %H:%M')},

                                ]},

                                {'type': 'TextBlock', 'text': f'Asunto: {asunto}',

                                 'size': 'Small', 'isSubtle': True, 'wrap': True},

                            ]

                        }

                    }]

                }

                data = _json.dumps(card).encode('utf-8')

                req = _ureq.Request(webhook, data=data, headers={'Content-Type': 'application/json'})

                with _ureq.urlopen(req, timeout=12) as resp:

                    if 200 <= getattr(resp, 'status', 200) < 300:

                        teams_ok = True

                    else:

                        teams_err = f'Webhook respondió {resp.status}'

            except Exception as ex:

                teams_err = str(ex)

        else:

            teams_err = 'Webhook Teams no configurado en pa_config.json'



    canales_txt = []

    if enviar_email:

        canales_txt.append(f'Email: {"OK" if smtp_ok else f"ERROR ({smtp_err})"}')

    if enviar_teams:

        canales_txt.append(f'Teams: {"OK" if teams_ok else f"ERROR ({teams_err})"}')



    return jsonify({

        'ok': True,

        'matricula': matricula,

        'nombre': nombre,

        'mensaje': f'Procesado para {nombre}. {" | ".join(canales_txt) if canales_txt else "Sin canales activos."}',

        'smtp_ok': smtp_ok,

        'smtp_error': smtp_err,

        'teams_ok': teams_ok,

        'teams_error': teams_err,

        'modo_prueba': modo_prueba,

        'contenido': {

            'bp_nombre': bp_nombre,

            'email_destino': email_destino,

        }

    })





@app.route('/api/enviar-a-supervisor', methods=['POST'])

def api_enviar_a_supervisor():

    """Envio individual: preconstruye mensaje real, guarda cola y envia SMTP."""

    payload = request.get_json(silent=True) or {}

    nombre = str(payload.get('nombre', '')).strip()

    nombre_objetivo = str(payload.get('nombre_objetivo', '') or '').strip()

    hrbp_objetivo = str(payload.get('hrbp_objetivo', '') or '').strip()

    email = _norm_email(payload.get('email', ''))

    email_objetivo = _norm_email(payload.get('email_objetivo', ''))

    overrides = {

        'asunto': payload.get('asunto', '').strip(),

        'mensaje': payload.get('mensaje', ''),

        'aviso': payload.get('aviso', ''),

        'recomendacion': payload.get('recomendacion', ''),

        '_modo_prueba': _cfg_bool(payload.get('modo_prueba', False), False),

    }

    enviar_smtp = _cfg_bool(payload.get('enviar_smtp', False), False)

    enviar_teams = _cfg_bool(payload.get('enviar_teams', True), True)

    encolar_pa = _cfg_bool(payload.get('encolar_para_pa', _ENCOLAR_ALERTAS_PA), _ENCOLAR_ALERTAS_PA)

    if enviar_teams:

        encolar_pa = True

    if not enviar_smtp and not encolar_pa and not enviar_teams:

        return jsonify({'ok': False, 'error': 'Selecciona al menos un canal de envio'}), 400



    contenido, err = _armar_contenido_supervisor(nombre, email, plantilla_override=overrides, nombre_objetivo=nombre_objetivo, hrbp_objetivo=hrbp_objetivo)

    if err:

        return jsonify({'ok': False, 'error': err}), 400

    if nombre_objetivo:

        contenido['nombre_objetivo'] = nombre_objetivo

    if email_objetivo:

        contenido['email_objetivo'] = email_objetivo

    if hrbp_objetivo:

        contenido['hrbp_nombre'] = hrbp_objetivo

    contenido = _registrar_confirmaciones_contenido(contenido, request.host_url)



    try:

        fname = None

        pa_meta = None

        if encolar_pa:

            fname, _, pa_meta = _guardar_json_cola([contenido], prefijo='alerta_supervisor')



        smtp_ok, smtp_err = False, None

        if enviar_smtp:

            smtp_ok, smtp_err = _enviar_correo_smtp(

                contenido.get('email_jefe', ''),

                contenido.get('nombre_jefe', ''),

                contenido.get('asunto', 'Alerta Vacaciones USIL'),

                contenido.get('mensaje_html', ''),

            )



        teams_ok, teams_err = False, None

        if enviar_teams:

            teams_ok, teams_err = _enviar_teams_webhook_supervisor(contenido)



        canales = []

        if enviar_teams:

            canales.append('Teams/Webhook')

        if enviar_smtp:

            canales.append('SMTP')

        if encolar_pa:

            canales.append('PA')

        canales_txt = ' + '.join(canales) if canales else '-'



        teams_estado = (

            'No enviado' if not enviar_teams else ('OK' if teams_ok else 'Error en envio')

        )

        smtp_estado = (

            'No enviado' if not enviar_smtp else ('OK' if smtp_ok else 'Error en envio')

        )

        cola_estado = (

            'No encolado' if not encolar_pa else ('OK' if fname else 'Error encolando')

        )



        try:

            threading.Thread(

                target=lambda: _notificar_teams_personal(

                    titulo='Envio manual por jefe ejecutado',

                    detalle_items=[

                        {'titulo': 'Jefe objetivo:', 'valor': f"{contenido.get('nombre_objetivo', contenido.get('nombre_jefe', '-'))} <{contenido.get('email_objetivo', contenido.get('email_destino_real', contenido.get('email_jefe', '-')))}>"},

                        {'titulo': 'Entrega:', 'valor': (f"{contenido.get('email_jefe', '-')} (modo prueba)" if bool(contenido.get('modo_prueba', False)) else contenido.get('email_jefe', '-'))},

                        {'titulo': 'Canales:', 'valor': canales_txt},

                        {'titulo': 'Teams:', 'valor': teams_estado},

                        {'titulo': 'SMTP:', 'valor': smtp_estado},

                        {'titulo': 'Cola PA:', 'valor': cola_estado},

                        {'titulo': 'Archivo cola:', 'valor': fname or '-'},

                    ],

                    color='Good' if (smtp_ok or teams_ok or bool(fname)) else 'Warning'

                ),

                daemon=True

            ).start()

        except Exception:

            pass



        return jsonify({

            'ok': smtp_ok or teams_ok or bool(fname),

            'mensaje': f'Envio preparado para {contenido.get("email_jefe", "")}',

            'archivo': fname,

            'ruta': COLA_DIR,

            'retraso_seguridad_segundos': int(_PA_DELAY_SECONDS),

            'teams_enviado': teams_ok,

            'teams_error': teams_err,

            'smtp_enviado': smtp_ok,

            'smtp_error': smtp_err,

            'error': (teams_err or smtp_err) if (not teams_ok and not smtp_ok and not fname) else None,

            'json_guardado': bool(fname),

            'json_meta': pa_meta,

            'encolar_pa': encolar_pa,

            'enviar_teams': enviar_teams,

            'enviar_smtp': enviar_smtp,

            'canales': {

                'teams_webhook': enviar_teams,

                'smtp_correo': enviar_smtp,

                'flujo_power_automate': encolar_pa,

            },

            'en_retraso': contenido.get('en_retraso', 0),

            'proximos': contenido.get('proximos', 0),

            'sin_cumplir': contenido.get('sin_cumplir', 0),

            'total_colaboradores': contenido.get('total_colaboradores', 0),

            'modo_prueba': bool(contenido.get('modo_prueba', False)),

            'metodo': ('teams+smtp+json_cola' if teams_ok and smtp_ok and fname else

                       'teams+smtp' if teams_ok and smtp_ok else

                       'teams+json_cola' if teams_ok and fname else

                       'teams_directo' if teams_ok else

                       'smtp+json_cola' if smtp_ok and fname else

                       'smtp_directo' if smtp_ok else

                       'json_cola_pa' if fname else 'fallido')

        })

    except Exception as e:

        return jsonify({'ok': False, 'error': f'No se pudo generar cola/envio: {e}', 'ruta': COLA_DIR}), 500







# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# ALERTAS ONEDRIVE (Power Automate)

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _escribir_alertas_onedrive():

    df, err = cargar_datos()

    if err: print(f'[OD] {err}'); return

    if '_dias_aniv' not in df.columns: return

    hoy = date.today()

    hitos = {90, 30, 7}

    alertas = [_persona(r, hoy) for _, r in df.iterrows()

               if pd.notna(r.get('_dias_aniv')) and int(r['_dias_aniv']) in hitos]

    payload = {'fecha':hoy.isoformat(),'generado_en':datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),

               'total_alertas':len(alertas),'alertas':alertas}

    try:

        with open(_ONEDRIVE_JSON,'w',encoding='utf-8') as f:

            json.dump(payload, f, ensure_ascii=False, indent=2)

        print(f'[OD] alertas_pa_onedrive.json: {len(alertas)} alertas')

    except Exception as e:

        print(f'[OD] Error: {e}')



@app.route('/api/forzar-alertas-onedrive', methods=['POST'])

def api_forzar():

    threading.Thread(target=_escribir_alertas_onedrive, daemon=True).start()

    return jsonify({'ok':True,'mensaje':'Generando alertas_pa_onedrive.json...'})



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# LOG

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _registrar_log(canal, metodo, resultados):

    entrada = {'fecha':datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),'canal':canal,'metodo':metodo,

               'envios':[{'nombre':r.get('nombre',''),'email':r.get('email',''),'ok':r.get('ok',False),

                          'error':r.get('error',''),'retraso':r.get('retraso',0),'proximos':r.get('proximos',0)}

                         for r in resultados]}

    with _log_lock:

        try:

            hist = []

            if os.path.exists(_LOG_ENVIOS_PATH):

                with open(_LOG_ENVIOS_PATH,'r',encoding='utf-8') as f: hist = json.load(f)

            hist.insert(0, entrada); hist = hist[:200]

            with open(_LOG_ENVIOS_PATH,'w',encoding='utf-8') as f:

                json.dump(hist, f, ensure_ascii=False, indent=2)

        except Exception as e: print(f'[LOG] {e}')


@app.route('/confirmar-vacaciones', methods=['GET'])

def confirmar_vacaciones():

    token = str(request.args.get('token', '') or '').strip()

    if not token:

        return _render_confirmacion_html('Confirmacion invalida', 'El enlace no contiene un token valido.', '#991b1b'), 400, {'Content-Type': 'text/html; charset=utf-8'}

    registros = _leer_confirmaciones_vacaciones()
    if not registros:

        return _render_confirmacion_html('Confirmacion no encontrada', 'No se encontro el registro asociado al enlace.', '#991b1b'), 404, {'Content-Type': 'text/html; charset=utf-8'}

    encontrado = None
    for registro in registros:
        if str(registro.get('token', '') or '').strip() == token:
            encontrado = registro
            break

    if not encontrado:

        return _render_confirmacion_html('Confirmacion no encontrada', 'El enlace ya no existe o fue generado para otro entorno.', '#991b1b'), 404, {'Content-Type': 'text/html; charset=utf-8'}

    nombre = encontrado.get('colaborador_nombre', 'el colaborador')
    jefe = encontrado.get('supervisor_nombre', 'la jefatura')
    if encontrado.get('status') == 'confirmado':

        mensaje = f'La confirmacion para {nombre} ya estaba registrada anteriormente por {jefe}.'
        return _render_confirmacion_html('Confirmacion ya registrada', mensaje, '#1d4ed8'), 200, {'Content-Type': 'text/html; charset=utf-8'}

    encontrado['status'] = 'confirmado'
    encontrado['confirmed_at'] = datetime.now().isoformat()
    encontrado['confirmed_ip'] = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    encontrado['confirmed_user_agent'] = request.headers.get('User-Agent', '')
    _guardar_confirmaciones_vacaciones(registros)

    mensaje = f'Se registro la salida de vacaciones de {nombre}. Este dato ya puede contarse en los KPIs de seguimiento de {jefe}.'
    return _render_confirmacion_html('Confirmacion registrada', mensaje), 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/api/confirmaciones-vacaciones/resumen', methods=['GET'])

def api_confirmaciones_vacaciones_resumen():

    registros = _leer_confirmaciones_vacaciones()
    incluir_prueba = _cfg_bool(request.args.get('incluir_prueba', False), False)
    if not incluir_prueba:
        registros = [r for r in registros if not r.get('modo_prueba')]

    total = len(registros)
    confirmadas = sum(1 for r in registros if r.get('status') == 'confirmado')
    pendientes = max(total - confirmadas, 0)
    por_jefe = {}
    for r in registros:
        jefe = str(r.get('supervisor_nombre', '') or 'Sin jefatura').strip()
        bucket = por_jefe.setdefault(jefe, {'total': 0, 'confirmadas': 0})
        bucket['total'] += 1
        if r.get('status') == 'confirmado':
            bucket['confirmadas'] += 1

    ranking = [
        {
            'jefe': jefe,
            'total': vals['total'],
            'confirmadas': vals['confirmadas'],
            'pendientes': max(vals['total'] - vals['confirmadas'], 0),
        }
        for jefe, vals in por_jefe.items()
    ]
    ranking.sort(key=lambda x: (-x['confirmadas'], -x['total'], x['jefe']))

    return jsonify({
        'ok': True,
        'total': total,
        'confirmadas': confirmadas,
        'pendientes': pendientes,
        'avance_pct': round((confirmadas / total) * 100, 1) if total else 0.0,
        'por_jefe': ranking[:100],
    })



@app.route('/api/log-envios', methods=['GET'])

def api_log():

    try:

        if os.path.exists(_LOG_ENVIOS_PATH):

            with open(_LOG_ENVIOS_PATH,'r',encoding='utf-8') as f:

                return jsonify({'ok':True,'historial':json.load(f)})

        return jsonify({'ok':True,'historial':[]})

    except Exception as e:

        return jsonify({'ok':False,'error':str(e)})





@app.route('/api/cola-pa/ultima-pendiente', methods=['GET'])

def api_cola_pa_ultima_pendiente():

    try:

        os.makedirs(COLA_PENDIENTES_DIR, exist_ok=True)

        files = [f for f in os.listdir(COLA_PENDIENTES_DIR) if f.lower().endswith('.json')]

        if not files:

            return jsonify({'ok': True, 'pendiente': None})

        files.sort(key=lambda n: os.path.getmtime(os.path.join(COLA_PENDIENTES_DIR, n)), reverse=True)

        nombre = files[0]

        liberar_en = _resolver_liberar_por_archivo(nombre)

        segundos = 0

        if liberar_en:

            segundos = max(0, int((liberar_en - datetime.now()).total_seconds()))

        return jsonify({'ok': True, 'pendiente': {

            'archivo': nombre,

            'liberar_en': liberar_en.isoformat() if liberar_en else None,

            'segundos_restantes': segundos,

            'retraso_segundos': int(_PA_DELAY_SECONDS),

        }})

    except Exception as e:

        return jsonify({'ok': False, 'error': str(e)}), 500





@app.route('/api/cola-pa/cancelar', methods=['POST'])

def api_cola_pa_cancelar():

    payload = request.get_json(silent=True) or {}

    archivo = str(payload.get('archivo', '') or '').strip()

    motivo = str(payload.get('motivo', 'cancelado por usuario') or 'cancelado por usuario').strip()

    ok, err, info = _cancelar_pendiente_pa(archivo, motivo=motivo)

    if not ok:

        return jsonify({'ok': False, 'error': err or 'No se pudo cancelar'}), 400

    return jsonify({'ok': True, 'cancelado': info, 'mensaje': f'Envio cancelado: {archivo}'})





@app.route('/api/cola-pa/cancelar-todos', methods=['POST'])

def api_cola_pa_cancelar_todos():

    payload = request.get_json(silent=True) or {}

    motivo = str(payload.get('motivo', 'cancelado global por usuario') or 'cancelado global por usuario').strip()

    try:

        os.makedirs(COLA_PENDIENTES_DIR, exist_ok=True)

        files = [f for f in os.listdir(COLA_PENDIENTES_DIR) if f.lower().endswith('.json')]

        if not files:

            return jsonify({'ok': True, 'cancelados': [], 'total_cancelados': 0, 'mensaje': 'No habia pendientes'})



        cancelados = []

        errores = []

        for nombre in files:

            ok, err, info = _cancelar_pendiente_pa(nombre, motivo=motivo)

            if ok and info:

                cancelados.append(info)

            else:

                errores.append({'archivo': nombre, 'error': err or 'No se pudo cancelar'})



        return jsonify({

            'ok': len(errores) == 0,

            'cancelados': cancelados,

            'errores': errores,

            'total_cancelados': len(cancelados),

            'mensaje': f'Se cancelaron {len(cancelados)} pendiente(s)'

        })

    except Exception as e:

        return jsonify({'ok': False, 'error': str(e)}), 500



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# SCHEDULER

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _leer_ultimo_dia():

    try:

        if os.path.exists(_NOTIF_STATE_FILE):

            with open(_NOTIF_STATE_FILE,'r',encoding='utf-8') as f:

                return date.fromisoformat(json.load(f).get('ultimo_dia_enviado',''))

    except: pass

    return None



def _guardar_ultimo_dia(d):

    try:

        with open(_NOTIF_STATE_FILE,'w',encoding='utf-8') as f:

            json.dump({'ultimo_dia_enviado':d.isoformat()}, f)

    except: pass



@app.route('/api/scheduler/estado', methods=['GET'])

def api_sched_est():

    return jsonify({'ok': True, 'activo': False, 'mensaje': 'Envio automatico deshabilitado'})



@app.route('/api/scheduler/toggle', methods=['POST'])

def api_sched_tog():

    return jsonify({'ok': False, 'error': 'Envio automatico deshabilitado'}), 410



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# SUBIR EXCEL

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _clasificar_excel_subido(nombre_archivo):

    nombre = os.path.basename(str(nombre_archivo or '').strip())

    norm = _norm(nombre).upper()

    if 'PERSONALMAESTROREPORTE' in norm or 'MAESTRO' in norm:

        return {'tipo': 'maestra', 'trimestre': None}



    tri = _extraer_trimestre_de_nombre(nombre)

    if not tri:

        tri = _normalizar_trimestre_txt(str(_PA_CONFIG.get('vacaciones_trimestre_actual', '') or '').strip())

    if not tri:

        tri = _trimestre_desde_fecha()

    return {'tipo': 'objetivos', 'trimestre': tri}



def _nombre_activo_excel(info_carga):

    ahora = datetime.now()

    if info_carga.get('tipo') == 'maestra':

        return f"PersonalMaestroReporte_{ahora.strftime('%m_%d_%Y %H_%M_%S')}.xlsx"



    tri = info_carga.get('trimestre') or {}

    tri_label = str(tri.get('label', 'Q?-????')).strip() or 'Q?-????'

    return f"Copia de Reporte Vacaciones Objetivo {tri_label} {ahora.strftime('%Y%m%d_%H%M%S')}.xlsx"



def _archivar_excel_en_carpeta_compartida(ruta_origen, nombre_original, info_carga):

    ahora = datetime.now()

    if info_carga.get('tipo') == 'maestra':

        destino_dir = os.path.join(COLA_FUENTES_MAESTRA_DIR, str(ahora.year))

    else:

        tri = info_carga.get('trimestre') or {}

        anio = str(tri.get('anio', ahora.year))

        q = f"Q{tri.get('q', 0) or '?'}"

        destino_dir = os.path.join(COLA_FUENTES_OBJETIVOS_DIR, anio, q)



    os.makedirs(destino_dir, exist_ok=True)

    base = os.path.basename(str(nombre_original or '')).strip() or os.path.basename(ruta_origen)

    destino = os.path.join(destino_dir, f"{ahora.strftime('%Y%m%d_%H%M%S')}_{base}")

    shutil.copy2(ruta_origen, destino)

    return destino

@app.route('/api/subir_excel', methods=['POST'])

def api_subir():

    archivo = request.files.get('excel') or request.files.get('file')

    if not archivo or not archivo.filename:

        return jsonify({'ok': False, 'error': 'Adjunta un archivo Excel (.xlsx)'}), 400



    nombre = os.path.basename(str(archivo.filename or '').strip())

    if not nombre.lower().endswith('.xlsx'):

        return jsonify({'ok': False, 'error': 'Solo se admite formato .xlsx'}), 400



    info_carga = _clasificar_excel_subido(nombre)

    nombre_activo = _nombre_activo_excel(info_carga)

    destino = os.path.join(DATAS_DIR, nombre_activo)

    historial_path = ''

    try:

        archivo.save(destino)

        historial_path = _archivar_excel_en_carpeta_compartida(destino, nombre, info_carga)

        # Invalidar cache para que el siguiente load tome la fuente mas reciente

        _df_cache.update({'df': None, 'mtime': None, 'ruta': None, 'fecha': None})

        _MAESTRO_CONTACTOS_CACHE.update({'ruta': None, 'mtime': None, 'df': None})

        globals()['_CANDIDATOS_CACHE_DATA'] = None

        globals()['_CANDIDATOS_CACHE_TIME'] = 0

    except Exception as e:

        return jsonify({'ok': False, 'error': f'No se pudo guardar archivo: {e}'}), 500



    return jsonify({

        'ok': True,

        'archivo': nombre,

        'archivo_activo': nombre_activo,

        'tipo_carga': info_carga.get('tipo'),

        'trimestre': (info_carga.get('trimestre') or {}).get('label', ''),

        'carpeta': DATAS_DIR,

        'historial_path': historial_path,

        'fuente_datos': _resumen_fuente_datos(),

    })



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ═══════════════════════════════════════════════════════════════════════════════
# STARTUP
# ═══════════════════════════════════════════════════════════════════════════════

_PRELOAD_DONE = False
_PRELOAD_ELAPSED = 0

@app.route('/api/status/ready', methods=['GET'])
def api_status_ready():
    """Indica si la precarga de datos ya terminó."""
    return jsonify({
        'ok': True,
        'ready': _PRELOAD_DONE,
        'elapsed': round(_PRELOAD_ELAPSED, 1),
    })

def _pre_cargar_todo_sincrono():
    """Carga SINCRONA de caches antes de aceptar requests HTTP.
    Esto evita que el frontend haga timeout esperando datos."""
    global _PRELOAD_DONE, _PRELOAD_ELAPSED
    print("[BOOT-PRELOAD] Iniciando carga sincrona de caches...")
    t0 = time.time()

    # 1. Datos principales (el más pesado)
    try:
        df, err = cargar_datos()
        if not err:
            print(f"[BOOT-PRELOAD][DATOS] OK ({len(df)} registros)")
        else:
            print(f"[BOOT-PRELOAD][DATOS] WARN: {err}")
    except Exception as e:
        print(f"[BOOT-PRELOAD][DATOS] ERR: {e}")

    # 2. Objetivos
    try:
        dfo, err_o = cargar_objetivos()
        if not err_o:
            print(f"[BOOT-PRELOAD][OBJ] OK ({len(dfo)} registros)")
        else:
            print(f"[BOOT-PRELOAD][OBJ] WARN: {err_o}")
    except Exception as e:
        print(f"[BOOT-PRELOAD][OBJ] ERR: {e}")

    # 3. Jefes y supervisores
    try:
        jefes, info, err_j = _cargar_tabla_maestra_jefes()
        if not err_j:
            print(f"[BOOT-PRELOAD][JEFES] OK ({len(jefes)} relaciones)")
        else:
            print(f"[BOOT-PRELOAD][JEFES] WARN: {err_j}")

        sups, err_s = _supervisores_automaticos_desde_maestra()
        if not err_s:
            print(f"[BOOT-PRELOAD][SUPS] OK ({len(sups)} supervisores)")
        else:
            print(f"[BOOT-PRELOAD][SUPS] WARN: {err_s}")
    except Exception as e:
        print(f"[BOOT-PRELOAD][JEFES/SUPS] ERR: {e}")

    _PRELOAD_ELAPSED = time.time() - t0
    _PRELOAD_DONE = True
    print(f"[BOOT-PRELOAD] Sistema listo y pre-calentado en {_PRELOAD_ELAPSED:.1f}s 🔥")


# ═══════════════════════════════════════════════════════════════════════════════
# INTEGRACION CON EL PIPELINE DE VACACIONES (boton "Actualizar" + KPIs en vivo)
# ═══════════════════════════════════════════════════════════════════════════════
import subprocess as _subprocess

_PIPELINE_MOTOR   = os.path.join(SCRIPT_DIR, 'PIPELINE', 'motor')
_PIPELINE_SCRIPT  = os.path.join(_PIPELINE_MOTOR, 'pipeline.py')
_PIPELINE_CONFIG  = os.path.join(_PIPELINE_MOTOR, 'config.json')
_BOT_ADRYAN       = os.path.join(SCRIPT_DIR, 'PIPELINE', 'bot_adryan', 'bot_adryan.py')
_BOT_ADRYAN_CFG   = os.path.join(SCRIPT_DIR, 'PIPELINE', 'bot_adryan', 'config_bot.json')

_pipeline_jobs       = {}
_pipeline_lock       = threading.Lock()
_pipeline_corriendo  = {'flag': False}

# Marcadores del log del pipeline -> (porcentaje, etiqueta para el front)
_PIPE_MARCAS = [
    ('Crudo:',        15, 'Leyendo y limpiando el crudo de Adryan'),
    ('Base volcada',  40, 'Volcando la base de vacaciones'),
    ('RefreshAll #1', 55, 'Refrescando la tabla dinamica'),
    ('Espejo',        65, 'Calculando dias registrados por persona'),
    ('BASE GENERAL',  80, 'Recalculando meta y obligatorio'),
    ('RefreshAll #2', 88, 'Actualizando el cumplimiento'),
    ('Guardado OK',   93, 'Guardando copia fechada'),
    ('Publicado en',  97, 'Publicando al dashboard'),
    ('PIPELINE OK',  100, 'Listo'),
]

def _pipeline_python():
    """Python que corre el pipeline (el que tiene xlwings). Lo toma del config."""
    try:
        with open(_PIPELINE_CONFIG, 'r', encoding='utf-8') as f:
            _cfg = json.load(f)
        py = (_cfg.get('integracion_front', {}) or {}).get('python_pipeline')
        if py and os.path.isfile(py):
            return py
    except Exception:
        pass
    return sys.executable

# ─── Cache en memoria por mtime para los endpoints lentos de vacaciones ──────
# El archivo vivo (Reporte ... Q2.xlsx, ~5.5 MB) tarda ~10-25 seg en parsear con
# openpyxl, asi que cualquier click en el front que pegue a este endpoint
# bloquea al usuario. Solucion: leer una vez por mtime y servir desde RAM. Cuando
# el pipeline publica una nueva version, el mtime cambia y el siguiente click
# regenera. Esto da: 1ra carga lenta (~10s), siguientes <10ms; al actualizar
# el archivo, los datos refrescan automaticamente. Es "tiempo real" para el
# usuario sin pagar el costo de Excel en cada request.
_AVANCE_CACHE = {'mtime': None, 'data': None}
_KPIS_CACHE   = {'mtime': None, 'data': None}
_AVANCE_LOCK  = threading.Lock()
_KPIS_LOCK    = threading.Lock()
_META_CACHE   = {'mtime': None, 'data': None}
_META_LOCK    = threading.Lock()

def _meta_vac_data():
    """Lee BASE GENERAL del archivo vivo y arma conteos de colaboradores + filas por
    segmento (mismo origen que el avance, para que TODO sea consistente). Cache por mtime.

    Columnas BASE GENERAL (0-based): 0 Matricula, 1 Nombre, 2 Tipo, 3 Puesto, 6 HRBP,
    7 Departamento, 8 Area, 12 Vencidas, 13 Pendiente, 15 Suma Dias Total, 16 Objetivo,
    19 Registradas."""
    cur_mt = _vac_mtime()
    with _META_LOCK:
        if _META_CACHE['mtime'] == cur_mt and _META_CACHE['data'] is not None:
            return _META_CACHE['data']
    try:
        import openpyxl
        ruta = VACACIONES_DATA_FILE
        if not os.path.isfile(ruta):
            return None
        tmpp = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmpp = tmp.name
            for _i in range(4):
                try:
                    shutil.copy2(ruta, tmpp); break
                except PermissionError:
                    time.sleep(0.6)
            else:
                return None
            wb = openpyxl.load_workbook(tmpp, data_only=True, read_only=True)
            rows = []
            base_vacs = {}
            if 'base' in wb.sheetnames:
                b_sheet = wb['base']
                idx_mat, idx_dias, idx_nombre = 0, 7, 1
                for r_idx, r in enumerate(b_sheet.iter_rows(min_row=1, values_only=True)):
                    if r_idx == 0:
                        for i, val in enumerate(r):
                            c_name = str(val).strip().lower() if val else ''
                            if c_name in ('matrícula', 'matricula'): idx_mat = i
                            elif c_name in ('días', 'dias'): idx_dias = i
                            elif c_name == 'nombre': idx_nombre = i
                        continue
                    if not r: continue
                    mat = r[idx_mat]
                    if mat is None or str(mat).strip() == '': continue
                    mat_str = str(int(float(mat))) if str(mat).replace('.','').isdigit() else str(mat).strip()
                    try: dias = float(r[idx_dias]) if r[idx_dias] is not None else 0.0
                    except: dias = 0.0
                    if dias <= 0: continue
                    nombre = str(r[idx_nombre]) if r[idx_nombre] else ''
                    if mat_str not in base_vacs: base_vacs[mat_str] = {'dias': 0, 'nombre': nombre}
                    base_vacs[mat_str]['dias'] += dias

            bg_mats = set()
            if 'BASE GENERAL' in wb.sheetnames:
                g = wb['BASE GENERAL']
                def _num(v):
                    try: return float(v) if v is not None else 0.0
                    except Exception: return 0.0
                for r in g.iter_rows(min_row=3, values_only=True):
                    if not r:
                        continue
                    mat = r[0] if len(r) > 0 else None
                    if mat is None or str(mat).strip() == '':
                        continue
                    mat_str = str(int(float(mat))) if str(mat).replace('.','').isdigit() else str(mat).strip()
                    bg_mats.add(mat_str)
                    obj = _num(r[16]) if len(r) > 16 else 0.0
                    reg  = _num(r[19]) if len(r) > 19 else 0.0
                    if obj <= 0:
                        if reg > 0:
                            rows.append({
                                'matricula': str(mat).strip(),
                                'nombre': r[1] if len(r) > 1 and r[1] else '',
                                'tipo': r[2] if len(r) > 2 and r[2] else '',
                                'puesto': r[3] if len(r) > 3 and r[3] else '',
                                'unidad_negocio': str(r[4]).strip() if len(r) > 4 and r[4] else '',
                                'hrbp': r[6] if len(r) > 6 and r[6] else '',
                                'departamento': r[7] if len(r) > 7 and r[7] else '',
                                'area': r[8] if len(r) > 8 and r[8] else '',
                                'objetivo': 0.0, 'dias_gozados': reg,
                                'total_dias': _num(r[15]) if len(r) > 15 else 0.0,
                                'vencidas': _num(r[12]) if len(r) > 12 else 0.0,
                                'pendientes': _num(r[13]) if len(r) > 13 else 0.0,
                                'meta_pct': 0.0,
                                'cumplio': False, 'oblig_pend': False,
                                'sin_iniciar': False, 'casi_listo': False,
                                'sin_meta_con_vac': True,
                            })
                        continue
                    venc = _num(r[12]) if len(r) > 12 else 0.0
                    pend = _num(r[13]) if len(r) > 13 else 0.0
                    cumplio = reg >= obj
                    oblig_pend = (not cumplio) and (venc > 0 or pend > 0)
                    meta_pct = (reg / obj) if obj > 0 else 0.0
                    sin_iniciar = (not cumplio) and reg == 0
                    casi_listo = (not cumplio) and 0.75 <= meta_pct < 1.0
                    rows.append({
                        'matricula': str(mat).strip(),
                        'nombre': r[1] if len(r) > 1 and r[1] else '',
                        'tipo': r[2] if len(r) > 2 and r[2] else '',
                        'puesto': r[3] if len(r) > 3 and r[3] else '',
                        'unidad_negocio': str(r[4]).strip() if len(r) > 4 and r[4] else '',
                        'hrbp': r[6] if len(r) > 6 and r[6] else '',
                        'departamento': r[7] if len(r) > 7 and r[7] else '',
                        'area': r[8] if len(r) > 8 and r[8] else '',
                        'objetivo': obj, 'dias_gozados': reg,
                        'total_dias': _num(r[15]) if len(r) > 15 else 0.0,
                        'vencidas': venc, 'pendientes': pend,
                        'meta_pct': meta_pct,
                        'cumplio': cumplio, 'oblig_pend': oblig_pend,
                        'sin_iniciar': sin_iniciar, 'casi_listo': casi_listo,
                        'sin_meta_con_vac': False,
                    })

            fantasmas = [m for m in base_vacs if m not in bg_mats]
            if fantasmas:
                try:
                    df_mae, _ = _cargar_maestro_universo()
                    mae_dict = {}
                    if df_mae is not None and not df_mae.empty:
                        def _cln(x): return str(int(float(x))) if str(x).replace('.','').isdigit() else str(x).strip()
                        df_mae['mat_clean'] = df_mae['matricula'].apply(_cln)
                        df_mae = df_mae.drop_duplicates(subset=['mat_clean'], keep='first')
                        mae_dict = df_mae.set_index('mat_clean').to_dict('index')
                    
                    for m in fantasmas:
                        info = mae_dict.get(m, {})
                        rows.append({
                            'matricula': str(m).zfill(10) if m.isdigit() else m,
                            'nombre': base_vacs[m]['nombre'] or info.get('nombre completo', ''),
                            'tipo': info.get('tipo de trabajador', ''),
                            'puesto': info.get('puesto', ''),
                            'unidad_negocio': info.get('sucursal', ''),
                            'hrbp': info.get('hrbp', ''),
                            'departamento': info.get('departamento', ''),
                            'area': info.get('area', ''),
                            'objetivo': 0.0, 'dias_gozados': base_vacs[m]['dias'],
                            'total_dias': 0.0,
                            'vencidas': 0.0, 'pendientes': 0.0,
                            'meta_pct': 0.0,
                            'cumplio': False, 'oblig_pend': False,
                            'sin_iniciar': False, 'casi_listo': False,
                            'sin_meta_con_vac': True,
                        })
                except Exception as e:
                    print('[META-VAC] Error cargando fantasmas desde maestro:', e)

            wb.close()

            # Cruzar con PersonalMaestroReporte para excluir cesados/fantasmas.
            # Mismo filtro que /api/objetivos para que todos los datos sean consistentes.
            try:
                df_mae, _ = _cargar_maestro_universo()
                if df_mae is not None and not df_mae.empty:
                    def _cln_mat(x): return str(int(float(x))) if str(x).replace('.','').isdigit() else str(x).strip()
                    c_mat_m = _col(df_mae, 'Matricula')
                    if c_mat_m:
                        mats_activas = set(df_mae[c_mat_m].dropna().apply(_cln_mat))
                        antes = len(rows)
                        rows = [r for r in rows if _cln_mat(r.get('matricula', '')) in mats_activas]
                        excl = antes - len(rows)
                        if excl > 0:
                            print(f'[META-VAC] {excl} registros excluidos por no estar en el maestro (cesados/fantasmas)')
            except Exception as e:
                print(f'[META-VAC] Advertencia al cruzar con maestro: {e}')

            con_meta_rows   = [x for x in rows if not x.get('sin_meta_con_vac')]
            cumplieron_rows = [x for x in rows if x.get('cumplio')]
            sin_ini_rows    = [x for x in rows if x.get('sin_iniciar')]
            sin_meta_rows   = [x for x in rows if x.get('sin_meta_con_vac')]
            parciales_rows  = [x for x in rows if not x.get('cumplio') and not x.get('sin_iniciar') and not x.get('sin_meta_con_vac')]
            counts = {
                # conteos existentes
                'con_meta':        len(con_meta_rows),
                'cumplieron':      len(cumplieron_rows),
                'sin_iniciar':     len(sin_ini_rows),
                'casi_listos':     sum(1 for x in rows if x.get('casi_listo')),
                'sin_meta_con_vac': len(sin_meta_rows),
                'parciales':       len(parciales_rows),
                # días totales de meta (universo con meta)
                'dias_meta_total':        round(sum(x.get('objetivo', 0) or 0 for x in con_meta_rows)),
                # días gozados por segmento
                'dias_gozados_con_meta':  round(sum(x.get('dias_gozados', 0) or 0 for x in con_meta_rows)),
                'dias_gozados_cumplieron': round(sum(x.get('dias_gozados', 0) or 0 for x in cumplieron_rows)),
                'dias_meta_cumplieron':   round(sum(x.get('objetivo', 0) or 0 for x in cumplieron_rows)),
                'dias_gozados_parciales': round(sum(x.get('dias_gozados', 0) or 0 for x in parciales_rows)),
                'dias_meta_parciales':    round(sum(x.get('objetivo', 0) or 0 for x in parciales_rows)),
                'dias_gozados_sin_meta':  round(sum(x.get('dias_gozados', 0) or 0 for x in sin_meta_rows)),
            }
            data = {'counts': counts, 'rows': rows}
            with _META_LOCK:
                _META_CACHE['mtime'] = cur_mt
                _META_CACHE['data'] = data
            return data
        finally:
            if tmpp:
                try: os.unlink(tmpp)
                except Exception: pass
    except Exception as e:
        print('[META-VAC] err:', e)
        return None

def _vac_mtime():
    try: return os.path.getmtime(VACACIONES_DATA_FILE) if os.path.isfile(VACACIONES_DATA_FILE) else None
    except OSError: return None


def _kpis_vacaciones():
    """KPIs de avance de meta. Prefiere el JSON local del pipeline (rapido, sin lock de
    OneDrive); si no existe, lee el Excel vivo con cache por mtime."""
    jpath = os.path.join(SCRIPT_DIR, 'PIPELINE', 'estado_pipeline.json')
    try:
        if os.path.isfile(jpath):
            with open(jpath, 'r', encoding='utf-8') as f:
                kp = json.load(f)
            out = {k: kp.get(k) for k in ('meta_total', 'registrado_total', 'avance', 'avance_cumplimiento')}
            if out.get('avance') is not None or out.get('registrado_total') is not None:
                return out
    except Exception as e:
        print('[PIPE-KPI] json err:', e)
    # Fallback: Excel vivo. Aqui cae el costo de ~10-25s la PRIMERA vez por mtime.
    cur_mt = _vac_mtime()
    with _KPIS_LOCK:
        if _KPIS_CACHE['mtime'] == cur_mt and _KPIS_CACHE['data'] is not None:
            return _KPIS_CACHE['data']
    try:
        import openpyxl
        ruta = VACACIONES_DATA_FILE
        if not os.path.isfile(ruta):
            return None
        tmpp = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmpp = tmp.name
            for _i in range(4):
                try:
                    shutil.copy2(ruta, tmpp); break
                except PermissionError:
                    time.sleep(0.6)
            else:
                return None
            # read_only=True acelera la carga (no carga estilos ni pivot caches)
            wb = openpyxl.load_workbook(tmpp, data_only=True, read_only=True)
            kp = {}
            if 'BASE GENERAL' in wb.sheetnames:
                g = wb['BASE GENERAL']
                # leer la fila 1 una sola vez en vez de 3 accesos a cell()
                row1 = next(g.iter_rows(min_row=1, max_row=1, values_only=True))
                kp['meta_total']       = row1[16]  # Q1
                kp['registrado_total'] = row1[19]  # T1
                kp['avance']           = row1[22]  # W1
            if 'R_Cumplimiento' in wb.sheetnames:
                rc = wb['R_Cumplimiento']
                # E9 = fila 9 col 5
                for i, r in enumerate(rc.iter_rows(min_row=9, max_row=9, values_only=True)):
                    kp['avance_cumplimiento'] = r[4] if len(r) > 4 else None
                    break
            wb.close()
            with _KPIS_LOCK:
                _KPIS_CACHE['mtime'] = cur_mt
                _KPIS_CACHE['data'] = kp
            return kp
        finally:
            if tmpp:
                try: os.unlink(tmpp)
                except Exception: pass
    except Exception as e:
        print('[PIPE-KPI] err:', e)
        return None

def _bot_adryan_descargar(job):
    """Corre el bot que descarga el crudo de Adryan. Devuelve True si OK."""
    if not os.path.isfile(_BOT_ADRYAN):
        job['lineas'].append('[BOT] No existe bot_adryan.py; se omite la descarga.')
        return True  # no bloquea: usa el ultimo crudo de Descargas
    job['paso'] = 'Descargando de Adryan'; job['pct'] = max(job['pct'], 5)
    try:
        cmd = [_pipeline_python(), _BOT_ADRYAN]
        if job.get('fecha_inicio'):
            cmd += ['--fecha-inicio', job['fecha_inicio']]
        if job.get('fecha_termino'):
            cmd += ['--fecha-termino', job['fecha_termino']]
        proc = _subprocess.Popen(
            cmd,
            cwd=os.path.dirname(_BOT_ADRYAN),
            stdout=_subprocess.PIPE, stderr=_subprocess.STDOUT,
            text=True, encoding='utf-8', errors='replace', bufsize=1,
        )
        for line in proc.stdout:
            line = (line or '').rstrip('\n')
            if line:
                job['lineas'].append('[BOT] ' + line)
        proc.wait()
        if proc.returncode != 0:
            lineas_str = ' '.join(job['lineas'])
            if 'playwright' in lineas_str.lower() or 'No module named' in lineas_str:
                job['estado'] = 'error'
                job['error'] = 'Esta funcion requiere playwright, disponible solo en la maquina administrativa. Sube el archivo VACRptMotivo manualmente.'
            else:
                job['estado'] = 'error'; job['error'] = 'Fallo la descarga de Adryan (revisa logs del bot).'
            return False
        job['paso'] = 'Crudo descargado'; job['pct'] = max(job['pct'], 12)
        return True
    except Exception as e:
        job['estado'] = 'error'; job['error'] = f'Error corriendo el bot: {e}'
        return False

def _pipeline_worker(job_id):
    job = _pipeline_jobs[job_id]
    if job.get('descargar_adryan'):
        if not _bot_adryan_descargar(job):
            job['fin'] = time.time(); _pipeline_corriendo['flag'] = False
            return
    cmd = [_pipeline_python(), _PIPELINE_SCRIPT, '--oculto']
    if job.get('crudo'):
        cmd += ['--crudo', job['crudo']]
    job['cmd'] = ' '.join(cmd)
    try:
        proc = _subprocess.Popen(
            cmd, cwd=_PIPELINE_MOTOR,
            stdout=_subprocess.PIPE, stderr=_subprocess.STDOUT,
            text=True, encoding='utf-8', errors='replace', bufsize=1,
        )
        for line in proc.stdout:
            line = (line or '').rstrip('\n')
            if not line:
                continue
            job['lineas'].append(line)
            for marca, pct, etiqueta in _PIPE_MARCAS:
                if marca in line:
                    if pct > job['pct']:
                        job['pct'] = pct
                    job['paso'] = etiqueta
            if 'RECONCIL' in line.upper():
                job['paso'] = 'Reconciliando nuevos ingresos'
            if len(job['lineas']) > 500:
                job['lineas'] = job['lineas'][-500:]
        proc.wait()
        job['rc'] = proc.returncode
        if proc.returncode == 0:
            job['estado'] = 'done'; job['pct'] = 100; job['paso'] = 'Listo'
            # el archivo recien publicado puede estar bloqueado un instante por OneDrive
            kp = None
            for _ in range(10):
                kp = _kpis_vacaciones()
                if kp:
                    break
                time.sleep(1.0)
            job['kpis_despues'] = kp
        else:
            job['estado'] = 'error'
    except Exception as e:
        job['estado'] = 'error'; job['error'] = str(e)
    finally:
        job['fin'] = time.time()
        _pipeline_corriendo['flag'] = False

@app.route('/api/vacaciones/pipeline/run', methods=['POST'])
def api_pipeline_run():
    """Dispara el pipeline. Metodo 1 (sin archivo): usa el ultimo VACRptMotivo de Descargas.
    Metodo 2 (multipart con 'archivo'): usa el .xlsx subido como crudo."""
    if not os.path.isfile(_PIPELINE_SCRIPT):
        return jsonify({'ok': False, 'error': f'No existe el pipeline: {_PIPELINE_SCRIPT}'}), 404
    archivo = request.files.get('archivo') or request.files.get('excel') or request.files.get('file')
    if archivo and archivo.filename and not archivo.filename.lower().endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'Solo se admite formato .xlsx'}), 400
    with _pipeline_lock:
        if _pipeline_corriendo['flag']:
            return jsonify({'ok': False, 'error': 'Ya hay una actualizacion en curso.'}), 409
        _pipeline_corriendo['flag'] = True
    crudo = None
    try:
        if archivo and archivo.filename:
            subdir = os.path.join(SCRIPT_DIR, 'PIPELINE', 'SUBIDOS')
            os.makedirs(subdir, exist_ok=True)
            crudo = os.path.join(subdir, os.path.basename(archivo.filename))
            archivo.save(crudo)
    except Exception as e:
        _pipeline_corriendo['flag'] = False
        return jsonify({'ok': False, 'error': f'No se pudo guardar el archivo: {e}'}), 500
    # Si no se subio archivo, se puede pedir que el bot descargue de Adryan primero.
    descargar = False
    if not (archivo and archivo.filename):
        val = (request.form.get('descargar') or request.values.get('descargar') or '').strip().lower()
        descargar = val in ('1', 'true', 'si', 'sí', 'on', 'yes')
    fecha_inicio = (request.form.get('fecha_inicio') or '').strip()
    fecha_termino = (request.form.get('fecha_termino') or '').strip()
    job_id = uuid.uuid4().hex[:12]
    _pipeline_jobs[job_id] = {
        'estado': 'running', 'lineas': [], 'pct': 3, 'paso': 'Iniciando...',
        'inicio': time.time(), 'fin': None, 'rc': None, 'crudo': crudo,
        'descargar_adryan': descargar,
        'fecha_inicio': fecha_inicio, 'fecha_termino': fecha_termino,
        'kpis_antes': _kpis_vacaciones(), 'kpis_despues': None,
    }
    threading.Thread(target=_pipeline_worker, args=(job_id,), daemon=True).start()
    modo = 'subido' if crudo else ('adryan' if descargar else 'descargas')
    return jsonify({'ok': True, 'job_id': job_id, 'modo': modo})

def _pipeline_maestro_worker(job_id):
    job = _pipeline_jobs[job_id]
    try:
        _BOT_MAESTRO = os.path.join(SCRIPT_DIR, 'PIPELINE', 'bot_adryan', 'bot_maestro.py')
        if not os.path.isfile(_BOT_MAESTRO):
            job['estado'] = 'error'; job['error'] = 'No existe bot_maestro.py'
            return
        cmd = [_pipeline_python(), _BOT_MAESTRO]
        proc = _subprocess.Popen(
            cmd, cwd=os.path.dirname(_BOT_MAESTRO),
            stdout=_subprocess.PIPE, stderr=_subprocess.STDOUT,
            text=True, encoding='utf-8', errors='replace', bufsize=1,
        )
        for line in proc.stdout:
            line = (line or '').rstrip('\n')
            if line:
                job['lineas'].append('[BOT] ' + line)
                if 'Descargando maestro' in line:
                    job['pct'] = 20; job['paso'] = 'Descargando maestro'
                elif 'Sanitizando' in line:
                    job['pct'] = 60; job['paso'] = 'Sanitizando base'
                elif 'ARCHIVO_MAESTRO=' in line:
                    job['pct'] = 90; job['paso'] = 'Copiando archivo'
                    src = line.split('=')[1].strip()
                    if os.path.isfile(src):
                        import shutil
                        dest = os.path.join(DATAS_DIR, os.path.basename(src))
                        shutil.copy2(src, dest)
                        job['lineas'].append(f'[SYS] Copiado a {dest}')
        proc.wait()
        job['rc'] = proc.returncode
        if proc.returncode == 0:
            _cargar_tabla_maestra_jefes()
            job['estado'] = 'done'; job['pct'] = 100; job['paso'] = 'Maestro actualizado'
        else:
            lineas_str = ' '.join(job['lineas'])
            if 'playwright' in lineas_str.lower() or 'No module named' in lineas_str:
                job['error'] = 'Esta funcion requiere playwright, disponible solo en la maquina administrativa.'
            else:
                job['error'] = 'Fallo la descarga del maestro.'
            job['estado'] = 'error'
    except Exception as e:
        job['estado'] = 'error'; job['error'] = str(e)
    finally:
        job['fin'] = time.time()
        _pipeline_corriendo['flag'] = False

@app.route('/api/vacaciones/pipeline/maestro', methods=['POST'])
def api_pipeline_maestro():
    with _pipeline_lock:
        if _pipeline_corriendo['flag']:
            return jsonify({'ok': False, 'error': 'Ya hay una actualizacion en curso.'}), 409
        _pipeline_corriendo['flag'] = True
    job_id = uuid.uuid4().hex[:12]
    _pipeline_jobs[job_id] = {
        'estado': 'running', 'lineas': [], 'pct': 5, 'paso': 'Iniciando descarga de maestro...',
        'inicio': time.time(), 'fin': None, 'rc': None, 'kpis_antes': None, 'kpis_despues': None,
    }
    threading.Thread(target=_pipeline_maestro_worker, args=(job_id,), daemon=True).start()
    return jsonify({'ok': True, 'job_id': job_id, 'modo': 'maestro'})

@app.route('/api/vacaciones/pipeline/reset', methods=['POST'])
def api_pipeline_reset():
    """Fuerza el desbloqueo del flag de pipeline cuando un job quedo colgado."""
    with _pipeline_lock:
        estaba = _pipeline_corriendo['flag']
        _pipeline_corriendo['flag'] = False
    return jsonify({'ok': True, 'estaba_bloqueado': estaba, 'msg': 'Flag de pipeline reseteado.'})

@app.route('/api/vacaciones/pipeline/estado/<job_id>', methods=['GET'])
def api_pipeline_estado(job_id):
    """Progreso en vivo + KPIs antes/despues."""
    job = _pipeline_jobs.get(job_id)
    if not job:
        return jsonify({'ok': False, 'error': 'job no encontrado'}), 404
    return jsonify({
        'ok': True,
        'estado': job['estado'],
        'pct': job['pct'],
        'paso': job.get('paso'),
        'lineas': job['lineas'][-60:],
        'rc': job.get('rc'),
        'error': job.get('error'),
        'kpis_antes': job.get('kpis_antes'),
        'kpis_despues': job.get('kpis_despues'),
        'duracion': round((job.get('fin') or time.time()) - job['inicio'], 1),
    })

@app.route('/api/vacaciones/kpis', methods=['GET'])
def api_vacaciones_kpis():
    """KPIs del archivo vivo: avance de meta (dias) + conteos de colaboradores
    (con meta / ya cumplieron / obligatorios pendientes), del mismo BASE GENERAL."""
    kp = _kpis_vacaciones() or {}
    md = _meta_vac_data()
    if md and md.get('counts'):
        kp = {**kp, **md['counts']}
    return jsonify({'ok': True, 'kpis': kp})

@app.route('/api/vacaciones/meta_detalle', methods=['GET'])
def api_vacaciones_meta_detalle():
    """Detalle de colaboradores por segmento de meta (mismo origen que los KPIs).
    seg = con_meta | cumplieron | sin_iniciar | casi_listos"""
    seg = (request.args.get('seg', 'con_meta') or 'con_meta').strip()
    md = _meta_vac_data()
    if not md:
        return jsonify({'ok': False, 'error': 'Sin datos de meta', 'registros': []})
    rows = md['rows']
    if seg == 'cumplieron':
        rows = [r for r in rows if r.get('cumplio')]
    elif seg == 'sin_iniciar':
        rows = [r for r in rows if r.get('sin_iniciar')]
    elif seg == 'parciales':
        rows = [r for r in rows if not r.get('cumplio') and not r.get('sin_iniciar') and not r.get('sin_meta_con_vac')]
    elif seg == 'casi_listos':
        rows = [r for r in rows if r.get('casi_listo')]
    elif seg == 'ya_iniciaron':
        rows = [r for r in rows if not r.get('sin_iniciar') and not r.get('sin_meta_con_vac')]
    elif seg == 'con_meta':
        rows = [r for r in rows if not r.get('sin_meta_con_vac')]
    elif seg == 'sin_meta_con_vac':
        rows = [r for r in rows if r.get('sin_meta_con_vac')]
    rows = sorted(rows, key=lambda r: (-(r['total_dias'] or 0), str(r['nombre'] or '')))
    def _estado(r):
        if r['cumplio']: return 'Cumplió meta'
        if r['sin_iniciar']: return 'Sin iniciar'
        if r['casi_listo']: return f"Casi listo ({r['meta_pct']*100:.0f}%)"
        return f"En proceso ({r['meta_pct']*100:.0f}%)"
    out = [{
        'matricula': r['matricula'], 'nombre': r['nombre'],
        'unidad_negocio': r.get('unidad_negocio', ''),
        'gerencia': r['departamento'], 'departamento': r['departamento'],
        'area': r['area'], 'puesto': r['puesto'], 'hrbp': r['hrbp'],
        'dias_gozados': r['dias_gozados'], 'total_dias': r['total_dias'],
        'objetivo': r['objetivo'], 'meta_pct': round(r['meta_pct'] * 100, 1),
        'estado': _estado(r),
    } for r in rows]
    return jsonify({'ok': True, 'total': len(out), 'registros': out})

@app.route('/api/vacaciones/exportar_areas', methods=['GET'])
def api_exportar_areas():
    """Genera y descarga un Excel con el detalle de cumplimiento por área/gerencia.
    Parámetros opcionales: gerencia=X, area=Y, buscar=texto"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl no disponible'}), 500

    md = _meta_vac_data()
    if not md:
        return jsonify({'ok': False, 'error': 'Sin datos'}), 500

    f_udn    = (request.args.get('unidad_negocio', '') or '').strip().lower()
    f_ger    = (request.args.get('gerencia', '') or '').strip().lower()
    f_area   = (request.args.get('area',     '') or '').strip().lower()
    f_buscar = (request.args.get('buscar',   '') or '').strip().lower()

    def _norm_s(s): return (s or '').strip().lower()

    rows = md['rows']
    # Aplicar filtros
    if f_udn:
        rows = [r for r in rows if f_udn in _norm_s(r.get('unidad_negocio', ''))]
    if f_ger:
        rows = [r for r in rows if f_ger in _norm_s(r.get('departamento', ''))]
    if f_area:
        if f_area in ('(sin área asignada)', '(sin area asignada)'):
            rows = [r for r in rows if not (r.get('area') or '').strip()]
        else:
            rows = [r for r in rows if f_area in _norm_s(r.get('area', ''))]
    if f_buscar:
        rows = [r for r in rows if f_buscar in _norm_s(r.get('nombre', '')) or
                f_buscar in _norm_s(r.get('area', '')) or
                f_buscar in _norm_s(r.get('departamento', ''))]

    rows = sorted(rows, key=lambda r: (
        _norm_s(r.get('unidad_negocio', '')),
        _norm_s(r.get('departamento', '')),
        _norm_s(r.get('area', '') or 'sin área'),
        _norm_s(r.get('nombre', ''))
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cumplimiento por Área'

    # Estilos
    hdr_fill   = PatternFill('solid', fgColor='0F6EA5')
    hdr_font   = Font(bold=True, color='FFFFFF', size=10)
    ok_fill    = PatternFill('solid', fgColor='D4F0E0')
    warn_fill  = PatternFill('solid', fgColor='FFF0D4')
    err_fill   = PatternFill('solid', fgColor='FCE8E8')
    bold_font  = Font(bold=True, size=10)
    normal_font= Font(size=10)
    center_al  = Alignment(horizontal='center', vertical='center')
    left_al    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_side  = Side(style='thin', color='CCCCCC')
    thin_bdr   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = ['Matrícula','Nombre','Unidad de Negocio','Gerencia / Depto.','Área','Puesto','HRBP',
               'Meta (días)','Gozados','Saldo','% Avance','Estado']
    col_w   = [12, 32, 28, 28, 26, 28, 18, 11, 10, 10, 10, 14]

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center_al; cell.border = thin_bdr
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    for ri, r in enumerate(rows, 2):
        obj  = round(r.get('objetivo', 0) or 0, 1)
        goz  = round(r.get('dias_gozados', 0) or 0, 1)
        saldo= round(max(obj - goz, 0), 1)
        pct  = round((r.get('meta_pct', 0) or 0) * 100, 1)
        if r.get('cumplio'):      estado = 'Cumplió'
        elif r.get('sin_iniciar'): estado = 'Sin iniciar'
        elif r.get('sin_meta_con_vac'): estado = 'Sin meta (gozó)'
        else:                     estado = 'En proceso'

        vals = [r.get('matricula',''), r.get('nombre',''), r.get('unidad_negocio', ''), r.get('departamento',''),
                r.get('area','') or 'Sin área', r.get('puesto',''), r.get('hrbp',''),
                obj, goz, saldo, pct, estado]
        fill = ok_fill if r.get('cumplio') else (err_fill if r.get('sin_iniciar') else warn_fill)

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin_bdr; cell.font = normal_font
            cell.alignment = center_al if ci in (1,8,9,10,11) else left_al
            if ci == 12: cell.fill = fill

    # Hoja resumen por área
    ws2 = wb.create_sheet('Resumen por Área')
    h2 = ['Unidad de Negocio','Gerencia','Área','Personas','Meta (días)','Gozados','Saldo','Avance %','Cumplieron','En proceso','Sin iniciar']
    for ci, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center_al; cell.border = thin_bdr
        ws2.column_dimensions[get_column_letter(ci)].width = [28,26,24,10,12,10,10,10,12,12,12][ci-1]

    area_map = {}
    for r in rows:
        key = (r.get('unidad_negocio','') or '', r.get('departamento','') or '', r.get('area','') or 'Sin área')
        if key not in area_map:
            area_map[key] = {'n':0,'meta':0,'goz':0,'cumpl':0,'proceso':0,'sini':0}
        d = area_map[key]
        d['n']  += 1
        d['meta'] += r.get('objetivo',0) or 0
        d['goz']  += r.get('dias_gozados',0) or 0
        if r.get('cumplio'): d['cumpl'] += 1
        elif r.get('sin_iniciar'): d['sini'] += 1
        elif not r.get('sin_meta_con_vac'): d['proceso'] += 1

    for ri2, (key, d) in enumerate(sorted(area_map.items()), 2):
        avp = round(d['goz']/d['meta']*100, 1) if d['meta'] > 0 else 0
        row_vals = [key[0], key[1], key[2], d['n'], round(d['meta'],1), round(d['goz'],1),
                    round(max(d['meta']-d['goz'],0),1), avp, d['cumpl'], d['proceso'], d['sini']]
        for ci, v in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri2, column=ci, value=v)
            cell.border = thin_bdr; cell.font = normal_font
            cell.alignment = center_al if ci > 3 else left_al

    ws2.freeze_panes = 'A2'

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fecha_str = __import__('datetime').date.today().strftime('%Y%m%d')
    nombre_archivo = f'CumplimientoAreas_{fecha_str}.xlsx'
    if f_area:
        area_tag = 'sin_area_asignada' if f_area in ('(sin área asignada)', '(sin area asignada)') else f_area[:30].replace(' ', '_')
        nombre_archivo = f'Area_{area_tag}_{fecha_str}.xlsx'
    elif f_ger:
        nombre_archivo = f'Gerencia_{f_ger[:30].replace(" ","_")}_{fecha_str}.xlsx'
    elif f_udn:
        nombre_archivo = f'UdN_{f_udn[:30].replace(" ","_")}_{fecha_str}.xlsx'

    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=nombre_archivo)


@app.route('/api/vacaciones/config', methods=['GET'])
def api_vacaciones_config():
    """Config para el front (p.ej. URL de Adryan del metodo 1)."""
    url = ''
    try:
        with open(_PIPELINE_CONFIG, 'r', encoding='utf-8') as f:
            c = json.load(f)
        url = (c.get('integracion_front', {}) or {}).get('url_adryan', '') or ''
    except Exception:
        pass
    fecha_inicio = ''
    fecha_termino = ''
    try:
        with open(_BOT_ADRYAN_CFG, 'r', encoding='utf-8') as f:
            bc = json.load(f)
        fi = bc.get('fecha_inicio', '')
        ft = bc.get('fecha_termino', '')
        if fi:
            p = fi.split('/')
            if len(p) == 3:
                fecha_inicio = f'{p[2]}-{p[1]}-{p[0]}'
        if ft:
            p = ft.split('/')
            if len(p) == 3:
                fecha_termino = f'{p[2]}-{p[1]}-{p[0]}'
    except Exception:
        pass
    return jsonify({'ok': True, 'url_adryan': url,
                    'fecha_inicio': fecha_inicio, 'fecha_termino': fecha_termino})

# Lista canonica de Business Partners (HRBP) que deben mostrarse en el panel.
# Se construye desde la columna HRBP de BASE GENERAL (no del pivote R_Cumplimiento,
# que omite a Cesar Reyes y Gabriel Chang). Orden = (tokens a buscar, nombre a mostrar).
_BP_CANON = [
    (('CARLOS', 'JARA'),    'Carlos Jara'),
    (('CESAR', 'REYES'),    'César Reyes'),
    (('GABRIEL', 'CHANG'),  'Gabriel Chang'),
    (('FATIMA', 'SALAZAR'), 'Fatima Salazar'),
    (('LESL', 'REYES'),     'Lesley Reyes'),
    (('MELISSA', 'HIGA'),   'Melissa Higa'),
]


def _canon_bp(valor):
    """Normaliza el texto de HRBP a uno de los nombres canonicos. '' si no coincide."""
    n = _norm(valor).upper()
    if not n:
        return ''
    for toks, nombre in _BP_CANON:
        if all(t in n for t in toks):
            return nombre
    return ''


def _fmt_fecha(v):
    """Formatea una fecha de Excel (datetime o texto) a DD/MM/YYYY."""
    if v is None:
        return ''
    try:
        return v.strftime('%d/%m/%Y')
    except Exception:
        s = str(v).strip()
        return s.split(' ')[0] if s else ''


def _compute_avance(ruta):
    """Lee el Excel vivo y arma el payload de /api/vacaciones/avance.

    El avance por BP se calcula directamente desde la hoja 'BASE GENERAL'
    (columna HRBP) para incluir a TODOS los Business Partners — incluidos
    Cesar Reyes y Gabriel Chang, que el pivote 'R_Cumplimiento' del Excel
    no contempla. Ademas arma el detalle por persona (con sus registros de
    vacaciones de la hoja 'base') para la vista de drill-down. Lento (~10-25s)."""
    import openpyxl
    tmpp = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmpp = tmp.name
        for _i in range(4):
            try:
                shutil.copy2(ruta, tmpp); break
            except PermissionError:
                time.sleep(0.6)
        else:
            return None, 'Archivo bloqueado'
        # read_only=True corta a la mitad el tiempo de open en este libro
        wb = openpyxl.load_workbook(tmpp, data_only=True, read_only=True)

        def _n(x):
            try: return float(x)
            except Exception: return None

        def _num(x):
            try: return float(x)
            except Exception: return 0.0

        # --- Detalle de registros reales por matricula (hoja 'base') ---------
        # Cada fila = un periodo de vacaciones efectivamente tomado.
        registros_por_mat = {}
        if 'base' in wb.sheetnames:
            b = wb['base']
            for r in b.iter_rows(min_row=2, values_only=True):
                mat = r[0]
                if mat is None or str(mat).strip() == '':
                    continue
                mat = str(mat).strip()
                dias = _num(r[7]) if len(r) > 7 else 0.0
                registros_por_mat.setdefault(mat, []).append({
                    'inicio':    _fmt_fecha(r[5]) if len(r) > 5 else '',
                    'termino':   _fmt_fecha(r[6]) if len(r) > 6 else '',
                    'dias':      dias,
                    'motivo':    str(r[8]).strip() if len(r) > 8 and r[8] is not None else '',
                    'periodo':   str(r[4]).strip() if len(r) > 4 and r[4] is not None else '',
                    'situacion': str(r[11]).strip() if len(r) > 11 and r[11] is not None else '',
                })

        glob = {}
        # Acumuladores por BP canonico
        acc = {nombre: {'hrbp': nombre, 'n': 0, 'meta': 0.0, 'registro': 0.0,
                        'personas': []} for _, nombre in _BP_CANON}

        # Pre-cargar maestro para filtrar fantasmas/cesados
        mats_activas = None
        try:
            df_mae, _ = _cargar_maestro_universo()
            if df_mae is not None and not df_mae.empty:
                def _cln_mat(x): return str(int(float(x))) if str(x).replace('.','').isdigit() else str(x).strip()
                c_mat_m = _col(df_mae, 'Matricula')
                if c_mat_m:
                    mats_activas = set(df_mae[c_mat_m].dropna().apply(_cln_mat))
        except Exception as e:
            print('[COMPUTE-AVANCE] Error cargando maestro:', e)

        if 'BASE GENERAL' in wb.sheetnames:
            g = wb['BASE GENERAL']
            # Fila 1 = totales globales (toda la base, incluido colegio sin meta)
            row1 = next(g.iter_rows(min_row=1, max_row=1, values_only=True))
            glob['meta_total']        = _n(row1[16])   # Q1
            glob['registrado_total']  = _n(row1[19])   # T1
            glob['avance_todo']       = _n(row1[22])   # W1 (T1/Q1)
            # Indices de columna en BASE GENERAL (cabecera fila 2, datos fila 3+)
            # 0 Matricula · 1 Apellidos y Nombres · 3 Nombre Puesto · 6 HRBP
            # 7 Nombre Departamento · 8 Nombre Area · 16 Objetivo · 19 Registradas
            for row in g.iter_rows(min_row=3, values_only=True):
                bp = _canon_bp(row[6] if len(row) > 6 else '')
                if not bp:
                    continue
                mat = str(row[0]).strip() if row[0] is not None else ''
                if not mat: continue
                
                # Excluir si no está en maestro (cesado)
                if mats_activas is not None:
                    mat_cln = str(int(float(mat))) if str(mat).replace('.','').isdigit() else mat
                    if mat_cln not in mats_activas:
                        continue

                meta_p = _num(row[16]) if len(row) > 16 else 0.0
                reg_p  = _num(row[19]) if len(row) > 19 else 0.0
                a = acc[bp]
                a['n'] += 1
                a['meta'] += meta_p
                a['registro'] += reg_p
                a['personas'].append({
                    'matricula':    mat,
                    'nombre':       str(row[1]).strip() if len(row) > 1 and row[1] is not None else '',
                    'puesto':       str(row[3]).strip() if len(row) > 3 and row[3] is not None else '',
                    'unidad_negocio': str(row[4]).strip() if len(row) > 4 and row[4] is not None else '',
                    'departamento': str(row[7]).strip() if len(row) > 7 and row[7] is not None else '',
                    'gerencia':     str(row[7]).strip() if len(row) > 7 and row[7] is not None else '',
                    'area':         str(row[8]).strip() if len(row) > 8 and row[8] is not None else '',
                    'meta':         meta_p,
                    'gozado':       reg_p,
                    'avance':       (reg_p / meta_p) if meta_p else None,
                    'registros':    registros_por_mat.get(mat, []),
                })
        wb.close()

        # Resumen por BP + detalle. Solo BPs con al menos una persona.
        por_bp, detalle = [], {}
        tot_meta = tot_reg = 0.0
        for _, nombre in _BP_CANON:
            a = acc[nombre]
            if a['n'] == 0:
                continue
            av = (a['registro'] / a['meta']) if a['meta'] else None
            por_bp.append({'hrbp': nombre, 'n': a['n'], 'meta': a['meta'],
                           'registro': a['registro'], 'avance': av})
            # Personas ordenadas por avance ascendente (los mas atrasados primero)
            a['personas'].sort(key=lambda p: (p['avance'] if p['avance'] is not None else 9, -p['meta']))
            detalle[nombre] = {'hrbp': nombre, 'n': a['n'], 'meta': a['meta'],
                               'registro': a['registro'], 'avance': av,
                               'personas': a['personas']}
            tot_meta += a['meta']; tot_reg += a['registro']

        glob['bp_meta_total']     = tot_meta
        glob['bp_registro_total'] = tot_reg
        glob['avance_bp']         = (tot_reg / tot_meta) if tot_meta else None

        por_bp.sort(key=lambda x: (x['avance'] if x['avance'] is not None else 9))
        return {'ok': True, 'global': glob, 'por_bp': por_bp, '_detalle': detalle}, None
    except Exception as e:
        return None, str(e)
    finally:
        if tmpp:
            try: os.unlink(tmpp)
            except Exception: pass


@app.route('/api/vacaciones/avance', methods=['GET'])
def api_vacaciones_avance():
    """Avance de la META de vacaciones: global (toda la base y sin colegio) + por BP (HRBP).

    Sirve desde cache en memoria invalidando por mtime del archivo vivo. Primera
    carga: ~10s (parsear el xlsx). Siguientes: <10ms hasta que el pipeline publique
    una version nueva (cambia el mtime y el siguiente click recomputa). Permite
    cliquear las graficas sin esperar."""
    ruta = VACACIONES_DATA_FILE
    if not os.path.isfile(ruta):
        return jsonify({'ok': False, 'error': 'Archivo de vacaciones no encontrado'}), 404
    cur_mt = _vac_mtime()
    # Cache hit: instantaneo
    with _AVANCE_LOCK:
        if _AVANCE_CACHE['mtime'] == cur_mt and _AVANCE_CACHE['data'] is not None:
            resp = {k: v for k, v in _AVANCE_CACHE['data'].items() if k != '_detalle'}
            resp['_cache'] = 'hit'
            return jsonify(resp)
    # Miss: leer Excel (caro). Tomamos el lock SOLO al final para no serializar lecturas.
    data, err = _compute_avance(ruta)
    if data is None:
        return jsonify({'ok': False, 'error': err or 'error desconocido'}), 503 if err == 'Archivo bloqueado' else 500
    with _AVANCE_LOCK:
        _AVANCE_CACHE['mtime'] = cur_mt
        _AVANCE_CACHE['data']  = data
    # El detalle por persona (pesado) no viaja en el resumen; se sirve aparte.
    resp = {k: v for k, v in data.items() if k != '_detalle'}
    resp['_cache'] = 'miss'
    return jsonify(resp)


@app.route('/api/vacaciones/bp_detalle', methods=['GET'])
def api_vacaciones_bp_detalle():
    """Detalle por Business Partner: lista de colaboradores asignados a ese HRBP,
    con su meta, dias realmente gozados y el detalle de cada registro de vacaciones.

    Reutiliza el mismo cache que /avance (un solo parseo del Excel por mtime)."""
    bp = (request.args.get('bp') or '').strip()
    if not bp:
        return jsonify({'ok': False, 'error': 'Falta el parametro bp'}), 400
    ruta = VACACIONES_DATA_FILE
    if not os.path.isfile(ruta):
        return jsonify({'ok': False, 'error': 'Archivo de vacaciones no encontrado'}), 404
    cur_mt = _vac_mtime()
    data = None
    with _AVANCE_LOCK:
        if _AVANCE_CACHE['mtime'] == cur_mt and _AVANCE_CACHE['data'] is not None:
            data = _AVANCE_CACHE['data']
    if data is None:
        data, err = _compute_avance(ruta)
        if data is None:
            return jsonify({'ok': False, 'error': err or 'error desconocido'}), 503 if err == 'Archivo bloqueado' else 500
        with _AVANCE_LOCK:
            _AVANCE_CACHE['mtime'] = cur_mt
            _AVANCE_CACHE['data']  = data
    detalle = (data.get('_detalle') or {})
    # Match por nombre canonico (tolerante a tildes/mayusculas)
    item = detalle.get(bp)
    if item is None:
        objetivo = _canon_bp(bp)
        item = detalle.get(objetivo) if objetivo else None
    if item is None:
        return jsonify({'ok': False, 'error': 'BP no encontrado: ' + bp}), 404
    return jsonify({'ok': True, 'detalle': item})


def _warmup_avance_cache():
    """Llama una vez al boot para que el primer click del usuario sea instantaneo."""
    try:
        if not os.path.isfile(VACACIONES_DATA_FILE):
            return
        cur_mt = _vac_mtime()
        with _AVANCE_LOCK:
            if _AVANCE_CACHE['mtime'] == cur_mt and _AVANCE_CACHE['data'] is not None:
                return
        t0 = time.perf_counter()
        data, err = _compute_avance(VACACIONES_DATA_FILE)
        if data is not None:
            with _AVANCE_LOCK:
                _AVANCE_CACHE['mtime'] = cur_mt
                _AVANCE_CACHE['data']  = data
            print(f'[BOOT] avance precalculado en {time.perf_counter()-t0:.1f}s '
                  f'(BPs={len(data.get("por_bp", []))})')
        else:
            print(f'[BOOT] avance no precalculado: {err}')
    except Exception as e:
        print('[BOOT] warmup avance err:', e)
# ═══════════════════════════════════════════════════════════════════════════════


if __name__ == '__main__':

    # === PRECARGA EN SEGUNDO PLANO ===
    # El servidor HTTP arranca de inmediato; la precarga corre en paralelo.
    # /api/status/ready devuelve {"ready": false} hasta que termine y el front
    # muestra un spinner ligero en lugar de quedar bloqueado 20-40 seg.
    def _recuperar_errores_al_inicio():
        """Mueve archivos de errores/ → in/ al arrancar para que el loop los reintente."""
        err_dir = os.path.join(COLA_DIR, 'errores')
        if not os.path.isdir(err_dir) or not os.path.isdir(COLA_IN_DIR):
            return
        count = 0
        for nombre in os.listdir(err_dir):
            if not nombre.lower().endswith('.json'):
                continue
            src = os.path.join(err_dir, nombre)
            dst = os.path.join(COLA_IN_DIR, nombre)
            if os.path.isfile(src) and not os.path.exists(dst):
                try:
                    shutil.move(src, dst)
                    count += 1
                except Exception:
                    pass
        if count:
            print(f'[BOOT] {count} archivo(s) recuperados de errores/ → in/ para reintento via Outlook COM')

    def _boot_async():
        print('[BOOT] Precargando datos en segundo plano...')
        _recuperar_errores_al_inicio()
        _pre_cargar_todo_sincrono()
        _warmup_avance_cache()
    threading.Thread(target=_boot_async, daemon=True).start()
    threading.Thread(target=_cola_pa_loop, daemon=True).start()

    try:
        from waitress import serve
        print('[OK] Servidor iniciado con waitress en http://127.0.0.1:5002')
        serve(app, host='0.0.0.0', port=5002, threads=8)
    except ImportError:
        print('[WARN] waitress no disponible, usando Flask dev server')
        app.run(host='0.0.0.0', port=5002, debug=False, threaded=True)
